"""El supermercado deja de validar el índice y pasa a integrarlo (ADR-0225).

Era el ancla de validación externa del cinturón. Mide condiciones materiales del
hogar, así que le corresponde ser componente y no juez: es la misma regla que
había desplazado al ICC.

Estos tests cuidan cinco cosas que pueden volver a romperse:

1. **No puede volver al panel.** Un indicador que compone el índice y a la vez
   lo valida vuelve circular la validación, y el camino de vuelta es de una
   línea en `panel_validacion.FAMILIA`.
2. **El colector revienta si falta la base.** Sin los tres meses del 4T-2023 el
   rebase mide contra otra cosa, y un índice rebaseado contra vaya a saber qué
   no se distingue a ojo de uno sano.
3. **Card y serie son el mismo número**, para que no haya dos caminos que
   puedan divergir.
4. **No lleva promedio móvil de 12 meses.** La fuente ya publica la serie
   desestacionalizada y volver a suavizarla la atrasaría medio año — el error
   que ADR-0155 documentó midiéndolo.
5. **El cinturón sigue sin ancla única**, que es la otra mitad de la decisión:
   si alguien vuelve a poner un par suelto de titular, el gráfico y la matriz
   cruzada dejan de decir lo que el texto afirma.
"""
import datetime as dt
import importlib.util
import io
import json
import sys
from pathlib import Path

import pytest

ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT / "scripts"))
import itvc                      # noqa: E402
import panel_validacion          # noqa: E402
import validacion_externa        # noqa: E402

# El colector se carga POR RUTA, igual que el de autos: `scripts/vida_cotidiana/`
# tiene su propio `config.py` y ponerlo al frente del path tapa al del proyecto.
_spec = importlib.util.spec_from_file_location(
    "indec_supermercados",
    ROOT / "scripts" / "vida_cotidiana" / "collectors" / "indec_supermercados.py")
_col = importlib.util.module_from_spec(_spec)
_spec.loader.exec_module(_col)

SNAPSHOT = json.loads(
    (ROOT / "web" / "src" / "data" / "informe.json").read_text(encoding="utf-8"))
SERIES = json.loads(
    (ROOT / "web" / "src" / "data" / "series.json").read_text(encoding="utf-8"))
VIDA = SNAPSHOT["cinturones"]["vida_cotidiana"]
IND = VIDA["indicadores"]["consumo_supermercados"]
CLAVE = "consumo_supermercados"


TITULO = ("Encuesta de Supermercados. Índice de ventas totales a precios constantes. "
          "Serie original, desestacionalizada y tendencia-ciclo, base año 2017=100, "
          "en números índice y variación porcentual. Enero 2017 – {ultimo}")

_MES_ES = ("enero", "febrero", "marzo", "abril", "mayo", "junio", "julio", "agosto",
           "septiembre", "octubre", "noviembre", "diciembre")


def _libro(filas, titulo=None, columnas=("Serie original",
                                         "Serie desestacionalizada (1)",
                                         "Serie tendencia-ciclo")) -> bytes:
    """Un «Cuadro 1» sintético, con la forma real de la planilla del INDEC.

    Se construye un xlsx de verdad en vez de mockear `openpyxl`: el colector
    lee encabezados combinados y filas con `datetime` en la primera columna, y
    un mock de esa estructura probaría el mock, no el parser.

    `filas` es [(mes 'YYYY-MM', valor desestacionalizado)]. Las otras dos
    series se rellenan con valores distintos a propósito, para que leer la
    columna equivocada dé un número distinto y los tests puedan verlo.
    """
    import openpyxl
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Cuadro 1"
    if titulo is None:
        ultimo = max(m for m, _ in filas) if filas else "2017-01"
        a, m = ultimo.split("-")
        titulo = TITULO.format(ultimo=f"{_MES_ES[int(m) - 1]} {a}")
    ws.cell(1, 1).value = titulo
    ws.cell(3, 1).value = "Período"
    for i, nombre in enumerate(columnas):
        ws.cell(3, 2 + i * 4).value = nombre
        ws.cell(6, 2 + i * 4).value = "Números índice"
    for f, (mes, valor) in enumerate(sorted(filas), start=7):
        ws.cell(f, 1).value = dt.datetime(int(mes[:4]), int(mes[5:7]), 1)
        for i in range(len(columnas)):
            # original = ×1,15 · desestacionalizada = valor · tendencia = ×1,02
            factor = (1.15, 1.0, 1.02)[i] if i < 3 else 1.0
            ws.cell(f, 2 + i * 4).value = valor * factor * (1.03 if i == 0 and f % 2 else 1)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


BASE_OK = [("2023-10", 100.0), ("2023-11", 102.0), ("2023-12", 98.0)]

# La función REAL, guardada antes de que la autouse de abajo la tape. Sin esto,
# el test que comprueba que el espejo se traga sus fallos se prueba a sí mismo:
# llama al lambda del fixture, que devuelve None pase lo que pase. Verificado
# rompiéndolo — la mutación «el espejo propaga la excepción» pasaba entera.
_ESPEJO_REAL = _col._serie_del_espejo


@pytest.fixture(autouse=True)
def _sin_red(monkeypatch):
    """Ningún test de este archivo sale a la red. El espejo devuelve `None`
    salvo que el test lo pise: así el contraste queda fuera del camino de los
    tests que no lo están mirando."""
    monkeypatch.setattr(_col, "_serie_del_espejo", lambda: None)


def _con_libro(monkeypatch, filas, **kw):
    monkeypatch.setattr(_col, "_descargar_xlsx", lambda: _libro(filas, **kw))


# ── 1 · Componente sí, ancla no ─────────────────────────────────────────────
def test_es_componente_del_indice():
    comp = {i for d in itvc.DIMENSIONES_ITVC.values() for i in d["indicadores"]}
    assert CLAVE in comp, "dejó de ser componente del ITCIS"
    assert CLAVE in itvc.DIMENSIONES_ITVC["ingresos"]["indicadores"]


def test_no_puede_estar_en_el_panel_de_validacion():
    """El camino de vuelta es de una línea, y volvería circular la validación:
    el índice se estaría contrastando contra una parte de sí mismo."""
    assert CLAVE not in panel_validacion.FAMILIA
    assert CLAVE not in panel_validacion.ETIQUETAS
    assert CLAVE not in panel_validacion.FUENTES
    for claves in panel_validacion.FACTOR.values():
        assert CLAVE not in claves


def test_no_quedo_plumbing_de_ancla_en_validacion_externa():
    """Dos caminos de descarga para la misma serie es como el índice y su
    contraste terminan leyendo bases distintas sin que nada avise."""
    fuente = (ROOT / "scripts" / "validacion_externa.py").read_text(encoding="utf-8")
    assert "fetch_consumo_supermercados_mensual" not in fuente
    assert "consumo_supermercados_mensual" not in fuente
    assert "ITVC vs consumo" not in fuente


def test_puntua_y_no_es_card_de_contexto():
    """ADR-0216/0153: o integra el índice, o no es card."""
    assert IND.get("en_indice") is True
    assert IND.get("peso_efectivo", 0) > 0
    assert IND.get("dimension") == "ingresos"


# ── 2 · El colector revienta en vez de mentir ───────────────────────────────
def test_sin_los_meses_de_la_base_el_colector_levanta(monkeypatch):
    _con_libro(monkeypatch, [("2024-01", 100.0), ("2024-02", 101.0)])
    with pytest.raises(ValueError, match="base 4T-2023"):
        _col.fetch_consumo_supermercados()


def test_una_planilla_vacia_levanta(monkeypatch):
    _con_libro(monkeypatch, [])
    with pytest.raises(ValueError, match="ningún punto"):
        _col.fetch_consumo_supermercados()


def test_con_la_base_completa_devuelve_el_ultimo_punto(monkeypatch):
    _con_libro(monkeypatch, BASE_OK + [("2024-01", 91.5)])
    d = _col.fetch_consumo_supermercados()
    assert d["consumo_supermercados"] == {
        "valor": 91.5, "fecha": "2024-01", "anio_base": 2017,
        "unidad": "índice (2017 = 100, desestacionalizado)"}
    assert len(d["serie"]) == 4


def test_los_meses_de_la_base_son_los_del_cinturon():
    assert tuple(_col.BASE_MESES) == tuple(validacion_externa.BASE_MESES)


# ── 3 · Card y serie, un solo número ────────────────────────────────────────
def test_la_card_es_el_ultimo_punto_de_la_serie():
    serie = [p for p in SERIES.get(CLAVE, []) if p.get("valor") is not None]
    assert serie, "la serie no llegó al snapshot"
    ult = serie[-1]
    assert round(float(ult["valor"]), 1) == round(float(IND["valor"]), 1), (
        f"card {IND['valor']} contra último punto {ult['valor']}")
    assert ult["fecha"][:7] == IND["fecha_dato"][:7]


def test_la_serie_llega_a_la_base():
    meses = {p["fecha"][:7] for p in SERIES.get(CLAVE, [])}
    faltan = [m for m in validacion_externa.BASE_MESES if m not in meses]
    assert not faltan, f"sin base no hay rebase posible: faltan {faltan}"


# ── 4 · Sin promedio móvil, que la fuente ya desestacionalizó ───────────────
def test_no_entra_por_movil_12m():
    """Motos y autos sí; éste no. La ventana móvil existe para sacarle el
    calendario a un flujo crudo, y acá el INDEC ya lo sacó — volver a suavizar
    atrasa la serie medio año, que es el error medido en ADR-0155."""
    assert CLAVE not in validacion_externa.MOVIL12
    itvc_py = (ROOT / "scripts" / "itvc.py").read_text(encoding="utf-8")
    linea = next(l for l in itvc_py.splitlines() if f'idx["{CLAVE}"]' in l)
    assert "rebase_de_serie" in linea and "rebase_movil12" not in linea


def test_la_reconstruccion_lee_la_misma_serie_que_la_card():
    skey, invertido, anual, ya_rebaseada = validacion_externa.COMPONENTES[CLAVE]
    assert skey == CLAVE
    assert invertido is False, "más volumen comprado es mejor, no se invierte"
    assert anual is False and ya_rebaseada is False


# ── 5 · Los pesos: la regla, no los decimales ───────────────────────────────
def test_entra_con_20_por_ciento_y_los_previos_cedieron():
    ind = itvc.DIMENSIONES_ITVC["ingresos"]["indicadores"]
    assert ind[CLAVE] == 0.20
    assert abs(sum(ind.values()) - 1.0) < 1e-9
    # los CUATRO previos, contra lo que tenían antes de esta alta. Eran cinco
    # cuando el alta se escribió; ADR-0224 fundió los dos vehículos en
    # `motorizacion_total` mientras la rama estaba abierta, y la cesión se
    # recalculó sola sobre lo que quedó — que es para lo que existe
    # `alta_proporcional`.
    previos = {"brecha_salario_cbt": 0.5959, "pobreza_nowcast": 0.3253,
               "consumo_carnes_total": 0.0392, "motorizacion_total": 0.0396}
    for k, v in previos.items():
        assert abs(ind[k] - round(v * 0.80, 4)) < 1e-9, f"{k} no cedió ×0,80"


def test_el_peso_nominal_de_la_dimension_no_se_toco():
    assert itvc.DIMENSIONES_ITVC["ingresos"]["peso"] == 0.2806


# ── 6 · El cinturón sigue sin ancla única ───────────────────────────────────
def test_la_validacion_no_titula_contra_una_serie_suelta():
    """La otra mitad de la decisión. Si alguien vuelve a poner un par suelto de
    titular, el gráfico dibuja el factor y el encabezado informa otra cosa."""
    val = VIDA["itvc"]["validacion"]
    assert "factor" in val["externa_label"].lower(), val["externa_label"]
    assert val["plot"] == "minmax", "el factor es un puntaje estandarizado, cruza el cero"
    for prohibido in ("supermercado", "consumo en supermercados"):
        assert prohibido not in val["externa_label"].lower()


def test_la_conclusion_explica_por_que_no_hay_ancla_unica():
    """El motivo es parte del resultado: si el texto sólo dijera «se usa un
    panel», el lector no tendría cómo saber que hay una referencia identificada
    y esperando muestra."""
    txt = VIDA["itvc"]["validacion"]["conclusion"].lower()
    assert "cuentas nacionales" in txt
    assert "trimestr" in txt


def test_la_matriz_cruzada_no_se_contrasta_contra_un_componente():
    """`validacion.pares` alimenta la matriz de ADR-0031. Dejarla apuntando al
    supermercado la habría vuelto circular sin que ningún test avisara."""
    fila = next(f for f in SNAPSHOT["validacion_cruzada"]["filas"]
                if f["indice"] == "ITCIS")
    assert fila["propio"] == "volumen_hogar"
    externas = dict(SNAPSHOT["validacion_cruzada"]["externas"])
    assert "supermercado" not in externas["volumen_hogar"].lower()


# ── 6 · La base y las revisiones (ADR-0243) ─────────────────────────────────
# La card rotulaba «índice (2004 = 100)». La Encuesta de Supermercados vigente
# usa base 2017=100 y la serie ni siquiera tiene puntos antes de enero de 2017:
# el rótulo no describía nada. Sobrevivió porque era un literal repetido en
# tres archivos y no había contra qué compararlo.

def test_la_base_sale_de_la_fuente_y_no_de_un_literal(monkeypatch):
    _con_libro(monkeypatch, BASE_OK + [("2024-01", 91.5)])
    d = _col.fetch_consumo_supermercados()["consumo_supermercados"]
    assert d["anio_base"] == 2017
    assert "2017 = 100" in d["unidad"]


def test_la_base_erronea_no_puede_volver(monkeypatch):
    """2004 no es una base plausible para una serie que empieza en 2017."""
    _con_libro(monkeypatch, BASE_OK + [("2024-01", 91.5)])
    d = _col.fetch_consumo_supermercados()["consumo_supermercados"]
    assert "2004" not in d["unidad"]


def test_si_la_fuente_cambia_de_base_el_rotulo_la_sigue(monkeypatch):
    """El punto de leerla: que no haya que acordarse de actualizar el rótulo."""
    _con_libro(monkeypatch, BASE_OK + [("2024-01", 91.5)],
               titulo="Encuesta de Supermercados. Índice de ventas totales a "
                      "precios constantes, base año 2021=100. Enero 2017 – enero 2024")
    d = _col.fetch_consumo_supermercados()["consumo_supermercados"]
    assert d["anio_base"] == 2021
    assert "2021 = 100" in d["unidad"]


def test_sin_base_declarada_el_colector_levanta(monkeypatch):
    """Antes que inventar un rótulo, fallar. El anterior no vino de ningún lado
    verificable y estuvo publicado igual."""
    _con_libro(monkeypatch, BASE_OK,
               titulo="Encuesta de Supermercados. Números índice. "
                      "Enero 2017 – diciembre 2023")
    with pytest.raises(ValueError, match="base"):
        _col.fetch_consumo_supermercados()


def test_una_serie_anterior_a_su_propia_base_levanta(monkeypatch):
    """Si la serie tuviera puntos de antes del año base, la base declarada no
    es la de esos puntos: hay dos bases empalmadas y el nivel no es comparable."""
    _con_libro(monkeypatch, [("2010-01", 70.0)] + BASE_OK)
    with pytest.raises(ValueError, match="base declarada"):
        _col.fetch_consumo_supermercados()


def test_una_revision_del_indec_reemplaza_el_valor_anterior(monkeypatch):
    """El INDEC revisó mayo-2026 de 83,2 a 83,0 al publicar junio.

    El colector lee la planilla entera en cada corrida, así que la revisión
    entra sola. Lo que este test cuida es que nadie meta un cache o un
    `setdefault` que conserve la versión vieja: la serie desestacionalizada se
    recalcula cada mes, y guardar la primera lectura la congelaría."""
    _con_libro(monkeypatch, BASE_OK + [("2026-05", 83.222)])
    primera = _col.fetch_consumo_supermercados()
    assert primera["consumo_supermercados"]["valor"] == 83.2

    _con_libro(monkeypatch, BASE_OK + [("2026-05", 83.0), ("2026-06", 82.1)])
    segunda = _col.fetch_consumo_supermercados()
    assert segunda["serie"]["2026-05"] == 83.0, "quedó congelada la lectura vieja"
    assert segunda["consumo_supermercados"] == {
        "valor": 82.1, "fecha": "2026-06", "anio_base": 2017,
        "unidad": "índice (2017 = 100, desestacionalizado)"}


def test_el_ultimo_periodo_es_el_ultimo_publicado_no_el_penultimo(monkeypatch):
    """Con junio disponible, la card tiene que ser junio."""
    _con_libro(monkeypatch, BASE_OK + [("2026-05", 83.0), ("2026-06", 82.1)])
    d = _col.fetch_consumo_supermercados()["consumo_supermercados"]
    assert d["fecha"] == "2026-06" and d["valor"] == 82.1


# ── 7 · La fuente es el INDEC, no su espejo (ADR-0256) ──────────────────────
# La card mostró mayo mientras el INDEC ya había publicado junio, porque la
# serie venía de la API de datos.gob.ar, que espeja esta misma planilla con dos
# semanas de atraso. Cinco guardas para que no vuelva a pasar y para que el
# camino nuevo no traiga sus propias formas de romperse en silencio.

def test_el_valor_no_sale_del_espejo(monkeypatch):
    """Si alguien volviera a poner la API como fuente, el espejo mandaría.

    El test lo fuerza: la planilla trae junio y el espejo se quedó en mayo con
    otro número. La card tiene que ser la de la planilla."""
    _con_libro(monkeypatch, BASE_OK + [("2026-05", 83.0), ("2026-06", 82.1)])
    monkeypatch.setattr(_col, "_serie_del_espejo",
                        lambda: {"2023-10": 100.0, "2026-05": 83.2})
    d = _col.fetch_consumo_supermercados()
    assert d["consumo_supermercados"]["fecha"] == "2026-06"
    assert "2026-06" in d["serie"]


def test_el_espejo_caido_no_impide_publicar(monkeypatch):
    """Es un contraste, no una dependencia: que datos.gob.ar esté caído no es
    motivo para no publicar lo que el INDEC ya publicó. Son dos mitades — que
    `_serie_del_espejo` se trague el fallo, y que sin contraste igual haya card."""
    monkeypatch.setattr(_col, "requests", _RequestsQueExplota())
    assert _ESPEJO_REAL() is None

    _con_libro(monkeypatch, BASE_OK + [("2026-06", 82.1)])
    d = _col.fetch_consumo_supermercados()
    assert d["consumo_supermercados"]["fecha"] == "2026-06"


def test_la_cascara_html_del_indec_no_pasa_por_xlsx(monkeypatch):
    """`indec.gob.ar` contesta 200 con 37 KB de HTML para cualquier ruta que no
    exista, así que `raise_for_status()` no distingue el archivo de la nada.
    Si esto no estuviera, el fallo aparecería como un error de openpyxl en
    medio del pipeline en vez de decir que movieron el archivo."""
    monkeypatch.setattr(_col, "requests", _RequestsConHTML())
    with pytest.raises(_col.FuenteINDECError, match="no es un xlsx"):
        _col._descargar_xlsx()


def test_la_columna_se_busca_por_encabezado_y_no_por_posicion(monkeypatch):
    """Una columna agregada río arriba correría la desestacionalizada de lugar,
    y un índice fijo publicaría la de al lado: seguiría siendo un índice de
    ventas plausible."""
    filas = BASE_OK + [("2026-06", 82.1)]
    _con_libro(monkeypatch, filas,
               columnas=("Serie original", "Serie corregida por calendario",
                         "Serie desestacionalizada (1)", "Serie tendencia-ciclo"))
    d = _col.fetch_consumo_supermercados()
    # con cuatro columnas, la desestacionalizada quedó tercera y el factor de
    # `_libro` para la tercera es 1,02 — o sea, se leyó ESA y no la segunda.
    assert d["consumo_supermercados"]["valor"] == round(82.1 * 1.02, 1)


def test_sin_la_columna_declarada_el_colector_levanta(monkeypatch):
    _con_libro(monkeypatch, BASE_OK + [("2026-06", 82.1)],
               columnas=("Serie original", "Serie tendencia-ciclo"))
    with pytest.raises(ValueError, match="Serie desestacionalizada"):
        _col.fetch_consumo_supermercados()


def test_leer_la_columna_de_al_lado_lo_agarra_el_contraste(monkeypatch):
    """La guarda de arriba cubre que la columna DESAPAREZCA. Ésta cubre que se
    lea otra que sí existe y trae números plausibles — el caso que ningún
    chequeo de forma puede ver. Se calibró con la superposición real: la
    columna correcta diverge del espejo 0,041 pp de mediana y «serie original»
    4,92, veinte veces el tope."""
    import math
    meses = [f"2024-{m:02d}" for m in range(1, 13)] + \
            [f"2025-{m:02d}" for m in range(1, 13)]
    # una serie con estacionalidad marcada: original y desestacionalizada se
    # separan en la VARIACIÓN, que es lo que el contraste mira.
    desest = {m: 100.0 + i * 0.2 for i, m in enumerate(meses)}
    original = {m: v * (1 + 0.08 * math.sin(i)) for i, (m, v) in enumerate(desest.items())}
    serie = dict(BASE_OK) | desest
    monkeypatch.setattr(_col, "_serie_del_espejo", lambda: dict(BASE_OK) | original)
    with pytest.raises(ValueError, match="es una columna distinta"):
        _col._contrastar(serie, _col._serie_del_espejo())


def test_el_contraste_tolera_una_revision_pero_no_una_serie_distinta():
    """Lo que el contraste mide cuando todo está bien es la revisión del ajuste
    estacional, que mueve la serie unas centésimas."""
    meses = [f"{a}-{m:02d}" for a in (2024, 2025, 2026) for m in range(1, 13)]
    a = {m: 100.0 + i * 0.2 for i, m in enumerate(meses)}
    b = {m: v * (1 + 0.0005 * (-1) ** i) for i, (m, v) in enumerate(a.items())}
    assert _col._contrastar(a, b) < _col.CONTRASTE_MAX_PP


def test_un_rebase_de_la_fuente_no_dispara_el_contraste():
    """Compara variaciones, no niveles: si el INDEC cambia la base, los niveles
    se corren enteros y el contraste no tiene por qué saltar — el rebase a
    4T-2023 absorbe el cambio solo."""
    meses = [f"{a}-{m:02d}" for a in (2024, 2025, 2026) for m in range(1, 13)]
    a = {m: 100.0 + i * 0.2 for i, m in enumerate(meses)}
    b = {m: v * 1.37 for m, v in a.items()}          # misma serie, otra base
    assert _col._contrastar(a, b) == 0


def test_una_lectura_truncada_choca_contra_la_cobertura_declarada(monkeypatch):
    """El título declara hasta qué mes llega la planilla. Es una segunda
    afirmación de la fuente, independiente de las filas: si el parser se comiera
    las últimas, el número seguiría siendo plausible y nadie lo vería."""
    _con_libro(monkeypatch, BASE_OK + [("2026-05", 83.0)],
               titulo=TITULO.format(ultimo="junio 2026"))
    with pytest.raises(ValueError, match="cobertura hasta 2026-06"):
        _col.fetch_consumo_supermercados()


def test_el_tope_de_rezago_cubre_el_ciclo_medido_y_agarra_un_mes_salteado():
    """Medido sobre las 14 publicaciones del calendario del INDEC entre
    julio-2025 y agosto-2026: el último punto llega como mucho a 116 días la
    víspera de la publicación siguiente, y un mes salteado lo llevaría a ~146.
    El tope tiene que quedar entre esos dos números."""
    import config
    tope = config.MAX_DIAS[CLAVE]
    assert 116 < tope < 146, f"tope {tope} fuera del ciclo medido"


class _RequestsConHTML:
    """El sitio del INDEC contestando 200 con su cáscara para una ruta muerta."""
    class _R:
        content = b"<!DOCTYPE html>\n<html><head><title>INDEC</title></head>" + b" " * 200
        def raise_for_status(self): pass
    RequestException = Exception
    def get(self, *a, **k): return self._R()


class _RequestsQueExplota:
    RequestException = Exception
    def get(self, *a, **k): raise RuntimeError("timeout")
