"""Contratos de las dos dimensiones que dejaron de depender de una sola señal."""
import io
import sys
from datetime import date, datetime
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT))
sys.path.insert(0, str(ROOT / "scripts"))

import config
import descargar_series
import gate_calidad
import itcp
import itvc
import politica
import publicar


class _Respuesta:
    def __init__(self, *, content=b"", text=""):
        self.content = content
        self.text = text

    def raise_for_status(self):
        return None


def _xlsx_deuda() -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "carga"
    ws["A2"] = "Estimación de la carga mensual de los servicios de deuda de las familias"
    ws.append([])
    ws.append([])
    ws.append([])
    ws.append(["Período / Period", "CDF / MS", "CDF / MSA", "CDF / PIB"])
    ws.append([datetime(2023, 10, 1), 10.0, 7.0, 3.0])
    ws.append([datetime(2026, 4, 1), 24.0758, 16.8, 4.6])
    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()


def _xlsx_jornadas() -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "1"
    ws.append(["C1. Conflictos con paro, huelguistas y jornadas de paro"])
    ws.append(["Mes", None, None, None, None, None, None, "Jornadas de paro"])
    ws.append([None, None, None, None, None, None, None, "Total"])
    for mes in range(1, 13):
        ws.append([datetime(2025, mes, 1), None, None, None, None, None, None, mes * 100])
    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()


def _xlsx_jornadas_con_columnas_reordenadas() -> bytes:
    """El mismo cuadro C1 con los tres grupos abiertos por sector, que es como
    la Secretaría de Trabajo puede publicarlo: el total de jornadas deja de
    caer en la columna 7."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "1"
    ws.append(["C1. Conflictos con paro, huelguistas y jornadas de paro"])
    ws.append(["Mes", "Conflictos con paro", None, None,
               "Huelguistas", None, None, None, "Jornadas de paro", None, None])
    ws.append([None, "Total", "Privado", "Estatal",
               "Total", "Privado", "Estatal", "Sin clasificar",
               "Total", "Privado", "Estatal"])
    for mes in range(1, 13):
        ws.append([datetime(2025, mes, 1), 5, 3, 2, 900, 500, 400, 0,
                   mes * 100, mes * 60, mes * 40])
    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()


def test_parser_carga_deuda_busca_rotulos_no_numero_de_hoja(monkeypatch):
    monkeypatch.setattr(descargar_series.requests, "get",
                        lambda *args, **kwargs: _Respuesta(content=_xlsx_deuda()))
    serie = descargar_series.fetch_carga_servicio_deuda_serie()
    assert serie == [["2023-10-01", 10.0], ["2026-04-01", 24.076]]


def test_carga_deuda_lee_la_planilla_de_la_edicion_vigente(monkeypatch):
    """La planilla del IEF cambia de nombre cada semestre: si el colector la
    tuviera clavada, la edición nueva daría 404 y la serie quedaría congelada
    sin ruido. Se descubre desde la página de la edición más reciente."""
    edicion_vigente = descargar_series._bcra_ief_ediciones()[0]
    pedidas = []

    def fake_get(url, **_kwargs):
        pedidas.append(url)
        if url == edicion_vigente:
            return _Respuesta(
                text='<a href="/archivos/informes/IEF-2027-01-serie.xlsx">series</a>')
        if url.endswith(".xlsx"):
            return _Respuesta(content=_xlsx_deuda())
        raise descargar_series.requests.ConnectionError(f"sin edición: {url}")

    monkeypatch.setattr(descargar_series.requests, "get", fake_get)
    assert descargar_series.fetch_carga_servicio_deuda_serie()[-1] == ["2026-04-01", 24.076]
    assert pedidas[-1] == "https://www.bcra.gob.ar/archivos/informes/IEF-2027-01-serie.xlsx"
    assert pedidas[-1] != descargar_series.BCRA_IEF_CARGA_DEUDA


def test_carga_deuda_usa_planilla_directa_si_ninguna_edicion_responde(monkeypatch):
    pedidas = []

    def fake_get(url, **_kwargs):
        pedidas.append(url)
        if url == descargar_series.BCRA_IEF_CARGA_DEUDA:
            return _Respuesta(content=_xlsx_deuda())
        raise descargar_series.requests.ConnectionError("BCRA caído")

    monkeypatch.setattr(descargar_series.requests, "get", fake_get)
    assert descargar_series.fetch_carga_servicio_deuda_serie()[-1] == ["2026-04-01", 24.076]
    assert pedidas[-1] == descargar_series.BCRA_IEF_CARGA_DEUDA


def test_ediciones_del_ief_van_de_la_mas_nueva_a_la_mas_vieja():
    """El IEF es semestral y su página usa un slug regular, así que la edición
    a probar se deduce de la fecha en vez de fijarse a mano."""
    from datetime import date as _date
    assert descargar_series._bcra_ief_ediciones(_date(2026, 8, 21)) == [
        "https://www.bcra.gob.ar/publicaciones/"
        "informe-de-estabilidad-financiera-segundo-semestre-2026/",
        "https://www.bcra.gob.ar/publicaciones/"
        "informe-de-estabilidad-financiera-primer-semestre-2026/",
        "https://www.bcra.gob.ar/publicaciones/"
        "informe-de-estabilidad-financiera-segundo-semestre-2025/",
    ]
    assert descargar_series._bcra_ief_ediciones(_date(2026, 3, 1))[0].endswith(
        "informe-de-estabilidad-financiera-primer-semestre-2026/")


def test_parser_jornadas_suma_exactamente_doce_meses(monkeypatch):
    contenido = _xlsx_jornadas()
    pedidas = []

    def fake_get(url, **_kwargs):
        pedidas.append(url)
        if url == politica.CONFLICTOS_LABORALES_URL:
            return _Respuesta(
                text='<a href="/sites/default/files/'
                     'evolucion_mensual_de_la_conflictividad_laboral._datos_a_junio_2026.xlsx">'
                     'mensual</a>')
        return _Respuesta(content=contenido)

    monkeypatch.setattr(politica.requests, "get", fake_get)
    serie = politica.fetch_jornadas_individuales_no_trabajadas_serie()
    assert serie == [["2025-12-01", 7_800]]
    # La planilla que se bajó es la que anuncia la página, no la clavada.
    assert pedidas[-1] == (
        "https://www.argentina.gob.ar/sites/default/files/"
        "evolucion_mensual_de_la_conflictividad_laboral._datos_a_junio_2026.xlsx")
    assert pedidas[-1] != politica.CONFLICTOS_LABORALES_XLSX_FALLBACK


def test_jornadas_usa_planilla_directa_si_falla_la_pagina_indice(monkeypatch):
    contenido = _xlsx_jornadas()

    def fake_get(url, **_kwargs):
        if url == politica.CONFLICTOS_LABORALES_URL:
            raise politica.requests.ConnectionError("landing caída")
        assert url == politica.CONFLICTOS_LABORALES_XLSX_FALLBACK
        return _Respuesta(content=contenido)

    monkeypatch.setattr(politica.requests, "get", fake_get)
    assert politica.fetch_jornadas_individuales_no_trabajadas_serie()[-1] == [
        "2025-12-01", 7_800]


def test_jornadas_ubica_su_columna_por_rotulo_y_no_por_posicion(monkeypatch):
    """Si el cuadro abre una columna más antes del grupo de jornadas, tomar la
    posición fija devuelve otra magnitud —huelguistas sin clasificar— con un
    valor plausible que ni la frescura ni la coherencia card/serie detectan."""
    monkeypatch.setattr(
        politica.requests, "get",
        lambda *_a, **_k: _Respuesta(content=_xlsx_jornadas_con_columnas_reordenadas()))
    assert politica.fetch_jornadas_individuales_no_trabajadas_serie() == [
        ["2025-12-01", 7_800]]


def test_jornadas_falla_ruidoso_si_el_cuadro_pierde_la_columna(monkeypatch):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["C1. Conflictos con paro y huelguistas"])
    ws.append(["Mes", "Conflictos con paro", "Huelguistas"])
    ws.append([None, "Total", "Total"])
    for mes in range(1, 13):
        ws.append([datetime(2025, mes, 1), 5, 900])
    out = io.BytesIO()
    wb.save(out)
    monkeypatch.setattr(politica.requests, "get",
                        lambda *_a, **_k: _Respuesta(content=out.getvalue()))
    try:
        politica.fetch_jornadas_individuales_no_trabajadas_serie()
    except ValueError as e:
        assert "jornadas de paro" in str(e)
    else:
        raise AssertionError("el parser aceptó un cuadro sin columna de jornadas")
    # y el colector lo traduce en ausencia de card, no en un número inventado
    assert politica.fetch_jornadas_individuales_no_trabajadas() is None


def test_una_fuente_laboral_demorada_con_fetch_exitoso_no_es_cache(monkeypatch):
    monkeypatch.setattr(politica.requests, "get",
                        lambda *_a, **_k: _Respuesta(content=_xlsx_jornadas()))
    monkeypatch.setitem(config.MAX_DIAS,
                        "jornadas_individuales_no_trabajadas_12m", 1)
    card = politica.fetch_jornadas_individuales_no_trabajadas()
    fecha = datetime.strptime(card["fecha_dato"], "%Y-%m-%d").date()
    assert (date.today() - fecha).days > 1
    assert card["desactualizado"] is False
    assert gate_calidad.MAX_DIAS["jornadas_individuales_no_trabajadas_12m"] == 1


def test_vulnerabilidad_combina_incumplimiento_y_carga():
    indices = {"mora_familias": 20.0, "carga_servicio_deuda_hogares": 50.0}
    dim = itvc.calcular_itvc(indices)["dimensiones"]["vulnerabilidad"]
    assert dim["puntaje"] == 29.0
    assert dim["indicadores"]["mora_familias"]["peso"] == 0.70
    assert dim["indicadores"]["carga_servicio_deuda_hogares"]["peso"] == 0.30


def test_conflicto_social_combina_calle_e_intensidad_laboral():
    valores = {
        "conflictividad_nacional": -40.0,  # 100 puntos
        "jornadas_individuales_no_trabajadas_12m": 11_000_000,  # 10 puntos
    }
    dim = itcp.calcular_itcp(valores)["dimensiones"]["conflicto_social"]
    assert dim["puntaje"] == 64.0
    assert dim["indicadores"]["conflictividad_nacional"]["peso"] == 0.60
    assert dim["indicadores"]["jornadas_individuales_no_trabajadas_12m"]["peso"] == 0.40


def test_el_tope_de_frescura_de_la_carga_tiene_un_solo_dueno(monkeypatch):
    """El rezago del dato se configura para el gate sin redefinir su tope."""
    monkeypatch.setitem(config.MAX_DIAS, "carga_servicio_deuda_hogares", 42)
    assert config.rezago_maximo_tolerado("carga_servicio_deuda_hogares") == 42
    assert gate_calidad.MAX_DIAS.get(
        "carga_servicio_deuda_hogares", gate_calidad.MAX_DIAS_DEFAULT) == 42


def test_una_serie_de_deuda_demorada_no_se_publica_como_cache():
    enriquecido = {}
    publicar.agregar_carga_servicio_deuda(enriquecido, {
        "carga_servicio_deuda_hogares": [
            {"fecha": "2020-01-01", "valor": 12.5},
        ],
    })
    card = enriquecido["carga_servicio_deuda_hogares"]
    assert card["fecha_dato"] == "2020-01"
    assert card["desactualizado"] is False
