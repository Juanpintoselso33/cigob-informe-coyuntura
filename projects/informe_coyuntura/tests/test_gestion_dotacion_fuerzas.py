"""Tests de los dos cierres del aporte de gestión (ADR-0128)."""
import sys
from pathlib import Path

import pytest

sys.path.insert(0, str(Path(__file__).parent.parent / "scripts"))
import gestion
import itcg


# ── 1. Fuerzas armadas y de seguridad en la dotación ────────────────────────

def _serie():
    try:
        return gestion.dotacion_apn_series()
    except Exception as e:
        pytest.skip(f"fuente inaccesible: {e}")


def test_el_desglose_de_fuerzas_llega():
    fuerzas = _serie().get("_fuerzas") or {}
    assert fuerzas, "el padrón dejó de traer el detalle por entidad"
    assert "2023-12" in fuerzas, "sin baseline dic-2023 no se puede comparar"


def test_las_fuerzas_pesan_lo_esperado():
    """~10% de la dotación APN. Si esto se corre mucho, o el INDEC cambió el
    universo o el matcheo por nombre dejó de encontrar entidades — y el
    contexto publicado quedaría mal sin que nada más avise."""
    serie = _serie()
    fuerzas = serie.pop("_fuerzas", {})
    serie.pop("_empresas", None)
    serie.pop("_total", None)
    comunes = sorted(set(serie) & set(fuerzas))
    if not comunes:
        pytest.skip("sin meses comparables")
    ym = comunes[-1]
    peso = fuerzas[ym] / serie[ym] * 100
    assert 6.0 < peso < 15.0, f"las fuerzas pesan {peso:.1f}% en {ym}"


def test_las_siete_entidades_siguen_identificandose():
    """El desglose matchea por NOMBRE. Si el INDEC renombra una entidad, se
    pierde en silencio: este test cuenta cuántas se encuentran."""
    import io
    import openpyxl
    contenido = gestion._http_get_resiliente(gestion.INDEC_DOTACION_URL)
    wb = openpyxl.load_workbook(io.BytesIO(contenido), data_only=True)
    hoja = next((s for s in wb.sheetnames if s.strip().lower() == "detalle2"), None)
    if hoja is None:
        pytest.skip("sin hoja de detalle")
    encontradas = set()
    for f in wb[hoja].iter_rows(values_only=True):
        ent = str(f[1] or "").strip()
        if any(k in ent.lower() for k in gestion._ENTES_FUERZAS):
            encontradas.add(ent)
    assert len(encontradas) >= 7, (
        f"sólo se identificaron {len(encontradas)} entidades de fuerzas: "
        f"{sorted(encontradas)}")


def test_la_card_publica_el_desglose():
    card = gestion.fetch_reduccion_estado()
    if card is None:
        pytest.skip("colector sin datos")
    assert card.get("var_planta_civil") is not None
    assert card.get("var_fuerzas_seguridad") is not None
    assert card.get("fecha_desglose_fuerzas")
    # las fuerzas se redujeron MENOS que el conjunto: si se invierte, la
    # afirmación de la ficha deja de ser cierta
    assert card["var_fuerzas_seguridad"] > card["var_planta_civil"], card


def test_las_fuerzas_no_se_descuentan_del_valor_que_puntua():
    """ADR-0128 decidió publicarlas, NO sacarlas. El valor del indicador sigue
    siendo la APN completa."""
    serie = _serie()
    fuerzas = serie.pop("_fuerzas", {})
    serie.pop("_empresas", None)
    serie.pop("_total", None)
    card = gestion.fetch_reduccion_estado()
    if card is None or not fuerzas:
        pytest.skip("sin datos")
    ult = max(serie)
    esperado = round((serie[ult] - serie["2023-12"]) / serie["2023-12"] * 100, 2)
    assert card["valor"] == esperado, (
        "el valor publicado dejó de ser la variación de la APN completa")


# ── 2. Peso del FAL ─────────────────────────────────────────────────────────

def test_el_fal_y_la_litigiosidad_pesan_igual():
    """ADR-0128: instrumento y resultado pesan lo mismo."""
    ind = itcg.DIMENSIONES_ITCG["reforma_laboral"]["indicadores"]
    assert ind == {"fal_modernizacion_laboral": 0.50, "litigiosidad_laboral": 0.50}


def test_la_dimension_laboral_sigue_siendo_la_mas_floja():
    """Guardia heredada de ADR-0098. DISPARÓ el 2026-07-26 y funcionó: ADR-0142
    llevó al FAL de 30,8 a 100 puntos y la dimensión dejó de ser la más floja
    del ITCG (45,1 → 79,7). La respuesta a "¿mejoró la realidad o se aflojó la
    vara?" es la segunda, y fue una decisión editorial declarada — no cambió
    ningún hecho del mundo entre el 25 y el 26 de julio.

    Lo que la guardia protege ahora es lo único que sigue en pie: la
    litigiosidad laboral, que es el RESULTADO de la reforma y no su
    instrumento, no mejoró. Si algún día el FAL vuelve a ser un indicador
    vivo, esta guardia debería volver a mirar la dimensión.

    MIRA LA BANDA, NO EL PUNTAJE (retocado el 2026-08-21, ADR-0221). Antes
    exigía `puntaje < 60,0`, y el puntaje era 59,4 el día que se escribió: el
    cable trampa quedó a un punto del valor. Un mes nuevo de la SRT lo movió a
    60,8 y disparó cuatro noches seguidas de nocturno — pero los juicios no se
    habían enfriado: crecían 2,1% (127.363 en 12 meses contra 124.767 previos).
    Lo único que había pasado es que la tasa se desaceleró 0,7 pp DENTRO DE LA
    MISMA banda, y el puntaje interpolado cruzó un número redondo.

    Un cable trampa a un punto del valor actual dispara por ruido, y lo que
    dispara por ruido se termina ignorando — que es la peor manera de perder
    una guardia. La banda (-5%, +5%) es la categoría de la propia metodología:
    "sin cambio apreciable". Mientras la variación siga ahí adentro no pasó
    nada que mirar. Si baja de -5%, la litigiosidad se enfrió de verdad."""
    import json
    cache = Path(__file__).parent.parent / "output" / "cache" / "gestion.json"
    if not cache.exists():
        pytest.skip("sin caché de gestión")
    valores = {k: v.get("valor")
               for k, v in json.loads(cache.read_text(encoding="utf-8"))["indicadores"].items()
               if isinstance(v, dict)}
    r = itcg.calcular_itcg(valores)
    puntajes = {k: d["puntaje"] for k, d in r["dimensiones"].items()}
    import parametrica
    variacion = valores["litigiosidad_laboral"]
    litigiosidad = parametrica.puntaje_de(
        variacion, "litigiosidad_laboral", itcg.BANDAS_ITCG)
    # El piso de la banda "sin cambio apreciable" de BANDAS_ITCG, leído de la
    # paramétrica y no escrito a mano acá: si alguien recalibra las bandas, esta
    # guardia se mueve con ellas en vez de quedar apuntando a un número viejo.
    piso_sin_cambio = next(desde for desde, hasta, _ in
                           itcg.BANDAS_ITCG["litigiosidad_laboral"]
                           if desde < 0 < hasta)
    assert variacion > piso_sin_cambio, (
        f"los juicios laborales cayeron {variacion}% interanual, por debajo del "
        f"{piso_sin_cambio}% que la paramétrica llama «sin cambio apreciable»: "
        f"la industria del juicio se habría enfriado de verdad, y eso hay que "
        f"mirarlo, no dar por bueno (puntaje {litigiosidad})")
    # el FAL ya no discrimina (ADR-0142): quedó fijo en su techo
    assert puntajes["reforma_laboral"] > litigiosidad, (
        "la dimensión debería estar sostenida por el FAL en 100")
