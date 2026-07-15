"""Regresiones offline para la presión de dolarización del ITCM."""
import io
import sys
from datetime import date, datetime
from pathlib import Path

import openpyxl
import pytest

ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT / "scripts"))

import descargar_series
import macro
import presion_dolarizacion as pdol


def _m2_y_tc(*meses: str) -> tuple[dict[str, float], dict[str, float]]:
    return ({mes: 100_000.0 for mes in meses}, {mes: 1_000.0 for mes in meses})


@pytest.mark.parametrize(
    ("valor", "anclas", "esperado"),
    [
        (-10.0, pdol.ANCLAS_PRECIO, 0.0),
        (5.0, pdol.ANCLAS_PRECIO, 0.0),
        (10.0, pdol.ANCLAS_PRECIO, 12.5),
        (30.0, pdol.ANCLAS_PRECIO, 50.0),
        (80.0, pdol.ANCLAS_PRECIO, 100.0),
        (0.0, pdol.ANCLAS_PUNTAJE, 100.0),
        (25.0, pdol.ANCLAS_PUNTAJE, 85.0),
        (50.0, pdol.ANCLAS_PUNTAJE, 60.0),
        (75.0, pdol.ANCLAS_PUNTAJE, 35.0),
        (100.0, pdol.ANCLAS_PUNTAJE, 10.0),
    ],
)
def test_interpolar_respeta_anclas_y_satura(valor, anclas, esperado):
    assert pdol.interpolar(valor, anclas) == pytest.approx(esperado)


def test_regimen_restringido_usa_brecha_ccl_contigua_de_tres_meses():
    serie = pdol.construir_serie(
        brecha_ccl={"2023-10": 40.0, "2023-11": 50.0, "2023-12": 60.0},
        demanda_neta_usd={},
        m2_privado_ars={},
        tc_a3500={},
    )

    assert serie == [
        {
            "mes": "2023-12",
            "regimen": "precio",
            "metrica": 50.0,
            "presion": 75.0,
            "puntaje_itcm": 35.0,
            "ventana_meses": 3,
            "ventana_parcial": False,
            "presion_formal": None,
            "presion_informal": None,
            "brecha_informal": None,
        }
    ]


def test_regimen_restringido_omite_ventanas_con_huecos():
    assert pdol.construir_serie(
        brecha_ccl={"2023-10": 40.0, "2023-12": 60.0},
        demanda_neta_usd={},
        m2_privado_ars={},
        tc_a3500={},
    ) == []


def test_regimen_abierto_suma_flujo_y_m2_antes_de_dividir():
    meses = ("2025-04", "2025-05", "2025-06")
    m2, tc = _m2_y_tc(*meses)

    serie = pdol.construir_serie(
        brecha_ccl={},
        demanda_neta_usd={"2025-04": 5.0, "2025-05": 7.0, "2025-06": 12.0},
        m2_privado_ars=m2,
        tc_a3500=tc,
        desde="2025-04",
    )

    assert serie == [
        {
            "mes": "2025-04",
            "regimen": "flujo",
            "metrica": 5.0,
            "presion": 41.67,
            "puntaje_itcm": 68.33,
            "ventana_meses": 1,
            "ventana_parcial": True,
            "presion_formal": 41.67,
            "presion_informal": None,
            "brecha_informal": None,
        },
        {
            "mes": "2025-05",
            "regimen": "flujo",
            "metrica": 6.0,
            "presion": 50.0,
            "puntaje_itcm": 60.0,
            "ventana_meses": 2,
            "ventana_parcial": True,
            "presion_formal": 50.0,
            "presion_informal": None,
            "brecha_informal": None,
        },
        {
            "mes": "2025-06",
            "regimen": "flujo",
            "metrica": 8.0,
            "presion": 62.5,
            "puntaje_itcm": 47.5,
            "ventana_meses": 3,
            "ventana_parcial": False,
            "presion_formal": 62.5,
            "presion_informal": None,
            "brecha_informal": None,
        },
    ]


def test_regimen_abierto_combina_formal_e_informal_70_30():
    """Con dólar cripto disponible, la presión del régimen abierto es el
    promedio ponderado 70/30 (formal/informal) sobre la MISMA ventana que ya
    usa el flujo formal (ADR-0057), no un promedio de brechas diarias aparte."""
    meses = ("2025-04", "2025-05", "2025-06")
    m2, tc = _m2_y_tc(*meses)
    # brecha cripto/A3500 (tc=1000): 3%, 4%, 6% por mes
    cripto = {"2025-04": 1030.0, "2025-05": 1040.0, "2025-06": 1060.0}

    serie = pdol.construir_serie(
        brecha_ccl={},
        demanda_neta_usd={"2025-04": 5.0, "2025-05": 7.0, "2025-06": 12.0},
        m2_privado_ars=m2,
        tc_a3500=tc,
        cripto_mensual=cripto,
        desde="2025-04",
    )

    assert serie == [
        {
            "mes": "2025-04", "regimen": "flujo", "metrica": 5.0,
            "presion": 36.67, "puntaje_itcm": 73.33,
            "ventana_meses": 1, "ventana_parcial": True,
            "presion_formal": 41.67, "presion_informal": 25.0,
            "brecha_informal": 3.0,
        },
        {
            "mes": "2025-05", "regimen": "flujo", "metrica": 6.0,
            "presion": 43.75, "puntaje_itcm": 66.25,
            "ventana_meses": 2, "ventana_parcial": True,
            "presion_formal": 50.0, "presion_informal": 29.17,
            "brecha_informal": 3.5,
        },
        {
            "mes": "2025-06", "regimen": "flujo", "metrica": 8.0,
            "presion": 54.58, "puntaje_itcm": 55.42,
            "ventana_meses": 3, "ventana_parcial": False,
            "presion_formal": 62.5, "presion_informal": 36.11,
            "brecha_informal": 4.33,
        },
    ]


def test_regimen_abierto_sin_cripto_para_toda_la_ventana_se_degrada_a_formal():
    """Si falta cripto para AL MENOS un mes de la ventana, no se mezcla nada
    a medias: la presión de ese mes queda 100% formal (sin inventar el dato
    faltante ni omitir el mes)."""
    meses = ("2025-04", "2025-05", "2025-06")
    m2, tc = _m2_y_tc(*meses)
    cripto_incompleto = {"2025-05": 1040.0, "2025-06": 1060.0}  # falta 2025-04

    serie = pdol.construir_serie(
        brecha_ccl={},
        demanda_neta_usd={"2025-04": 5.0, "2025-05": 7.0, "2025-06": 12.0},
        m2_privado_ars=m2,
        tc_a3500=tc,
        cripto_mensual=cripto_incompleto,
        desde="2025-04",
    )
    por_mes = {fila["mes"]: fila for fila in serie}

    # 2025-04: ventana = [2025-04] sola, falta cripto ahí → degrada a formal
    assert por_mes["2025-04"]["presion"] == 41.67
    assert por_mes["2025-04"]["presion_informal"] is None
    # 2025-06: ventana = [04,05,06], falta 04 en cripto → degrada a formal
    # también, aunque 05 y 06 sí tengan dato
    assert por_mes["2025-06"]["presion"] == 62.5
    assert por_mes["2025-06"]["presion_informal"] is None


def test_regimen_abierto_no_une_meses_no_contiguos():
    meses = ("2025-04", "2025-06")
    m2, tc = _m2_y_tc(*meses)

    serie = pdol.construir_serie(
        brecha_ccl={},
        demanda_neta_usd={"2025-04": 5.0, "2025-06": 12.0},
        m2_privado_ars=m2,
        tc_a3500=tc,
        desde="2025-04",
    )

    assert [fila["mes"] for fila in serie] == ["2025-04"]


def test_hitos_historicos_no_confunden_cera_con_corrida():
    meses_flujo = (
        "2025-07", "2025-08", "2025-09", "2025-10",
        "2026-03", "2026-04", "2026-05",
    )
    m2, tc = _m2_y_tc(*meses_flujo)

    serie = pdol.construir_serie(
        brecha_ccl={
            "2024-07": 10.0,
            "2024-08": 12.0,
            "2024-09": 14.0,
        },
        demanda_neta_usd={
            "2025-07": 10.0,
            "2025-08": 12.0,
            "2025-09": 17.504,
            "2025-10": 10.954,
            "2026-03": 5.0,
            "2026-04": 5.0,
            "2026-05": 6.2864,
        },
        m2_privado_ars=m2,
        tc_a3500=tc,
        desde="2024-09",
    )
    por_mes = {fila["mes"]: fila for fila in serie}

    assert por_mes["2024-09"]["presion"] == 17.5
    assert por_mes["2025-09"]["presion"] == 90.84
    assert por_mes["2025-09"]["puntaje_itcm"] == 19.16
    assert por_mes["2025-10"]["presion"] == 92.43
    assert por_mes["2025-10"]["puntaje_itcm"] == 17.57
    assert por_mes["2026-05"]["presion"] == 45.24
    assert por_mes["2026-05"]["puntaje_itcm"] == 64.76


class _Respuesta:
    def __init__(self, *, json_data=None, content=b""):
        self._json_data = json_data
        self.content = content

    def raise_for_status(self):
        return None

    def json(self):
        return self._json_data


def test_parser_xlsx_filtra_personas_concepto_agrupa_e_invierte_signo():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Datos Mercado de Cambios"
    ws.append(["Monto", "B", "Mes", "Sector"])
    concepto = "03- Compra-venta de billetes y divisas sin fines específicos"
    ws.append([-2_000_000, concepto, datetime(2025, 4, 1), "Personas Humanas"])
    ws.append([500_000, f" {concepto} ", datetime(2025, 4, 1), "Personas Humanas"])
    ws.append([-99_000_000, concepto, datetime(2025, 4, 1), "Sector Real"])
    ws.append([-7_000_000, "04- Viajes", datetime(2025, 4, 1), "Personas Humanas"])
    ws.append([-3_250_000, concepto, datetime(2025, 5, 1), "Personas Humanas"])
    contenido = io.BytesIO()
    wb.save(contenido)

    assert pdol.parsear_demanda_neta_personas_humanas(contenido.getvalue()) == {
        "2025-04": 1.5,
        "2025-05": 3.25,
    }


def test_parser_xlsx_acepta_variaciones_de_mayusculas():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Datos Mercado de Cambios"
    ws.append(["Monto", "B", "Mes", "Sector"])
    ws.append([
        -2_000_000,
        "03- COMPRA-VENTA DE BILLETES Y DIVISAS SIN FINES ESPECÍFICOS",
        datetime(2025, 4, 1),
        "Personas humanas",
    ])
    contenido = io.BytesIO()
    wb.save(contenido)

    assert pdol.parsear_demanda_neta_personas_humanas(contenido.getvalue()) == {
        "2025-04": 2.0,
    }


def test_parser_xlsx_falla_si_no_encuentra_filas_del_concepto():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Datos Mercado de Cambios"
    ws.append(["Monto", "B", "Mes", "Sector"])
    ws.append([-2_000_000, "04- Viajes", datetime(2025, 4, 1), "Personas Humanas"])
    contenido = io.BytesIO()
    wb.save(contenido)

    with pytest.raises(ValueError, match="sin filas"):
        pdol.parsear_demanda_neta_personas_humanas(contenido.getvalue())


def test_fetch_brecha_ccl_promedia_cada_serie_por_mes(monkeypatch):
    respuestas = iter([
        _Respuesta(json_data=[
            {"fecha": "2024-01-02", "venta": 120.0},
            {"fecha": "2024-01-31", "venta": 130.0},
            {"fecha": "2024-02-01", "venta": 150.0},
            {"fecha": "2024-02-29", "venta": 170.0},
        ]),
        _Respuesta(json_data={"results": [{"detalle": [
            {"fecha": "2024-01-02", "valor": 100.0},
            {"fecha": "2024-01-31", "valor": 110.0},
            {"fecha": "2024-02-01", "valor": 120.0},
            {"fecha": "2024-02-29", "valor": 130.0},
        ]}]}),
    ])
    llamadas = []

    def get(url, **kwargs):
        llamadas.append((url, kwargs))
        return next(respuestas)

    monkeypatch.setattr(pdol.requests, "get", get)

    assert pdol.fetch_brecha_ccl_mensual(desde="2024-01") == {
        "2024-01": 19.05,
        "2024-02": 28.0,
    }
    assert llamadas[0][0].endswith("/contadoconliqui")
    assert llamadas[1][0].endswith("/5")


def test_fetch_cripto_promedia_por_mes(monkeypatch):
    respuesta = _Respuesta(json_data=[
        {"fecha": "2024-01-02", "venta": 120.0},
        {"fecha": "2024-01-31", "venta": 130.0},
        {"fecha": "2024-02-01", "venta": 150.0},
        {"fecha": "2024-02-29", "venta": 170.0},
    ])
    llamadas = []

    def get(url, **kwargs):
        llamadas.append(url)
        return respuesta

    monkeypatch.setattr(pdol.requests, "get", get)

    assert pdol.fetch_cripto_mensual(desde="2024-01") == {
        "2024-01": 125.0,
        "2024-02": 160.0,
    }
    assert llamadas[0].endswith("/cripto")


def test_fetch_cripto_falla_si_viene_vacio(monkeypatch):
    monkeypatch.setattr(
        pdol.requests, "get", lambda *_a, **_k: _Respuesta(json_data=[]),
    )
    with pytest.raises(ValueError, match="dólar cripto"):
        pdol.fetch_cripto_mensual(desde="2023-10")


def test_fetch_brecha_falla_si_una_fuente_viene_vacia(monkeypatch):
    respuestas = iter([
        _Respuesta(json_data=[]),
        _Respuesta(json_data={"results": []}),
    ])
    monkeypatch.setattr(
        pdol.requests,
        "get",
        lambda _url, **_kwargs: next(respuestas),
    )

    with pytest.raises(ValueError, match="sin datos"):
        pdol.fetch_brecha_ccl_mensual(desde="2023-10")


def test_fetch_demanda_descarga_el_anexo_oficial(monkeypatch):
    monkeypatch.setattr(
        pdol.requests,
        "get",
        lambda url, **kwargs: _Respuesta(content=b"xlsx"),
    )
    monkeypatch.setattr(
        pdol,
        "parsear_demanda_neta_personas_humanas",
        lambda content: {"2025-04": 4_000.0} if content == b"xlsx" else {},
    )

    assert pdol.fetch_demanda_neta_personas_humanas() == {"2025-04": 4_000.0}


def test_obtener_serie_comparte_fuentes_y_recorta_el_historial():
    pedidos_bcra = []

    def fetch_bcra_fin_mes(var_id, meses):
        pedidos_bcra.append((var_id, meses))
        if var_id == pdol.BCRA_M2_PRIV_ID:
            return {
                "2025-04": 100_000.0,
                "2025-05": 100_000.0,
                "2025-06": 100_000.0,
            }
        return {
            "2025-04": 1_000.0,
            "2025-05": 1_000.0,
            "2025-06": 1_000.0,
        }

    serie = pdol.obtener_serie(
        meses_hist=2,
        fetch_brecha=lambda desde: {
            "2023-10": 40.0,
            "2023-11": 50.0,
            "2023-12": 60.0,
        },
        fetch_demanda=lambda: {
            "2025-04": 5.0,
            "2025-05": 7.0,
            "2025-06": 12.0,
        },
        fetch_cripto=lambda desde: {},
        fetch_bcra_fin_mes=fetch_bcra_fin_mes,
    )

    assert [fila["mes"] for fila in serie] == ["2025-05", "2025-06"]
    assert serie[-1]["presion"] == 62.5
    assert pedidos_bcra == [
        (pdol.BCRA_M2_PRIV_ID, 6),
        (pdol.BCRA_TC_A3500_ID, 6),
    ]


def test_obtener_serie_no_tumba_si_falla_solo_el_canal_informal():
    """El dólar cripto es complementario (ADR-0057): si su fetch explota, el
    indicador se sigue calculando con la presión 100% formal, no se pierde."""
    def fetch_bcra_fin_mes(var_id, _meses):
        if var_id == pdol.BCRA_M2_PRIV_ID:
            return {"2025-04": 100_000.0, "2025-05": 100_000.0, "2025-06": 100_000.0}
        return {"2025-04": 1_000.0, "2025-05": 1_000.0, "2025-06": 1_000.0}

    def fetch_cripto_roto(desde):
        raise RuntimeError("ArgentinaDatos caído")

    serie = pdol.obtener_serie(
        meses_hist=2,
        fetch_brecha=lambda desde: {
            "2023-10": 40.0, "2023-11": 50.0, "2023-12": 60.0,
        },
        fetch_demanda=lambda: {"2025-04": 5.0, "2025-05": 7.0, "2025-06": 12.0},
        fetch_cripto=fetch_cripto_roto,
        fetch_bcra_fin_mes=fetch_bcra_fin_mes,
    )

    assert [fila["mes"] for fila in serie] == ["2025-05", "2025-06"]
    assert serie[-1]["presion"] == 62.5
    assert serie[-1]["presion_informal"] is None


def test_obtener_serie_falla_si_falta_una_familia_de_fuentes():
    def fetch_bcra_fin_mes(var_id, _meses):
        if var_id == pdol.BCRA_M2_PRIV_ID:
            return {"2025-04": 100_000.0}
        return {"2025-04": 1_000.0}

    with pytest.raises(ValueError, match="demanda neta"):
        pdol.obtener_serie(
            meses_hist=2,
            fetch_brecha=lambda desde: {
                "2023-10": 40.0,
                "2023-11": 50.0,
                "2023-12": 60.0,
            },
            fetch_demanda=lambda: {},
            fetch_bcra_fin_mes=fetch_bcra_fin_mes,
        )


def test_macro_fetch_publica_presion_y_metadatos_del_regimen(monkeypatch):
    class FechaFija(date):
        @classmethod
        def today(cls):
            return cls(2026, 7, 14)

    monkeypatch.setattr(macro, "date", FechaFija)
    fila = {
        "mes": "2026-05",
        "regimen": "flujo",
        "metrica": 5.43,
        "presion": 45.24,
        "puntaje_itcm": 64.76,
        "ventana_meses": 3,
        "ventana_parcial": False,
    }
    monkeypatch.setattr(
        macro,
        "_presion_dolarizacion_serie_mensual",
        lambda meses_hist=36: [fila],
    )

    indicador = macro.fetch_presion_dolarizacion()

    assert indicador["valor"] == 45.24
    assert indicador["unidad"] == "pts (0-100)"
    assert indicador["fecha_dato"] == "2026-05-01"
    assert indicador["regimen"] == "flujo"
    assert indicador["metrica"] == 5.43
    assert indicador["ventana_meses"] == 3
    assert indicador["ventana_parcial"] is False
    assert indicador["puntaje_itcm"] == 64.76


def test_macro_no_publica_como_fresca_una_serie_fuera_de_rezago(monkeypatch):
    class FechaFija(date):
        @classmethod
        def today(cls):
            return cls(2026, 7, 14)

    monkeypatch.setattr(macro, "date", FechaFija)
    monkeypatch.setattr(
        macro,
        "_presion_dolarizacion_serie_mensual",
        lambda meses_hist=36: [{
            "mes": "2025-06",
            "regimen": "flujo",
            "metrica": 7.0,
            "presion": 56.25,
            "puntaje_itcm": 53.75,
            "ventana_meses": 3,
            "ventana_parcial": False,
        }],
    )

    assert macro.fetch_presion_dolarizacion() is None


def test_macro_registra_clave_nueva_y_purga_la_anterior():
    indicadores = {
        "idm": {"valor": 1.0},
        "dolarizacion_depositos": {"valor": 29.07},
    }

    macro._purgar_indicadores_obsoletos(indicadores)

    assert indicadores == {"idm": {"valor": 1.0}}
    assert "presion_dolarizacion" in macro.INDICADORES_ESPERADOS
    assert "dolarizacion_depositos" not in macro.INDICADORES_ESPERADOS


def test_descargar_series_reutiliza_la_serie_de_macro(monkeypatch):
    monkeypatch.setattr(
        macro,
        "_presion_dolarizacion_serie_mensual",
        lambda meses_hist: [
            {"mes": "2026-04", "presion": 40.0},
            {"mes": "2026-05", "presion": 45.24},
        ],
    )

    assert descargar_series.fetch_presion_dolarizacion_serie(meses=2) == [
        ["2026-04-01", 40.0],
        ["2026-05-01", 45.24],
    ]


def test_write_csv_merge_purga_la_clave_sustituida(tmp_path, monkeypatch):
    monkeypatch.setattr(descargar_series, "OUTPUT_DIR", tmp_path)
    path = tmp_path / "macro.csv"
    path.write_text(
        "fecha,indicador,valor,unidad,fuente\n"
        "2026-05-01,dolarizacion_depositos,29.07,pp,anterior\n"
        "2026-05-01,idm,1.0,pp,BCRA\n",
        encoding="utf-8",
    )

    descargar_series.write_csv(
        "macro",
        [["2026-05-01", "presion_dolarizacion", 45.24, "pts", "BCRA"]],
        merge=True,
        eliminar_indicadores={"dolarizacion_depositos"},
    )

    contenido = path.read_text(encoding="utf-8")
    assert "presion_dolarizacion" in contenido
    assert "dolarizacion_depositos" not in contenido
    assert ",idm," in contenido
