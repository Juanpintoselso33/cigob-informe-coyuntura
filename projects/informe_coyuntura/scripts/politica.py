"""
Colector Cinturón Político — CIGOB
Capital político según Carlos Matus: capacidad de gobernar (no popularidad).
Ejecutar desde projects/informe_coyuntura/: python scripts/politica.py

Indicadores:
  votometro_ventaja_lla     — Brecha LLA−PJ en intención de voto (Votómetro CIGOB, auto)
  ratio_dnu                 — DNUs / leyes sancionadas, ventana móvil 365 días (InfoLeg, auto;
                               ADR-0058 — antes acumulado del año calendario, resetaba en enero)
  conflictividad_nacional   — % var. eventos de protesta y disturbios en TODO el país vs. base
                               2023 (ACLED, 12m completos; ADR-0052 — reemplaza a
                               movilizacion_cepa en la dimensión conflicto_social)
  jornadas_individuales_no_trabajadas_12m — intensidad laboral oficial: huelguistas ×
                               duración del paro, acumulado móvil de 12 meses (ADR-0232)
  movilizacion_cepa         — Conflictividad social CEPA 0–100 (scrape centrocepa.com.ar, auto).
                               SEGUIMIENTO INTERNO desde 2026-07-11 (ADR-0052): sin backfill
                               posible (CEPA publica desde fines de 2025) y con fórmula de
                               acumulado YTD no comparable mes a mes — queda como contraste
                               oculto del snapshot
  iaf_transferencias        — Variación real YoY transferencias federales RON (Hacienda, auto)
  eficacia_legislativa      — % proyectos PE aprobados, ventana 12m (datos.hcdn.gob.ar CKAN, auto)
  cohesion_bloque           — % cohesión (Rice) del bloque LLA, COMPUESTO bicameral desde
                               2026-07-10 (ADR-0048): Diputados 65% (scrape votaciones.hcdn.gob.ar)
                               + Senado 35% (scrape senado.gob.ar), renormalizado si falta una cámara
  rotacion_gabinete         — salidas de rango ministerial (JGM + ministros) acumuladas 12m
                               (registro curado data/politica/gabinete_salidas.json + detector
                               de alerta InfoLeg, semiauto — patrón privatizaciones del ITCG).
                               SEGUIMIENTO INTERNO desde 2026-07-10 (ADR-0048): no puntúa
                               y publicar.py lo oculta del snapshot; se sigue relevando
  alineamiento_senadores_prov — % votos de senadores no-LLA alineados con LLA, por provincia
                               (scrape senado.gob.ar, auto — reemplaza a gobernadores_alineamiento,
                               placeholder manual congelado desde 2026-04, ver manuales.json)
  veto_quorum               — % sesiones frustradas por falta de quórum (datos.hcdn.gob.ar CKAN, auto)
  comisiones_caidas         — % proyectos con dictamen que no llegan al recinto (datos.hcdn.gob.ar CKAN, auto)
  adhesion_reformas_provincial — % provincias adheridas al RIGI (MAGyP, auto)
  derrotas_legislativas     — derrotas del Ejecutivo en el recinto, 12m: vetos insistidos +
                               decretos rechazados bajo la ley 26.122 (InfoLeg + actas Senado, auto)
  bloqueo_sostenido         — % de normas desafiadas en el recinto que siguen en pie, 12m
                               (ADR-0069): la cara ganada del pulso que derrotas no acredita —
                               insistencias de veto votadas + decretos bajo la 26.122 (actas de
                               AMBAS cámaras + InfoLeg, auto; comparte el registro de eventos
                               con derrotas_legislativas)
  protestas_caba            — % var. eventos de protesta en CABA vs. base 2023 (ACLED, reutiliza
                               gestion.py). SEGUIMIENTO INTERNO desde 2026-07-10 (ADR-0048): no
                               puntúa y publicar.py lo oculta del snapshot de política (la card
                               de gestión sigue siendo su lectura pública)

Nota: ICG UTDT removido (mide confianza ciudadana, no capacidad de gobernar con actores
políticos). Reemplazado por ratio_dnu según framework Luis Babino / reunión 12-may-2026.
IAF (Índice de Armonía Federal): iaf_transferencias captura cumplimiento fiscal federal
(Babino: Agregados de Poder). alineamiento_senadores_prov captura la dimensión territorial
(reemplaza a gobernadores_alineamiento, ver docstring de fetch_alineamiento_senadores_prov).
veto_quorum y comisiones_caidas capturan la eficacia legislativa de la oposición y el
bloqueo en comisiones — candidatos a fusionarse en índice compuesto legislativo.

Ojo con el «auto» de esta lista: acá significa que el colector busca el dato solo,
sin entrada en manuales.json. NO significa que el valor avance sin una persona:
varios de estos indicadores dejan lo ambiguo en triage (`pendientes_de_codificar`,
`pendientes_triage`). Lo que la web declara sobre la procedencia de cada dato sale
de `METODO_OBTENCION_EXCEPCIONES` en publicar.py (ADR-0234), que es el único lugar
donde eso se define.
"""
import sys
import io
import json
import re
import math
import calendar
import logging
import time
import unicodedata
import warnings
import requests
import urllib3
import pdfplumber
import openpyxl
import gestion  # reutiliza el fetcher ACLED ya construido para protestas_caba (ADR-0017)
import itcp
from html import unescape
from bs4 import BeautifulSoup
from datetime import datetime, date, timedelta
from pathlib import Path

sys.stdout.reconfigure(encoding="utf-8")
urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)

# ── Paths ─────────────────────────────────────────────────────────────────────
SCRIPT_DIR     = Path(__file__).parent
PROJECT_DIR    = SCRIPT_DIR.parent
CACHE_PATH     = PROJECT_DIR / "output" / "cache" / "politica.json"
MANUALES_PATH  = PROJECT_DIR / "data" / "politica" / "manuales.json"
AJUSTES_ITCP_PATH = PROJECT_DIR / "data" / "politica" / "ajustes_itcp.json"
# Registro versionado de derrotas legislativas (semilla verificada a mano +
# detección incremental de fetch_derrotas_legislativas) — está en el git add
# del cron (data-pipeline.yml): un caché que no se commitea no sobrevive a la
# corrida nocturna (lección 2026-07-09, ver CLAUDE.md).
DERROTAS_EVENTOS_PATH = PROJECT_DIR / "data" / "politica" / "derrotas_legislativas_eventos.json"
GABINETE_SALIDAS_PATH = PROJECT_DIR / "data" / "politica" / "gabinete_salidas.json"
GABINETE_DECRETOS_CACHE_PATH = PROJECT_DIR / "data" / "politica" / "gabinete_decretos_cache.json"
CSJN_NOVEDADES_PATH = PROJECT_DIR / "data" / "politica" / "csjn_novedades.json"
VOTOMETRO_URL  = "https://cigob.github.io/Votometro/"  # Votómetro live (embebido en cigob.org/votometro)
VOTOMETRO_HTML = PROJECT_DIR / "data" / "politica" / "votometro_fallback.html"  # fallback local

CINTURON              = "politica"
INDICADORES_ESPERADOS = [
    "votometro_ventaja_lla",
    "ratio_dnu",
    "brecha_obra_publica",
    "apoyo_empresario",
    "conflictividad_nacional",
    "jornadas_individuales_no_trabajadas_12m",
    "movilizacion_cepa",
    "iaf_transferencias",
    "cohesion_bloque",
    "rotacion_gabinete",
    "eficacia_legislativa",
    "alineamiento_senadores_prov",
    "adhesion_reformas_provincial",
    "veto_quorum",
    "comisiones_caidas",
    "derrotas_legislativas",
    "bloqueo_sostenido",
    "desafios_legislativos",
    "protestas_caba",
    "cobertura_judicial",
]

# ── Poder judicial (ADR-0126) ────────────────────────────────────────────────
# Padrón de cargos de juez con marca de vacante + registros de designaciones y
# renuncias, los tres en datos.jus.gob.ar (Ministerio de Justicia), formato CSV.
#
# El aporte externo proponía scrapear el archivo de "Concursos" del Consejo de
# la Magistratura. No hace falta: estos datasets ya publican el padrón completo
# con `cargo_vacante` SI/NO, que es la magnitud que el indicador necesita, y en
# CSV estructurado en vez de HTML. El scraper del Consejo queda como opción para
# los OTROS indicadores del bloque, que sí requieren datos de concursos.
JUS_API           = "https://datos.jus.gob.ar/api/3/action/package_search"
JUS_PADRON_Q      = "Magistrados de la Justicia Federal y de la Justicia Nacional"
JUS_DESIGNACIONES_Q = "Designaciones de magistrados de la Justicia Federal"
JUS_RENUNCIAS_Q   = "Renuncias de magistrados de la Justicia Federal"

STALE_MANUAL_DAYS    = 45
STALE_VOTOMETRO_DAYS = 60
STALE_DNU_DAYS       = 30
STALE_IAF_DAYS       = 365  # dato anual — válido todo el año

HTTP_TIMEOUT = 20
HTTP_HEADERS = {"User-Agent": "CIGOB-InformeCoyuntura/1.0"}

CEPA_INFORMES_URL       = "https://centrocepa.com.ar/documentos/informes"
CEPA_MAX_CASOS_MES      = 80.0
CEPA_MAX_CONFLICTOS_TOT = 200.0

# InfoLeg — leyes y DNUs
# tipoNorma: 1=Ley, 2=Decreto. DNUs se identifican con texto="necesidad y urgencia"
INFOLEG_HOME    = "https://servicios.infoleg.gob.ar/infolegInternet/"
INFOLEG_BUSCAR  = "https://servicios.infoleg.gob.ar/infolegInternet/buscarNormas.do"

# Hacienda RON — transferencias federales históricas (serie anual 2003–año_actual)
RON_CSV_URL = "https://www.argentina.gob.ar/sites/default/files/serie_ron_2003_2025.csv"
# El CSV distribuye los recursos entre TODAS las jurisdicciones del régimen,
# no solo provincias: la porción del Tesoro Nacional (~35% del total) y de la
# Seguridad Social/ANSES no son transferencias a provincias, y el Fondo ATN es
# una retención que queda en la Nación hasta repartirse discrecionalmente
# (ADR-0066 — sin este filtro el nivel daba $104B cuando IARAF/DNAP reportan
# ~$60B, y la variación mezclaba la dinámica de la porción nacional). El
# fdo.compensador SÍ se gira a provincias y se conserva.
RON_NO_PROVINCIA = {"tesoro nacional", "seguridad social", "fondo a.t.n."}

# HCDN CKAN — datos.hcdn.gob.ar (open data portal de la Cámara de Diputados)
HCDN_CKAN            = "https://datos.hcdn.gob.ar/api/3/action/datastore_search"
HCDN_PROYECTOS_RID   = "22b2d52c-7a0e-426b-ac0a-a3326c388ba6"   # proyectos-parlamentarios
HCDN_MOVIMIENTOS_RID = "6108ea83-3f12-423c-a136-df1ae9cb2972"   # movimientos-de-proyectos
HCDN_LEYES_SANC_RID  = "68dfd7f8-91f3-4ecf-aebf-a860d1ca1a98"   # leyes-sancionadas (ADR-0062)
HCDN_SESIONES_RID    = "4ac70a51-a82d-428b-966a-0a203dd0a7e3"   # sesiones plenarias
HCDN_DICTAMENES_RID  = "59595a93-5a5e-4ba6-a3db-c1044e2f949e"   # dictámenes de comisión
# Expedientes del Poder Ejecutivo: -PE- (Presidencia) y -JGM- (Jefatura de
# Gabinete — el Presupuesto anual entra SIEMPRE por esta vía, art. 100 inc. 6
# CN; sin JGM el indicador de eficacia era ciego a la ley más importante del
# año, ADR-0063).
_RE_PE_EXP           = re.compile(r"\d+-(?:PE|JGM)-\d{4}")

# HCDN Votaciones — votaciones.hcdn.gob.ar (portal de votaciones nominales)
HCDN_VOTACIONES_BASE = "https://votaciones.hcdn.gob.ar"
_HCDN_VOTACIONES_DELAY = 0.3  # segundos entre requests — evita el WAF F5 BIG-IP
                              # (confirmado: Como_voto corre a diario con este patrón)

# Senado — senado.gob.ar (portal de votaciones nominales, sitio distinto de HCDN)
SENADO_BASE = "https://www.senado.gob.ar"

# MAGyP — tabla de provincias adheridas al RIGI (Título VII, Ley 27.742)
MAGYP_RIGI_URL = "https://www.magyp.gob.ar/desarrollo-foresto-industrial/provincias-adheridas.php"

# IPC interanual diciembre (INDEC). Actualizar en enero de cada año.
# 2024: 117.06% acumulado anual. 2025: 38.3% acumulado anual (estimado previo a cierre).

INDEC_SERIES_URL = "https://apis.datos.gob.ar/series/api/series/"
INDEC_IPC_INDICE = "148.3_INIVELNAL_DICI_M_26"   # IPC nivel general, índice (base dic-2016=100)


# ── RON mensual — transferencias federales a precios de cada mes (ADR-0239) ───
# El CSV anual sólo permite comparar dos sumas nominales contra un deflactor
# único, y eso subdeflacta: las transferencias se devengan mes a mes, con
# estacionalidad propia, así que el deflactor correcto está ponderado por el
# flujo de cada mes y no por el calendario. La misma planilla de Hacienda que
# publica el consolidado anual trae una hoja por mes; de ahí sale la serie.

RON_CONSOLIDADO_PAGINA = "https://www.argentina.gob.ar/economia/sechacienda/asuntosprovinciales/ron"

# Las tres filas del cuadro que SON transferencias a jurisdicciones. El resto
# —Tesoro Nacional, Seguridad Social, Fondo A.T.N.— queda en la Nación
# (ADR-0066). Los rótulos vienen espaciados letra por letra en los años viejos
# ("P R O V I N C I A S"), así que se comparan sin espacios.
RON_FILAS_JURISDICCION = ("provincias", "c.a.b.a", "fdo.compensador")

_MESES_RON = ("enero", "febrero", "marzo", "abril", "mayo", "junio", "julio",
              "agosto", "septiembre", "octubre", "noviembre", "diciembre")

_RON_MENSUAL_MEMO: dict = {}      # {año: {YYYY-MM: monto}}
_RON_CSV_MEMO: dict = {}          # {año: total anual, en la unidad del CSV}
_RON_URLS_MEMO: dict = {}


def _norm_ron(s) -> str:
    """Sin acentos, sin espacios repetidos, en minúsculas."""
    import unicodedata
    s = unicodedata.normalize("NFKD", str(s))
    s = "".join(c for c in s if not unicodedata.combining(c))
    return re.sub(r"\s+", " ", s).strip().lower()


def _rotulo_ron(s) -> str:
    """Como `_norm_ron` pero sin NINGÚN espacio: hasta 2023 los rótulos de fila
    venían espaciados letra por letra («P R O V I N C I A S»)."""
    return _norm_ron(s).replace(" ", "")


def _ron_consolidado_urls() -> dict:
    """{año: url} de las planillas «Información consolidada» de Hacienda.

    El nombre de archivo no es estable —`informacion_consolidada_2024.xlsx`,
    `informacion_consolidada2025_5.xlsx`, `informacion_consolidada_2026_4.xlsx`—
    así que el año se lee del nombre y la URL se resuelve desde la página."""
    if _RON_URLS_MEMO:
        return _RON_URLS_MEMO
    r = requests.get(RON_CONSOLIDADO_PAGINA, headers=HTTP_HEADERS, timeout=HTTP_TIMEOUT)
    r.raise_for_status()
    patron = (r'((?:https://www\.argentina\.gob\.ar)?'
              r'/sites/default/files/informacion_consolidada[^"]*?\.xlsx?)')
    for href in re.findall(patron, r.text):
        m = re.search(r"(20\d{2})", href.rsplit("/", 1)[-1])
        if not m:
            continue
        _RON_URLS_MEMO[int(m.group(1))] = (
            href if href.startswith("http")
            else f"https://www.argentina.gob.ar{href}")
    if not _RON_URLS_MEMO:
        raise ValueError("no se encontraron planillas consolidadas de RON")
    return _RON_URLS_MEMO


def _hojas_de_planilla(contenido: bytes, es_xlsx: bool) -> dict:
    """{hoja: filas}. Hacienda pasó a .xlsx recién en 2024; todo lo anterior
    sigue en el .xls OLE2, que openpyxl no abre."""
    if es_xlsx:
        import io
        import openpyxl
        wb = openpyxl.load_workbook(io.BytesIO(contenido), read_only=True, data_only=True)
        return {_norm_ron(h): [list(f) for f in wb[h].iter_rows(values_only=True)]
                for h in wb.sheetnames}
    import xlrd
    wb = xlrd.open_workbook(file_contents=contenido)
    hojas = {}
    for h in wb.sheet_names():
        sh = wb.sheet_by_name(h)
        hojas[_norm_ron(h)] = [
            [sh.cell_value(i, j) if sh.cell_value(i, j) != "" else None
             for j in range(sh.ncols)]
            for i in range(sh.nrows)
        ]
    return hojas


def _columnas_ron(filas: list) -> tuple:
    """(columna del total RON, columna de la compensación del Consenso Fiscal).

    Se buscan por encabezado, no por posición: el cuadro fue ganando columnas
    con los años y una posición fija se desplaza sin que nada falle. El
    encabezado está partido en varias filas, así que se concatena en vertical."""
    ancho = max((len(f) for f in filas[:16]), default=0)
    encabezados = {}
    for j in range(ancho):
        # sólo las celdas de TEXTO: en las planillas viejas el encabezado y los
        # datos comparten las primeras filas, y arrastrar los números haría que
        # ningún encabezado "termine" donde se lo espera
        encabezados[j] = " ".join(
            _norm_ron(filas[i][j])
            for i in range(min(16, len(filas)))
            if j < len(filas[i]) and isinstance(filas[i][j], str) and filas[i][j].strip()
        )
    total = consenso = None
    for j, txt in encabezados.items():
        if total is None and "total" in txt and "origen nacional" in txt:
            total = j
        if consenso is None and "consenso fiscal" in txt:
            consenso = j
    if total is None:
        # Hasta 2017 la columna se llamaba sólo «T O T A L», sin decir de qué.
        # Se toma la última que TERMINA en total y no es un subtotal: hay dos
        # columnas «Sub-total» antes, y quedarse con una de ellas dejaría afuera
        # media planilla sin que nada fallara.
        candidatas = [j for j, txt in encabezados.items()
                      if txt.replace(" ", "").endswith("total")
                      and "subtotal" not in txt.replace(" ", "")
                      and "sub-total" not in txt.replace(" ", "")]
        total = candidatas[-1] if candidatas else None
    if total is None:
        raise ValueError("no está la columna de total de recursos de origen nacional")
    return total, consenso


def _total_jurisdicciones(filas: list) -> float:
    """Suma del cuadro para las filas que son transferencias a jurisdicciones.

    «Provincias» aparece dos veces —encabezado y subtotal—; se toma la que trae
    números. La compensación del Consenso Fiscal va aparte en su propia columna
    y sí forma parte de lo girado: sin ella el total no cierra contra el CSV
    anual (2024: 41,13 B contra 42,13 B)."""
    total_col, consenso_col = _columnas_ron(filas)
    suma = 0.0
    vistas = set()
    for f in filas:
        if not f or f[0] is None:
            continue
        rot = _rotulo_ron(f[0])
        if rot not in RON_FILAS_JURISDICCION or rot in vistas:
            continue
        val = f[total_col] if total_col < len(f) else None
        if not isinstance(val, (int, float)):
            continue                       # el encabezado homónimo, sin números
        vistas.add(rot)
        suma += float(val)
        if consenso_col is not None and consenso_col < len(f):
            c = f[consenso_col]
            if isinstance(c, (int, float)):
                suma += float(c)
    if "provincias" not in vistas:
        raise ValueError("el cuadro no trae la fila de subtotal de provincias")
    return suma


def _ron_mensual(desde: int) -> dict:
    """{YYYY-MM: monto girado a jurisdicciones}, desde el año `desde`.

    Mismo universo que el CSV anual —lo reconcilia peso por peso— pero abierto
    por mes, que es lo que hace falta para deflactar cada flujo a los precios en
    los que se devengó. Un mes sin hoja (el año en curso) simplemente no está."""
    urls = _ron_consolidado_urls()
    for y in sorted(y for y in urls if y >= desde):
        if y in _RON_MENSUAL_MEMO:
            continue
        r = requests.get(urls[y], headers=HTTP_HEADERS, timeout=120)
        r.raise_for_status()
        hojas = _hojas_de_planilla(r.content, urls[y].endswith(".xlsx"))
        del_anio = {}
        for k, mes in enumerate(_MESES_RON, 1):
            if mes not in hojas:
                continue
            try:
                del_anio[f"{y}-{k:02d}"] = _total_jurisdicciones(hojas[mes])
            except ValueError:
                continue
        _RON_MENSUAL_MEMO[y] = del_anio
    out = {}
    for y, meses in _RON_MENSUAL_MEMO.items():
        if y >= desde:
            out.update(meses)
    return out


def _ipc_indice_mensual() -> dict:
    """{YYYY-MM: índice IPC nivel general} de INDEC, base dic-2016=100."""
    r = requests.get(INDEC_SERIES_URL,
                     params={"ids": INDEC_IPC_INDICE, "format": "json",
                             "limit": 1000, "sort": "desc"},
                     headers=HTTP_HEADERS, timeout=HTTP_TIMEOUT)
    r.raise_for_status()
    return {str(f)[:7]: float(v) for f, v in r.json()["data"] if v is not None}


def _ron_total_anual_csv() -> dict:
    """{año: total girado a jurisdicciones} del CSV anual de Hacienda.

    Es el mismo universo que las hojas mensuales y cubre 2003-2025 en UNA sola
    unidad, así que hace de ancla: las planillas mensuales cambiaron de miles a
    millones de pesos entre 2022 y 2023 y no declaran su unidad en ningún lado
    (sin el ancla, 2023 daba −99,9% real)."""
    import csv
    import io
    if _RON_CSV_MEMO:
        return _RON_CSV_MEMO
    r = requests.get(RON_CSV_URL, headers=HTTP_HEADERS, timeout=HTTP_TIMEOUT)
    r.raise_for_status()
    rd = csv.reader(io.StringIO(r.text), delimiter=";")
    next(rd)
    for row in rd:
        if len(row) < 5:
            continue
        if row[1].strip().lower() in RON_NO_PROVINCIA:      # ADR-0066
            continue
        try:
            y, v = int(row[0]), float(row[4].replace(",", "."))
        except ValueError:
            continue
        _RON_CSV_MEMO[y] = _RON_CSV_MEMO.get(y, 0.0) + v
    if not _RON_CSV_MEMO:
        raise ValueError("el CSV anual de RON no trajo ningún año")
    return _RON_CSV_MEMO


def _factor_unidad(suma_mensual: float, total_csv: float) -> float:
    """Potencia de 1000 que lleva las hojas del año a la unidad del CSV.

    No se "ajusta" nada: el factor tiene que ser exactamente una potencia de mil
    y el residuo, menor al 1%. Si no lo es, algo cambió en el cuadro —una fila
    nueva, una columna corrida— y es preferible que el colector falle a que
    publique una variación real construida sobre dos unidades distintas."""
    import math
    if suma_mensual <= 0 or total_csv <= 0:
        raise ValueError("suma mensual o total anual no positivos")
    factor = 1000.0 ** round(math.log(total_csv / suma_mensual, 1000.0))
    residuo = abs(suma_mensual * factor / total_csv - 1.0)
    if residuo > 0.01:
        raise ValueError(
            f"las hojas mensuales no reconcilian con el CSV anual: "
            f"factor {factor:g}, residuo {residuo:.3%}")
    return factor


def _iaf_real_por_anio(desde: int = 2016) -> dict:
    """{año: (var_real, var_nominal, deflactor, total_ref, total_ant)}.

    Cada flujo mensual se lleva a precios de una base común dividiéndolo por el
    IPC de SU mes, y recién ahí se suman los doce. Es la diferencia entre medir
    la variación real y medir la variación nominal con un deflactor promedio:
    para 2025 la primera da +1,6% —lo mismo que IARAF y Politikon— y la segunda
    daba +0,8% (ADR-0239).

    Sólo entran los años con los doce meses publicados: un año a medias
    compararía nueve meses contra doce."""
    mensual = _ron_mensual(desde)
    ipc = _ipc_indice_mensual()
    anual_csv = _ron_total_anual_csv()
    completos = {}
    for y in sorted({int(ym[:4]) for ym in mensual}):
        meses = [f"{y}-{k:02d}" for k in range(1, 13)]
        if not all(m in mensual and m in ipc for m in meses):
            continue
        if y not in anual_csv:
            continue                       # sin ancla no se sabe en qué unidad está
        factor = _factor_unidad(sum(mensual[m] for m in meses), anual_csv[y])
        completos[y] = (meses, factor)
    out = {}
    for y in sorted(completos):
        if y - 1 not in completos:
            continue
        (meses_ref, f_ref), (meses_ant, f_ant) = completos[y], completos[y - 1]
        nom_ref = sum(mensual[m] * f_ref for m in meses_ref)
        nom_ant = sum(mensual[m] * f_ant for m in meses_ant)
        real_ref = sum(mensual[m] * f_ref / ipc[m] for m in meses_ref)
        real_ant = sum(mensual[m] * f_ant / ipc[m] for m in meses_ant)
        if not nom_ant or not real_ant:
            continue
        var_real = real_ref / real_ant - 1.0
        var_nom = nom_ref / nom_ant - 1.0
        deflactor = (1.0 + var_nom) / (1.0 + var_real) - 1.0
        out[y] = (var_real, var_nom, deflactor, nom_ref, nom_ant)
    return out


logging.basicConfig(level=logging.WARNING, format="%(message)s")


# ── HCDN Votaciones session helpers ──────────────────────────────────────────────

def _hcdn_votaciones_session() -> requests.Session:
    """Sesión persistente con headers estables. El WAF del sitio devuelve 403
    ante ráfagas o headers que varían entre requests — reusar la misma sesión
    y no variar el UA es lo que lo evita."""
    s = requests.Session()
    s.headers.update(HTTP_HEADERS)
    return s


def _paced_get(session: requests.Session, base_url: str, path: str, aceptar_404: bool = False, **kwargs):
    """GET con pacing fijo y retry/backoff ante 403 (hasta 3 intentos).
    Generaliza el helper de HCDN para reusar sesión/pacing contra Senado.
    None si se agotan los reintentos o hay un error de red. Con
    `aceptar_404=True` un 404 devuelve la Response (para que el caller pueda
    distinguir "no existe" de un fallo transitorio -- lo necesita el walk de
    actas de Diputados, donde un hueco de id es normal pero un fallo de red
    no debe congelarse como hueco)."""
    url = f"{base_url}{path}"
    for intento in range(3):
        time.sleep(_HCDN_VOTACIONES_DELAY)
        try:
            r = session.get(url, timeout=HTTP_TIMEOUT, **kwargs)
        except requests.RequestException:
            return None
        if r.status_code == 200:
            return r
        if r.status_code == 403:
            continue
        if r.status_code == 404 and aceptar_404:
            return r
        return None
    return None


def _paced_post(session: requests.Session, base_url: str, path: str, data: dict, **kwargs):
    """POST con el mismo pacing/retry que _paced_get (backoff ante 403, hasta
    3 intentos). Usado por _descubrir_actas_senado: el listado de actas de
    Senado requiere POST con busqueda_actas[anio] en el form -- un GET plano
    siempre devuelve el año en curso (bug encontrado en auditoría 2026-07-08,
    ver commit)."""
    url = f"{base_url}{path}"
    for intento in range(3):
        time.sleep(_HCDN_VOTACIONES_DELAY)
        try:
            r = session.post(url, data=data, timeout=HTTP_TIMEOUT, **kwargs)
        except requests.RequestException:
            return None
        if r.status_code == 200:
            return r
        if r.status_code == 403:
            continue
        return None
    return None


def _hcdn_votaciones_get(session: requests.Session, path: str, **kwargs):
    return _paced_get(session, HCDN_VOTACIONES_BASE, path, **kwargs)


# ── Cache helpers ─────────────────────────────────────────────────────────────

def load_cache() -> dict:
    if CACHE_PATH.exists():
        with open(CACHE_PATH, encoding="utf-8") as f:
            return json.load(f)
    return {}


def save_cache(data: dict) -> None:
    CACHE_PATH.parent.mkdir(parents=True, exist_ok=True)
    with open(CACHE_PATH, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2, default=str)


def _warn(indicador: str, msg: str) -> None:
    print(f"[WARN] {CINTURON}.{indicador}: {msg}. Usando cache.")


def _sellar(resultado: dict) -> dict:
    """Sella el momento en que este valor se obtuvo de la fuente en vivo (ADR-0191).

    Se aplica SOLO a resultados frescos. El carry-forward hace
    `{**anterior, "desactualizado": True}`, asi que arrastra el sello viejo
    intacto: es esa fecha que deja de moverse la que mide hace cuanto que la
    fuente no contesta. `fecha_dato` no sirve para eso — en las series anuales
    no se mueve aunque el fetch ande perfecto.
    """
    return {**resultado, "obtenido_en": datetime.now().isoformat(timespec="seconds")}


def _days_old(fecha_str: str) -> int:
    try:
        fecha = date.fromisoformat(str(fecha_str)[:10])
        return (date.today() - fecha).days
    except Exception:
        return 999


# ── Votómetro parser ──────────────────────────────────────────────────────────

def _cargar_votometro_html() -> str:
    """HTML del Votómetro LIVE (cigob.github.io/Votometro, embebido en cigob.org).
    Cae al archivo local si la URL falla o no trae encuestasRaw."""
    try:
        r = requests.get(VOTOMETRO_URL, headers=HTTP_HEADERS, timeout=HTTP_TIMEOUT)
        r.raise_for_status()
        if "encuestasRaw" in r.text:
            return r.text
        raise ValueError("encuestasRaw ausente en la respuesta del Votómetro live")
    except Exception as e:
        if VOTOMETRO_HTML.exists():
            _warn("votometro (URL live falló, uso archivo local)", str(e))
            return VOTOMETRO_HTML.read_text(encoding="utf-8")
        raise


def fetch_votometro() -> dict | None:
    """
    Parsea encuestasRaw del Votómetro CIGOB y calcula la brecha ponderada LLA−PJ.

    Filtros:
    - tipo='espacio' (porcentajes de espacio político, no candidatos individuales)
    - últimos STALE_VOTOMETRO_DAYS días desde la encuesta más reciente

    Peso = exp(−0.015 × días) × calidad_mult  donde A=3, B=2, C=1
    """
    try:
        html = _cargar_votometro_html()

        m = re.search(r"const\s+encuestasRaw\s*=\s*\[(.*?)\];", html, re.DOTALL)
        if not m:
            raise ValueError("No se encontró encuestasRaw en el HTML")

        raw_block = m.group(1)

        entries = []
        for obj in re.finditer(r"\{([^}]+)\}", raw_block):
            fields = {}
            for kv in re.finditer(r"(\w+)\s*:\s*'([^']*)'|(\w+)\s*:\s*([\d.]+)", obj.group(1)):
                if kv.group(1):
                    fields[kv.group(1)] = kv.group(2)
                else:
                    fields[kv.group(3)] = float(kv.group(4))
            if fields:
                entries.append(fields)

        espacios = [e for e in entries if str(e.get("tipo", "")).strip() == "espacio"]
        if not espacios:
            raise ValueError("Sin encuestas tipo='espacio' en Votómetro")

        fechas = []
        for e in espacios:
            try:
                fechas.append(date.fromisoformat(str(e["fecha"])[:10]))
            except Exception:
                pass
        if not fechas:
            raise ValueError("Sin fechas válidas en encuestas espacio")
        fecha_max = max(fechas)

        cutoff = (fecha_max - timedelta(days=STALE_VOTOMETRO_DAYS)).isoformat()
        recientes = [e for e in espacios if str(e.get("fecha", "")) >= cutoff]
        if not recientes:
            raise ValueError("Sin encuestas espacio recientes en ventana de tiempo")

        CALIDAD_MULT = {"A": 3.0, "B": 2.0, "C": 1.0}
        LAMBDA = 0.015

        suma_peso = 0.0
        suma_lla  = 0.0
        suma_pj   = 0.0

        for e in recientes:
            try:
                fecha_enc = date.fromisoformat(str(e["fecha"])[:10])
                dias = (date.today() - fecha_enc).days
                wT = math.exp(-LAMBDA * dias)
                cal = str(e.get("calidad", "B")).strip().upper()
                wC = CALIDAD_MULT.get(cal, 2.0)
                w  = wT * wC

                lla = float(e.get("LLA", 0))
                pj  = float(e.get("PJ", 0))

                suma_peso += w
                suma_lla  += w * lla
                suma_pj   += w * pj
            except Exception:
                continue

        if suma_peso == 0:
            raise ValueError("Suma de pesos = 0")

        lla_pond = round(suma_lla / suma_peso, 1)
        pj_pond  = round(suma_pj / suma_peso, 1)
        gap      = round(lla_pond - pj_pond, 1)

        return {
            "valor": gap,
            "lla_ponderado": lla_pond,
            "pj_ponderado": pj_pond,
            "n_encuestas": len(recientes),
            "unidad": "Puntos porcentuales",
            "fuente": "Votómetro CIGOB",
            "fecha_dato": str(fecha_max),
            "desactualizado": _days_old(str(fecha_max)) > STALE_VOTOMETRO_DAYS,
        }

    except Exception as e:
        _warn("votometro_ventaja_lla", str(e))
        return None


def votometro_serie_mensual() -> list:
    """Serie histórica mensual de la brecha LLA−PJ del Votómetro, reconstruida desde
    encuestasRaw (que trae todos los sondeos desde dic-2023). Para cada mes se aplica
    la MISMA ponderación que fetch_votometro (recencia exp(−0,015·días) × calidad,
    ventana de STALE_VOTOMETRO_DAYS anclada en el último sondeo del mes), evaluada al
    cierre del mes. Devuelve [(YYYY-MM, gap)] ascendente."""
    html = _cargar_votometro_html()
    m = re.search(r"const\s+encuestasRaw\s*=\s*\[(.*?)\];", html, re.DOTALL)
    if not m:
        raise ValueError("No se encontró encuestasRaw en el HTML")
    esp = []
    for obj in re.finditer(r"\{([^}]+)\}", m.group(1)):
        f = {}
        for kv in re.finditer(r"(\w+)\s*:\s*'([^']*)'|(\w+)\s*:\s*([\d.]+)", obj.group(1)):
            f[kv.group(1) or kv.group(3)] = kv.group(2) if kv.group(1) else float(kv.group(4))
        if f and str(f.get("tipo", "")).strip() == "espacio" and f.get("fecha"):
            esp.append(f)
    fechas = sorted(date.fromisoformat(str(e["fecha"])[:10]) for e in esp)
    if not fechas:
        raise ValueError("Sin encuestas tipo='espacio'")
    CAL = {"A": 3.0, "B": 2.0, "C": 1.0}; LAMBDA = 0.015

    def gap_al(asof: date):
        fmax = max((f for f in fechas if f <= asof), default=None)
        if not fmax:
            return None
        cutoff = fmax.toordinal() - STALE_VOTOMETRO_DAYS
        sw = sl = sp = 0.0
        for e in esp:
            fe = date.fromisoformat(str(e["fecha"])[:10])
            if fe > asof or fe.toordinal() < cutoff:
                continue
            w = math.exp(-LAMBDA * (asof - fe).days) * CAL.get(str(e.get("calidad", "B")).strip().upper(), 2.0)
            sw += w; sl += w * float(e.get("LLA", 0)); sp += w * float(e.get("PJ", 0))
        return round((sl - sp) / sw, 1) if sw else None

    out = []
    y, mo = fechas[0].year, fechas[0].month
    while (y, mo) <= (date.today().year, date.today().month):
        asof = min(date(y, mo, calendar.monthrange(y, mo)[1]), date.today())
        g = gap_al(asof)
        if g is not None:
            out.append((f"{y}-{mo:02d}", g))
        mo += 1
        if mo > 12:
            mo = 1; y += 1
    return out


# ── Ratio DNU ─────────────────────────────────────────────────────────────────

def _infoleg_session_count(session: requests.Session, action_url: str,
                            tipo: str, desde: date, hasta: date, texto: str = "") -> int:
    """
    POST a InfoLeg buscarNormas.do dentro de una sesión activa.
    tipo: "1"=Ley, "2"=Decreto. DNUs se identifican con texto="necesidad y urgencia".
    Ventana [desde, hasta] explícita (ADR-0058): antes tomaba un año calendario
    completo (o el año corriente hasta hoy); ahora cualquier rango, para poder
    pedir ventanas móviles de 365 días además del corte anual.
    """
    post_data = {
        "tipoNorma": tipo,
        "numero": "",
        "anioSancion": "",
        "dependencia": "",
        "diaPubDesde": f"{desde.day:02d}",
        "mesPubDesde": f"{desde.month:02d}",
        "anioPubDesde": str(desde.year),
        "diaPubHasta": f"{hasta.day:02d}",
        "mesPubHasta": f"{hasta.month:02d}",
        "anioPubHasta": str(hasta.year),
        "texto": texto,
    }
    r = session.post(action_url, data=post_data, headers=HTTP_HEADERS, timeout=HTTP_TIMEOUT)
    r.raise_for_status()

    m = re.search(r"Encontradas?[:\s]+(\d+)", r.text, re.IGNORECASE)
    if m:
        return int(m.group(1))
    raise ValueError(f"Conteo no encontrado en InfoLeg (tipo={tipo}, texto={texto!r})")


# ── DNU: el tipo jurídico, no la coincidencia textual (ADR-0241) ─────────────

_RE_INFOLEG_DNU = re.compile(r"^decreto\s+dnu\b", re.IGNORECASE)
_INFOLEG_FILAS_POR_PAGINA = 50


def _infoleg_listado_completo(session: requests.Session, action_url: str, tipo: str,
                              desde: date, hasta: date, texto: str = "") -> list[dict]:
    """Todas las filas del listado de InfoLeg, paginando.

    La grilla devuelve 50 por página y el resto queda detrás de un submit del
    mismo formulario con `desplazamiento=AP` e `irAPagina`. Contar sólo la
    primera página daría 50 y no fallaría: la ventana de 365 días trae 48, o
    sea que el techo estaba a dos normas de distancia."""
    vistos, items, total = set(), [], None
    for pagina in range(1, 60):
        data = {
            "tipoNorma": tipo, "numero": "", "anioSancion": "", "dependencia": "",
            "diaPubDesde": f"{desde.day:02d}", "mesPubDesde": f"{desde.month:02d}",
            "anioPubDesde": str(desde.year),
            "diaPubHasta": f"{hasta.day:02d}", "mesPubHasta": f"{hasta.month:02d}",
            "anioPubHasta": str(hasta.year),
            "texto": texto,
        }
        if pagina > 1:
            data["desplazamiento"] = "AP"
            data["irAPagina"] = str(pagina)
        time.sleep(_DERROTAS_PAUSA_INFOLEG)
        r = session.post(action_url, data=data, headers=HTTP_HEADERS, timeout=HTTP_TIMEOUT)
        r.raise_for_status()
        if "Encontradas" not in r.text and "No se encontraron normas" not in r.text:
            raise ValueError(f"respuesta InfoLeg sin listado (tipo={tipo}, texto={texto!r})")
        m = re.search(r"Encontradas?[:\s]+(\d+)", r.text, re.IGNORECASE)
        if total is None:
            total = int(m.group(1)) if m else 0
        antes = len(items)
        for it in _parsear_listado_infoleg(r.text):
            if it["infoleg_id"] not in vistos:
                vistos.add(it["infoleg_id"])
                items.append(it)
        if len(items) >= total or len(items) == antes:
            # una página que no aporta filas nuevas no va a aportarlas en la
            # siguiente: se corta acá y decide el control de total de abajo
            break
    if len(items) != total:
        raise ValueError(
            f"InfoLeg devolvió {len(items)} filas de {total} declaradas "
            f"(tipo={tipo}, texto={texto!r}): el listado quedó truncado")
    return items


def _infoleg_contar_dnus(session: requests.Session, action_url: str,
                         desde: date, hasta: date) -> tuple[int, list[dict]]:
    """(cantidad de DNU publicados en la ventana, inventario).

    El conteo anterior era el de la búsqueda de texto completo `necesidad y
    urgencia` sobre los decretos, y esa frase aparece también en decretos que
    **no** son DNU: los que prorrogan una intervención dispuesta por un DNU,
    los reglamentarios de una ley sancionada por DNU, los vetos que la citan.
    En la ventana auditada eso daba 48 donde había 37 (ADR-0241).

    El tipo jurídico está en la propia grilla: InfoLeg rotula `Decreto DNU 771 /
    2026` frente a `Decreto 710 / 2026` o `Decreto Reglamentario 58 / 2026`. La
    búsqueda de texto se conserva porque acota la grilla —sin ella habría que
    traer todos los decretos del año— pero ya no decide."""
    items = _infoleg_listado_completo(session, action_url, "2", desde, hasta,
                                      texto="necesidad y urgencia")
    dnus = [it for it in items if _RE_INFOLEG_DNU.match(it["norma"])]
    return len(dnus), dnus


# ── Sector privado: brecha de expectativas obra pública vs. privada ───────────
# INDEC publica la Encuesta Cualitativa de la Construcción dentro del ISAC. El
# Cuadro 7.1 pregunta a las grandes empresas constructoras cómo esperan que
# evolucione su actividad en los próximos tres meses, y —esto es lo que lo hace
# útil acá— responde por separado para OBRA PRIVADA y OBRA PÚBLICA.
#
# Se puntúa la BRECHA (pública − privada), no el nivel. Las dos submuestras son
# el mismo sector, con el mismo costo de insumos y el mismo ciclo macro: lo
# único que las distingue es quién les paga. La diferencia entre ambas aísla el
# componente que viene del Estado, depurado del ciclo económico.
#
# Que eso no es una racionalización se ve en la serie: la brecha marca 2024
# como el peor momento de los diez años disponibles (−29,8), mientras que el
# nivel de obra pública marca 2019 (−58,2), que fue la recesión de Macri y no un
# conflicto entre el gobierno y sus contratistas. La brecha separa las dos cosas
# y el nivel no (ADR-0088).

ISAC_URL = "https://www.indec.gob.ar/ftp/cuadros/economia/sh_isac_{anio}.xls"
ISAC_HOJA = "Cuadro 7.1"
ISAC_MESES = {"enero": 1, "febrero": 2, "marzo": 3, "abril": 4, "mayo": 5,
              "junio": 6, "julio": 7, "agosto": 8, "septiembre": 9,
              "setiembre": 9, "octubre": 10, "noviembre": 11, "diciembre": 12}
ISAC_VENTANA = 12   # meses del promedio móvil


def _isac_descargar() -> bytes:
    """El archivo rota de nombre cada enero (`sh_isac_<año>.xls`), así que se
    prueba el año corriente y se cae al anterior.

    INDEC devuelve **HTTP 200 con HTML** cuando el archivo no existe (soft-404),
    de modo que el status code no alcanza para saber si vino un XLS: se valida
    el content-type."""
    intentos = []
    for anio in (date.today().year, date.today().year - 1):
        url = ISAC_URL.format(anio=anio)
        try:
            r = requests.get(url, headers=HTTP_HEADERS, timeout=HTTP_TIMEOUT)
            tipo = r.headers.get("content-type", "")
            if r.status_code == 200 and "excel" in tipo.lower():
                return r.content
            intentos.append(f"{anio}: HTTP {r.status_code} content-type={tipo!r}")
        except Exception as e:
            intentos.append(f"{anio}: {e}")
    raise ValueError("ISAC no descargable — " + " · ".join(intentos))


def brecha_obra_publica_serie() -> list:
    """Serie mensual de la brecha de expectativas, promedio móvil de 12 meses.

    Es la ÚNICA implementación del cálculo: la card de `fetch_brecha_obra_publica`
    devuelve el último punto de esta misma lista. Card y serie no pueden
    divergir porque son el mismo número, que es la propiedad que el gate G3
    verifica y que dos bugs de esta jornada (ADR-0086/0087) violaron por tener
    el cálculo escrito dos veces.

    [[YYYY-MM-01, brecha]] ascendente."""
    import xlrd

    ws = xlrd.open_workbook(file_contents=_isac_descargar()).sheet_by_name(ISAC_HOJA)
    # Columnas: 1/3 = privada aumentará/disminuirá · 5/7 = pública. La 4 y la 8
    # son separadoras vacías, no datos.
    mensual = []
    for i in range(ws.nrows):
        etiqueta = str(ws.cell(i, 0).value)
        # La fila rotula la ventana completa ("Junio 2026 - agosto 2026"); el
        # dato corresponde al mes de CIERRE, que es el segundo.
        m = re.search(r"-\s*([a-zA-Zéí]+)\s+(\d{4})", etiqueta)
        if not m or m.group(1).lower() not in ISAC_MESES:
            continue
        try:
            priv = float(ws.cell(i, 1).value) - float(ws.cell(i, 3).value)
            publ = float(ws.cell(i, 5).value) - float(ws.cell(i, 7).value)
        except (ValueError, TypeError):
            continue
        ym = (int(m.group(2)), ISAC_MESES[m.group(1).lower()])
        mensual.append((ym, publ - priv))

    # El cuadro trae filas de notas al pie que también matchean el patrón de
    # fecha; deduplicar por mes y ordenar deja sólo la grilla real.
    por_mes = dict(mensual)
    meses = sorted(por_mes)
    if len(meses) < ISAC_VENTANA:
        raise ValueError(f"ISAC {ISAC_HOJA}: {len(meses)} meses, se necesitan {ISAC_VENTANA}")

    out = []
    for i in range(ISAC_VENTANA - 1, len(meses)):
        ventana = [por_mes[meses[j]] for j in range(i - ISAC_VENTANA + 1, i + 1)]
        y, mm = meses[i]
        out.append([f"{y}-{mm:02d}-01", round(sum(ventana) / ISAC_VENTANA, 1)])
    return out


def fetch_brecha_obra_publica() -> dict | None:
    """
    Brecha de expectativas entre empresas constructoras de obra pública y de
    obra privada (INDEC, Encuesta Cualitativa de la Construcción, Cuadro 7.1),
    en promedio móvil de 12 meses.

    Saldo de respuesta = %"aumentará" − %"disminuirá", por submuestra; la brecha
    es el saldo de obra pública menos el de obra privada. Más negativa = las
    empresas que dependen del Estado esperan peor que sus pares privadas =
    mayor tensión con el gobierno.

    Dimensión: sector privado (ADR-0088).
    """
    try:
        serie = brecha_obra_publica_serie()
        fecha, valor = serie[-1]
        anterior = serie[-13][1] if len(serie) > 13 else None
        return {
            "valor":          valor,
            "unidad":         "pp de brecha (obra pública − privada, 12m)",
            "fuente":         "INDEC · Encuesta Cualitativa de la Construcción (ISAC, Cuadro 7.1)",
            "fecha_dato":     fecha,
            "desactualizado": False,
            "variacion_12m":  None if anterior is None else round(valor - anterior, 1),
            "detalle_txt": (
                f"Brecha de {str(valor).replace('.', ',')} pp entre lo que esperan las "
                f"constructoras de obra pública y las de obra privada (promedio de 12 meses). "
                f"Negativa = las que dependen del Estado esperan peor que sus pares privadas."),
        }
    except Exception as e:
        _warn("brecha_obra_publica", str(e))
        return None


def fetch_ratio_dnu() -> dict | None:
    """
    Ratio DNU = count(DNUs) / count(leyes sancionadas) en los últimos 365 días
    (ventana móvil, ADR-0058 — antes era acumulado del año calendario en
    curso, resetado cada 1° de enero, igual que el defecto que hizo sacar a
    movilizacion_cepa del tablero por ADR-0052; el criterio no se le había
    aplicado a este indicador hasta esta revisión).
    Mayor ratio = mayor dependencia del decreto → debilidad legislativa y exposición judicial.
    Dimensión: capacidad legislativa del Ejecutivo (Luis Babino: Agregados de Poder).

    Los dos lados usan la MISMA convención jurídica: **publicación en el
    Boletín Oficial** (ADR-0241). Mezclar leyes sancionadas con DNU publicados
    compararía dos momentos distintos del trámite.

    Fuente: servicios.infoleg.gob.ar
    - Leyes: tipoNorma=1 (Ley), por fecha de publicación
    - DNUs: tipoNorma=2 (Decreto) acotado con texto="necesidad y urgencia" y
      filtrado por el tipo que declara la grilla (`Decreto DNU`). La frase sola
      no alcanza: la dicen también decretos que no son DNU.
    Requiere GET previo para obtener jsessionid del formulario.
    """
    try:
        hasta = date.today()
        desde = hasta - timedelta(days=365)

        session = requests.Session()
        r_home = session.get(INFOLEG_HOME, headers=HTTP_HEADERS, timeout=HTTP_TIMEOUT)
        r_home.raise_for_status()

        action_m = re.search(r'action="(/infolegInternet/[^"]+)"', r_home.text)
        if not action_m:
            raise ValueError("No se encontró la URL del formulario InfoLeg")
        action_url = "https://servicios.infoleg.gob.ar" + action_m.group(1)

        leyes = _infoleg_session_count(session, action_url, "1", desde, hasta)
        if leyes == 0:
            raise ValueError("0 leyes — posible fallo en búsqueda InfoLeg (tipoNorma=1)")

        dnus, inventario = _infoleg_contar_dnus(session, action_url, desde, hasta)

        ratio = round(dnus / leyes, 3)

        return {
            "valor": ratio,
            "dnu_count": dnus,
            "leyes_count": leyes,
            "ventana_dias": 365,
            "unidad": "DNUs publicados por ley publicada",
            "fuente": "InfoLeg — Ministerio de Justicia (base de normas nacionales)",
            "fecha_dato": str(date.today()),
            "desactualizado": False,
            "ventana_desde": desde.isoformat(),
            "ventana_hasta": hasta.isoformat(),
            "inventario_dnu": [{"norma": it["norma"].split(" PODER")[0].strip(),
                                "fecha_pub": it["fecha_pub"]} for it in inventario],
            "detalle_txt": (
                f"{dnus} DNU y {leyes} leyes publicados en el Boletín Oficial "
                f"entre {desde.isoformat()} y {hasta.isoformat()} → {ratio:.2f} "
                f"DNU por ley"),
        }

    except Exception as e:
        _warn("ratio_dnu", str(e))
        return None


# ── CEPA conflictividad ───────────────────────────────────────────────────────

_RE_CEPA_FECHA = re.compile(r'property="datePublished"\s+content="(\d{4}-\d{2}-\d{2})')


def _fecha_informe_cepa(html: str) -> str:
    """Fecha real de publicación del informe (meta datePublished), NO la
    fecha de la corrida del scraper -- necesaria para el backfill (cada
    informe histórico debe fechar SU período, no 'hoy'). Fallback a hoy si
    el informe no trae el tag (no esperado: confirmado presente en los 4
    informes históricos verificados en vivo, 2026-07-08)."""
    m = _RE_CEPA_FECHA.search(html)
    return m.group(1) if m else str(date.today())


def _extraer_cifra_cepa(html: str) -> dict | None:
    """Cifra de conflictividad de UN informe CEPA (índice 0-100
    normalizado), con el mismo regex que fetch_cepa_movilizacion() usa para
    el informe vigente. Reconoce dos patrones textuales, mutuamente
    excluyentes, cada uno con su propia escala de normalización:

      - "X casos por mes" / "promedio de X casos mensuales" -> TASA mensual
        (rama "m_mes", escala 0-CEPA_MAX_CASOS_MES).
      - "al menos N conflictos" / "se registraron N conflictos" -> CONTEO
        acumulado (rama "m_tot", escala 0-CEPA_MAX_CONFLICTOS_TOT).

    El dict devuelto incluye "rama": "m_mes" o "m_tot" para que el llamador
    pueda distinguir de forma explícita qué patrón matcheó (en vez de
    inspeccionar el string "metrica"). Devuelve None si el HTML no matchea
    ninguno de los dos patrones.

    Esta función NO sabe nada sobre qué informes deben excluirse de una
    serie histórica ni por qué -- eso es responsabilidad de cada llamador.
    En particular, un informe puede citar una cifra bajo un ancla temporal
    distinta a la del indicador vigente (p.ej. acumulado "desde enero 2024"
    en vez de "desde inicios del año en curso") y aun así matchear (rama
    m_mes o m_tot) si el patrón textual está presente en algún lugar del
    HTML -- ver docstring de fetch_cepa_movilizacion_serie() en
    descargar_series.py para el filtro real que excluye lecturas no
    comparables de la serie."""
    m_mes = re.search(
        r"(\d+(?:[.,]\d+)?)\s+casos?\s+por\s+mes"
        r"|promedio\s+de\s+(\d+(?:[.,]\d+)?)\s+casos?\s+mensuales?",
        html, re.IGNORECASE
    )
    m_tot = re.search(
        r"(?:al menos,?\s+|se registraron,?\s+al menos,?\s+|se registraron\s+)"
        r"(\d+)\s+conflictos?",
        html, re.IGNORECASE
    )
    if m_mes:
        raw = (m_mes.group(1) or m_mes.group(2)).replace(",", ".")
        cifra = float(raw)
        return {"valor": round(min(100.0, (cifra / CEPA_MAX_CASOS_MES) * 100.0), 1),
                "cifra_cruda": cifra, "metrica": f"{cifra} casos/mes", "rama": "m_mes"}
    if m_tot:
        cifra = float(m_tot.group(1))
        return {"valor": round(min(100.0, (cifra / CEPA_MAX_CONFLICTOS_TOT) * 100.0), 1),
                "cifra_cruda": cifra, "metrica": f"{cifra} conflictos acumulados", "rama": "m_tot"}
    return None


def fetch_cepa_movilizacion() -> dict | None:
    """
    Conflictividad social CEPA — índice 0–100 normalizado.
    Estrategia: listar centrocepa.com.ar/informes → encontrar el último informe
    con "conflictividad" en la URL → parsear HTML del informe buscando
    "X casos por mes" o "al menos N conflictos".
    Dimensión: conflicto social (Matus).
    """
    try:
        from bs4 import BeautifulSoup
    except ImportError:
        _warn("movilizacion_cepa", "beautifulsoup4 no disponible")
        return None

    try:
        # La sección de informes está paginada (start=0, 10, 20...).
        # Buscar en hasta 5 páginas (50 informes) para encontrar el más reciente
        # con "conflictividad" o "conflictos-laborales" en la URL.
        links = []
        for page in range(5):
            page_url = CEPA_INFORMES_URL if page == 0 else f"{CEPA_INFORMES_URL}?start={page * 10}"
            r = requests.get(page_url, headers=HTTP_HEADERS, timeout=HTTP_TIMEOUT)
            r.raise_for_status()
            soup = BeautifulSoup(r.text, "html.parser")
            page_links = [
                a for a in soup.find_all("a", href=True)
                if any(kw in a.get("href", "").lower() for kw in ("conflictividad", "conflictos-laborales"))
            ]
            links.extend(page_links)
            if links:
                break

        if not links:
            raise ValueError("No se encontraron links de conflictividad en las primeras 5 páginas de informes CEPA")

        def url_num(a):
            m = re.search(r"/(\d+)[/-]", a["href"])
            return int(m.group(1)) if m else 0

        links.sort(key=url_num, reverse=True)
        href = links[0]["href"]
        informe_url = ("https://centrocepa.com.ar" + href) if href.startswith("/") else href

        r2 = requests.get(informe_url, headers=HTTP_HEADERS, timeout=HTTP_TIMEOUT)
        r2.raise_for_status()

        cifra_info = _extraer_cifra_cepa(r2.text)
        if cifra_info is None:
            raise ValueError(f"No se encontró patrón de conflictividad en {informe_url}")
        if cifra_info["rama"] != "m_tot":
            # Hallazgo de revisión final (2026-07-08): el informe más reciente
            # puede matchear la rama m_mes (tasa mensual, escala 0-80) en vez de
            # m_tot (conteo acumulado "desde inicios del año", escala 0-200) --
            # ver fetch_cepa_movilizacion_serie() para el caso real (informe
            # 748). Si eso pasara acá, BANDAS_ITCP["movilizacion_cepa"] (0-200)
            # puntuaría mal una tasa mensual. Se trata igual que "sin patrón":
            # degrada a cache en vez de publicar un valor en la escala
            # equivocada.
            raise ValueError(
                f"informe más reciente ({informe_url}) es rama={cifra_info['rama']!r}, "
                "no comparable con la escala m_tot ya publicada"
            )

        return {
            **cifra_info,
            "unidad": "Índice (0–100)",
            "fuente": informe_url,
            "fecha_dato": _fecha_informe_cepa(r2.text),
            "desactualizado": False,
        }

    except Exception as e:
        _warn("movilizacion_cepa", str(e))
        return None


# ── Protestas CABA (ACLED) — reutilizado de gestión ──────────────────────────

def fetch_protestas_caba() -> dict | None:
    """Reutiliza el fetcher ACLED ya construido en gestion.py (ADR-0017) para
    eventos de protesta en CABA. En gestión es CONTEXTO y no puntúa
    ('premiaría menos marchas'); en política fue indicador del ITCP hasta
    ADR-0048 (revisión editorial: "no pertinente en este cinturón") — sigue
    relevándose como seguimiento interno oculto del snapshot."""
    return gestion.fetch_protestas_caba()


# ── Conflictividad social nacional (ACLED, país entero) ──────────────────────

STALE_CONFLICTIVIDAD_DAYS = 30   # el agregado ACLED es semanal con rezago corto
CONFLICTOS_LABORALES_URL = (
    "https://www.argentina.gob.ar/trabajo/estadisticas/relaciones-laborales/"
    "conflictos-laborales")
CONFLICTOS_LABORALES_XLSX_FALLBACK = (
    "https://www.argentina.gob.ar/sites/default/files/"
    "evolucion_mensual_de_la_conflictividad_laboral._datos_a_mayo_2026.xlsx")


def fetch_conflictividad_nacional() -> dict | None:
    """
    Conflictividad social nacional (ADR-0052): eventos de protesta y
    disturbios (Protests+Riots, ACLED) en TODO el país, acumulados en 12
    meses completos, expresados como % de variación contra el total 2023
    (la base del mandato). Reemplaza a movilizacion_cepa como el indicador
    puntuante de la dimensión conflicto_social del ITCP.

    "valor" ES la variación % (puntúa directo sobre BANDAS_ITCP, sin el
    caso especial que necesitaba protestas_caba, que exponía el conteo
    crudo). El conteo y la base quedan en el detalle (acum_12m,
    eventos_2023).

    El último mes del archivo se excluye si está parcial (la semana final
    no llega a fin de mes — el corte semanal + rezago de carga de ACLED lo
    dejan incompleto) — mismo criterio que gestion.fetch_protestas_caba.
    La serie usable arranca en dic-2023 (primera ventana 12m íntegramente
    comparable con la base); la cobertura ACLED pre-2020 NO es confiable
    (2019 promedia 102 eventos/mes vs 240 de 2020 — artefacto de expansión
    de cobertura, ver ADR-0052) y no se usa ni para calibrar ni para el
    gráfico público.
    """
    try:
        store = gestion.actualizar_protestas_caba()
        if store is None and gestion.PROTESTAS_STORE_PATH.exists():
            store = json.loads(gestion.PROTESTAS_STORE_PATH.read_text(encoding="utf-8"))
        if not store or "mensual_nacional" not in store:
            raise ValueError("store ACLED sin serie nacional (¿corrida vieja sin ADR-0052?)")
        mensual = store["mensual_nacional"]
        hasta = store.get("_meta", {}).get("hasta_semana", "")
        yms = sorted(mensual)
        if hasta and yms and hasta[:7] == yms[-1]:
            import calendar as _cal
            a, m = int(hasta[:4]), int(hasta[5:7])
            if int(hasta[8:10]) < _cal.monthrange(a, m)[1]:
                yms = yms[:-1]
        if len(yms) < 12:
            raise ValueError("serie ACLED nacional demasiado corta")
        ult12 = sum(mensual[ym] for ym in yms[-12:])
        base_2023 = sum(v for ym, v in mensual.items() if ym.startswith("2023"))
        if not base_2023:
            raise ValueError("store ACLED sin base 2023")
        var = round((ult12 / base_2023 - 1.0) * 100.0, 1)
        return {
            "valor":          var,
            "acum_12m":       ult12,
            "eventos_2023":   base_2023,
            "unidad":         "% vs 2023",
            "fuente":         "ACLED — agregado semanal por provincia (acleddata.com)",
            "fecha_dato":     f"{yms[-1]}-01",
            "desactualizado": bool(hasta) and _days_old(str(hasta)) > STALE_CONFLICTIVIDAD_DAYS,
            "detalle_txt": (f"{ult12} eventos de protesta y disturbios en el país en 12m "
                            f"(hasta {yms[-1]}) vs {base_2023} en todo 2023 "
                            + f"({var:+.1f}%)".replace(".", ",")
                            + " — cuenta marchas, concentraciones y disturbios de ACLED en "
                              "las 24 jurisdicciones; CABA es ~9% del total del país"),
        }
    except Exception as e:
        _warn("conflictividad_nacional", str(e))
        return None


def _columna_jornadas_de_paro(ws) -> int:
    """Columna del total de jornadas de paro del cuadro C1, buscada por rótulo.

    El cuadro apila tres grupos —conflictos con paro, huelguistas y jornadas—
    y cada uno abre sus propias columnas de total y de sector. Tomar la columna
    por posición fija no falla el día que la Secretaría de Trabajo agrega o
    reordena una: devuelve otra magnitud, con un valor plausible que ningún
    control de frescura ni de coherencia card/serie puede distinguir.
    """
    filas = list(ws.iter_rows(min_row=1, max_row=8, values_only=True))
    for i, fila in enumerate(filas[1:], start=1):   # la fila 1 es el título del cuadro
        for j in range(1, len(fila)):               # la columna 0 es el período
            if "jornada" not in str(fila[j] or "").strip().lower():
                continue
            fin = next((k for k in range(j + 1, len(fila))
                        if str(fila[k] or "").strip()), len(fila))
            sub = filas[i + 1] if i + 1 < len(filas) else ()
            for k in range(j, min(fin, len(sub))):
                if str(sub[k] or "").strip().lower().startswith("total"):
                    return k
            if fin - j == 1:
                return j
            raise ValueError("conflictos laborales: el grupo de jornadas de paro "
                             "no declara columna de total")
    raise ValueError("conflictos laborales: columna de jornadas de paro no encontrada")


def fetch_jornadas_individuales_no_trabajadas_serie() -> list:
    """Serie rolling-12 de jornadas individuales no trabajadas, total país.

    La Secretaría de Trabajo define la métrica como huelguistas por duración
    del paro y declara expresamente que, a diferencia de conflictos y
    huelguistas, sus valores mensuales sí pueden sumarse para períodos largos.
    Se descubre el XLSX vigente desde la página oficial; el enlace directo se
    conserva sólo como fallback ante cambios transitorios del HTML.
    """
    url_xlsx = CONFLICTOS_LABORALES_XLSX_FALLBACK
    try:
        r_page = requests.get(CONFLICTOS_LABORALES_URL, headers=HTTP_HEADERS,
                              timeout=HTTP_TIMEOUT)
        r_page.raise_for_status()
        soup = BeautifulSoup(r_page.text, "html.parser")
        hrefs = [a.get("href", "") for a in soup.find_all("a")]
        candidato = next((h for h in hrefs
                          if h.lower().endswith(".xlsx")
                          and "evolucion_mensual" in h.lower()), None)
        if candidato:
            from urllib.parse import urljoin
            url_xlsx = urljoin(CONFLICTOS_LABORALES_URL, candidato)
    except requests.RequestException:
        # La planilla directa puede seguir disponible aunque falle la página
        # índice; el fallback permite conservar esa independencia.
        pass

    r = requests.get(url_xlsx, headers=HTTP_HEADERS, timeout=HTTP_TIMEOUT * 3)
    r.raise_for_status()
    # La planilla oficial trae rangos de impresión inválidos en varias hojas;
    # openpyxl los ignora correctamente, pero emite catorce warnings por
    # corrida. No son una condición del dato y ensucian el log operativo.
    with warnings.catch_warnings():
        warnings.filterwarnings("ignore", message="Print area cannot be set")
        wb = openpyxl.load_workbook(io.BytesIO(r.content), read_only=True,
                                    data_only=True)
    ws = next((wb[s] for s in wb.sheetnames
               if str(wb[s].cell(1, 1).value or "").startswith("C1.")), None)
    if ws is None:
        raise ValueError("conflictos laborales: cuadro C1 no encontrado")

    col = _columna_jornadas_de_paro(ws)
    mensual = []
    for row in ws.iter_rows(values_only=True):
        fecha = row[0] if row else None
        valor = row[col] if len(row) > col else None
        if not hasattr(fecha, "year") or not isinstance(valor, (int, float)):
            continue
        mensual.append((f"{fecha.year:04d}-{fecha.month:02d}", int(valor)))
    mensual.sort()
    if len(mensual) < 12:
        raise ValueError("conflictos laborales: menos de 12 meses de jornadas")
    return [[f"{mensual[i][0]}-01", sum(v for _, v in mensual[i - 11:i + 1])]
            for i in range(11, len(mensual))]


def fetch_jornadas_individuales_no_trabajadas() -> dict | None:
    """Intensidad laboral oficial acumulada en los últimos doce meses."""
    try:
        serie = fetch_jornadas_individuales_no_trabajadas_serie()
        fecha, valor = serie[-1]
        return {
            "valor": valor,
            "unidad": "jornadas individuales no trabajadas (12m)",
            "fuente": "Secretaría de Trabajo — Estadísticas de conflictos laborales",
            "fecha_dato": fecha,
            "desactualizado": False,
            "detalle_txt": (
                f"{valor:,}".replace(",", ".")
                + " jornadas en los últimos 12 meses. La fuente las calcula como "
                  "huelguistas × duración del paro y permite sumar los meses; "
                  "mide intensidad laboral, no cantidad de protestas."),
        }
    except Exception as e:
        _warn("jornadas_individuales_no_trabajadas_12m", str(e))
        return None


# ── IAF — Índice de Armonía Federal (transferencias) ─────────────────────────

def fetch_iaf_transferencias() -> dict | None:
    """
    Variación real i.a. de las transferencias federales (RON Hacienda).
    Dimensión: armonía fiscal federal (Luis Babino: Agregados de Poder — IAF).

    Universo (ADR-0066): lo girado a jurisdicciones —Provincias, C.A.B.A. y
    Fondo Compensador, incluida la compensación del Consenso Fiscal—. Quedan
    afuera Tesoro Nacional, Seguridad Social y Fondo A.T.N., que no salen de la
    Nación.

    Deflactor (ADR-0239): **cada flujo mensual a precios de su propio mes**. El
    método anterior dividía el cociente de dos sumas nominales por un único IPC
    promedio anual, y eso subdeflacta cuando el gasto no se reparte parejo por
    el calendario: para 2025 publicaba +0,8% donde IARAF y Politikon informaban
    +1,6/1,7%. Mes a mes da +1,64%.

    La serie es de transferencias EJECUTADAS del año calendario: el año de
    referencia es el último cerrado, no el presupuesto del siguiente.
    """
    try:
        year_ref = date.today().year - 1   # último año completo
        por_anio = _iaf_real_por_anio()
        if year_ref not in por_anio:
            disponibles = sorted(por_anio)
            if not disponibles:
                raise ValueError("sin años completos en la planilla mensual de RON")
            year_ref = disponibles[-1]
        var_real, var_nominal, deflactor, tot_ref, tot_ant = por_anio[year_ref]

        return {
            "valor": round(var_real * 100.0, 1),
            "var_nominal_pct": round(var_nominal * 100.0, 1),
            "total_ref_mm": round(tot_ref / 1e6, 0),
            "total_ant_mm": round(tot_ant / 1e6, 0),
            "periodo": f"{year_ref} vs {year_ref - 1}",
            "ipc_aplicado_pct": round(deflactor * 100.0, 1),
            "unidad": "% interanual real",
            "fuente": ("Sec. Hacienda — RON, planilla mensual consolidada + IPC "
                       "INDEC (deflactado mes a mes)"),
            "detalle_txt": (
                f"{year_ref}: {var_nominal * 100:+.1f}% nominal contra un deflactor "
                f"de {deflactor * 100:.1f}% ponderado por el flujo de cada mes "
                f"→ {var_real * 100:+.1f}% real"
            ),
            # La fecha del DATO es el cierre del año de referencia, no la de la
            # corrida. Antes acá iba `date.today()`: la card se mostraba fresca
            # todos los días mientras comparaba dos años calendario cerrados —en
            # julio de 2026 estaba informando 2025 contra 2024— y G2 no podía
            # avisar nada porque la fecha declaraba hoy. Con la fecha real, el
            # rezago queda a la vista y su tolerancia se declara en
            # gate_calidad.MAX_DIAS, que es donde va el criterio por indicador.
            "fecha_dato": f"{year_ref}-12-31",
            "desactualizado": False,
        }

    except Exception as e:
        _warn("iaf_transferencias", str(e))
        return None


# ── HCDN CKAN — eficacia legislativa ─────────────────────────────────────────

def _hcdn_paginate(resource_id: str, *, q: str = "") -> list[dict]:
    """Fetch all records from a CKAN datastore resource, handling pagination."""
    records: list[dict] = []
    offset = 0
    while True:
        params: dict = {"resource_id": resource_id, "limit": 500, "offset": offset}
        if q:
            params["q"] = q
        r = requests.get(HCDN_CKAN, params=params, headers=HTTP_HEADERS, timeout=HTTP_TIMEOUT)
        r.raise_for_status()
        result = r.json().get("result")
        if not result:
            raise ValueError(f"CKAN sin result para resource_id={resource_id[:8]}…")
        batch = result.get("records", [])
        records.extend(batch)
        if len(batch) < 500:
            break
        offset += 500
    return records


def _leyes_sancionadas_ids(hasta: str | None = None) -> set[str]:
    """PROYECTO_IDs con sanción definitiva según el dataset oficial
    leyes-sancionadas de HCDN (cada fila trae número de ley, fecha de
    sanción definitiva y cámara sancionadora — cubre las sanciones que
    ocurren en el SENADO, invisibles en movimientos-de-proyectos, que solo
    registra la vida del expediente en Diputados; ADR-0062). Reemplaza a la
    detección por texto de movimiento (q='SANCION' + filtro de medias
    sanciones): el registro de leyes trae solo sanciones definitivas por
    construcción, así que las ambigüedades de etiquetas ('TEXTO DE LA MEDIA
    SANCION' vs 'CONSIDERACION Y SANCION', auditoría 2026-07-09) dejan de
    existir para eficacia_legislativa.

    `hasta` (ISO YYYY-MM-DD) acota por SANCION_DEFINITIVA — lo usa la serie
    histórica para que un punto ya publicado no cambie retroactivamente
    cuando un proyecto se sanciona más tarde. Filas con fecha "NA" solo
    entran sin cota (no se puede verificar su timing)."""
    filas = _hcdn_paginate(HCDN_LEYES_SANC_RID)
    return {
        str(r["PROYECTO_ID"]).strip()
        for r in filas
        if r.get("PROYECTO_ID")
        and (hasta is None or str(r.get("SANCION_DEFINITIVA", ""))[:10] <= hasta)
    }


def fetch_eficacia_legislativa() -> dict | None:
    """
    % de proyectos ejecutivos MADUROS que se sancionaron alguna vez (ADR-0061,
    reemplaza a ADR-0050 — mismo defecto en espejo que motivó, ese mismo día,
    la corrección de ratio_dnu en ADR-0058/0059).

    ADR-0050 exigía que PUBLICACION_FECHA y la sanción cayeran en la MISMA
    ventana de 365 días: un proyecto recién enviado casi nunca alcanza a
    sancionarse todavía, así que el % quedaba estructuralmente deprimido. Esa
    ADR asumió que ni un Ejecutivo con mayoría sólida podría superar 35-55%,
    pero nunca lo verificó contra ningún caso real — verificado hoy contra
    Directorio Legislativo (tasas de éxito eventual: CFK 2do mandato 82%,
    Alberto Fernández primer año 67% con 63 días de trámite promedio): el
    trámite típico histórico (63-112 días) deja tiempo de sobra dentro de una
    ventana de 12 meses salvo para lo publicado en sus últimos 2-3 meses — el
    sesgo era real pero mucho menor al asumido.

    Cohorte MADURA (en vez de ventana compartida): proyectos PE publicados
    entre hoy−730 y hoy−365 días — un tramo de 12 meses de publicaciones,
    desplazado un año atrás para que CADA proyecto de la cohorte haya tenido
    AL MENOS 365 días de margen antes de evaluarlo. Elimina el sesgo de raíz
    en vez de compensarlo con anclas más generosas.

    Identificación PE: EXP_DIPUTADOS o EXP_SENADO con patrón NNNN-PE-AAAA o
    NNNN-JGM-AAAA (ADR-0063: el Presupuesto anual entra siempre por la
    Jefatura de Gabinete, art. 100 inc. 6 CN — sin la sigla JGM el indicador
    era ciego a la ley más importante del año), y TIPO con "PROYECTO DE LEY"
    (ADR-0062: los TIPO "MENSAJE" a secas son comunicaciones administrativas
    — avisos de vetos, resoluciones, decisiones administrativas de JGM — que
    jamás pueden sancionarse y contaminaban el denominador).

    Aprobación: PROYECTO_ID presente en el dataset oficial leyes-sancionadas
    de HCDN (ADR-0062). El dataset movimientos-de-proyectos que se usaba
    antes solo registra la vida del expediente EN DIPUTADOS: para un
    proyecto con origen en Diputados la sanción definitiva ocurre en el
    SENADO y jamás aparece como movimiento "SANCION" — el numerador era
    ciego a toda ley sancionada por la cámara revisora (verificado: leyes
    27.783, 27.799 y 27.801, las tres PE de la cohorte vigente, las tres
    invisibles para la métrica anterior; el 0,0% publicado era en realidad
    18,8%).

    Fuente: datos.hcdn.gob.ar CKAN
      - proyectos-parlamentarios: 22b2d52c-7a0e-426b-ac0a-a3326c388ba6
      - leyes-sancionadas:        68dfd7f8-91f3-4ecf-aebf-a860d1ca1a98
    """
    try:
        hoy = date.today()
        cohorte_hasta = (hoy - timedelta(days=365)).isoformat()[:10]
        cohorte_desde = (hoy - timedelta(days=730)).isoformat()[:10]

        # q= del CKAN es búsqueda full-text por token: una consulta por sigla
        # (los sets de abajo deduplican solos)
        raw_pe = (_hcdn_paginate(HCDN_PROYECTOS_RID, q="-PE-")
                  + _hcdn_paginate(HCDN_PROYECTOS_RID, q="-JGM-"))
        pe_cohorte: set[str] = {
            r["PROYECTO_ID"]
            for r in raw_pe
            if cohorte_desde <= str(r.get("PUBLICACION_FECHA", ""))[:10] <= cohorte_hasta
            and "PROYECTO DE LEY" in str(r.get("TIPO", "")).upper()
            and (
                _RE_PE_EXP.search(r.get("EXP_DIPUTADOS", "") or "")
                or _RE_PE_EXP.search(r.get("EXP_SENADO", "") or "")
            )
        }
        if not pe_cohorte:
            raise ValueError("Sin proyectos de ley PE en la cohorte madura (hoy-730d a hoy-365d)")

        sancionados = _leyes_sancionadas_ids()

        aprobados = pe_cohorte & sancionados
        total     = len(pe_cohorte)
        count     = len(aprobados)
        pct       = round(count / total * 100.0, 1) if total else 0.0

        return {
            "valor":             pct,
            "aprobados_n":       count,
            "enviados_n":        total,
            "cohorte_desde":     cohorte_desde,
            "cohorte_hasta":     cohorte_hasta,
            "dias_madurado_min": 365,
            "unidad":            "% de proyectos",
            "fuente":            "datos.hcdn.gob.ar — proyectos-parlamentarios + leyes-sancionadas",
            "fecha_dato":        str(hoy),
            "desactualizado":    False,
        }

    except Exception as e:
        _warn("eficacia_legislativa", str(e))
        return None


# ── HCDN CKAN — veto por quórum ──────────────────────────────────────────────

# ── Sesiones de Diputados: qué cuenta como fracaso de quórum (ADR-0091) ──────
# Auditoría de los registros crudos del dataset de sesiones HCDN, 2026-07-19.
# El campo REUNION_TIPO tiene estos valores (Diputados, 2024-2026):
#
#   Especial                                    29   SESION_NO asignado  ·  11,8 h medianas
#   Minoría                                     11   SESION_NO = 0 en 11/11  ·  2,0 h
#   Informativa                                  4   SESION_NO asignado  ·   7,0 h
#   Informativa Art. 71 CN - Citada - Fracasada  2   SESION_NO = 0  ·  0,0 h
#   Preparatoria / Presupuesto / Homenaje        4
#
# "Minoría" es el fracaso de quórum: la sesión se convocó, esperó unas dos horas,
# nunca se constituyó y no recibió número de sesión. La versión anterior de este
# indicador buscaba la subcadena "fracasada", con lo cual (a) omitía las once
# sesiones caídas y (b) contaba como fracaso las dos informativas del art. 71 CN
# que no se realizaron — que duran 0,0 h y son otro fenómeno: el jefe de
# Gabinete que no concurre, no el quórum que no se junta.
#
# El denominador son las sesiones convocadas para tratar temas: especiales
# (incluidas continuación y homenaje) más las que quedaron en minoría. Quedan
# afuera las informativas, la preparatoria y la presentación de presupuesto, que
# no son instancias donde el oficialismo necesite juntar quórum para avanzar.

def _hcdn_sesiones_legislativas() -> list[tuple[str, bool]]:
    """[(YYYY-MM-DD, fracaso_de_quorum)] de Diputados, ascendente."""
    out = []
    for anio in range(2023, date.today().year + 1):
        for r in _hcdn_paginate(HCDN_SESIONES_RID, q=str(anio)):
            if str(r.get("SESION_CAMARA", "")).upper() != "DIPUTADOS":
                continue
            inicio = str(r.get("REUNION_INICIO") or "")[:10]
            if not inicio.startswith(str(anio)):
                continue
            # El dataset llega con mojibake ("MinorÃ­a"): se compara sobre el
            # prefijo sin tildes, que sobrevive a la codificación rota.
            tipo = str(r.get("REUNION_TIPO", "")).lower()
            es_minoria = tipo.startswith("minor")
            if tipo.startswith("especial") or es_minoria:
                out.append((inicio, es_minoria))
    return sorted(set(out))


def _veto_quorum_tasa_12m(sesiones: list, referencia: date):
    """(pct, total, fracasadas) en los 12 meses calendario que terminan en el
    mes de `referencia`. None si no hubo sesiones en la ventana — sin
    denominador no hay tasa y el motor renormaliza, igual que bloqueo_sostenido.

    La ventana móvil reemplaza al período legislativo (ADR-0091): con el
    período, el denominador arrancaba en cero cada marzo y el indicador
    publicaba puntaje máximo sobre dos o tres sesiones. Con 0 de 5 —el estado
    del 19-jul-2026— la tasa real de fracaso podía llegar al 60% con 95% de
    confianza, que puntúa 10; se estaba publicando 100."""
    meses = referencia.year * 12 + (referencia.month - 1)
    desde = meses - 11
    ym_desde = f"{desde // 12}-{desde % 12 + 1:02d}"
    ym_hasta = f"{referencia.year}-{referencia.month:02d}"
    en_ventana = [f for d, f in sesiones if ym_desde <= d[:7] <= ym_hasta]
    if not en_ventana:
        return None
    total = len(en_ventana)
    fracasadas = sum(1 for f in en_ventana if f)
    return round(fracasadas / total * 100.0, 1), total, fracasadas


def fetch_veto_quorum() -> dict | None:
    """
    % sesiones plenarias (Diputados) frustradas por falta de quórum en el período corriente.
    Detección: REUNION_TIPO contiene "Fracasada" en dataset de sesiones HCDN.
    Período corriente: PERIODO_ID con prefijo HCDN{periodo_num} (144 = 2026).
    Fórmula período: 144 + (año_actual − 2026).

    Nota: sesiones que nunca abren ("desactivadas") NO aparecen en HCDN — solo
    sesiones formalmente iniciadas y luego fracasadas por quórum son registradas.

    Fuente: datos.hcdn.gob.ar CKAN — sesiones (4ac70a51-...)
    Score: 0%→0, 15%→5, 30%+→10  (formula: valor / 3)
    """
    try:
        sesiones = _hcdn_sesiones_legislativas()
        tasa = _veto_quorum_tasa_12m(sesiones, date.today())
        if tasa is None:
            raise ValueError("sin sesiones legislativas en la ventana de 12 meses")
        pct, total_n, fracasadas_n = tasa
        return {
            "valor":        pct,
            "fracasadas_n": fracasadas_n,
            "total_n":      total_n,
            "unidad":       "% de sesiones",
            "fuente":       "Cámara de Diputados (datos abiertos) — sesiones",
            "fecha_dato":   str(date.today()),
            "desactualizado": False,
            "detalle_txt": (f"{fracasadas_n} de {total_n} sesiones legislativas convocadas en los "
                            f"últimos 12 meses quedaron en minoría (no reunieron quórum)"),
        }

    except Exception as e:
        _warn("veto_quorum", str(e))
        return None


# ── HCDN CKAN — comisiones caídas ─────────────────────────────────────────────

def fetch_comisiones_caidas() -> dict | None:
    """
    % proyectos con dictamen 'Orden del Día' en los últimos 12 meses que no fueron sancionados.
    Identifica proyectos con dictamen listo para el recinto (OD) pero sin SANCION en movimientos.

    Si dictámenes incluye PROYECTO_ID → join directo a movimientos.
    Si solo incluye EXPEDIENTE → map a PROYECTO_ID vía proyectos-parlamentarios.

    Fuente: datos.hcdn.gob.ar CKAN
      - dictámenes:           59595a93-5a5e-4ba6-a3db-c1044e2f949e
      - movimientos-de-proyectos: 6108ea83-3f12-423c-a136-df1ae9cb2972
      - proyectos-parlamentarios: 22b2d52c-7a0e-426b-ac0a-a3326c388ba6 (si hace falta join)

    Score: 20%→0, 40%→5, 60%+→10  (formula: (valor − 20) / 4)
    """
    try:
        year   = date.today().year
        cutoff = (date.today() - timedelta(days=365)).isoformat()[:10]

        # Dictámenes: dos pasadas para cubrir la ventana de 12 meses
        raw_cur  = _hcdn_paginate(HCDN_DICTAMENES_RID, q=str(year))
        raw_prev = _hcdn_paginate(HCDN_DICTAMENES_RID, q=str(year - 1))
        all_dict = raw_cur + raw_prev

        # Python-side: FECHA >= cutoff AND TIPO contiene "orden" (Orden del Día)
        od_records = [
            r for r in all_dict
            if str(r.get("FECHA", ""))[:10] >= cutoff
            and "orden" in str(r.get("TIPO", "")).lower()
        ]

        if not od_records:
            muestra = list(all_dict[0].keys()) if all_dict else []
            raise ValueError(
                f"Sin dictámenes 'Orden del Día' en ventana 12m "
                f"({len(all_dict)} registros totales, campos: {muestra})"
            )

        # dictámenes.EXPEDIENTE = HCDN project ID (same format as movimientos.PROYECTO_ID)
        od_ids: set[str] = {
            str(r["EXPEDIENTE"]).strip()
            for r in od_records if r.get("EXPEDIENTE")
        }

        if not od_ids:
            raise ValueError("Sin EXPEDIENTE válidos en dictámenes OD")

        # Sancionados en la ventana
        raw_san = _hcdn_paginate(HCDN_MOVIMIENTOS_RID, q="SANCION")
        sancionados: set[str] = {
            str(r["PROYECTO_ID"]).strip()
            for r in raw_san
            if str(r.get("FECHA", ""))[:10] >= cutoff and r.get("PROYECTO_ID")
        }

        aprobados_n = len(od_ids & sancionados)
        total_n     = len(od_ids)
        caidas_n    = total_n - aprobados_n
        pct         = round(caidas_n / total_n * 100.0, 1) if total_n else 0.0

        return {
            "valor":        pct,
            "dictamen_n":   total_n,
            "aprobados_n":  aprobados_n,
            "caidas_n":     caidas_n,
            "ventana_dias": 365,
            "unidad":       "% de proyectos",
            "fuente":       "datos.hcdn.gob.ar — dictámenes + movimientos-de-proyectos",
            "fecha_dato":   str(date.today()),
            "desactualizado": False,
        }

    except Exception as e:
        _warn("comisiones_caidas", str(e))
        return None


# ── Poder judicial — cobertura de cargos de juez (ADR-0126) ──────────────────

def _jus_csv(consulta: str) -> list[dict]:
    """Filas del primer recurso CSV del dataset que matchea `consulta`.

    La URL del CSV lleva la fecha de actualización en el nombre
    (`...-jueces-20260605.csv`) y cambia con cada publicación, así que se
    resuelve por API y NUNCA se arma a mano — mismo criterio que los informes
    de desregulación (ADR-0125).
    """
    import csv

    r = requests.get(JUS_API, params={"q": consulta, "rows": 5},
                     headers=HTTP_HEADERS, timeout=HTTP_TIMEOUT)
    r.raise_for_status()
    for paquete in r.json()["result"]["results"]:
        if not paquete["title"].lower().startswith(consulta.split(" de ")[0].lower()):
            continue
        for recurso in paquete.get("resources", []):
            if recurso.get("format", "").upper() != "CSV":
                continue
            rr = requests.get(recurso["url"], headers=HTTP_HEADERS, timeout=60)
            rr.raise_for_status()
            texto = rr.content.decode("utf-8-sig", "replace")
            return list(csv.DictReader(io.StringIO(texto)))
    raise ValueError(f"sin recurso CSV para '{consulta}'")


def _jus_fechas(filas: list[dict], campo: str, tipo: str = "Juez") -> list[str]:
    """Fechas ISO de los eventos de `tipo` (los registros mezclan jueces,
    fiscales y defensores; el padrón que ancla la serie es sólo de jueces)."""
    out = []
    for f in filas:
        if f.get("cargo_tipo") != tipo:
            continue
        v = (f.get(campo) or "").strip()
        if len(v) >= 10 and v[:2] == "20":
            out.append(v[:10])
    return sorted(out)


# ── Detector de novedades de postura empresaria (ADR-0149) ──────────────────
# Vigila las dos cámaras del registro de ADR-0148 y marca los comunicados
# nuevos como pendientes de codificar. NO puntúa y NO toca el ITCP: la postura
# la asigna una persona, y hasta que no haya segunda pasada con kappa ≥ 0,70 el
# indicador no se publica. Lo que esto evita es que el registro se quede viejo
# entre tanto — sin vigilancia, 103 comunicados codificados a mano dejan de
# servir apenas la cámara publica el siguiente.
#
# Mismo patrón que ADR-0129 (privatizaciones) y ADR-0141 (fallos de la CSJN):
# automatiza la omisión, no el juicio.
APOYO_NOVEDADES_PATH = PROJECT_DIR / "data" / "politica" / "apoyo_empresario_novedades.json"
APOYO_CODIFICACION_PATH = PROJECT_DIR / "data" / "politica" / "apoyo_empresario_codificacion.json"

# La UIA migró en agosto de 2026 de /prensa/{id}/ a un sitio Next con slugs.
# El HTML inicial trae el listado de novedades y cada ficha trae fecha, título
# y cuerpo en el stream RSC, por lo que no hace falta ejecutar JavaScript.
UIA_NOVEDADES_URL = "https://www.uia.org.ar/uia/novedades"
AEA_PRENSA_URL = "https://www.aeanet.net/prensa.html"

_MESES_AEA = {m: i for i, m in enumerate(
    ["enero", "febrero", "marzo", "abril", "mayo", "junio", "julio",
     "agosto", "septiembre", "octubre", "noviembre", "diciembre"], 1)}


def _limpio(html: str) -> str:
    return unescape(re.sub(r"<[^>]+>|\s+", " ", html)).strip()


def _uia_cuerpo(html: str) -> str:
    """Extrae el rich text completo del stream RSC del sitio Next de UIA."""
    bloques = re.findall(
        r'(?:\\u003cp\b.*?\\u003c/p\\u003e)+', html, re.S | re.I)
    if not bloques:
        return ""
    cuerpo = max(bloques, key=len)
    cuerpo = cuerpo.replace(r"\u003c", "<").replace(r"\u003e", ">")
    return _limpio(cuerpo.replace(r'\"', '"'))


def _uia_comunicado(url: str, session: requests.Session) -> dict | None:
    """Una novedad del sitio vigente de UIA, identificada por su slug."""
    r = session.get(url, timeout=HTTP_TIMEOUT)
    if r.status_code != 200:
        return None
    fecha = re.search(
        r'<meta\s+property="article:published_time"\s+content="(20\d{2}-\d{2}-\d{2})"',
        r.text, re.I)
    titulo = re.search(r'<meta\s+property="og:title"\s+content="([^"]+)"', r.text, re.I)
    if not (fecha and titulo):
        return None
    slug = url.rstrip("/").rsplit("/", 1)[-1]
    return {"camara": "UIA", "id": slug,
            "fecha": fecha.group(1), "titulo": unescape(titulo.group(1)),
            "texto": _uia_cuerpo(r.text),
            "url": url}


def _uia_comunicados(session: requests.Session, omitidos: set[str] | None = None) -> list[dict]:
    """Novedades visibles en el listado vigente; no re-fetch las ya vistas."""
    r = session.get(UIA_NOVEDADES_URL, timeout=HTTP_TIMEOUT)
    r.raise_for_status()
    slugs = sorted(set(re.findall(r'href="/uia/novedades/([^"/?#]+)', r.text)))
    salida, omitidos = [], omitidos or set()
    for slug in slugs:
        if slug in omitidos:
            continue
        url = f"{UIA_NOVEDADES_URL}/{slug}"
        try:
            c = _uia_comunicado(url, session)
        except Exception as e:
            print(f"  [WARN] apoyo/UIA {slug}: {e}")
            continue
        if c:
            salida.append(c)
    return salida


def _aea_comunicados(session: requests.Session) -> list[dict]:
    """Comunicados de AEA. La fecha vive en un bloque `post-meta` POSTERIOR al
    título, así que se parte por ahí en vez de buscar hacia adelante."""
    r = session.get(AEA_PRENSA_URL, timeout=HTTP_TIMEOUT)
    r.raise_for_status()
    html = r.text
    if "Ã" in html:                      # el sitio declara latin-1 y sirve utf-8
        html = html.encode("latin1", "ignore").decode("utf-8", "ignore")
    salida, partes = [], re.split(r'<div class="post-meta">', html)
    for i in range(len(partes) - 1):
        titulos = re.findall(r"<h2[^>]*>(.*?)</h2>", partes[i], re.S)
        fecha = re.search(r"(\d{1,2})\s+de\s+(\w+)\s+de\s+(20\d{2})",
                          partes[i + 1][:400], re.I)
        if not (titulos and fecha):
            continue
        mes = _MESES_AEA.get(fecha.group(2).lower())
        titulo = _limpio(titulos[-1])
        if not mes or len(titulo) < 8:
            continue
        iso = f"{fecha.group(3)}-{mes:02d}-{int(fecha.group(1)):02d}"
        salida.append({"camara": "AEA", "id": _apoyo_clave_aea(iso, titulo),
                       "fecha": iso, "titulo": titulo, "url": AEA_PRENSA_URL})
    return salida


def _apoyo_clave_aea(fecha_iso: str, titulo: str) -> str:
    """AEA no numera sus comunicados y publica más de uno el mismo día, así que
    la fecha sola no identifica: la clave lleva además el título normalizado."""
    slug = re.sub(r"[^a-z0-9]+", "-",
                  unicodedata.normalize("NFKD", titulo.lower())
                  .encode("ascii", "ignore").decode()).strip("-")
    return f"{fecha_iso}-{slug[:40]}"


def _apoyo_ya_codificados() -> set[str]:
    """Claves del registro de ADR-0148, para no re-avisar lo ya clasificado."""
    try:
        d = json.loads(APOYO_CODIFICACION_PATH.read_text(encoding="utf-8-sig"))
    except (OSError, json.JSONDecodeError):
        return set()
    claves = set()
    for c in d.get("casos", []):
        cam = c.get("camara", "AEA")
        if cam == "UIA":
            claves.add(f"UIA|{c.get('id')}")
        else:
            claves.add(f"AEA|{_apoyo_clave_aea(c['fecha'], c.get('titulo', ''))}")
    return claves


def _apoyo_firmas_codificadas() -> set[tuple[str, str]]:
    """Fecha+título evita reavisar notas migradas cuyo id viejo desapareció."""
    try:
        casos = json.loads(APOYO_CODIFICACION_PATH.read_text(
            encoding="utf-8-sig")).get("casos", [])
    except (OSError, json.JSONDecodeError):
        return set()
    return {(c.get("fecha", ""), _apoyo_clave_aea("", c.get("titulo", "")))
            for c in casos}


APOYO_DESDE = "2023-12"        # arranque del período (asunción Milei)


def apoyo_empresario_serie() -> list:
    """Saldo de postura empresaria hacia el Ejecutivo nacional, ventana móvil
    de 12 meses: (apoyos − críticas) / (apoyos + críticas).

    Es la ÚNICA implementación del cálculo: `fetch_apoyo_empresario` devuelve el
    último punto de esta misma lista, así que card y serie no pueden divergir
    (misma disciplina que `brecha_obra_publica_serie`; escribir la cuenta dos
    veces causó ADR-0086 y ADR-0087).

    Sólo cuentan los comunicados con destinatario `ejecutivo_nacional` y postura
    `apoyo` o `critica`: un neutro no toma posición y un dudoso no se pudo
    determinar, así que ninguno de los dos entra (reglas v2, ADR-0150). Un mes
    sin comunicados computables NO se rellena con cero — cero significa
    equilibrio entre apoyo y crítica, que es una afirmación distinta de «no se
    pronunció» — así que ese mes simplemente no tiene punto.

    [[YYYY-MM-01, saldo]] ascendente, saldo en [−1, +1].
    """
    d = json.loads(APOYO_CODIFICACION_PATH.read_text(encoding="utf-8-sig"))
    comp = [(c["fecha"], c["postura"]) for c in d["casos"]
            if c["destinatario"] == "ejecutivo_nacional"
            and c["postura"] in ("apoyo", "critica")]
    if not comp:
        raise ValueError("apoyo_empresario: el registro no tiene casos computables")

    y, m = map(int, APOYO_DESDE.split("-"))
    hoy = date.today()
    out = []
    while (y, m) <= (hoy.year, hoy.month):
        ini, fin = f"{y - 1:04d}-{m:02d}-01", f"{y:04d}-{m:02d}-31"
        v = [p for f, p in comp if ini <= f <= fin]
        a, c = v.count("apoyo"), v.count("critica")
        if a + c:
            out.append([f"{y:04d}-{m:02d}-01", round((a - c) / (a + c), 3)])
        m += 1
        if m == 13:
            y, m = y + 1, 1
    return out


def _apoyo_pendientes() -> int:
    """Cuántos comunicados detectados esperan codificación humana (ADR-0149).

    Este indicador es el único del cinturón cuyo dato lo actualiza una PERSONA,
    no el cron: si nadie codifica, la serie se congela sin que falle nada. Por
    eso el número va a la card, a la vista."""
    try:
        store = json.loads(APOYO_NOVEDADES_PATH.read_text(encoding="utf-8-sig"))
        return len(store.get("pendientes", {}))
    except (OSError, json.JSONDecodeError):
        return 0


def fetch_apoyo_empresario() -> dict | None:
    """
    Postura pública de AEA y UIA hacia el Ejecutivo nacional, saldo en ventana
    móvil de 12 meses (ADR-0150).

    Más negativo = las dos cámaras empresarias de referencia critican más de lo
    que apoyan = mayor tensión con el sector privado organizado.

    Fuente: comunicados institucionales fechados de aeanet.net/prensa.html y
    uia.org.ar/uia/novedades, codificados a mano con el protocolo de ADR-0131 —
    dos codificadores ciegos entre sí, kappa 1,000 (postura) y 0,955
    (destinatario), desacuerdos adjudicados por el autor del manual.

    Dimensión: sector privado (ADR-0088).
    """
    try:
        serie = apoyo_empresario_serie()
        fecha, valor = serie[-1]
        anterior = serie[-13][1] if len(serie) > 13 else None
        d = json.loads(APOYO_CODIFICACION_PATH.read_text(encoding="utf-8-sig"))
        comp = [c for c in d["casos"]
                if c["destinatario"] == "ejecutivo_nacional"
                and c["postura"] in ("apoyo", "critica")
                and c["fecha"] >= f"{int(fecha[:4]) - 1}{fecha[4:7]}-01"]
        apoyos = sum(1 for c in comp if c["postura"] == "apoyo")
        pend = _apoyo_pendientes()
        return {
            "valor":          valor,
            "unidad":         "saldo de postura (−1 a +1, 12m móviles)",
            # Sin número de ADR: este string se publica (G6).
            "fuente":         "Comunicados de AEA y UIA — codificación CIGOB",
            "fecha_dato":     fecha,
            "desactualizado": False,
            "variacion_12m":  None if anterior is None else round(valor - anterior, 3),
            "comunicados_ventana": len(comp),
            "apoyos_ventana":      apoyos,
            "criticas_ventana":    len(comp) - apoyos,
            "pendientes_de_codificar": pend,
            "detalle_txt": (
                f"En los últimos doce meses AEA y UIA se pronunciaron {len(comp)} veces "
                f"sobre medidas del Gobierno nacional: {apoyos} de apoyo y "
                f"{len(comp) - apoyos} de crítica. Saldo "
                # signo menos tipográfico, como en la unidad y en las anclas
                f"{str(valor).replace('-', '−').replace('.', ',')} en una escala de "
                f"−1 (todo crítica) a +1 (todo apoyo)."
                + (f" Hay {pend} comunicado{'s' if pend != 1 else ''} detectado"
                   f"{'s' if pend != 1 else ''} sin codificar." if pend else "")),
        }
    except Exception as e:
        _warn("apoyo_empresario", str(e))
        return None


def detectar_novedades_empresarias() -> dict:
    """Comunicados nuevos de UIA y AEA, pendientes de codificar.

    NO puntúa ni alimenta el ITCP (ADR-0149). Cada comunicado se avisa una sola
    vez: los ya codificados en `apoyo_empresario_codificacion.json` entran como
    revisados de arranque, y los nuevos quedan en `pendientes` hasta que alguien
    los saque.
    """
    try:
        store = json.loads(APOYO_NOVEDADES_PATH.read_text(encoding="utf-8-sig"))
    except (OSError, json.JSONDecodeError):
        store = {}
    revisadas = store.setdefault("revisadas", {})
    pendientes = store.setdefault("pendientes", {})
    for clave in _apoyo_ya_codificados():
        revisadas.setdefault(clave, {"origen": "codificacion ADR-0148"})

    session = requests.Session()
    session.headers.update(HTTP_HEADERS)
    nuevas, vistos = 0, 0
    firmas_codificadas = _apoyo_firmas_codificadas()

    def anotar(c: dict) -> None:
        nonlocal nuevas, vistos
        vistos += 1
        clave = f"{c['camara']}|{c['id']}"
        if clave in revisadas:
            return
        firma = (c.get("fecha", ""), _apoyo_clave_aea("", c.get("titulo", "")))
        if firma in firmas_codificadas:
            revisadas[clave] = {"fecha": c["fecha"], "titulo": c["titulo"][:160],
                                "origen": "migración de URL; ya codificado"}
            return
        revisadas[clave] = {"fecha": c["fecha"], "titulo": c["titulo"][:160]}
        pendientes[clave] = {**c, "postura": None, "destinatario": None,
                             "nota": "sin codificar — ver apoyo_empresario_reglas.json"}
        nuevas += 1

    try:
        omitidos = {k.split("|", 1)[1] for k in revisadas if k.startswith("UIA|")}
        for c in _uia_comunicados(session, omitidos):
            anotar(c)
    except Exception as e:
        print(f"  [WARN] apoyo/UIA: {e}")

    try:
        for c in _aea_comunicados(session):
            anotar(c)
    except Exception as e:
        print(f"  [WARN] apoyo/AEA: {e}")

    store["_meta"] = {
        "descripcion": ("Comunicados nuevos de UIA y AEA pendientes de codificar "
                        "(ADR-0149). Detección automática; la postura y el "
                        "destinatario los asigna una persona con las reglas de "
                        "apoyo_empresario_reglas.json. NO puntúa: el indicador no "
                        "se publica hasta que haya segunda pasada con kappa ≥ 0,70. "
                        "Sacar de 'pendientes' lo ya codificado."),
        "fuentes": {"UIA": UIA_NOVEDADES_URL, "AEA": AEA_PRENSA_URL},
        "ultima_corrida": date.today().isoformat(),
        "comunicados_vistos": vistos,
        "nuevos_en_la_corrida": nuevas,
    }
    APOYO_NOVEDADES_PATH.parent.mkdir(parents=True, exist_ok=True)
    APOYO_NOVEDADES_PATH.write_text(
        json.dumps(store, indent=1, ensure_ascii=False, sort_keys=True),
        encoding="utf-8")
    return store


# ── Detector de novedades judiciales de la CSJN (ADR-0140) ──────────────────
# Qué es y qué NO es: es un AVISO, no un contador. El endpoint abierto de la
# Secretaría de Jurisprudencia devuelve como máximo 10 registros por consulta y
# no pagina (verificado: jtStartIndex/jtPageSize, start/length y pagina no hacen
# efecto), así que con él no se puede construir una serie. Lo que sí puede es
# garantizar que no se escape un fallo relevante entre corrida y corrida. Mismo
# patrón que el detector de privatizaciones de ADR-0129: elimina el riesgo de
# omisión, no el juicio del analista.
#
# El buscador COMPLETO (fallos 1994-2026, con la casilla «Sentencias que
# declaran Inconstitucionalidad», búsqueda por partes y sentido del
# pronunciamiento) está detrás de CAPTCHA y no se toca. Ver ADR-0140 y el mapa
# de acceso en data/politica/csjn_jurisprudencia_mapa_de_acceso.json.
CSJN_SJ_BASE = "https://sjconsulta.csjn.gov.ar/sjconsulta"

# Términos que se barren. No pretenden ser exhaustivos: son las puertas de
# entrada a los tres fenómenos que el aporte externo pidió seguir.
CSJN_TERMINOS = (
    "inconstitucionalidad",
    "medida cautelar",
    "estado nacional",
    "amparo",
)

# El Estado como parte en la carátula. Deliberadamente amplio: un detector que
# sobre-avisa cuesta un vistazo del analista; uno que sub-avisa pierde el fallo
# y nadie se entera. Las formas salen de carátulas reales del propio endpoint
# ("EN-M ECONOMIA Y OTRO c/ BUNGE ARGENTINA SA", "SANTA CRUZ, PROVINCIA DE c/
# ESTADO NACIONAL", "BAEZ, CARMEN ALICIA c/ SECRETARIA NACIONAL...").
_CSJN_ESTADO = re.compile(
    r"\bEN\s*-|\bE\.\s?N\.|ESTADO\s+NACIONAL|PODER\s+EJECUTIVO|"
    r"\bANSES\b|\bAFIP\b|\bARCA\b|\bDGI\b|\bDGA\b|"
    r"MINISTERIO\s+DE|SECRETARIA\s+(?:DE|NACIONAL)|ADMINISTRACION\s+FEDERAL",
    re.I,
)


def _csjn_sesion() -> requests.Session:
    """Sesión con la cookie que el endpoint JSON exige (la entrega el GET previo)."""
    s = requests.Session()
    s.headers.update({**HTTP_HEADERS, "X-Requested-With": "XMLHttpRequest",
                      "Referer": f"{CSJN_SJ_BASE}/novedades/consulta.html"})
    s.get(f"{CSJN_SJ_BASE}/novedades/consulta.html", timeout=HTTP_TIMEOUT)
    return s


def _csjn_buscar(texto: str, session: requests.Session) -> list[dict]:
    """Consulta el módulo de novedades. Devuelve a lo sumo 10 registros."""
    r = session.post(f"{CSJN_SJ_BASE}/novedades/buscar.html",
                     data={"texto": texto}, timeout=HTTP_TIMEOUT)
    r.raise_for_status()
    payload = r.json()
    if payload.get("Result") != "OK":
        raise ValueError(f"CSJN respondió {payload.get('Result')}: "
                         f"{str(payload.get('Message'))[:80]}")
    return payload.get("Records") or []


def _csjn_fuero(identificador: str) -> str:
    """Prefijo del expediente: CAF = Contencioso Administrativo Federal, etc."""
    partes = (identificador or "").strip().split()
    return partes[0] if partes else ""


def detectar_novedades_judiciales(terminos: tuple[str, ...] = CSJN_TERMINOS) -> dict:
    """Fallos nuevos de la CSJN que declaran inconstitucionalidad o tienen al
    Estado como parte.

    NO puntúa, NO produce serie y NO alimenta ningún índice: el endpoint abierto
    topea en 10 registros por consulta (ADR-0140). Cada `idAnalisis` se evalúa
    una sola vez y queda anotado en `revisadas`; los que pasan el filtro van a
    `pendientes` para que el analista los lea. Sacar de `pendientes` lo revisado.
    """
    try:
        store = json.loads(CSJN_NOVEDADES_PATH.read_text(encoding="utf-8-sig"))
    except (OSError, json.JSONDecodeError):
        store = {}
    revisadas = store.setdefault("revisadas", {})
    pendientes = store.setdefault("pendientes", {})

    session = _csjn_sesion()
    nuevas, vistos = 0, 0
    for termino in terminos:
        try:
            registros = _csjn_buscar(termino, session)
        except Exception as e:
            print(f"  [WARN] csjn/novedades '{termino}': {e}")
            continue
        for reg in registros:
            vistos += 1
            id_analisis = str(reg.get("idAnalisis") or "").strip()
            if not id_analisis or id_analisis in revisadas:
                continue
            caratula = (reg.get("caratula") or "").strip()
            inconstitucional = reg.get("inconstitucional") is True
            estado_parte = bool(_CSJN_ESTADO.search(caratula))
            # el fallo publicado es inmutable: se evalúa una vez y el veredicto
            # queda cacheado, pase o no el filtro
            revisadas[id_analisis] = {"fecha": reg.get("fecha", ""),
                                      "termino": termino,
                                      "marcado": inconstitucional or estado_parte}
            if not (inconstitucional or estado_parte):
                continue
            motivos = []
            if inconstitucional:
                motivos.append("declara inconstitucionalidad")
            if estado_parte:
                motivos.append("el Estado es parte")
            pendientes[id_analisis] = {
                "fecha": reg.get("fecha", ""),
                "fuero": _csjn_fuero(reg.get("identificadorExpediente", "")),
                "expediente": reg.get("identificadorExpediente", ""),
                "caratula": caratula,
                "materia": reg.get("materia", ""),
                "titulo": (reg.get("titulo") or "")[:300],
                "inconstitucional": inconstitucional,
                "sentencia_arbitraria": reg.get("sentenciaArbitraria") is True,
                "motivos": motivos,
                "url": f"{CSJN_SJ_BASE}/fallos/getSintesisAnalisis.html"
                       f"?idAnalisis={id_analisis}",
            }
            nuevas += 1

    store["_meta"] = {
        "descripcion": ("Fallos de la CSJN que declaran inconstitucionalidad o "
                        "tienen al Estado como parte. Detección automática; la "
                        "lectura y la clasificación las hace el analista "
                        "(ADR-0140). NO es un contador: el endpoint abierto "
                        "topea en 10 registros por consulta y no pagina. Sacar "
                        "de 'pendientes' lo ya revisado."),
        "fuente": f"{CSJN_SJ_BASE}/novedades/buscar.html",
        "ultima_corrida": date.today().isoformat(),
        "registros_vistos_en_la_corrida": vistos,
        "nuevas_en_la_corrida": nuevas,
    }
    CSJN_NOVEDADES_PATH.parent.mkdir(parents=True, exist_ok=True)
    CSJN_NOVEDADES_PATH.write_text(
        json.dumps(store, indent=1, ensure_ascii=False, sort_keys=True),
        encoding="utf-8")
    return store


def cobertura_judicial_serie() -> tuple[dict, dict]:
    """({YYYY-MM: % de cargos con juez designado}, metadatos del padrón).

    El padrón es una FOTO fechada, no una serie. La serie se reconstruye
    moviéndose desde esa foto con los registros de designaciones y renuncias:

        hacia atrás   vacantes(t) = vacantes(P) + designaciones(t,P] − renuncias(t,P]
        hacia adelante vacantes(t) = vacantes(P) − designaciones(P,t] + renuncias(P,t]

    Importa el orden: entre t y P cada designación cubrió una vacante y cada
    renuncia creó una, así que hacia atrás los signos se invierten.

    **El numerador es `cargo_vacante = NO`**, no `cargo_cobertura = Titular`.
    Son dos campos distintos y no dan lo mismo: en el padrón del 5-jun-2026 hay
    610 cargos no vacantes y 604 con titular, porque 6 tienen titular designado
    pero con licencia y figuran cubiertos por subrogante. Es `cargo_vacante` el
    que se corresponde con lo que dice la unidad —cargos *con juez designado*— y
    el único que los registros de designaciones y renuncias saben mover.

    La card publicaba el porcentaje del primero y lo explicaba con el conteo del
    segundo, **y a la fecha del padrón en vez de la del corte**: 69,63% arriba y
    «604 de 955» abajo, que es 63,25% (ADR-0240). Por eso ahora los metadatos
    devuelven el numerador y el corte del valor que se publica, además de la
    foto del padrón.
    """
    padron = _jus_csv(JUS_PADRON_Q)
    habilitados = [f for f in padron
                   if (f.get("organo_habilitado") or "").strip().upper() == "SI"]
    if not habilitados:
        raise ValueError("padrón sin cargos habilitados")
    total = len(habilitados)
    vac_p = sum(1 for f in habilitados
                if (f.get("cargo_vacante") or "").strip().upper() == "SI")

    # fecha del padrón: la más reciente de las juras registradas no sirve
    # (quedan viejas), así que se toma del nombre del recurso vía su dataset.
    fecha_padron = _jus_fecha_padron()

    desig = _jus_fechas(_jus_csv(JUS_DESIGNACIONES_Q), "fecha_desginacion")
    renun = _jus_fechas(_jus_csv(JUS_RENUNCIAS_Q), "fecha_renuncia")

    hoy = date.today().isoformat()
    # Un registro fechado en el futuro no describe el presente: el dataset trae
    # designaciones con fecha posterior a hoy y contarlas adelantaría cobertura
    # que todavía no ocurrió.
    desig = [x for x in desig if x <= hoy]
    renun = [x for x in renun if x <= hoy]

    serie, ym = {}, "2023-12"
    corte_ultimo, mov_ultimo = None, None
    while ym <= hoy[:7]:
        ultimo_dia = calendar.monthrange(int(ym[:4]), int(ym[5:7]))[1]
        corte = min(f"{ym}-{ultimo_dia:02d}", hoy)
        if corte <= fecha_padron:
            d = sum(1 for x in desig if corte < x <= fecha_padron)
            r = sum(1 for x in renun if corte < x <= fecha_padron)
            vac = vac_p + d - r
        else:
            d = sum(1 for x in desig if fecha_padron < x <= corte)
            r = sum(1 for x in renun if fecha_padron < x <= corte)
            vac = vac_p - d + r
        serie[ym] = round(100.0 * (total - vac) / total, 2)
        corte_ultimo, mov_ultimo = corte, {"designaciones": d, "renuncias": r,
                                           "cargos_con_juez": total - vac}
        anio, mes = int(ym[:4]), int(ym[5:7]) + 1
        ym = f"{anio + 1}-01" if mes == 13 else f"{anio}-{mes:02d}"

    cobertura = {c: sum(1 for f in habilitados if f.get("cargo_cobertura") == c)
                 for c in ("Titular", "Subrogante", "Sin subrogante designado")}
    return serie, {"total_cargos": total, "vacantes_padron": vac_p,
                   "fecha_padron": fecha_padron, "composicion": cobertura,
                   "fecha_corte": corte_ultimo, **(mov_ultimo or {})}


def _jus_fecha_padron() -> str:
    """Fecha de corte del padrón, leída del nombre del recurso
    (`...-jueces-20260605.csv`)."""
    r = requests.get(JUS_API, params={"q": JUS_PADRON_Q, "rows": 5},
                     headers=HTTP_HEADERS, timeout=HTTP_TIMEOUT)
    r.raise_for_status()
    for paquete in r.json()["result"]["results"]:
        for recurso in paquete.get("resources", []):
            m = re.search(r"(20\d{2})(\d{2})(\d{2})\.csv", recurso.get("url", ""))
            if m:
                return f"{m.group(1)}-{m.group(2)}-{m.group(3)}"
    raise ValueError("no se pudo leer la fecha del padrón")


def fetch_cobertura_judicial() -> dict | None:
    """% de cargos de juez habilitados que tienen juez designado (ADR-0126).

    Mide la capacidad del Gobierno de completar el Poder Judicial, que exige
    acuerdo del Senado: es una capacidad NEGOCIADA, no una decisión propia, y
    por eso pertenece a este cinturón y no al de gestión.

    La card publica **numerador, denominador y la fecha de cada uno** (ADR-0240).
    Antes mostraba 69,63% arriba y «604 de 955 cargos» abajo —que es 63,25%—
    porque el porcentaje salía de `cargo_vacante` al corte de hoy y el conteo de
    `cargo_cobertura` a la fecha del padrón: dos definiciones y dos cortes en la
    misma card, y ningún gate mira si el texto reproduce el número.
    """
    try:
        serie, meta = cobertura_judicial_serie()
        ym = max(serie)
        total = meta["total_cargos"]
        numerador = meta["cargos_con_juez"]
        comp = meta["composicion"]
        titular = comp.get("Titular", 0)
        subrog = comp.get("Subrogante", 0)
        sin_nadie = comp.get("Sin subrogante designado", 0)
        padron_con_juez = total - meta["vacantes_padron"]
        return {
            "valor":          serie[ym],
            "unidad":         "% de cargos de juez con juez designado",
            "fuente":         "Ministerio de Justicia — padrón de magistrados, "
                              "designaciones y renuncias (datos.jus.gob.ar)",
            "fecha_dato":     f"{ym}-01",
            "desactualizado": False,
            # numerador, denominador y corte DEL VALOR publicado
            "cargos_con_juez": numerador,
            "cargos_totales":  total,
            "fecha_corte":     meta["fecha_corte"],
            # la foto de la que parte, con su propia fecha
            "fecha_padron":       meta["fecha_padron"],
            "padron_con_juez":    padron_con_juez,
            "padron_titular":     titular,
            "padron_subrogante":  subrog,
            "padron_sin_cubrir":  sin_nadie,
            # el inventario que explica la distancia entre las dos fechas
            "designaciones_desde_padron": meta["designaciones"],
            "renuncias_desde_padron":     meta["renuncias"],
            "detalle_txt": (
                f"{numerador} de {total} cargos de juez habilitados tienen juez "
                f"designado al {meta['fecha_corte']} · sale del padrón al "
                f"{meta['fecha_padron']} —{padron_con_juez} de {total} no "
                f"vacantes— más {meta['designaciones']} designaciones y menos "
                f"{meta['renuncias']} renuncias posteriores · en ese padrón, los "
                f"{total} cargos se repartían en {titular} con titular en "
                f"funciones, {subrog} con subrogante y {sin_nadie} sin cubrir"),
        }
    except Exception as e:
        _warn("cobertura_judicial", e)
        return None


# ── Bloque judicial y producción legislativa (ADR-0168) ──────────────────────
#
# Los cuatro indicadores que ADR-0166 desbloqueó al fijar la orientación. Tres
# leen el relevamiento ya versionado y uno consulta la fuente en vivo; cuál es
# cuál está declarado en cada colector, porque un store curado que nunca se
# refresca es un indicador viejo sin que ningún gate lo note.

CAUTELARES_PATH  = PROJECT_DIR / "data" / "politica" / "cautelares_saij_relevamiento.json"
DENUNCIAS_PATH   = PROJECT_DIR / "data" / "politica" / "denuncias_comisiones_universo.json"
CSJN_FUENTES_PATH = PROJECT_DIR / "data" / "politica" / "correccion_fuentes_judicial_empresario.json"


def _leer_store(path) -> dict:
    with open(path, encoding="utf-8") as fh:
        return json.load(fh)


def produccion_legislativa_serie() -> dict[str, int]:
    """Leyes con sanción definitiva en la ventana móvil de 12 meses, por mes.

    EN VIVO contra el CKAN de HCDN, el mismo recurso que ya usa
    eficacia_legislativa (ADR-0062). Se cuenta el TOTAL, sin mirar de dónde
    nace el expediente: el cociente de origen se mueve por el denominador
    (ADR-0137) y por eso lo que puntúa es este número (ADR-0168).
    """
    filas = _hcdn_paginate(HCDN_LEYES_SANC_RID)
    fechas: list[str] = []
    for r in filas:
        f = str(r.get("SANCION_DEFINITIVA", ""))[:10]
        if len(f) == 10 and f[:4].isdigit():
            fechas.append(f)
    if not fechas:
        raise ValueError("leyes-sancionadas sin fechas de sanción parseables")

    fechas.sort()
    desde = date(2023, 12, 1)
    hoy = date.today()
    serie: dict[str, int] = {}
    cur = desde
    while cur <= hoy:
        fin = cur
        ini = date(fin.year - 1, fin.month, 1)
        n = sum(1 for f in fechas if ini.isoformat() <= f <= fin.isoformat())
        serie[f"{fin.year}-{fin.month:02d}"] = n
        cur = date(cur.year + (cur.month // 12), (cur.month % 12) + 1, 1)
    return serie


def fetch_produccion_legislativa() -> dict | None:
    """Cuántas leyes sanciona el Congreso en 12 meses (ADR-0168)."""
    try:
        serie = produccion_legislativa_serie()
        ym = max(serie)
        valor = serie[ym]
        return {
            "valor": valor,
            "unidad": "leyes sancionadas (12m)",
            "fuente": "Cámara de Diputados — dataset de leyes sancionadas",
            "fecha_dato": f"{ym}-01",
            "desactualizado": False,
            "detalle_txt": (
                f"{valor} leyes sancionadas en los últimos 12 meses. El promedio "
                f"histórico del dataset (2008-2025, cuatro presidencias) es de 74 "
                f"leyes por año."),
        }
    except Exception as e:
        _warn("produccion_legislativa", e)
        return None


def fetch_judicializacion() -> dict | None:
    """Densidad cautelar en jurisdicción Federal + Nacional (ADR-0135/0168).

    EN VIVO contra SAIJ (ADR-0170). El relevamiento versionado queda como
    contraste: la consulta automatizada lo reproduce exacto en 10 de sus 11
    años, y el único que difiere es el año en curso, donde el vivo es el fresco.
    """
    try:
        puntos = judicializacion_serie()
        anio = max(puntos)
        rango = f"fecha-rango:[{anio}0101 TO {anio}1231]"
        crudo = [_saij_federal_y_nacional(f'(texto:"medida cautelar" AND {rango})'),
                 _saij_federal_y_nacional(f"({rango})")]
        return {
            "valor": puntos[anio],
            "unidad": "% de sumarios con medida cautelar",
            "fuente": "SAIJ — buscador de jurisprudencia (Federal + Nacional)",
            "fecha_dato": f"{anio}-01-01",
            "desactualizado": False,
            "detalle_txt": (
                f"{crudo[0]} de {crudo[1]} sumarios de jurisdicción federal y "
                f"nacional publicados en {anio} mencionan una medida cautelar. "
                f"Se publica la proporción y no el conteo porque el volumen que "
                f"la base publica varía por razones editoriales."),
        }
    except Exception as e:
        _warn("judicializacion", e)
        return None


def fetch_velocidad_resolucion() -> dict | None:
    """Expedientes resueltos sobre ingresados en la CSJN (ADR-0139/0168).

    Lee el relevamiento versionado. La fuente son los tableros estáticos y el
    anuario de la Corte, que no admiten consulta automática (ADR-0140): el
    refresco es anual y manual, y está declarado.
    """
    try:
        h = _leer_store(CSJN_FUENTES_PATH)["velocidad_de_resolucion"]["serie_historica_completa"]
        tasas = h["tasa_resolucion_pct"]
        anio = max(tasas)
        return {
            "valor": tasas[anio],
            "unidad": "% de expedientes resueltos sobre ingresados",
            "fuente": "CSJN — anuario estadístico (sistema de gestión judicial)",
            # Cierre del año de referencia, como iaf_transferencias: el dato
            # describe el año completo, no su primer día.
            "fecha_dato": f"{anio}-12-31",
            "desactualizado": False,
            "detalle_txt": (
                f"En {anio} la Corte resolvió {h['resueltos'][anio]:,} expedientes "
                f"sobre {h['ingresos'][anio]:,} ingresados. Por encima de 100% "
                f"descarga atraso; por debajo lo acumula."
            ).replace(",", "."),
        }
    except Exception as e:
        _warn("velocidad_resolucion", e)
        return None


def fetch_paralisis_denuncias() -> dict | None:
    """Sesiones de las comisiones de Acusación y Disciplina en 12m (ADR-0134/0168).

    EN VIVO contra la API del sitio del Consejo (ADR-0170). Mide AMBAS
    comisiones: Disciplina sola tiene 8 sesiones en cuatro años, o sea un
    indicador de evento, que es la clase que ADR-0147 dejó suspendida.
    """
    try:
        puntos = paralisis_denuncias_serie()
        ym = max(puntos)
        s = _leer_store(DENUNCIAS_PATH)["serie_12m"]
        return {
            "valor": puntos[ym],
            "unidad": "sesiones de las comisiones de control (12m)",
            "fuente": "Consejo de la Magistratura — archivo de notas de las comisiones",
            "fecha_dato": f"{ym}-01",
            "desactualizado": False,
            "detalle_txt": (
                f"{puntos[ym]} sesiones ordinarias de las comisiones de Acusación y "
                f"Disciplina en los últimos 12 meses (rango de la serie: "
                f"{min(puntos.values())} a {max(puntos.values())}, promedio "
                f"{round(sum(puntos.values())/len(puntos), 1)})."),
        }
    except Exception as e:
        _warn("paralisis_denuncias", e)
        return None



# ── Automatización de las dos fuentes judiciales (ADR-0170) ──────────────────

SAIJ_BUSQUEDA = "https://www.saij.gob.ar/busqueda"
CONSEJO_WP    = "https://consejomagistratura.gov.ar/index.php/wp-json/wp/v2"
CONSEJO_CATS  = {"acusacion": 129, "disciplina": 130}
_RE_SESION    = re.compile(
    r"sesion[oó]?-?(?:la|las)?-?comisi[oó]n-de-(acusacion|disciplina)-(\d+)")


def _saij_federal_y_nacional(consulta: str) -> int:
    """Suma de sumarios en jurisdicción Federal + Nacional para una consulta.

    El conteo sale de la faceta 'Jurisdicción' de `categoriesResultList`, no de
    `totalSearchResults`, que viene topeado por el pageSize — la corrección que
    el relevamiento de ADR-0135 le hizo a ADR-0131.
    """
    r = requests.get(SAIJ_BUSQUEDA,
                     params={"o": 0, "p": 1, "f": "Jurisdicción", "r": consulta},
                     headers=HTTP_HEADERS, timeout=HTTP_TIMEOUT)
    r.raise_for_status()
    cats = r.json()["searchResults"]["categoriesResultList"]
    hijos = {c["facetName"]: c["facetHits"] for c in cats[0]["facetChildren"]}
    return hijos.get("Federal", 0) + hijos.get("Nacional", 0)


def judicializacion_serie() -> dict[str, float]:
    """Densidad cautelar anual: % de sumarios con 'medida cautelar' sobre el
    total, ambos en jurisdicción Federal + Nacional (ADR-0135).

    EN VIVO contra SAIJ. Se publica la proporción y no el conteo porque el
    volumen que la base publica varía por razones editoriales: el conteo crudo
    va de 69 (2016) a 350 (2021) sin que las cautelares se hayan quintuplicado.
    """
    serie: dict[str, float] = {}
    for anio in range(2016, date.today().year + 1):
        rango = f"fecha-rango:[{anio}0101 TO {anio}1231]"
        num = _saij_federal_y_nacional(f'(texto:"medida cautelar" AND {rango})')
        den = _saij_federal_y_nacional(f"({rango})")
        if den:
            serie[str(anio)] = round(100.0 * num / den, 2)
    if not serie:
        raise ValueError("SAIJ no devolvió ningún año con denominador")
    return serie


def _consejo_sesiones() -> list[tuple[str, str, int]]:
    """(fecha, comisión, número de sesión) de las notas del Consejo.

    La numeración secuencial es lo que valida la cobertura: si el Consejo
    sesionara sin publicar la nota, el número siguiente delataría el hueco.
    """
    out: list[tuple[str, str, int]] = []
    for comision, cat in CONSEJO_CATS.items():
        pagina = 1
        while True:
            r = requests.get(f"{CONSEJO_WP}/posts",
                             params={"categories": cat, "per_page": 100,
                                     "page": pagina, "_fields": "date,slug"},
                             headers=HTTP_HEADERS, timeout=HTTP_TIMEOUT)
            if r.status_code != 200:
                break
            lote = r.json()
            if not lote:
                break
            for p in lote:
                m = _RE_SESION.search(p.get("slug", ""))
                if m:
                    out.append((p["date"][:10], m.group(1), int(m.group(2))))
            if len(lote) < 100:
                break
            pagina += 1
    if not out:
        raise ValueError("el archivo del Consejo no devolvió sesiones numeradas")
    return out


def paralisis_denuncias_serie() -> dict[str, int]:
    """Sesiones de Acusación y Disciplina en la ventana móvil de 12 meses.

    EN VIVO contra la API del sitio del Consejo. Se suman las dos comisiones:
    cada una por separado sesiona pocas veces al año —Disciplina, ocho veces en
    cuatro años— y una serie sobre una sola quedaría dominada por el ruido de
    un evento aislado (ADR-0168).
    """
    fechas = sorted(f for f, _, _ in _consejo_sesiones())
    serie: dict[str, int] = {}
    cur, hoy = date(2023, 12, 1), date.today()
    while cur <= hoy:
        # Ventana de 12 meses CALENDARIO terminada en el mes informado: desde
        # el primer día de once meses atrás hasta el último del mes corriente.
        # Cerrarla el día 1 dejaba afuera las sesiones del propio mes, que es
        # lo que producía diferencias de ±1 contra el relevamiento manual.
        ini = date(cur.year - 1, cur.month, 1)
        sig = date(cur.year + (cur.month // 12), (cur.month % 12) + 1, 1)
        serie[f"{cur.year}-{cur.month:02d}"] = sum(
            1 for f in fechas if ini.isoformat() <= f < sig.isoformat())
        cur = sig
    return serie

# ── Colectores manuales ───────────────────────────────────────────────────────

def load_manuales() -> dict:
    if not MANUALES_PATH.exists():
        return {}
    with open(MANUALES_PATH, encoding="utf-8") as f:
        data = json.load(f)
    data.pop("_meta", None)
    return data


def fetch_manual(nombre: str, stale_days: int = STALE_MANUAL_DAYS) -> dict | None:
    manuales = load_manuales()
    entry = manuales.get(nombre)
    if entry is None:
        _warn(nombre, f"No encontrado en {MANUALES_PATH}")
        return None
    if entry.get("valor") is None:
        _warn(nombre, "valor = null en manuales.json")
        return None

    dias = _days_old(str(entry.get("fecha_dato", "")))
    return {
        **entry,
        "desactualizado": dias > stale_days,
    }


# ── Score ─────────────────────────────────────────────────────────────────────

def indice_rice(afirmativos: int, negativos: int) -> float | None:
    """Índice de Rice de cohesión (0-100): |afirm-neg|/(afirm+neg) * 100.
    Ausentes/abstenciones ya excluidos por el caller (no forman parte de la
    votación dividida). None si no hubo votos afirmativos ni negativos del
    bloque en esa acta (no aporta información de cohesión)."""
    total = afirmativos + negativos
    if total == 0:
        return None
    return round(abs(afirmativos - negativos) / total * 100.0, 2)


# Bloque propio de LLA en Diputados/Senado. Excluye DELIBERADAMENTE aliados
# ambiguos (ej. "Fuerzas del Cielo - Espacio Liberal F.C.E.") que no son el
# bloque propio — sumarlos infla artificialmente la cohesión medida.
BLOQUES_LLA = {"la libertad avanza", "libertad avanza"}


def es_bloque_lla(nombre_bloque: str) -> bool:
    return nombre_bloque.strip().lower() in BLOQUES_LLA


_RE_REDIRECT_ACTA = re.compile(r"redirectActa\((\d+),\s*(\d+),\s*'([^']*)'\)")
_RE_DISPLAY_NONE   = re.compile(r"display\s*:\s*none")  # el sitio real usa "display: none" (con
                                                          # espacio) — confirmado vía snapshot de
                                                          # Wayback Machine (ene-2026) del listado
                                                          # real de votaciones.hcdn.gob.ar; el
                                                          # substring "display:none" sin espacio
                                                          # (supuesto inicial del fixture) no
                                                          # matchea contra la marca real.

def _descubrir_actas(session: requests.Session, anio: int):
    """POST a /votaciones/search por año -> [{id, slug, fecha}] de cada acta
    nominal encontrada. Cada fila del listado trae la fecha en un
    <span style="display: none">YYYYMMDD</span> y el link de detalle en un
    onclick=redirectActa(id, ?, 'slug') — se emparejan por fila (no por regex
    global sobre toda la página) para no desalinear fecha/acta.
    Nota: en el sitio real (confirmado vía snapshot archivado) el 3er argumento
    de redirectActa (slug) viene SIEMPRE vacío ('') — no es un bug de parseo,
    es el dato real; no asumir slugs legibles aguas abajo.
    None si el request en sí falló (distinto de 'sin actas ese año' = [])."""
    try:
        r = session.post(f"{HCDN_VOTACIONES_BASE}/votaciones/search",
                          data={"anoSearch": str(anio)}, timeout=HTTP_TIMEOUT)
    except requests.RequestException:
        return None
    if r.status_code != 200:
        return None
    soup = BeautifulSoup(r.text, "html.parser")
    actas = []
    vistos = set()
    for fila in soup.select("tr"):
        m = _RE_REDIRECT_ACTA.search(str(fila))
        span_fecha = fila.find("span", style=lambda s: s and _RE_DISPLAY_NONE.search(s))
        if not m or span_fecha is None:
            continue
        id_acta, _, slug = m.groups()
        if id_acta in vistos:
            continue
        try:
            fecha = datetime.strptime(span_fecha.get_text(strip=True), "%Y%m%d")
        except ValueError:
            continue
        vistos.add(id_acta)
        actas.append({"id": id_acta, "slug": slug, "fecha": fecha})
    return actas


# DORMIDO desde 2026-07-09 (ADR-0040): _descubrir_actas/_url_acta/_parsear_acta
# construían el listado vía la SPA de votaciones.hcdn.gob.ar, bloqueada por
# anti-bot (ADR-0037) -- fetch_cohesion_bloque ya no los llama, usa el
# endpoint PDF directo (_descubrir_actas_diputados_pdf y compañía, más abajo).
# Se conservan sin borrar (mismo criterio que las bandas de
# gobernadores_alineamiento): si el endpoint PDF alguna vez deja de andar,
# es la referencia de qué ya se probó del lado de la SPA.
def _url_acta(acta: dict) -> str:
    """Construye la URL de detalle de una acta. En producción el slug viene
    VACÍO en el 100% de las filas observadas (Tarea 4, HTML real archivado
    2026-01-15) — con slug vacío la URL real es /votacion/{id} (confirmado
    contra un snapshot real de Wayback Machine, id 5840, status 200), NO
    /votacion//{id}. Se soporta igual el caso con slug por si la fuente
    cambia en el futuro."""
    if acta.get("slug"):
        return f"/votacion/{acta['slug']}/{acta['id']}"
    return f"/votacion/{acta['id']}"


def _parsear_acta(html: str) -> list[dict]:
    """Parsea el HTML de una acta de votación nominal de Diputados ->
    [{nombre, bloque, provincia, voto}]. Ignora filas sin las columnas esperadas.

    Estructura REAL confirmada (Tarea 5, snapshot real de Wayback Machine
    2026-01-15, acta id 5840, tabla #myTable con 257 filas = las 257 bancas
    de la Cámara): cada fila tiene 6 <td> — foto (índice 0, sin texto),
    DIPUTADO/nombre (1), BLOQUE (2), PROVINCIA (3), "¿CÓMO VOTÓ?"/voto (4,
    anidado en un <span class="label ..."> dentro de <center>, no como texto
    directo del <td>), "¿QUÉ DIJO?" (5). La fila de <thead> tiene solo <th>
    (0 <td>) y queda afuera sola por el chequeo de longitud.

    Esto reemplaza la suposición original de 3 columnas
    (`<td>nombre</td><td class="ocultar">bloque</td><td>voto</td>`) que
    JAMÁS fue observada en vivo para Diputados (era inferencia por analogía
    con Senado, donde sí se había confirmado esa estructura de 3 columnas, y
    con el scraper de terceros Como_voto) — el HTML real de Diputados no
    tiene ninguna clase "ocultar" (0 ocurrencias en la página real).
    get_text(strip=True) extrae igual el voto aunque esté anidado en <span>.

    parser="html.parser" (stdlib): lxml NO está en requirements.txt y
    rompería en CI (confirmado en la Tarea 4; mismo parser que ya usa
    fetch_cepa_movilizacion)."""
    soup = BeautifulSoup(html, "html.parser")
    filas = []
    for tr in soup.select("table tr"):
        celdas = tr.find_all("td")
        if len(celdas) < 5:
            continue
        nombre = celdas[1].get_text(strip=True)
        bloque = celdas[2].get_text(strip=True)
        provincia = celdas[3].get_text(strip=True)
        voto = celdas[4].get_text(strip=True).upper()
        if not nombre or not bloque:
            continue
        filas.append({"nombre": nombre, "bloque": bloque, "provincia": provincia, "voto": voto})
    return filas


# ── Diputados vía PDF directo (desbloquea ADR-0037, ver ADR-0040) ────────────
#
# votaciones.hcdn.gob.ar/pdf/acta/{id} sirve el PDF de cada acta SIN pasar
# por la SPA bloqueada -- verificado en vivo 2026-07-09: HTTP 200 directo,
# sin JS, sin anti-bot. El id es un contador GLOBAL secuencial (no por
# período): actas consecutivas de una misma sesión tienen ids consecutivos
# (confirmado: 5955-5959 son 5 votaciones de la sesión del 24-jun-2026); no
# hay listado público de ids↔fecha, así que el rango se descubre caminando
# el propio endpoint.
_DIPUTADOS_ACTA_PDF_PATH = "/pdf/acta/{id}"
_DIPUTADOS_VOTOS_VALIDOS = {"AFIRMATIVO", "NEGATIVO", "ABSTENCION", "AUSENTE"}
_DIPUTADOS_UMBRAL_COLUMNA = 15.0   # puntos pdfplumber: separa columnas (~50-120pt) de palabras dentro de una columna (~2pt)


def _parsear_acta_diputados_pdf(contenido: bytes) -> list[dict]:
    """Parsea el PDF de una acta de votación nominal de Diputados (endpoint
    directo) -> [{nombre, bloque, provincia, voto}].

    El PDF no tiene una tabla con bordes que pdfplumber pueda detectar como
    tal (extract_tables() solo encuentra el bloque de metadata/encabezado,
    no la lista de votantes) -- se agrupan palabras por fila (mismo `top`,
    redondeado a 1 decimal) y se cortan en columnas nuevas donde el hueco
    horizontal entre palabras supera _DIPUTADOS_UMBRAL_COLUMNA (~2pt dentro
    de una columna -- "GUILLERMO CESAR", "Union Civica Radical" -- vs.
    50-120pt entre columnas, confirmado en vivo contra el acta 5959).

    El título decorativo ("Honorable Cámara...") y la fila de encabezados de
    columna vienen con cada CARÁCTER duplicado (fuente en negrita simulada
    del generador de PDF) -- no se intenta arreglar eso, esas filas no
    forman 4 columnas limpias y quedan afuera solas. La línea de metadata
    repetida en cada página ("Acta Nº... Fecha:... Hora:...") SÍ arma 4
    columnas por casualidad -- se filtra exigiendo que la última columna sea
    un voto válido (_DIPUTADOS_VOTOS_VALIDOS), no por posición de página."""
    filas = []
    with pdfplumber.open(io.BytesIO(contenido)) as pdf:
        for pagina in pdf.pages:
            por_fila = {}
            for palabra in pagina.extract_words():
                clave = round(palabra["top"], 1)
                por_fila.setdefault(clave, []).append(palabra)
            for _, palabras in sorted(por_fila.items()):
                palabras.sort(key=lambda p: p["x0"])
                columnas, actual = [], [palabras[0]["text"]]
                for anterior, palabra in zip(palabras, palabras[1:]):
                    if palabra["x0"] - anterior["x1"] > _DIPUTADOS_UMBRAL_COLUMNA:
                        columnas.append(" ".join(actual))
                        actual = [palabra["text"]]
                    else:
                        actual.append(palabra["text"])
                columnas.append(" ".join(actual))
                if len(columnas) != 4:
                    continue
                nombre, bloque, provincia, voto = columnas
                voto = voto.upper()
                if voto not in _DIPUTADOS_VOTOS_VALIDOS:
                    continue
                filas.append({"nombre": nombre, "bloque": bloque, "provincia": provincia, "voto": voto})
    return filas


def _diputados_acta_fecha(contenido: bytes) -> datetime | None:
    """Extrae la fecha ('Fecha: DD/MM/YYYY') del encabezado de metadata del
    PDF de una acta -- esa línea extrae limpia (sin duplicado de carácter),
    a diferencia del título decorativo."""
    with pdfplumber.open(io.BytesIO(contenido)) as pdf:
        texto = pdf.pages[0].extract_text() or ""
    m = re.search(r"Fecha:\s*(\d{2}/\d{2}/\d{4})", texto)
    if not m:
        return None
    try:
        return datetime.strptime(m.group(1), "%d/%m/%Y")
    except ValueError:
        return None


# Sentinelas del walk de actas de Diputados: un 404 genuino es un hueco de
# id (tolerable, la numeración puede saltearse) -- un fallo transitorio (red,
# 403 agotado, 5xx) NO es un hueco, y el backfill anual no debe congelar un
# año como "cerrado" si alguna acta falló por esto (quedaría un agujero
# permanente en la serie histórica y en la reconstrucción del ITCP).
_ACTA_NO_EXISTE = object()
_ACTA_FALLO = object()


def _diputados_acta_pdf(session: requests.Session, id_acta: int):
    """GET pausado del PDF de una acta puntual. Devuelve los bytes del PDF,
    `_ACTA_NO_EXISTE` si el servidor respondió 404 (hueco de id genuino), o
    `_ACTA_FALLO` si el request falló de forma transitoria (error de red,
    403 agotado, 5xx) -- los callers que cachean por año necesitan la
    distinción para no congelar un fallo como si fuera un hueco."""
    r = _paced_get(session, HCDN_VOTACIONES_BASE, _DIPUTADOS_ACTA_PDF_PATH.format(id=id_acta),
                   aceptar_404=True)
    if r is None:
        return _ACTA_FALLO
    if r.status_code == 404:
        return _ACTA_NO_EXISTE
    return r.content


# Caché PERMANENTE por acta (id -> {fecha, rice}): a diferencia de los
# stores "por año" de Senado/alineamiento (donde el año en curso se
# re-pide siempre porque pueden aparecer actas nuevas), acá el caché es a
# nivel de ACTA INDIVIDUAL -- una vez publicada, un acta nunca cambia de
# fecha ni de resultado, así que cachearla para siempre es seguro y barato.
# Sin este caché, cada corrida (live O backfill) volvía a descargar y
# parsear el mismo PDF de siempre -- hallazgo real 2026-07-09: sin caché,
# el backfill anual completo (4 años) tardó ~35 minutos extra en el
# pipeline; con caché, cada acta se descarga una sola vez en la vida del
# proyecto, sea cual sea cuántas veces se llame a fetch_cohesion_bloque o
# fetch_cohesion_bloque_diputados_actas_anio después.
DIPUTADOS_COHESION_CACHE_PATH = Path(__file__).resolve().parents[1] / "data" / "politica" / "cohesion_bloque_diputados_actas_cache.json"


def _cargar_cache_cohesion_diputados() -> dict:
    try:
        return json.loads(DIPUTADOS_COHESION_CACHE_PATH.read_text(encoding="utf-8-sig"))
    except (OSError, json.JSONDecodeError):
        return {}


def _guardar_cache_cohesion_diputados(cache: dict) -> None:
    DIPUTADOS_COHESION_CACHE_PATH.write_text(
        json.dumps(cache, indent=2, ensure_ascii=False, sort_keys=True), encoding="utf-8")


def _acta_diputados_cacheada(session: requests.Session, id_acta: int, cache: dict) -> dict | None:
    """{"fecha": datetime, "rice": float | None} de una acta, mirando
    primero `cache` (dict mutable id-str -> {"fecha": "YYYY-MM-DD", "rice":
    float | None}) antes de pedir red. Si no está cacheada, descarga+parsea
    y la agrega a `cache`, persistiendo a disco DE INMEDIATO (no al final
    del walk que la llama) -- un backfill completo camina cientos de ids a
    ~0,3s/request (varios minutos reales), y si se corta a mitad de camino
    (timeout de CI, corrida manual interrumpida) el trabajo ya hecho no
    puede perderse: la próxima corrida debe retomar desde donde quedó, no
    desde cero. Se cachea SIEMPRE que se pudo leer la fecha, con rice=None
    cuando el bloque LLA no aporta señal (empate o sin presentes), para que
    un walk repetido nunca vuelva a descargar la misma acta. None si la
    acta no existe (404, hueco de id genuino) o no se pudo extraer la fecha
    del PDF; `_ACTA_FALLO` si la descarga falló de forma transitoria (red,
    403 agotado) -- el backfill anual usa la distinción para no cachear un
    año con agujeros. Ninguno de esos casos se cachea."""
    clave = str(id_acta)
    if clave in cache:
        entrada = cache[clave]
        return {"fecha": datetime.strptime(entrada["fecha"], "%Y-%m-%d"), "rice": entrada["rice"]}
    contenido = _diputados_acta_pdf(session, id_acta)
    if contenido is _ACTA_FALLO:
        return _ACTA_FALLO
    if contenido is _ACTA_NO_EXISTE or contenido is None:
        # None: tolerancia defensiva (el contrato actual de
        # _diputados_acta_pdf ya no lo produce) -- se trata como hueco.
        return None
    fecha = _diputados_acta_fecha(contenido)
    if fecha is None:
        return None
    filas = _parsear_acta_diputados_pdf(contenido)
    afirm = sum(1 for f in filas if es_bloque_lla(f["bloque"]) and f["voto"] == "AFIRMATIVO")
    neg = sum(1 for f in filas if es_bloque_lla(f["bloque"]) and f["voto"] == "NEGATIVO")
    rice = indice_rice(afirm, neg)
    cache[clave] = {"fecha": fecha.strftime("%Y-%m-%d"), "rice": rice}
    _guardar_cache_cohesion_diputados(cache)
    return {"fecha": fecha, "rice": rice}


# El avance de _diputados_acta_id_maximo tolera huecos de numeración: la
# numeración real de actas NO es contigua — 101 huecos internos en el caché
# 4694..5959, con anchos de hasta 202 ids (4897→5100; luego 72, 35, 34, 33)
# y verificado EN VIVO 2026-07-11 que esos rangos devuelven 404 real, no un
# PDF sin fecha. "Primer 404 = final" perdía toda acta nueva publicada
# después de un hueco. El margen dobla el hueco máximo observado; el costo
# (~450 probeos pacientes al agotar la búsqueda) se paga UNA vez por proceso
# gracias a la memoización de abajo.
_MARGEN_HUECOS_ID = 450
_TOPE_AVANCE_ID = 2000

# Memo por PROCESO del id máximo descubierto: politica.py (card live) y
# descargar_series.py (4 backfills anuales) llaman al descubrimiento varias
# veces por corrida — el máximo no cambia dentro de un mismo proceso, y sin
# memo el probeo terminal de ~450 requests se pagaría por cada llamada. Solo
# se memoiza un resultado exitoso (int): un None (fallo transitorio) debe
# reintentar en la próxima llamada.
_ID_MAXIMO_MEMO: dict = {}

# Persistencia del descubrimiento por DÍA: el cron corre politica.py y
# descargar_series.py como PROCESOS separados en el mismo job, así que el
# memo de proceso no evita pagar dos veces el probeo terminal (~450 requests
# ≈ 2-3 min cada uno) — y el job ya roza su timeout (corridas reales de
# 19m32s/20m2s contra una cota de 20). Solo se persiste un descubrimiento
# EXITOSO y con fecha: al día siguiente se re-descubre (la numeración
# avanza), y un None o una exploración parcial nunca se persisten.
DIPUTADOS_ID_MAXIMO_STORE = Path(__file__).resolve().parents[1] / "data" / "politica" / "id_maximo_diputados.json"


def _hoy_buenos_aires() -> str:
    """Fecha calendario de referencia del proyecto (ART, UTC-3 fijo, sin
    horario de verano) como YYYY-MM-DD. El runner del cron vive en UTC: con
    datetime.now() a secas, un workflow_dispatch entre 00:00 y 02:59 UTC
    (todavía el día ANTERIOR en Argentina) sellaría el store con la fecha
    UTC nueva y el cron de las 03:00 UTC reutilizaría ese máximo sin
    redescubrir (hallazgo de octava pasada de revisión). Fallback a la hora
    local de la máquina si la base de zonas horarias no está disponible
    (Windows sin el paquete tzdata)."""
    try:
        from zoneinfo import ZoneInfo
        return datetime.now(ZoneInfo("America/Argentina/Buenos_Aires")).strftime("%Y-%m-%d")
    except Exception:
        return datetime.now().strftime("%Y-%m-%d")


def _diputados_acta_id_maximo(session: requests.Session, desde_id: int = 5959) -> int | None:
    """Encuentra el id de acta más reciente disponible, caminando desde
    `desde_id` (semilla conocida: 5959 = 24-jun-2026, verificado en vivo
    2026-07-09). Si la semilla ya no existe (quedó vieja), retrocede de a
    50 hasta encontrar un id válido; desde ahí avanza de a uno tolerando
    hasta _MARGEN_HUECOS_ID ids consecutivos sin acta (la numeración real
    tiene huecos, ver arriba) antes de dar por encontrado el máximo.

    None si ni retrocediendo se encuentra un id válido, si CUALQUIER probeo
    falla de forma transitoria (_ACTA_FALLO: con un fallo en el medio no se
    puede saber dónde termina la numeración, y un máximo subestimado haría
    que el walk anual cachee un año sin sus actas de arriba), o si el avance
    supera _TOPE_AVANCE_ID ids nuevos (endpoint patológico que responde 200
    a cualquier id).

    El punto de partida efectivo es el MAYOR entre `desde_id` (semilla
    estática de respaldo) y el id más alto ya presente en el caché por acta:
    así el arranque, el probeo y su tope acompañan el crecimiento real de la
    numeración. Con la semilla fija sola, cada corrida re-descargaba todos
    los ids posteriores a la semilla y, con el paso de las sesiones, el
    flujo normal habría alcanzado el tope y devuelto None para siempre
    (hallazgo de cuarta pasada de revisión).

    Caso extremo documentado: si la numeración avanzara más de
    _TOPE_AVANCE_ID ids sin que ninguna corrida lo registre (colector caído
    por meses), el descubrimiento devolvería None con WARN y haría falta
    subir la semilla estática a mano."""
    if "valor" in _ID_MAXIMO_MEMO:
        return _ID_MAXIMO_MEMO["valor"]
    hoy_str = _hoy_buenos_aires()
    try:
        store = json.loads(DIPUTADOS_ID_MAXIMO_STORE.read_text(encoding="utf-8-sig"))
        if store.get("descubierto") == hoy_str and isinstance(store.get("maximo"), int):
            # Descubrimiento de HOY hecho por otro proceso de la misma
            # corrida (o una corrida manual previa del día): se reusa. Las
            # actas se publican con rezago, así que congelar el máximo por
            # un día calendario no pierde nada en la práctica.
            _ID_MAXIMO_MEMO["valor"] = store["maximo"]
            return store["maximo"]
    except (OSError, json.JSONDecodeError):
        pass
    cache = _cargar_cache_cohesion_diputados()
    ultimo_conocido = max((int(k) for k in cache if str(k).isdigit()), default=0)
    arranques = [max(desde_id, ultimo_conocido)]
    if arranques[0] > desde_id:
        # Respaldo: el máximo cacheado no se valida contra nada, así que una
        # entrada corrupta/manual con un id inverosímilmente alto agotaría el
        # retroceso sin tocar el rango real. En ese caso se reintenta UNA vez
        # desde la semilla estática, en vez de devolver None para siempre
        # (el caché persistente impediría la recuperación automática).
        arranques.append(desde_id)
    actual = 0
    for arranque in arranques:
        actual = arranque
        intentos = 0
        while actual > 0:
            contenido = _diputados_acta_pdf(session, actual)
            if contenido is _ACTA_FALLO:
                return None
            if isinstance(contenido, bytes):
                break
            actual -= 50
            intentos += 1
            if intentos > 200:   # ~10000 ids de margen -- corta un loop patológico
                actual = 0
                break
        if actual > 0:
            break
    if actual <= 0:
        return None

    maximo = actual
    sonda = actual
    huecos_seguidos = 0
    while huecos_seguidos < _MARGEN_HUECOS_ID:
        sonda += 1
        if sonda - actual > _TOPE_AVANCE_ID:
            print(f"  [WARN] _diputados_acta_id_maximo: más de {_TOPE_AVANCE_ID} ids nuevos "
                  f"desde {actual} -- sospechoso, no se determina el máximo")
            return None
        contenido = _diputados_acta_pdf(session, sonda)
        if contenido is _ACTA_FALLO:
            return None
        if isinstance(contenido, bytes):
            maximo = sonda
            huecos_seguidos = 0
        else:
            huecos_seguidos += 1
    _ID_MAXIMO_MEMO["valor"] = maximo
    try:
        DIPUTADOS_ID_MAXIMO_STORE.parent.mkdir(parents=True, exist_ok=True)
        DIPUTADOS_ID_MAXIMO_STORE.write_text(json.dumps(
            {"maximo": maximo, "sondeado_hasta": sonda, "descubierto": hoy_str},
            indent=2), encoding="utf-8")
    except OSError:
        pass   # persistencia best-effort: el memo del proceso igual sirve
    return maximo


def fetch_cohesion_bloque_diputados_actas_anio(anio: int) -> list | None:
    """Detalle CRUDO de cohesión por acta de Diputados de TODO el año dado
    (sin recortar por ventana de días), vía el endpoint PDF directo --
    mismo shape que fetch_cohesion_bloque_senado_actas_anio (Senado):
    [{"fecha": "YYYY-MM-DD", "rice": float}, ...], una fila por acta con
    señal (votos del bloque LLA, no empatados). Existe para el backfill
    mensual (descargar_series.py) con el mismo patrón que Senado.

    A diferencia de Senado (que tiene un listado liviano por año), acá no
    hay forma de pedir "solo este año" -- camina hacia atrás desde el id
    más reciente HOY, así que llamar esta función para 2023..2026 en
    cualquier orden vuelve a caminar por los años ya vistos en llamadas
    anteriores. Eso es barato gracias a _acta_diputados_cacheada (caché
    PERMANENTE por acta, ver ahí) -- cada acta se descarga una sola vez,
    las siguientes veces que el walk la vuelve a pisar es solo un lookup
    de diccionario.

    Si la descarga de ALGUNA acta falla de forma transitoria (_ACTA_FALLO,
    distinto de un hueco de id genuino), devuelve None (detalle incompleto):
    el caller cachea años cerrados como inmutables y un año con agujeros
    contaminaría la serie histórica para siempre. El caché por acta igual
    retiene lo ya descargado, así que la corrida siguiente solo paga las
    actas que faltan."""
    MARGEN_SALIDA = 5
    session = _hcdn_votaciones_session()
    id_maximo = _diputados_acta_id_maximo(session)
    if id_maximo is None:
        return None

    cache = _cargar_cache_cohesion_diputados()
    detalle = []
    fallidas = 0
    fuera_de_anio_seguidas = 0
    id_actual = id_maximo
    while id_actual > 0 and fuera_de_anio_seguidas < MARGEN_SALIDA:
        entrada = _acta_diputados_cacheada(session, id_actual, cache)
        id_actual -= 1
        if entrada is _ACTA_FALLO:
            # fallo transitorio: el acta existe pero no se pudo leer AHORA.
            # No se puede saber de qué año es, así que el resultado completo
            # queda incompleto -- se termina el walk igual (para aprovechar
            # el caché por acta de lo que sí se pudo leer) pero se devuelve
            # None: el caller NO debe congelar este año con un agujero.
            fallidas += 1
            continue
        if entrada is None:
            continue   # hueco de id genuino (404) o sin fecha -- no cuenta como "fuera de año"
        if entrada["fecha"].year > anio:
            continue   # todavía viajando desde HOY hacia el año pedido -- no cuenta como salida
        if entrada["fecha"].year < anio:
            fuera_de_anio_seguidas += 1
            continue
        fuera_de_anio_seguidas = 0
        if entrada["rice"] is not None:
            detalle.append({"fecha": entrada["fecha"].strftime("%Y-%m-%d"), "rice": entrada["rice"]})

    if fallidas:
        print(f"  [WARN] cohesion_bloque_diputados_actas_anio {anio}: "
              f"{fallidas} actas con descarga fallida -- detalle incompleto, no se devuelve")
        return None
    return detalle


def fetch_cohesion_bloque(anio: int | None = None, dias_ventana: int = 90) -> dict | None:
    """Cohesión del bloque LLA en Diputados: índice de Rice promedio sobre
    las actas nominales divididas de los últimos `dias_ventana` días.

    Desde 2026-07-09 (ADR-0040) usa el endpoint PDF directo
    (votaciones.hcdn.gob.ar/pdf/acta/{id}), NO la SPA bloqueada por
    anti-bot (ADR-0037) que usaban _descubrir_actas/_url_acta/_parsear_acta
    (dormidas más arriba). No hay listado id↔fecha para este endpoint, así
    que camina hacia atrás desde el id más reciente descargando cada acta
    -- se corta a los MARGEN_SALIDA actas seguidas fuera de la ventana en
    vez de solo por fecha de la última, para tolerar algún id fuera de
    orden sin perder actas más viejas que sí caen en la ventana.

    `anio`/`dias_ventana=366` para backfill de un año pasado usan
    fetch_cohesion_bloque_diputados_actas_anio + descargar_series (mismo
    patrón que cohesion_bloque_senado) -- ESTA función solo sirve bien el
    caso live (año en curso, ventana de 90 días); para años pasados
    `_diputados_acta_id_maximo` seguiría ancladando al id de HOY, no al de
    fin de `anio`, así que no se soporta acá.

    Desde 2026-07-09 usa la caché permanente por acta
    (_acta_diputados_cacheada) en vez de descargar cada acta sin registro:
    en una corrida repetida (cron nocturno incluido) solo las actas nuevas
    desde la última corrida pagan una descarga real.

    Devuelve None SOLO si no se pudo determinar el id más reciente (el
    endpoint en sí falló) — 'sin votos en la ventana' (receso legislativo)
    es un resultado válido con valor=None pero corrida_exitosa_en seteado."""
    anio = anio or datetime.now().year
    session = _hcdn_votaciones_session()
    id_maximo = _diputados_acta_id_maximo(session)
    if id_maximo is None:
        return None

    # Referencia SIEMPRE a medianoche: el backfill mensual ancla sus ventanas
    # a medianoche del fin de mes, y con la hora del día incluida una acta
    # fechada exactamente `dias_ventana` días atrás quedaba fuera de la card
    # pero dentro de la serie (borde inconsistente entre las dos vistas).
    hoy = datetime.now()
    referencia = (datetime(hoy.year, hoy.month, hoy.day)
                  if anio == hoy.year else datetime(anio, 12, 31))
    limite = referencia - timedelta(days=dias_ventana)

    MARGEN_SALIDA = 5
    cache = _cargar_cache_cohesion_diputados()
    detalle = []
    fuera_de_ventana_seguidas = 0
    id_actual = id_maximo
    while id_actual > 0 and fuera_de_ventana_seguidas < MARGEN_SALIDA:
        entrada = _acta_diputados_cacheada(session, id_actual, cache)
        id_actual -= 1
        if entrada is None or entrada is _ACTA_FALLO:
            # hueco en la numeración, PDF ilegible o fallo transitorio -- el
            # valor live es best-effort (se recalcula entero cada corrida,
            # no se congela), así que acá el fallo solo se saltea.
            continue
        if entrada["fecha"] > referencia:
            continue
        if entrada["fecha"] < limite:
            fuera_de_ventana_seguidas += 1
            continue
        fuera_de_ventana_seguidas = 0
        if entrada["rice"] is not None:
            detalle.append((entrada["fecha"], entrada["rice"]))

    fecha_max = max((f for f, _ in detalle), default=None)
    return {
        "valor": round(sum(r for _, r in detalle) / len(detalle), 1) if detalle else None,
        "unidad": "% cohesión (índice de Rice), promedio actas divididas últimos 90 días",
        "fuente": "Votaciones nominales de la Cámara de Diputados — elaboración CIGOB",
        "fecha_dato": fecha_max.strftime("%Y-%m-%d") if fecha_max else None,
        "n_actas": len(detalle),
        "corrida_exitosa_en": datetime.now().strftime("%Y-%m-%d"),
        "desactualizado": False,
    }


_RE_DETALLE_ACTA_SENADO = re.compile(r"/votaciones/detalleActa/(\d+)")


def _descubrir_actas_senado(session: requests.Session, anio: int):
    """POST a /votaciones/actas (listado con fecha en <span style="display:none">
    YYYYMMDD</span> y link <a href="/votaciones/detalleActa/{id}">) ->
    [{id, fecha}] del año dado. Estructura confirmada en vivo (Senado, HTML
    server-side, sin headless browser). parser="html.parser": lxml no está en
    requirements.txt (Tarea 4 del plan de cohesion_bloque). Reusa
    _RE_DISPLAY_NONE (mismo plan, Tarea 4) en vez de un match exacto de
    "display:none" — la Tarea 4 confirmó que el HTML real de HCDN usa
    "display: none" CON espacio; dado que Senado es la misma familia de sitios
    de gobierno, no asumir que acá sí será sin espacio sin verificarlo en vivo
    (ver Step de verificación más abajo).

    FIX (auditoría 2026-07-08): un GET plano a esta URL siempre devuelve el
    listado del año EN CURSO -- el filtro `fecha.year != anio` de abajo
    descartaba todo cuando `anio` era un año pasado, y el backfill
    'funcionaba' produciendo 1 solo punto real (el año en curso). Verificado
    en vivo: el sitio acepta POST con busqueda_actas[anio]=<año> y devuelve el
    listado real de ESE año (26 actas en 2023, 91 en 2024, 95 en 2025, sin
    bloqueo anti-bot, a diferencia de HCDN Diputados)."""
    r = _paced_post(session, SENADO_BASE, "/votaciones/actas",
                     data={"busqueda_actas[anio]": str(anio)})
    if r is None:
        return None
    soup = BeautifulSoup(r.text, "html.parser")
    actas = []
    vistos = set()
    for fila in soup.select("tr"):
        link = fila.find("a", href=_RE_DETALLE_ACTA_SENADO)
        span_fecha = fila.find("span", style=lambda s: s and _RE_DISPLAY_NONE.search(s))
        if link is None or span_fecha is None:
            continue
        m = _RE_DETALLE_ACTA_SENADO.search(link["href"])
        if not m:
            continue
        id_acta = m.group(1)
        try:
            fecha = datetime.strptime(span_fecha.get_text(strip=True), "%Y%m%d")
        except ValueError:
            continue
        if fecha.year != anio or id_acta in vistos:
            continue
        vistos.add(id_acta)
        actas.append({"id": id_acta, "fecha": fecha})
    return actas


def fetch_cohesion_bloque_senado(anio: int | None = None, dias_ventana: int = 90) -> dict | None:
    """Cohesión del bloque LLA en el Senado — mismo índice de Rice que
    fetch_cohesion_bloque, indicador COMPLEMENTARIO (otra cámara, otra
    composición de bloque): nunca reemplaza al de Diputados.

    La ventana de recencia se ancla a HOY para el año en curso, y al 31 de
    diciembre de `anio` para años pasados (backfill) — mismo fix aplicado en
    fetch_cohesion_bloque (Tarea 6, cross-fix del plan de Diputados): anclar
    siempre a datetime.now() hacía el backfill de años pasados estructuralmente
    imposible (hallazgo de revisión de esta tarea, mismo bug reintroducido por
    el brief)."""
    anio = anio or datetime.now().year
    session = _hcdn_votaciones_session()
    actas = _descubrir_actas_senado(session, anio)
    if actas is None:
        return None

    # Referencia SIEMPRE a medianoche: el backfill mensual ancla sus ventanas
    # a medianoche del fin de mes, y con la hora del día incluida una acta
    # fechada exactamente `dias_ventana` días atrás quedaba fuera de la card
    # pero dentro de la serie (borde inconsistente entre las dos vistas).
    hoy = datetime.now()
    referencia = (datetime(hoy.year, hoy.month, hoy.day)
                  if anio == hoy.year else datetime(anio, 12, 31))
    limite = referencia - timedelta(days=dias_ventana)
    indices = []
    fecha_max = None
    for acta in actas:
        if acta["fecha"] < limite:
            continue
        r = _paced_get(session, SENADO_BASE, f"/votaciones/detalleActa/{acta['id']}")
        if r is None:
            continue
        filas = _parsear_acta(r.text)
        afirm = sum(1 for f in filas if es_bloque_lla(f["bloque"]) and f["voto"] == "AFIRMATIVO")
        neg = sum(1 for f in filas if es_bloque_lla(f["bloque"]) and f["voto"] == "NEGATIVO")
        rice = indice_rice(afirm, neg)
        if rice is None:
            continue
        indices.append(rice)
        fecha_max = acta["fecha"] if fecha_max is None else max(fecha_max, acta["fecha"])

    return {
        "valor": round(sum(indices) / len(indices), 1) if indices else None,
        "unidad": "% cohesión (índice de Rice, Senado), promedio actas divididas últimos 90 días",
        "fuente": "Votaciones nominales del Senado — elaboración CIGOB",
        "fecha_dato": fecha_max.strftime("%Y-%m-%d") if fecha_max else None,
        "n_actas": len(indices),
        "corrida_exitosa_en": datetime.now().strftime("%Y-%m-%d"),
        "desactualizado": False,
    }


def fetch_cohesion_bloque_senado_actas_anio(anio: int) -> list | None:
    """Detalle CRUDO de cohesión por acta de TODO el año dado (sin recortar
    por ventana de días): [{"fecha": "YYYY-MM-DD", "rice": float}, ...], una
    fila por acta con señal (votos AFIRMATIVO/NEGATIVO del bloque LLA, no
    empatados). Mismo criterio que fetch_alineamiento_senadores_actas_anio:
    factorizada de fetch_cohesion_bloque_senado sin modificarla (el valor
    live sigue pidiendo solo las actas dentro de su ventana de 90 días) --
    existe para el backfill mensual (descargar_series.fetch_cohesion_bloque_senado_mensual),
    que necesita derivar múltiples ventanas de 90 días sin re-scrapear el
    Senado por cada mes.

    Si el detalle de ALGUNA acta del listado falla, devuelve None (detalle
    incompleto): el caller cachea años cerrados como inmutables y un año
    con agujeros contaminaría la serie histórica para siempre."""
    session = _hcdn_votaciones_session()
    actas = _descubrir_actas_senado(session, anio)
    if actas is None:
        return None
    detalle = []
    fallidas = 0
    for acta in actas:
        r = _paced_get(session, SENADO_BASE, f"/votaciones/detalleActa/{acta['id']}")
        if r is None:
            # Cada acta viene del listado oficial del año: su detalle DEBE
            # existir, así que esto es un fallo transitorio, no un hueco.
            fallidas += 1
            continue
        filas = _parsear_acta(r.text)
        afirm = sum(1 for f in filas if es_bloque_lla(f["bloque"]) and f["voto"] == "AFIRMATIVO")
        neg = sum(1 for f in filas if es_bloque_lla(f["bloque"]) and f["voto"] == "NEGATIVO")
        rice = indice_rice(afirm, neg)
        if rice is not None:
            detalle.append({"fecha": acta["fecha"].strftime("%Y-%m-%d"), "rice": rice})
    if fallidas:
        print(f"  [WARN] cohesion_bloque_senado_actas_anio {anio}: "
              f"{fallidas} actas con detalle fallido -- detalle incompleto, no se devuelve")
        return None
    return detalle


def _agregar_cohesion_ventana(detalle: list[dict], referencia: datetime, dias_ventana: int) -> dict | None:
    """Agrega el detalle crudo por acta (fetch_cohesion_bloque_senado_actas_anio,
    de uno o más años) dentro de la ventana [referencia − dias_ventana,
    referencia] y devuelve {valor, fecha_dato, n_actas}, o None si ninguna
    acta cae en la ventana. Misma fórmula que fetch_cohesion_bloque_senado
    (promedio simple del índice de Rice entre las actas con señal) --
    factorizada acá para que el backfill mensual pueda recalcularla para
    cada fin de mes sin duplicar la matemática ni volver a pedir red."""
    limite = referencia - timedelta(days=dias_ventana)
    indices = []
    fecha_max = None
    for fila in detalle:
        fecha = datetime.strptime(fila["fecha"], "%Y-%m-%d")
        if fecha < limite or fecha > referencia:
            continue
        indices.append(fila["rice"])
        fecha_max = fecha if fecha_max is None else max(fecha_max, fecha)
    if not indices:
        return None
    return {
        "valor": round(sum(indices) / len(indices), 1),
        "fecha_dato": fecha_max.strftime("%Y-%m-%d"),
        "n_actas": len(indices),
    }


def _alineamiento_por_provincia(filas: list[dict]) -> dict:
    """Dada la lista de filas de UNA acta (nombre/bloque/provincia/voto),
    devuelve {provincia: (coincidencias, total)} solo para las provincias que
    tienen AL MENOS 1 senador no-LLA en esa acta. La posición del oficialismo
    en esa acta es el voto mayoritario de los senadores del bloque LLA
    (cualquier provincia); si LLA no tiene votos claros (empate o sin
    senadores LLA presentes) esa acta no aporta señal, se devuelve {}."""
    afirm_lla = sum(1 for f in filas if es_bloque_lla(f["bloque"]) and f["voto"] == "AFIRMATIVO")
    neg_lla = sum(1 for f in filas if es_bloque_lla(f["bloque"]) and f["voto"] == "NEGATIVO")
    if afirm_lla == neg_lla:
        return {}
    posicion_lla = "AFIRMATIVO" if afirm_lla > neg_lla else "NEGATIVO"

    resultado = {}
    for f in filas:
        if es_bloque_lla(f["bloque"]) or f["voto"] not in ("AFIRMATIVO", "NEGATIVO"):
            continue
        coincide, total = resultado.get(f["provincia"], (0, 0))
        resultado[f["provincia"]] = (coincide + (1 if f["voto"] == posicion_lla else 0), total + 1)
    return resultado


def fetch_alineamiento_senadores_prov(anio: int | None = None, dias_ventana: int = 90) -> dict | None:
    """% de votos de senadores NO-LLA que coincide con la posición del
    bloque LLA en el Senado, agregado por provincia y promediado entre
    provincias con al menos 1 senador no-LLA (las 100% LLA se excluyen --
    su "alineamiento" con LLA es tautológico, no aporta señal).

    Reemplaza a gobernadores_alineamiento (placeholder manual congelado
    desde 2026-04, sin fuente automatizable encontrada tras 2 rondas de
    investigación). CAVEAT HONESTO: mide comportamiento de voto de
    SENADORES, no la postura pública del gobernador (Poder Ejecutivo
    provincial) -- un senador no depende del gobernador de turno. Es la
    mejor señal automatizable disponible hoy (2026-07-08), no una medición
    directa -- mismo tipo de proxy que adhesion_reformas_provincial/RIGI.

    Misma ventana/ancla que fetch_cohesion_bloque_senado (hoy para año en
    curso, 31-dic para backfill)."""
    anio = anio or datetime.now().year
    session = _hcdn_votaciones_session()
    actas = _descubrir_actas_senado(session, anio)
    if actas is None:
        return None

    # Referencia SIEMPRE a medianoche: el backfill mensual ancla sus ventanas
    # a medianoche del fin de mes, y con la hora del día incluida una acta
    # fechada exactamente `dias_ventana` días atrás quedaba fuera de la card
    # pero dentro de la serie (borde inconsistente entre las dos vistas).
    hoy = datetime.now()
    referencia = (datetime(hoy.year, hoy.month, hoy.day)
                  if anio == hoy.year else datetime(anio, 12, 31))
    limite = referencia - timedelta(days=dias_ventana)
    acumulado = {}
    fecha_max = None
    for acta in actas:
        if acta["fecha"] < limite:
            continue
        r = _paced_get(session, SENADO_BASE, f"/votaciones/detalleActa/{acta['id']}")
        if r is None:
            continue
        filas = _parsear_acta(r.text)
        resultado_acta = _alineamiento_por_provincia(filas)
        for provincia, (coincide, total) in resultado_acta.items():
            c0, t0 = acumulado.get(provincia, (0, 0))
            acumulado[provincia] = (c0 + coincide, t0 + total)
        if resultado_acta:
            fecha_max = acta["fecha"] if fecha_max is None else max(fecha_max, acta["fecha"])

    if not acumulado:
        return {
            "valor": None,
            "unidad": "% votos de senadores no-LLA alineados con LLA, por provincia",
            "fuente": "Votaciones nominales del Senado — elaboración CIGOB",
            "fecha_dato": None,
            "n_provincias": 0,
            "corrida_exitosa_en": datetime.now().strftime("%Y-%m-%d"),
            "desactualizado": False,
        }

    ratios_por_provincia = [c / t for c, t in acumulado.values() if t > 0]
    valor = round(100 * sum(ratios_por_provincia) / len(ratios_por_provincia), 1)
    return {
        "valor": valor,
        "unidad": "% votos de senadores no-LLA alineados con LLA, por provincia",
        "fuente": "Votaciones nominales del Senado — elaboración CIGOB",
        "fecha_dato": fecha_max.strftime("%Y-%m-%d") if fecha_max else None,
        "n_provincias": len(acumulado),
        "corrida_exitosa_en": datetime.now().strftime("%Y-%m-%d"),
        "desactualizado": False,
    }


def fetch_alineamiento_senadores_actas_anio(anio: int) -> list | None:
    """Detalle CRUDO de alineamiento por acta de TODO el año dado (sin
    recortar por ventana de días): [{"fecha": "YYYY-MM-DD", "provincias":
    {provincia: [coincide, total]}}, ...], una fila por acta CON señal (las
    actas donde el bloque LLA queda empatado no aportan fila).

    A diferencia de fetch_alineamiento_senadores_prov (que recorta por
    dias_ventana ANTES de pedir el detalle de cada acta, para no gastar
    requests de más en la corrida diaria), acá se pide el año COMPLETO de
    una sola pasada. Existe para el backfill mensual
    (descargar_series.fetch_alineamiento_senadores_prov_mensual): con el
    detalle crudo cacheado por año se pueden derivar múltiples ventanas de
    90 días (una por fin de mes) sin volver a scrapear el Senado por cada
    mes — sólo tiene sentido llamarla para años YA cerrados (o, para el año
    en curso, sabiendo que se re-pide entera en cada corrida, igual que el
    resto de las series de Senado).

    Si el detalle de ALGUNA acta del listado falla, devuelve None (detalle
    incompleto): el caller cachea años cerrados como inmutables y un año
    con agujeros contaminaría la serie histórica para siempre."""
    session = _hcdn_votaciones_session()
    actas = _descubrir_actas_senado(session, anio)
    if actas is None:
        return None
    detalle = []
    fallidas = 0
    for acta in actas:
        r = _paced_get(session, SENADO_BASE, f"/votaciones/detalleActa/{acta['id']}")
        if r is None:
            # Cada acta viene del listado oficial del año: su detalle DEBE
            # existir, así que esto es un fallo transitorio, no un hueco.
            fallidas += 1
            continue
        filas = _parsear_acta(r.text)
        resultado_acta = _alineamiento_por_provincia(filas)
        if resultado_acta:
            detalle.append({
                "fecha": acta["fecha"].strftime("%Y-%m-%d"),
                "provincias": {p: list(ct) for p, ct in resultado_acta.items()},
            })
    if fallidas:
        print(f"  [WARN] alineamiento_senadores_actas_anio {anio}: "
              f"{fallidas} actas con detalle fallido -- detalle incompleto, no se devuelve")
        return None
    return detalle


def _agregar_alineamiento_ventana(detalle: list[dict], referencia: datetime, dias_ventana: int) -> dict | None:
    """Agrega el detalle crudo por acta (fetch_alineamiento_senadores_actas_anio,
    de uno o más años) dentro de la ventana [referencia − dias_ventana,
    referencia] y devuelve {valor, fecha_dato, n_provincias}, o None si
    ninguna acta cae en la ventana. Misma fórmula que
    fetch_alineamiento_senadores_prov (promedio simple entre provincias con
    señal) — factorizada acá para que el backfill mensual pueda recalcularla
    para cada fin de mes sin duplicar la matemática ni volver a pedir red."""
    limite = referencia - timedelta(days=dias_ventana)
    acumulado = {}
    fecha_max = None
    for fila in detalle:
        fecha = datetime.strptime(fila["fecha"], "%Y-%m-%d")
        if fecha < limite or fecha > referencia:
            continue
        for provincia, (coincide, total) in fila["provincias"].items():
            c0, t0 = acumulado.get(provincia, (0, 0))
            acumulado[provincia] = (c0 + coincide, t0 + total)
        fecha_max = fecha if fecha_max is None else max(fecha_max, fecha)
    if not acumulado:
        return None
    ratios_por_provincia = [c / t for c, t in acumulado.values() if t > 0]
    return {
        "valor": round(100 * sum(ratios_por_provincia) / len(ratios_por_provincia), 1),
        "fecha_dato": fecha_max.strftime("%Y-%m-%d"),
        "n_provincias": len(acumulado),
    }


# ── MAGyP — adhesión provincial al RIGI ──────────────────────────────────────

def _provincias_adheridas_rigi() -> set[str] | None:
    """{NOMBRE_PROVINCIA en mayúsculas, ...} de la tabla MAGyP de provincias
    adheridas al RIGI (Título VII, Ley 27.742), o None si el fetch falló.
    parser="html.parser" (stdlib, lxml no está en requirements.txt — ver
    Tarea 4 del plan de cohesion_bloque): el sitio fuente tiene un <tr> vacío
    malformado que con html.parser produce una fila SANTA CRUZ duplicada
    (confirmado en vivo en la investigación previa) — no requiere lxml para
    resolverlo, el resultado ya es un `set()`, así que agregar el mismo
    nombre dos veces es un no-op y el conteo final no se infla. Factorizada
    de fetch_adhesion_reformas_provincial (que solo necesita el conteo) para
    que descargar_series.fetch_adhesion_reformas_provincial_serie pueda
    cruzar los NOMBRES contra las fechas de adhesión investigadas a mano
    (data/politica/adhesion_reformas_provincial_fechas.json, ADR-0044)."""
    try:
        r = requests.get(MAGYP_RIGI_URL, headers=HTTP_HEADERS, timeout=HTTP_TIMEOUT)
    except requests.RequestException:
        return None
    if r.status_code != 200:
        return None
    soup = BeautifulSoup(r.text, "html.parser")
    provincias = set()
    for fila in soup.select("table tr"):
        celdas = fila.find_all("td")
        if len(celdas) < 2:
            continue
        provincia = celdas[0].get_text(strip=True)
        if provincia:
            provincias.add(provincia.upper())
    return provincias or None


def fetch_adhesion_reformas_provincial() -> dict | None:
    """% de provincias (sobre 24) adheridas formalmente al RIGI (Título VII,
    Ley 27.742) — tabla MAGyP. Mide adhesión FISCAL a un régimen puntual, NO
    alineamiento político general — no reemplaza a gobernadores_alineamiento."""
    provincias = _provincias_adheridas_rigi()
    if not provincias:
        return None
    return {
        "valor": round(len(provincias) / 24.0 * 100.0, 1),
        "unidad": "% de jurisdicciones (sobre 24) adheridas al RIGI",
        "fuente": "Tabla de provincias adheridas — Ministerio de Agricultura, Ganadería y Pesca",
        "fecha_dato": datetime.now().strftime("%Y-%m-%d"),
        "n_provincias": len(provincias),
        "desactualizado": False,
    }


# ── Derrotas legislativas del Ejecutivo (vetos insistidos + decretos caídos) ──
#
# Indicador de conteo absoluto en ventana móvil de 12 meses (ADR-0046): cuántas
# veces el Congreso le volteó una norma al Ejecutivo en el recinto. Dos familias
# de eventos, las dos únicas derrotas legislativas TERMINALES del régimen:
#   (a) veto insistido: ambas cámaras insisten con 2/3 (art. 83 CN) y la ley se
#       promulga pese al veto — se detecta vía InfoLeg (el proyecto vetado
#       aparece publicado como Ley con fecha POSTERIOR al decreto de veto);
#   (b) decreto rechazado: una cámara rechaza un DNU/decreto delegado bajo la
#       ley 26.122 — se detecta vía las actas del Senado (título con la fórmula
#       estable "en los términos de la ley 26.122"; en esas actas se vota la
#       VALIDEZ del decreto, así que gana NEGATIVO = rechazo, verificado contra
#       los 8 casos reales 2024-2025).
# Cada norma cuenta UNA vez, fechada en el mes de la derrota consumada
# (insistencia de la segunda cámara / primer rechazo en recinto). El estado
# vive en DERROTAS_EVENTOS_PATH (semilla histórica verificada a mano +
# detección incremental); la serie mensual se deriva determinísticamente del
# registro en descargar_series.fetch_derrotas_legislativas_mensual().
#
# Limitación declarada (ficha): la detección automática de decretos mira solo
# el Senado — un rechazo que ocurra primero en Diputados se registra recién
# con el voto del Senado (o a mano en el registro); en los 32 meses de la
# semilla eso habría corrido la fecha de los 5 decretos de ago-2025 apenas
# dos semanas, sin cambiar ningún conteo mensual publicado salvo el de agosto.

_DERROTAS_FRASES_VETO = (
    # Frases EXACTAS (entre comillas dobles: sin comillas InfoLeg hace OR de
    # palabras y devuelve miles de resultados) que cubren las 3 variantes de
    # sumario observadas en los 10 vetos reales dic-2023→jul-2026, con 0
    # falsos positivos y 0 falsos negativos (verificado en vivo 2026-07-09).
    '"observase en su totalidad"',
    '"observa en su totalidad"',
    '"promulgacion parcial"',
)
_RE_VERNORMA = re.compile(r"verNorma\.do\?(?:[^\"']*?&)?id=(\d+)")
_RE_DECRETO_ITEM = re.compile(r"Decreto\s+(\d+)\s*/\s*(\d{4})", re.IGNORECASE)
_RE_PROYECTO_LEY = re.compile(r"\b(2[5-9])\.?(\d{3})\b")
# Número de decreto en el título de un acta del Senado ("… Nº 340/25 …"). El
# prefijo Nº/N° es obligatorio a propósito: sin él matchearía el expediente
# ("PE-44/25-DC") que acompaña al título en el listado.
_RE_DECRETO_TITULO = re.compile(r"N[º°]\s*(\d+)\s*/\s*(\d{2,4})")
_MESES_INFOLEG = {"ene": 1, "feb": 2, "mar": 3, "abr": 4, "may": 5, "jun": 6,
                  "jul": 7, "ago": 8, "sep": 9, "oct": 10, "nov": 11, "dic": 12}
_RE_FECHA_INFOLEG = re.compile(r"(\d{1,2})-([a-z]{3})-(\d{4})", re.IGNORECASE)
_DERROTAS_PAUSA_INFOLEG = 1.5   # segundos entre POSTs a InfoLeg (mismo pacing
                                # probado en vivo sin bloqueo, 2026-07-09)


def _cargar_derrotas_registro() -> dict | None:
    """Registro de eventos (DERROTAS_EVENTOS_PATH) o None si falta/ilegible.
    utf-8-sig: tolera el BOM de editores/PowerShell de Windows (mismo criterio
    que cargar_ajustes)."""
    try:
        registro = json.loads(DERROTAS_EVENTOS_PATH.read_text(encoding="utf-8-sig"))
    except (OSError, json.JSONDecodeError):
        return None
    if not isinstance(registro, dict) or "vetos" not in registro or "decretos" not in registro:
        return None
    return registro


def _guardar_derrotas_registro(registro: dict) -> None:
    DERROTAS_EVENTOS_PATH.write_text(
        json.dumps(registro, indent=2, ensure_ascii=False), encoding="utf-8")


def _infoleg_abrir_sesion() -> tuple:
    """(session, action_url) para postear a buscarNormas.do — misma mecánica
    de sesión que fetch_ratio_dnu (GET al home para el jsessionid + action URL
    del formulario)."""
    session = requests.Session()
    r = session.get(INFOLEG_HOME, headers=HTTP_HEADERS, timeout=HTTP_TIMEOUT)
    r.raise_for_status()
    m = re.search(r'action="(/infolegInternet/[^"]+)"', r.text)
    if not m:
        raise ValueError("No se encontró la URL del formulario InfoLeg")
    return session, "https://servicios.infoleg.gob.ar" + m.group(1)


def _infoleg_buscar(session: requests.Session, action_url: str, *, tipo: str,
                    texto: str = "", numero: str = "",
                    desde: date | None = None, hasta: date | None = None) -> str:
    """POST a buscarNormas.do dentro de la sesión y devuelve el HTML del
    listado. Valida que la página sea un resultado real (con conteo
    "Encontradas: N" o el texto de cero resultados) — cualquier otra cosa es
    un fallo de la fuente, no "0 vetos". Pacing fijo antes de cada POST."""
    time.sleep(_DERROTAS_PAUSA_INFOLEG)
    data = {
        "tipoNorma": tipo,
        "numero": numero,
        "anioSancion": "",
        "dependencia": "",
        "diaPubDesde": f"{desde.day:02d}" if desde else "",
        "mesPubDesde": f"{desde.month:02d}" if desde else "",
        "anioPubDesde": str(desde.year) if desde else "",
        "diaPubHasta": f"{hasta.day:02d}" if hasta else "",
        "mesPubHasta": f"{hasta.month:02d}" if hasta else "",
        "anioPubHasta": str(hasta.year) if hasta else "",
        "texto": texto,
    }
    r = session.post(action_url, data=data, headers=HTTP_HEADERS, timeout=HTTP_TIMEOUT)
    r.raise_for_status()
    if "Encontradas" not in r.text and "No se encontraron normas" not in r.text:
        raise ValueError(f"respuesta InfoLeg sin listado (tipo={tipo}, texto={texto!r}, numero={numero!r})")
    return r.text


def _fecha_infoleg(texto: str) -> str | None:
    """'12-sep-2025' (formato del listado InfoLeg) → '2025-09-12'."""
    m = _RE_FECHA_INFOLEG.search(texto)
    if not m:
        return None
    mes = _MESES_INFOLEG.get(m.group(2).lower())
    if not mes:
        return None
    return f"{m.group(3)}-{mes:02d}-{int(m.group(1)):02d}"


def _parsear_listado_infoleg(html: str) -> list[dict]:
    """Filas del listado de resultados de buscarNormas.do:
    [{infoleg_id, norma, fecha_pub, sumario}]. Cada resultado es un <tr> con 3
    <td> (número/dependencia con link a verNorma.do, fecha de publicación,
    descripción); el link "Ver Norma y Textos Resaltados" repite el id en la
    misma fila y se deduplica. Página sin resultados → [] (sin error: la
    validación de que la página es un listado real vive en _infoleg_buscar)."""
    soup = BeautifulSoup(html, "html.parser")
    items, vistos = [], set()
    for fila in soup.select("tr"):
        link = fila.find("a", href=_RE_VERNORMA)
        if link is None:
            continue
        m = _RE_VERNORMA.search(link["href"])
        celdas = fila.find_all("td")
        if not m or len(celdas) < 3:
            continue
        iid = m.group(1)
        if iid in vistos:
            continue
        vistos.add(iid)
        items.append({
            "infoleg_id": iid,
            "norma": re.sub(r"\s+", " ", celdas[0].get_text(" ", strip=True)),
            "fecha_pub": _fecha_infoleg(celdas[1].get_text(" ", strip=True)),
            "sumario": re.sub(r"\s+", " ", celdas[2].get_text(" ", strip=True)),
        })
    return items


def _proyectos_de_sumario(sumario: str) -> list[str]:
    """Números de proyecto de ley del sumario de un decreto de veto,
    normalizados con punto de miles ("27794"/"27.794" → "27.794"). Soporta el
    caso multiproyecto (el decreto 534/2025 vetó tres leyes de una vez)."""
    return sorted({f"{m.group(1)}.{m.group(2)}" for m in _RE_PROYECTO_LEY.finditer(sumario)})


def _derrotas_detectar_vetos(session, action_url, registro: dict) -> None:
    """Detecta decretos de veto (observación total/parcial) nuevos vía las 3
    frases exactas y los agrega al registro con insistencia_completa=null.
    Muta `registro` en memoria; el caller persiste al final si todo salió bien."""
    conocidos = {v["proyecto"] for v in registro["vetos"]}
    for frase in _DERROTAS_FRASES_VETO:
        html = _infoleg_buscar(session, action_url, tipo="2", texto=frase,
                               desde=date(2023, 12, 1), hasta=date.today())
        for item in _parsear_listado_infoleg(html):
            m = _RE_DECRETO_ITEM.search(item["norma"])
            proyectos = _proyectos_de_sumario(item["sumario"])
            if not m or not item["fecha_pub"] or not proyectos:
                continue
            decreto = f"{m.group(1)}/{m.group(2)}"
            es_parcial = "PROMULGACION PARCIAL" in item["sumario"].upper()
            for proyecto in proyectos:
                if proyecto in conocidos:
                    continue
                conocidos.add(proyecto)
                registro["vetos"].append({
                    "proyecto": proyecto,
                    "tema": item["sumario"][:200],
                    "decreto": decreto,
                    "infoleg_id_decreto": int(item["infoleg_id"]),
                    "fecha_veto": item["fecha_pub"],
                    "tipo": "parcial" if es_parcial else "total",
                    "insistencia_completa": None,
                    "fuente_insistencia": None,
                    "detalle": "Detectado automáticamente en InfoLeg.",
                })


def _derrotas_detectar_insistencias(session, action_url, registro: dict) -> None:
    """Verifica si algún veto aún no insistido flipeó: el proyecto vetado
    aparece publicado como Ley en InfoLeg con fecha POSTERIOR al decreto de
    veto (una ley publicada el MISMO día es la promulgación parcial del propio
    decreto, no una insistencia — caso real 27.739). Se re-chequean TODOS los
    vetos sin insistencia, estén o no en ventana: media insistencia pendiente
    (27.790/27.794) no caduca y puede completarse en cualquier momento; el
    evento nuevo se fecharía en el mes del flip, no en el del veto. La fecha
    usada es la de publicación en el B.O. (proxy de la insistencia de la
    segunda cámara: rezago observado 18-19 días, mismo mes en los 3 casos
    reales de 2025 — limitación declarada en la ficha)."""
    for veto in registro["vetos"]:
        if veto.get("insistencia_completa"):
            continue   # inmutable: una insistencia consumada no se des-consuma
        numero = veto["proyecto"].replace(".", "")
        html = _infoleg_buscar(session, action_url, tipo="1", numero=numero)
        for item in _parsear_listado_infoleg(html):
            if item["fecha_pub"] and item["fecha_pub"] > veto["fecha_veto"]:
                veto["insistencia_completa"] = item["fecha_pub"]
                veto["fuente_insistencia"] = ("fecha de publicación en el B.O. de la ley "
                                              "insistida (InfoLeg, detección automática)")
                break


def _actas_senado_26122(session: requests.Session, anio: int):
    """[{id, fecha, titulo}] de las actas del Senado del año cuyo título
    contiene la fórmula estable "en los términos de la ley 26.122" (presente
    en los 8 tratamientos reales de decretos 2024-2025 y ausente en los falsos
    amigos: mociones de orden, el decreto simple 681/25, la reforma de la
    propia ley). Mismo POST/parseo de filas que _descubrir_actas_senado, pero
    conservando el texto de la fila (el título de la votación).
    None si el request falló (distinto de 'sin actas con match' = [])."""
    r = _paced_post(session, SENADO_BASE, "/votaciones/actas",
                    data={"busqueda_actas[anio]": str(anio)})
    if r is None:
        return None
    soup = BeautifulSoup(r.text, "html.parser")
    actas, vistos = [], set()
    for fila in soup.select("tr"):
        link = fila.find("a", href=_RE_DETALLE_ACTA_SENADO)
        span_fecha = fila.find("span", style=lambda s: s and _RE_DISPLAY_NONE.search(s))
        if link is None or span_fecha is None:
            continue
        m = _RE_DETALLE_ACTA_SENADO.search(link["href"])
        if not m:
            continue
        id_acta = m.group(1)
        try:
            fecha = datetime.strptime(span_fecha.get_text(strip=True), "%Y%m%d")
        except ValueError:
            continue
        if fecha.year != anio or id_acta in vistos:
            continue
        titulo = re.sub(r"\s+", " ", fila.get_text(" ", strip=True))
        # el texto de la fila arrastra la fecha oculta y el bloque "Ver
        # Expedientes ..." del listado — se recorta lo segundo para guardar
        # un título legible (la fecha oculta no molesta al filtro)
        titulo = titulo.split("Ver Expedientes")[0].strip()
        if "26.122" not in titulo:
            continue
        vistos.add(id_acta)
        actas.append({"id": id_acta, "fecha": fecha, "titulo": titulo})
    return actas


def _clave_decreto_de_titulo(titulo: str) -> str | None:
    """'… Decreto … Nº 340/25 …' → '340/2025' (clave normalizada del registro).
    None si el título no trae un número con prefijo Nº/N° (se avisa y se deja
    para revisión manual — no se adivina)."""
    m = _RE_DECRETO_TITULO.search(titulo)
    if not m:
        return None
    numero, anio = m.group(1), m.group(2)
    if len(anio) == 2:
        anio = f"20{anio}"
    return f"{int(numero)}/{anio}"


def _tipo_decreto_de_titulo(titulo: str) -> str:
    t = titulo.lower()
    if "facultades delegadas" in t:
        return "delegado"
    if "necesidad" in t and "urgencia" in t:
        return "DNU"
    return "decreto"


def _derrotas_detectar_decretos(registro: dict) -> None:
    """Detecta tratamientos nuevos de decretos bajo la ley 26.122 en el recinto
    del Senado (año en curso + anterior; las actas ya procesadas quedan en
    actas_senado_vistas y no se re-piden). En estas actas se vota la VALIDEZ
    del decreto: gana NEGATIVO = rechazo (dirección verificada contra los 8
    casos reales 2024-2025; una votación sin votos A/N registrados es un acta
    anulada y se ignora). Decretos anteriores a 2023 (la bicameral a veces
    trata decretos viejos de gestiones anteriores, caso 829/19 en 2024) quedan
    fuera de alcance. Muta `registro`; el caller persiste al final."""
    session = _hcdn_votaciones_session()
    vistos = registro.setdefault("actas_senado_vistas", {})
    hoy = date.today()
    for anio in sorted({hoy.year - 1, hoy.year}):
        actas = _actas_senado_26122(session, anio)
        if actas is None:
            raise ValueError(f"listado de actas del Senado {anio} inaccesible")
        for acta in actas:
            if acta["id"] in vistos:
                continue
            clave = _clave_decreto_de_titulo(acta["titulo"])
            if clave is None:
                # no se marca vista: el aviso se repite cada corrida hasta que
                # alguien resuelva el caso a mano (preferible a perder una
                # derrota en silencio)
                print(f"[WARN] {CINTURON}.derrotas_legislativas: acta {acta['id']} con fórmula "
                      f"26.122 pero sin número de decreto legible — revisar a mano: {acta['titulo'][:140]}")
                continue
            if int(clave.split("/")[1]) < 2023:
                vistos[acta["id"]] = f"{clave}: decreto de gestión anterior — fuera de alcance (dic-2023→)"
                continue
            r = _paced_get(session, SENADO_BASE, f"/votaciones/detalleActa/{acta['id']}")
            if r is None:
                raise ValueError(f"detalleActa {acta['id']} del Senado inaccesible")
            filas = _parsear_acta(r.text)
            afirm = sum(1 for f in filas if f["voto"] == "AFIRMATIVO")
            neg = sum(1 for f in filas if f["voto"] == "NEGATIVO")
            if afirm == 0 and neg == 0:
                vistos[acta["id"]] = f"{clave}: votación sin votos registrados (anulada/rehecha) — sin efecto"
                continue
            if neg > afirm:
                entry = next((d for d in registro["decretos"] if d.get("clave") == clave), None)
                if entry is None:
                    entry = {"clave": clave, "etiqueta": f"Decreto {clave}",
                             "tipo": _tipo_decreto_de_titulo(acta["titulo"]),
                             "tema": acta["titulo"], "publicado_bo": None, "rechazos": [],
                             "estado": "rechazado por el Senado",
                             "detalle": "Detectado automáticamente en las actas del Senado."}
                    registro["decretos"].append(entry)
                fecha = acta["fecha"].strftime("%Y-%m-%d")
                if not any(rz.get("camara") == "Senado" and rz.get("fecha") == fecha
                           for rz in entry["rechazos"]):
                    entry["rechazos"].append({"fecha": fecha, "camara": "Senado",
                                              "acta": acta["id"],
                                              "votos": f"{neg} rechazo - {afirm} validez"})
                vistos[acta["id"]] = f"{clave}: RECHAZO ({neg}-{afirm})"
            else:
                vistos[acta["id"]] = f"{clave}: aprobación/validez ({afirm}-{neg}) — no es derrota"


def _derrotas_eventos(registro: dict) -> list[dict]:
    """Eventos de derrota consumada derivados del registro, cada norma UNA vez:
    [{fecha, tipo: veto_insistido | decreto_rechazado, nombre}] ascendente.
    Veto → fecha de la insistencia completa; decreto → fecha del PRIMER rechazo
    en recinto (el segundo consuma la derogación pero es la misma derrota)."""
    eventos = []
    for v in registro.get("vetos", []):
        if v.get("insistencia_completa"):
            eventos.append({"fecha": v["insistencia_completa"], "tipo": "veto_insistido",
                            "nombre": f"ley {v['proyecto']}"})
    for d in registro.get("decretos", []):
        fechas = [rz["fecha"] for rz in d.get("rechazos", []) if rz.get("fecha")]
        if fechas:
            eventos.append({"fecha": min(fechas), "tipo": "decreto_rechazado",
                            "nombre": d.get("etiqueta") or d.get("clave", "?")})
    return sorted(eventos, key=lambda e: e["fecha"])


def _derrotas_conteo_12m(eventos: list[dict], referencia: date) -> tuple:
    """(total, n_vetos, n_decretos, fecha_ultimo_evento) en la ventana de los
    12 meses calendario que terminan en el mes de `referencia` (inclusive) —
    misma ventana con la que descargar_series deriva la serie mensual, así
    card y serie cuentan igual."""
    meses = referencia.year * 12 + (referencia.month - 1)
    desde = meses - 11
    ym_desde = f"{desde // 12}-{desde % 12 + 1:02d}"
    ym_hasta = f"{referencia.year}-{referencia.month:02d}"
    en_ventana = [e for e in eventos if ym_desde <= e["fecha"][:7] <= ym_hasta]
    n_vetos = sum(1 for e in en_ventana if e["tipo"] == "veto_insistido")
    ultimo = max((e["fecha"] for e in en_ventana), default=None)
    return len(en_ventana), n_vetos, len(en_ventana) - n_vetos, ultimo


def _derrotas_detalle_txt(n_vetos: int, n_decretos: int) -> str:
    """Detalle legible de la card ("valor usado" del modal, vía publicar.py)."""
    if n_vetos == 0 and n_decretos == 0:
        return ("sin derrotas legislativas en los últimos 12 meses "
                "(ni vetos insistidos ni decretos rechazados en el recinto)")
    txt_vetos = (f"{n_vetos} veto{'s' if n_vetos != 1 else ''} "
                 f"insistido{'s' if n_vetos != 1 else ''} por el Congreso")
    txt_decretos = (f"{n_decretos} decreto{'s' if n_decretos != 1 else ''} "
                    f"rechazado{'s' if n_decretos != 1 else ''} en el recinto")
    return f"{txt_vetos} + {txt_decretos} en los últimos 12 meses"


def fetch_derrotas_legislativas() -> dict | None:
    """
    Derrotas legislativas del Ejecutivo en los últimos 12 meses: vetos
    insistidos por ambas cámaras + decretos (DNU/delegados) rechazados por al
    menos una cámara bajo la ley 26.122. Conteo absoluto, cada norma una vez,
    fechada en el mes de la derrota consumada. Menor = mejor.

    Fuentes: InfoLeg (decretos de veto + leyes promulgadas por insistencia,
    misma sesión que fetch_ratio_dnu) y actas de votación del Senado (misma
    mecánica de POST paginado que cohesion_bloque_senado). El estado vive en
    DERROTAS_EVENTOS_PATH; la corrida hace detección INCREMENTAL (≈5 POSTs a
    InfoLeg + 1-2 al Senado por corrida) y solo persiste el registro si las
    tres etapas llegaron a las fuentes — un fallo degrada al caché del
    snapshot anterior (desactualizado=True), sin corromper el registro.
    """
    try:
        registro = _cargar_derrotas_registro()
        if registro is None:
            raise ValueError(f"registro de eventos ausente o ilegible ({DERROTAS_EVENTOS_PATH})")
        session, action_url = _infoleg_abrir_sesion()
        _derrotas_detectar_vetos(session, action_url, registro)
        _derrotas_detectar_insistencias(session, action_url, registro)
        _derrotas_detectar_decretos(registro)
        registro.setdefault("_meta", {})["actualizado"] = str(date.today())
        _guardar_derrotas_registro(registro)

        eventos = _derrotas_eventos(registro)
        total, n_vetos, n_decretos, ultimo = _derrotas_conteo_12m(eventos, date.today())
        return {
            "valor": total,
            "n_vetos_12m": n_vetos,
            "n_decretos_12m": n_decretos,
            "n_eventos_historico": len(eventos),
            "ultimo_evento": ultimo,
            "detalle_txt": _derrotas_detalle_txt(n_vetos, n_decretos),
            "unidad": "Derrotas del Ejecutivo en el recinto, últimos 12 meses (vetos insistidos + decretos rechazados)",
            "fuente": "InfoLeg (vetos e insistencias) + actas de votación del Senado (decretos, ley 26.122) — elaboración CIGOB",
            "fecha_dato": str(date.today()),
            "desactualizado": False,
        }

    except Exception as e:
        _warn("derrotas_legislativas", str(e))
        return None

# ── Bloqueo sostenido (ADR-0069) ──────────────────────────────────────────────
#
# bloqueo_sostenido = % de normas propias DESAFIADAS en el recinto que el
# Ejecutivo mantuvo en pie, ventana de 12 meses calendario (la misma de
# derrotas_legislativas). Es el espejo de derrotas: aquél cuenta las derrotas
# consumadas en términos absolutos y nunca acredita los bloqueos GANADOS (los
# vetos sostenidos de sep/oct-2024 no puntúan en ningún lado); éste mide la
# TASA de supervivencia sobre el total de desafíos votados.
#
# Comparte el registro versionado de eventos (DERROTAS_EVENTOS_PATH) y sus
# detecciones ya existentes (vetos por InfoLeg, insistencias completas por
# B.O., rechazos de decretos por actas del Senado — fetch_derrotas_legislativas
# corre ANTES en el mismo main()) y le suma la detección de los DESAFÍOS:
#
#   * Diputados (clasificador de actas PDF, ids del caché permanente de
#     cohesión): una insistencia de veto se reconoce en dos formatos reales
#     — desde 2025, motivo "INSISTENCIA PROYECTO DE LEY 27.791." (la ley
#     directa, sin lookup); en 2024, motivo "EXPTE. N-PE-AAAA" a secas con
#     mayoría "Dos tercios" (el expediente del MENSAJE del PE que comunicó
#     el veto — se mapea a la ley vetada buscando ese expediente en
#     proyectos-parlamentarios del CKAN: TIPO "MENSAJE" con "OBSERVA" en el
#     título, verificado contra 0015-PE-2024→27.756 y 0017-PE-2024→27.757).
#     Las HABILITACIONES de tratamiento (2/3 procesal previo, mismo
#     expediente) se excluyen por palabra clave: en la sesión del
#     20-ago-2025 la habilitación de la 27.791 salió AFIRMATIVA y la
#     insistencia real NEGATIVA — confundirlas invierte el registro. El
#     "Resultado de Votación" impreso en el PDF decide (ya incorpora la
#     regla de los 2/3): NEGATIVO = insistencia rechazada (veto sostenido),
#     AFIRMATIVO = la cámara insistió (la caída solo se consuma con la
#     segunda cámara, vía InfoLeg como siempre). Una votación de decreto se
#     reconoce por motivo "DECRETO […] N° X/AA" o "DNU X/AAAA"; la moción
#     estándar de la bicameral es el RECHAZO (AFIRMATIVO lo consuma —
#     dirección verificada contra los 7 casos reales de Diputados
#     2024-2025), y un motivo con "APRUEBA"/"VALIDEZ" (dictamen de
#     aprobación, caso real DNU 179/2025 del FMI) queda PENDIENTE con
#     aviso: la dirección es ambigua y no se adivina. Solo se anotan
#     votaciones sobre decretos YA presentes en el registro (detección del
#     Senado o semilla): un número que el registro no conoce puede ser un
#     decreto simple fuera de la ley 26.122 (caso real 681/25) y queda
#     pendiente para triage manual.
#
#   * Senado (listado anual de actas, misma mecánica que las 26.122): las
#     insistencias se titulan "Insistencia …" — se excluyen las
#     habilitaciones de tratamiento sobre tablas — y casi siempre traen el
#     número de ley en el título (27.793/27.794/27.795/27.796 verificados);
#     un título sin número matcheable avisa en cada corrida hasta resolverse
#     a mano (caso real: 27.790 Bahía Blanca, resuelto en la semilla).
#     Insistió = afirmativos ≥ 2/3 de los votos emitidos (art. 83 CN).
#
# El avance por Diputados es un watermark por id de acta
# (registro["actas_diputados_bloqueo"]["clasificadas_hasta_id"]): cada acta
# se descarga y clasifica UNA vez en la vida del proyecto (las anteriores a
# la gestión ni se descargan: la fecha ya está en el caché de cohesión); las
# ambiguas quedan en "pendientes" y se reintentan/avisan en cada corrida.
# Limitaciones declaradas (ficha): el universo de ids es el caché del walk
# de cohesión (un acta publicada hoy se clasifica cuando el walk nocturno la
# cachea), y un mes sin desafíos votados en la ventana no genera dato — el
# motor renormaliza, igual que veto_quorum entre períodos.

_RE_EXPTE_PE_MOTIVO = re.compile(r"\bEXPTE\.?\s*0*(\d{1,4})-PE-(\d{4})\b", re.IGNORECASE)
_RE_DECRETO_MOTIVO = re.compile(
    r"(?:DECRETO(?:\s+DE\s+(?:NECESIDAD\s+Y\s+URGENCIA|FACULTADES\s+DELEGADAS))?\s*N\s*[º°]\s*"
    r"|DNU\s+)(\d{1,4})\s*/\s*(\d{2,4})\b", re.IGNORECASE)
_BLOQUEO_ERA_DESDE = "2023-12-10"   # asunción — actas anteriores ni se descargan


def _dedup_tokens_dobles(texto: str) -> str:
    """'MMááss ddee llaa mmiittaadd' → 'Más de la mitad'. La negrita simulada del
    generador de PDF duplica CADA carácter dentro de cada palabra (los
    espacios quedan simples); solo se colapsa un token si viene entero en
    pares, para no romper palabras legítimas con letras dobles."""
    def _colapsar(tok: str) -> str:
        if len(tok) >= 2 and len(tok) % 2 == 0 and all(
                tok[i] == tok[i + 1] for i in range(0, len(tok), 2)):
            return tok[::2]
        return tok
    return " ".join(_colapsar(t) for t in texto.split())


def _parsear_encabezado_acta_diputados(contenido: bytes) -> dict | None:
    """{fecha ("YYYY-MM-DD"), motivo, mayoria, resultado, votos_txt} del
    encabezado de la página 1 del PDF de un acta de Diputados. El motivo son
    las líneas entre la línea de sesión ("143° - Período … Reunión") y la
    línea "Acta Nº …" (títulos largos ocupan más de una línea y se unen);
    mayoría y resultado vienen con carácter duplicado y se normalizan token a
    token. None si falta algún campo (acta ilegible para clasificar)."""
    with pdfplumber.open(io.BytesIO(contenido)) as pdf:
        texto = pdf.pages[0].extract_text() or ""
    m_fecha = re.search(r"Fecha:\s*(\d{2}/\d{2}/\d{4})", texto)
    m_mayoria = re.search(r"Tipo Mayor[íi]a:\s*(.+?)\s*Miembros", texto)
    m_resultado = re.search(r"Resultado de Votaci[óo]n:\s*(\S+)", texto)
    lineas = [l.strip() for l in texto.split("\n")]
    ini = next((i for i, l in enumerate(lineas) if re.search(r"\d+°\s*-\s*Per[íi]odo", l)), None)
    fin = next((i for i, l in enumerate(lineas) if re.match(r"Acta\s*N", l)), None)
    if not (m_fecha and m_mayoria and m_resultado) or ini is None or fin is None or fin <= ini:
        return None
    try:
        fecha = datetime.strptime(m_fecha.group(1), "%d/%m/%Y").strftime("%Y-%m-%d")
    except ValueError:
        return None
    votos = {}
    for campo in ("Afirmativos", "Negativos", "Abstenciones"):
        m = re.search(rf"{campo}\s+(\d+)", texto)
        votos[campo] = int(m.group(1)) if m else None
    votos_txt = (f"{votos['Afirmativos']} afirmativos - {votos['Negativos']} negativos - "
                 f"{votos['Abstenciones']} abst.") if None not in votos.values() else None
    return {
        "fecha": fecha,
        "motivo": " ".join(lineas[ini + 1:fin]).strip(),
        "mayoria": _dedup_tokens_dobles(m_mayoria.group(1)),
        "resultado": _dedup_tokens_dobles(m_resultado.group(1)).upper(),
        "votos_txt": votos_txt,
    }


def _mensaje_pe_a_leyes(numero: int, anio: str) -> list[str] | None:
    """Leyes observadas por el MENSAJE del PE con expediente NNNN-PE-AAAA
    (dataset proyectos-parlamentarios del CKAN): ['27.756', …]. Lista vacía
    si el expediente existe pero no es un mensaje de veto (tratado, proyecto
    del PE, etc. — la votación 2/3 era otra cosa); None si el CKAN no
    devolvió ninguna fila para el expediente (no se puede clasificar)."""
    expte = f"{numero:04d}-PE-{anio}"
    filas = _hcdn_paginate(HCDN_PROYECTOS_RID, q=expte)
    encontrado = False
    leyes: set[str] = set()
    for r in filas:
        if str(r.get("EXP_DIPUTADOS", "")).strip() != expte:
            continue
        encontrado = True
        if str(r.get("TIPO", "")).strip().upper() != "MENSAJE":
            continue
        titulo = str(r.get("TITULO", ""))
        if "OBSERVA" not in titulo.upper():
            continue
        leyes.update(f"{m.group(1)}.{m.group(2)}" for m in _RE_PROYECTO_LEY.finditer(titulo))
    if not encontrado:
        return None
    return sorted(leyes)


def _bloqueo_clasificar_acta_diputados(session, registro: dict, id_acta: int) -> tuple:
    """Clasifica UNA acta de Diputados para bloqueo_sostenido, mutando el
    registro si produce un evento. Devuelve (resuelta, nota): resuelta=False
    deja el acta en pendientes (se reintenta y avisa en cada corrida);
    nota=None es un acta sin interés (solo avanza el watermark). Levanta
    ValueError ante un fallo transitorio de descarga (el caller corta el
    walk y reintenta mañana desde el watermark)."""
    contenido = _diputados_acta_pdf(session, id_acta)
    if contenido is _ACTA_FALLO:
        raise ValueError(f"acta {id_acta} de Diputados inaccesible (fallo transitorio)")
    if contenido is _ACTA_NO_EXISTE:
        return True, None   # hueco de numeración genuino
    enc = _parsear_encabezado_acta_diputados(contenido)
    if enc is None:
        return True, "encabezado ilegible — sin motivo/mayoría/resultado parseable"
    if enc["fecha"] < _BLOQUEO_ERA_DESDE:
        return True, None
    motivo_u = enc["motivo"].upper()

    # Las HABILITACIONES de tratamiento son la moción procesal previa (2/3
    # sobre el mismo expediente del mensaje del PE) — NO son la insistencia:
    # en la sesión del 20-ago-2025 la habilitación de la 27.791 salió
    # AFIRMATIVA (159-75, acta 5736) y la insistencia real fue NEGATIVA
    # (160-83, acta 5737). Confundirlas invierte el resultado registrado.
    if "HABILITACI" in motivo_u or "SOBRE TABLAS" in motivo_u:
        return True, None

    # ── votación de decreto (validez/rechazo, ley 26.122) ────────────────
    m_dec = _RE_DECRETO_MOTIVO.search(enc["motivo"])
    if m_dec:
        anio_dec = m_dec.group(2)
        if len(anio_dec) == 2:
            anio_dec = f"20{anio_dec}"
        clave = f"{int(m_dec.group(1))}/{anio_dec}"
        entry = next((d for d in registro["decretos"] if d.get("clave") == clave), None)
        if entry is None:
            print(f"[WARN] {CINTURON}.bloqueo_sostenido: acta {id_acta} de Diputados vota el "
                  f"decreto {clave} que el registro no conoce (¿decreto simple fuera de la "
                  f"26.122? ¿el Senado aún no lo trató?) — pendiente de triage manual")
            return False, f"decreto {clave} desconocido para el registro"
        ya = any(x.get("camara") == "Diputados" and x.get("fecha") == enc["fecha"]
                 for x in entry.get("rechazos", []) + entry.get("sostenimientos", []))
        if ya:
            return True, f"{clave}: votación ya reflejada en el registro"
        if "APRUEBA" in motivo_u or "VALIDEZ" in motivo_u:
            print(f"[WARN] {CINTURON}.bloqueo_sostenido: acta {id_acta} de Diputados sobre el "
                  f"decreto {clave} parece un dictamen de APROBACIÓN (dirección de la moción "
                  f"ambigua) — pendiente de triage manual: {enc['motivo'][:140]}")
            return False, f"decreto {clave}: posible dictamen de aprobación"
        # moción estándar de la bicameral: se vota el RECHAZO del decreto
        if enc["resultado"] == "AFIRMATIVO":
            entry["rechazos"].append({"fecha": enc["fecha"], "camara": "Diputados",
                                      "acta": str(id_acta), "votos": enc["votos_txt"]})
            return True, f"{clave}: RECHAZO de Diputados ({enc['votos_txt']})"
        entry.setdefault("sostenimientos", []).append({
            "fecha": enc["fecha"], "camara": "Diputados", "acta": str(id_acta),
            "votos": enc["votos_txt"], "tipo": "rechazo_fracasado"})
        return True, f"{clave}: rechazo FRACASADO en Diputados ({enc['votos_txt']}) — decreto en pie"

    # ── insistencia de veto, formato 2025+: la ley en el motivo ───────────
    # "INSISTENCIA PROYECTO DE LEY 27.791." (actas 5735/5737/5762/5766) —
    # asociación directa, sin CKAN. El resultado impreso ya incorpora la
    # regla de los 2/3 (la 27.791 salió NEGATIVO con 160-83, 65,8% > mitad
    # pero < 2/3).
    m_ley = _RE_PROYECTO_LEY.search(enc["motivo"])
    if "INSISTENCIA" in motivo_u and m_ley:
        ley = f"{m_ley.group(1)}.{m_ley.group(2)}"
        veto = next((v for v in registro["vetos"] if v.get("proyecto") == ley), None)
        if veto is None:
            print(f"[WARN] {CINTURON}.bloqueo_sostenido: acta {id_acta} insiste la ley {ley} "
                  f"pero el registro no tiene ese veto — pendiente de triage manual")
            return False, f"insistencia sobre veto no registrado ({ley})"
        if any(iv.get("camara") == "Diputados" and iv.get("fecha") == enc["fecha"]
               for iv in veto.get("insistencias_votadas", [])):
            return True, f"{ley}: insistencia ya reflejada en el registro"
        resultado = "insistio_camara" if enc["resultado"] == "AFIRMATIVO" else "insistencia_rechazada"
        veto.setdefault("insistencias_votadas", []).append({
            "fecha": enc["fecha"], "camara": "Diputados", "resultado": resultado,
            "fuente": f"acta {id_acta} votaciones.hcdn.gob.ar ({enc['resultado']}, "
                      f"{enc['votos_txt']})"})
        legible = "la cámara INSISTIÓ" if resultado == "insistio_camara" else "insistencia rechazada (veto sostenido)"
        return True, f"{ley}: {legible} ({enc['votos_txt']})"

    # ── insistencia de veto, formato 2024: solo el expediente del mensaje ─
    # Actas 5354/5400: motivo "EXPTE. 15-PE-2024" a secas, mayoría 2/3 — se
    # mapea a la ley vetada vía el mensaje del PE en el CKAN. RESTRINGIDO a
    # 2024 a propósito: desde 2025 este formato es el de las habilitaciones
    # (ya excluidas por palabra clave), así que un expediente pelado con 2/3
    # en 2025+ es un formato inesperado y queda pendiente, no se adivina.
    m_pe = _RE_EXPTE_PE_MOTIVO.search(enc["motivo"])
    if m_pe and "tercios" in enc["mayoria"].lower() and "MOCION" not in motivo_u \
            and "MOCIÓN" not in motivo_u:
        if enc["fecha"] >= "2025-01-01":
            print(f"[WARN] {CINTURON}.bloqueo_sostenido: acta {id_acta} ({enc['fecha']}) con "
                  f"2/3 sobre {m_pe.group(0)} sin palabra clave — formato inesperado, "
                  f"pendiente de triage manual: {enc['motivo'][:120]}")
            return False, f"2/3 sobre {m_pe.group(0)} sin palabra clave (formato inesperado)"
        leyes = _mensaje_pe_a_leyes(int(m_pe.group(1)), m_pe.group(2))
        if leyes is None:
            print(f"[WARN] {CINTURON}.bloqueo_sostenido: acta {id_acta} (2/3 sobre "
                  f"{m_pe.group(0)}) sin fila en el CKAN — pendiente de triage manual")
            return False, f"expediente {m_pe.group(0)} sin fila en el CKAN"
        if not leyes:
            return True, None   # 2/3 sobre un expediente PE que no es mensaje de veto
        if len(leyes) > 1:
            print(f"[WARN] {CINTURON}.bloqueo_sostenido: acta {id_acta} — el mensaje "
                  f"{m_pe.group(0)} observa varias leyes ({', '.join(leyes)}), no se puede "
                  f"atribuir la insistencia sola — pendiente de triage manual")
            return False, f"mensaje multiproyecto ({', '.join(leyes)})"
        ley = leyes[0]
        veto = next((v for v in registro["vetos"] if v.get("proyecto") == ley), None)
        if veto is None:
            print(f"[WARN] {CINTURON}.bloqueo_sostenido: acta {id_acta} insiste la ley {ley} "
                  f"pero el registro no tiene ese veto — pendiente de triage manual")
            return False, f"insistencia sobre veto no registrado ({ley})"
        if any(iv.get("camara") == "Diputados" and iv.get("fecha") == enc["fecha"]
               for iv in veto.get("insistencias_votadas", [])):
            return True, f"{ley}: insistencia ya reflejada en el registro"
        resultado = "insistio_camara" if enc["resultado"] == "AFIRMATIVO" else "insistencia_rechazada"
        veto.setdefault("insistencias_votadas", []).append({
            "fecha": enc["fecha"], "camara": "Diputados", "resultado": resultado,
            "fuente": f"acta {id_acta} votaciones.hcdn.gob.ar ({enc['resultado']}, "
                      f"{enc['votos_txt']})"})
        legible = "la cámara INSISTIÓ" if resultado == "insistio_camara" else "insistencia rechazada (veto sostenido)"
        return True, f"{ley}: {legible} ({enc['votos_txt']})"

    return True, None


def _bloqueo_clasificar_diputados(registro: dict) -> None:
    """Clasificación incremental de las actas de Diputados: procesa en orden
    ascendente los ids del caché de cohesión posteriores al watermark, más
    los pendientes de corridas anteriores. Persiste el registro cada 25 actas
    (el backfill inicial camina ~430 PDFs a 0,3s c/u y no debe perder el
    progreso si se corta — mismo criterio crash-safe que
    _acta_diputados_cacheada). Un fallo transitorio corta el walk con aviso:
    el watermark queda en el último id resuelto y mañana se retoma."""
    session = _hcdn_votaciones_session()
    cache = _cargar_cache_cohesion_diputados()
    estado = registro.setdefault(
        "actas_diputados_bloqueo", {"clasificadas_hasta_id": 0, "pendientes": {}})
    pendientes = estado.setdefault("pendientes", {})
    notas = estado.setdefault("notas", {})
    watermark = int(estado.get("clasificadas_hasta_id") or 0)
    fechas_cache = {int(k): (v.get("fecha") or "") for k, v in cache.items()
                    if str(k).isdigit() and isinstance(v, dict)}
    nuevos = sorted(i for i in fechas_cache if i > watermark)
    retries = sorted(int(i) for i in pendientes if str(i).isdigit())
    procesadas = 0
    try:
        for id_acta in retries + nuevos:
            es_nuevo = id_acta > watermark
            if es_nuevo and fechas_cache.get(id_acta, "") < _BLOQUEO_ERA_DESDE:
                watermark = id_acta   # anterior a la gestión: ni se descarga
                continue
            try:
                resuelta, nota = _bloqueo_clasificar_acta_diputados(session, registro, id_acta)
            except Exception as e:
                # transitorio (PDF inaccesible, CKAN caído): cortar acá — el
                # watermark queda en el último id resuelto y mañana se retoma
                print(f"  [WARN] bloqueo_sostenido: {e} — el walk corta acá y reintenta mañana")
                break
            if resuelta:
                pendientes.pop(str(id_acta), None)
                if nota:
                    notas[str(id_acta)] = nota
            else:
                pendientes[str(id_acta)] = nota or "pendiente de triage manual"
            if es_nuevo:
                watermark = id_acta
            procesadas += 1
            if procesadas % 25 == 0:
                estado["clasificadas_hasta_id"] = watermark
                _guardar_derrotas_registro(registro)
    finally:
        estado["clasificadas_hasta_id"] = watermark


def _actas_senado_insistencias(session: requests.Session, anio: int):
    """[{id, fecha, titulo}] de las actas del Senado del año cuyo título es
    una INSISTENCIA de veto (palabra "insistencia", excluyendo las
    habilitaciones de tratamiento sobre tablas — que son la moción procesal
    previa, no la insistencia misma). Mismo POST/parseo que
    _actas_senado_26122. None si el request falló."""
    r = _paced_post(session, SENADO_BASE, "/votaciones/actas",
                    data={"busqueda_actas[anio]": str(anio)})
    if r is None:
        return None
    soup = BeautifulSoup(r.text, "html.parser")
    actas, vistos = [], set()
    for fila in soup.select("tr"):
        link = fila.find("a", href=_RE_DETALLE_ACTA_SENADO)
        span_fecha = fila.find("span", style=lambda s: s and _RE_DISPLAY_NONE.search(s))
        if link is None or span_fecha is None:
            continue
        m = _RE_DETALLE_ACTA_SENADO.search(link["href"])
        if not m:
            continue
        id_acta = m.group(1)
        try:
            fecha = datetime.strptime(span_fecha.get_text(strip=True), "%Y%m%d")
        except ValueError:
            continue
        if fecha.year != anio or id_acta in vistos:
            continue
        titulo = re.sub(r"\s+", " ", fila.get_text(" ", strip=True))
        titulo = titulo.split("Ver Expedientes")[0].strip()
        t = titulo.lower()
        if not re.search(r"\binsistencia\b", t) or "sobre tablas" in t or "habilitaci" in t:
            continue
        vistos.add(id_acta)
        actas.append({"id": id_acta, "fecha": fecha, "titulo": titulo})
    return actas


def _bloqueo_detectar_insistencias_senado(registro: dict) -> None:
    """Detecta insistencias de veto votadas en el recinto del Senado (desde
    2024, backfill inherente: las actas ya vistas quedan en
    actas_senado_insistencia_vistas y no se re-piden). Insistió = afirmativos
    ≥ 2/3 de los votos emitidos (art. 83 CN). Un título sin número de ley
    matcheable avisa en cada corrida hasta resolverse a mano en la semilla
    (caso real 27.790 Bahía Blanca). Muta `registro`; el caller persiste."""
    session = _hcdn_votaciones_session()
    vistos = registro.setdefault("actas_senado_insistencia_vistas", {})
    for anio in range(2024, date.today().year + 1):
        actas = _actas_senado_insistencias(session, anio)
        if actas is None:
            raise ValueError(f"listado de actas del Senado {anio} inaccesible")
        for acta in actas:
            if acta["id"] in vistos:
                continue
            m = _RE_PROYECTO_LEY.search(acta["titulo"])
            if not m:
                # no se marca vista: el aviso se repite cada corrida hasta que
                # alguien la resuelva a mano (mismo criterio que las 26.122
                # sin número de decreto legible)
                print(f"[WARN] {CINTURON}.bloqueo_sostenido: acta {acta['id']} del Senado es "
                      f"una insistencia sin número de ley legible — revisar a mano: "
                      f"{acta['titulo'][:140]}")
                continue
            ley = f"{m.group(1)}.{m.group(2)}"
            veto = next((v for v in registro["vetos"] if v.get("proyecto") == ley), None)
            if veto is None:
                print(f"[WARN] {CINTURON}.bloqueo_sostenido: acta {acta['id']} del Senado "
                      f"insiste la ley {ley} pero el registro no tiene ese veto — revisar a mano")
                continue
            r = _paced_get(session, SENADO_BASE, f"/votaciones/detalleActa/{acta['id']}")
            if r is None:
                raise ValueError(f"detalleActa {acta['id']} del Senado inaccesible")
            filas = _parsear_acta(r.text)
            afirm = sum(1 for f in filas if f["voto"] == "AFIRMATIVO")
            neg = sum(1 for f in filas if f["voto"] == "NEGATIVO")
            if afirm == 0 and neg == 0:
                vistos[acta["id"]] = f"{ley}: votación sin votos registrados (anulada/rehecha) — sin efecto"
                continue
            fecha = acta["fecha"].strftime("%Y-%m-%d")
            if not any(iv.get("camara") == "Senado" and iv.get("fecha") == fecha
                       for iv in veto.get("insistencias_votadas", [])):
                resultado = "insistio_camara" if afirm >= 2 * neg else "insistencia_rechazada"
                veto.setdefault("insistencias_votadas", []).append({
                    "fecha": fecha, "camara": "Senado", "resultado": resultado,
                    "fuente": f"acta {acta['id']} senado.gob.ar ({afirm}-{neg})"})
            vistos[acta["id"]] = f"{ley}: insistencia votada ({afirm}-{neg})"


def _bloqueo_desafios(registro: dict) -> list[dict]:
    """Normas DESAFIADAS en el recinto, una fila por norma:
    [{nombre, fecha_desafio, fecha_caida | None}] ascendente por desafío.
    Veto: desafiado desde su primera insistencia votada (o desde la
    insistencia completa, si el registro no tiene la votación — defensivo);
    caído en la insistencia completa. Decreto: desafiado desde su primera
    votación en el recinto (rechazo o sostenimiento); caído cuando rechaza
    la SEGUNDA cámara distinta (ley 26.122, art. 24 — el rechazo de una sola
    cámara deja la norma en pie, caso real DNU 70/2023)."""
    out = []
    for v in registro.get("vetos", []):
        fechas = [iv["fecha"] for iv in v.get("insistencias_votadas", []) if iv.get("fecha")]
        if v.get("insistencia_completa"):
            fechas.append(v["insistencia_completa"])
        if not fechas:
            continue
        out.append({"nombre": f"ley {v['proyecto']}", "fecha_desafio": min(fechas),
                    "fecha_caida": v.get("insistencia_completa")})
    for d in registro.get("decretos", []):
        fechas = ([rz["fecha"] for rz in d.get("rechazos", []) if rz.get("fecha")]
                  + [s["fecha"] for s in d.get("sostenimientos", []) if s.get("fecha")])
        if not fechas:
            continue
        caida, camaras = None, set()
        for rz in sorted(d.get("rechazos", []), key=lambda r: r.get("fecha") or ""):
            if not rz.get("fecha"):
                continue
            camaras.add(rz.get("camara"))
            if len(camaras) >= 2:
                caida = rz["fecha"]
                break
        out.append({"nombre": d.get("etiqueta") or d.get("clave", "?"),
                    "fecha_desafio": min(fechas), "fecha_caida": caida})
    return sorted(out, key=lambda e: e["fecha_desafio"])


def _bloqueo_tasa_12m(desafios: list, referencia: date):
    """(pct_sostenidas, n_desafiadas, n_caidas, ultimo_desafio) en la ventana
    de los 12 meses calendario que terminan en el mes de `referencia` —
    misma ventana que _derrotas_conteo_12m, así card y serie cuentan igual.
    La caída se evalúa AL CIERRE de `referencia` (estado histórico
    reproducible: un punto ya publicado no cambia porque la norma cayera
    después). None si no hubo desafíos en la ventana (sin denominador no hay
    tasa — el motor renormaliza)."""
    meses = referencia.year * 12 + (referencia.month - 1)
    desde = meses - 11
    ym_desde = f"{desde // 12}-{desde % 12 + 1:02d}"
    ym_hasta = f"{referencia.year}-{referencia.month:02d}"
    corte = referencia.isoformat()
    en_ventana = [e for e in desafios if ym_desde <= e["fecha_desafio"][:7] <= ym_hasta]
    if not en_ventana:
        return None
    caidas = sum(1 for e in en_ventana if e["fecha_caida"] and e["fecha_caida"] <= corte)
    n = len(en_ventana)
    return (round((n - caidas) / n * 100.0, 1), n, caidas,
            max(e["fecha_desafio"] for e in en_ventana))


def fetch_desafios_legislativos() -> dict | None:
    """
    Cuántas normas propias del Ejecutivo fueron DESAFIADAS en el recinto en los
    últimos 12 meses calendario (insistencias de veto votadas + decretos puestos
    a votación bajo la ley 26.122). Menor = mejor: mide con qué frecuencia el
    Congreso decide dar la pelea, sin importar cómo termine.

    Reemplaza a `derrotas_legislativas` en el índice (ADR-0089), que salía del
    mismo registro de eventos y correlacionaba **−0,984** con la tasa de
    bloqueo. Desde mar-2025 los dos son, mes a mes, literalmente el mismo
    número: 16 lecturas consecutivas en que las derrotas igualan exactamente a
    las normas caídas del denominador del bloqueo.

    No es una identidad algebraica —difieren en 12 meses de 2024— porque cuentan
    la caída distinto: para las derrotas, que una cámara rechace un decreto ya
    es una derrota política; para el bloqueo, la norma cae recién cuando la
    rechaza la SEGUNDA cámara (ley 26.122, art. 24). El DNU 70/2023 es el caso
    testigo: el Ejecutivo perdió la votación y conservó la norma. Ambas
    lecturas son correctas, pero en el régimen actual convergieron.

    El par (desafíos, tasa) es en cambio la descomposición del fenómeno en sus
    dos preguntas —cuánto lo confrontan y cuánto aguanta— y baja el
    acoplamiento a −0,828. No lo elimina: la ventana tiene ~10 eventos y
    ninguna descomposición de un conjunto tan chico queda independiente.

    Las derrotas se siguen relevando y quedan a la vista como dato dentro de la
    card de bloqueo (`caidas_12m`); lo que sale es su puntaje propio.
    """
    try:
        registro = _cargar_derrotas_registro()
        if registro is None:
            raise ValueError(f"registro de eventos ausente o ilegible ({DERROTAS_EVENTOS_PATH})")
        # No reclasifica: corre después de fetch_bloqueo_sostenido, que ya dejó
        # el registro actualizado. Leer y contar, nada más.
        tasa = _bloqueo_tasa_12m(_bloqueo_desafios(registro), date.today())
        if tasa is None:
            raise ValueError("sin desafíos votados en la ventana de 12 meses")
        _pct, n, caidas, ultimo = tasa
        return {
            "valor":          float(n),
            "caidas_12m":     caidas,
            "sostenidas_12m": n - caidas,
            "ultimo_desafio": ultimo,
            "unidad":         "normas del Ejecutivo desafiadas en el recinto, últimos 12 meses",
            "fuente":         ("Actas de votación de Diputados y Senado + InfoLeg "
                               "(vetos e insistencias) — elaboración CIGOB"),
            "fecha_dato":     str(date.today()),
            "desactualizado": False,
            "detalle_txt": (f"{n} normas propias desafiadas en el recinto en los últimos 12 meses "
                            f"({caidas} cayeron, {n - caidas} siguen en pie)"),
        }
    except Exception as e:
        _warn("desafios_legislativos", str(e))
        return None


def fetch_bloqueo_sostenido() -> dict | None:
    """
    % de normas propias desafiadas en el recinto (insistencias de veto
    votadas + validez/rechazo de decretos bajo la ley 26.122) que el
    Ejecutivo mantuvo en pie, últimos 12 meses calendario. Mayor = mejor:
    es la capacidad de bloqueo (el tercio del art. 83 CN + la vigencia de
    los decretos), el recurso de poder central de un Ejecutivo sin mayoría.

    Corre DESPUÉS de fetch_derrotas_legislativas (que actualiza vetos,
    insistencias completas y rechazos del Senado en el registro compartido)
    y de fetch_cohesion_bloque (cuyo walk deja las actas nuevas de Diputados
    en el caché permanente que este clasificador usa como universo de ids).
    La clasificación de Diputados persiste su progreso de forma incremental
    (watermark) aunque la corrida falle después; un fallo en el listado del
    Senado degrada la card al caché del snapshot anterior, sin corromper el
    registro.
    """
    try:
        registro = _cargar_derrotas_registro()
        if registro is None:
            raise ValueError(f"registro de eventos ausente o ilegible ({DERROTAS_EVENTOS_PATH})")
        try:
            _bloqueo_clasificar_diputados(registro)
            _bloqueo_detectar_insistencias_senado(registro)
        finally:
            # el progreso de clasificación ya hecho es conocimiento válido
            # (cada acta clasificada es un hecho inmutable): se persiste
            # aunque una etapa posterior haya fallado
            _guardar_derrotas_registro(registro)

        desafios = _bloqueo_desafios(registro)
        tasa = _bloqueo_tasa_12m(desafios, date.today())
        if tasa is None:
            raise ValueError("sin desafíos votados en la ventana de 12 meses (sin denominador no hay tasa)")
        pct, n, caidas, ultimo = tasa
        pendientes = registro.get("actas_diputados_bloqueo", {}).get("pendientes", {})
        return {
            "valor": pct,
            "desafiadas_12m": n,
            "caidas_12m": caidas,
            "sostenidas_12m": n - caidas,
            "ultimo_desafio": ultimo,
            "pendientes_triage": len(pendientes),
            "detalle_txt": (f"{n - caidas} de {n} normas desafiadas en el recinto siguen en pie "
                            f"(vetos con insistencia votada + decretos votados bajo la ley 26.122, "
                            f"últimos 12 meses)"),
            "unidad": "% de normas desafiadas en el recinto que siguen en pie, últimos 12 meses",
            "fuente": ("Actas de votación de Diputados y Senado + InfoLeg (vetos e insistencias) "
                       "— elaboración CIGOB"),
            "fecha_dato": str(date.today()),
            "desactualizado": False,
        }

    except Exception as e:
        _warn("bloqueo_sostenido", str(e))
        return None


# ── Rotación del gabinete (registro curado + detector de alerta InfoLeg) ─────
#
# Modelo semiautomático, mismo patrón que privatizaciones (ITCG) y
# adhesion_reformas_provincial: la FUENTE DE VERDAD es un registro curado
# versionado (data/politica/gabinete_salidas.json — cada salida con persona,
# cargo, mes de cese efectivo, decreto BO y clasificación), y un detector
# automático contra InfoLeg AVISA si aparece un decreto de renuncia
# ministerial que el registro no refleja — el detector nunca modifica el
# registro solo (distinguir eyección de pase lateral es una clasificación
# política que ningún regex hace: el Dto. 548/2026 acepta la renuncia de
# Adorni a JGM Y la de Santilli a Interior en el mismo acto, pero solo la
# primera es una salida del gabinete).
#
# El detector es de dos etapas porque el buscador de InfoLeg es OR sobre
# palabras (no frase ni AND: "acéptase la renuncia ministro" devuelve MÁS
# resultados que "renuncia" sola) y trunca la síntesis del listado a ~150
# caracteres, casi siempre antes del cargo. E1: listado mensual con
# texto="renuncia" + tipoNorma=2 (7-20 filas/mes). E2: para las filas cuyo
# resumen truncado arranca con tag de dependencia ministerial y contiene
# RENUNCIA (2-8/mes), detalle verNorma.do (cacheado en disco, patrón
# cohesion_bloque_diputados_actas_cache) y regex final de cargo sobre el
# resumen completo. Probado end-to-end sobre dic-2023→jul-2026: recall 11/11
# salidas reales, cero renuncias no-ministeriales filtradas (jueces,
# fiscales, embajadores y directorios quedan afuera por el tag de
# dependencia; los "ministros plenipotenciarios" del servicio exterior, por
# la exigencia de "MINISTR[OA] DE").

_GABINETE_INFOLEG_PAUSA = 1.6   # seg entre requests al buscador (detector)
_RE_GABINETE_TAG = re.compile(r"^(JEFATURA DE GABINETE DE MINISTROS|MINISTERIO DE[L]? )")
_RE_GABINETE_CARGO = re.compile(
    r"RENUNCIA PRESENTADA POR (?:LA|EL) ([^(]{3,80}?)\s*\(D\.?N\.?I[^)]*\)\s*"
    r"AL CARGO DE (JEFE DE GABINETE DE MINISTROS|JEFA DE GABINETE DE MINISTROS|"
    r"MINISTR[OA] DE[L]? [A-ZÑ ,]{3,70})")
_RE_GABINETE_DECRETO = re.compile(r"Decreto\s+(\d+)\s*/\s*(\d{4})")


def _norm_mayusculas(s: str) -> str:
    """Mayúsculas sin acentos (el HTML de InfoLeg mezcla ambas formas)."""
    s = unicodedata.normalize("NFKD", s)
    return "".join(c for c in s if not unicodedata.combining(c)).upper()


def cargar_gabinete_salidas() -> dict | None:
    """Registro curado de salidas de rango ministerial, o None si falta/roto."""
    try:
        return json.loads(GABINETE_SALIDAS_PATH.read_text(encoding="utf-8-sig"))
    except (OSError, json.JSONDecodeError):
        return None


def _indice_mes(ym: str) -> int | None:
    """'YYYY-MM' → meses desde el año 0 (para aritmética de ventanas)."""
    try:
        return int(ym[:4]) * 12 + (int(ym[5:7]) - 1)
    except (ValueError, TypeError):
        return None


def salidas_gabinete_ventana_12m(salidas: list, ym: str) -> list:
    """Salidas del registro cuyo mes de cese efectivo cae en la ventana de 12
    meses calendario que termina en `ym` (YYYY-MM, inclusive) — la métrica
    del indicador. Los movimientos laterales y las reestructuraciones de la
    Ley de Ministerios no están en `salidas` (viven en sus propias claves del
    registro), así que quedan excluidos por construcción."""
    fin = _indice_mes(ym)
    if fin is None:
        return []
    out = []
    for s in salidas:
        idx = _indice_mes(str(s.get("mes", "")))
        if idx is not None and fin - 11 <= idx <= fin:
            out.append(s)
    return out


def _cargar_cache_gabinete_decretos() -> dict:
    try:
        return json.loads(GABINETE_DECRETOS_CACHE_PATH.read_text(encoding="utf-8-sig"))
    except (OSError, json.JSONDecodeError):
        return {}


def _guardar_cache_gabinete_decretos(cache: dict) -> None:
    GABINETE_DECRETOS_CACHE_PATH.write_text(
        json.dumps(cache, indent=1, ensure_ascii=False, sort_keys=True), encoding="utf-8")


def _gabinete_sesion_infoleg():
    """(session, action_url) contra el buscador de InfoLeg — misma mecánica
    que fetch_ratio_dnu / gestion._infoleg_post (GET home → form action)."""
    session = requests.Session()
    r_home = session.get(INFOLEG_HOME, headers=HTTP_HEADERS, timeout=HTTP_TIMEOUT)
    r_home.raise_for_status()
    action_m = re.search(r'action="(/infolegInternet/[^"]+)"', r_home.text)
    if not action_m:
        raise ValueError("No se encontró la URL del formulario InfoLeg")
    time.sleep(_GABINETE_INFOLEG_PAUSA)
    return session, "https://servicios.infoleg.gob.ar" + action_m.group(1)


def _gabinete_parsear_filas(html: str) -> list:
    """Filas del listado de resultados de InfoLeg → [{id, decreto, descripcion}].
    La síntesis viene truncada a ~150 caracteres — solo sirve para el
    prefiltro de E2, nunca para clasificar."""
    filas = []
    for tr in re.findall(r"<tr[^>]*>.*?</tr>", html, re.S):
        m_id = re.search(r"verNorma\.do\?id=(\d+)", tr)
        if not m_id:
            continue
        celdas = re.findall(r"<td[^>]*>(.*?)</td>", tr, re.S)
        limpio = [re.sub(r"\s+", " ", re.sub(r"<[^>]+>", " ", c)).strip() for c in celdas]
        titulo = limpio[0] if limpio else ""
        m_dec = _RE_GABINETE_DECRETO.search(titulo)
        filas.append({
            "id": m_id.group(1),
            "decreto": f"Decreto {m_dec.group(1)}/{m_dec.group(2)}" if m_dec else titulo[:40],
            "descripcion": limpio[-1] if len(limpio) >= 2 else "",
        })
    return filas


def _gabinete_listado_mes(session, action_url: str, anio: int, mes: int,
                           max_paginas: int = 4) -> list:
    """E1 del detector: decretos publicados en el mes con 'renuncia' en el
    texto (paginado con desplazamiento=AP, como el buscador real)."""
    ultimo_dia = calendar.monthrange(anio, mes)[1]
    data = {
        "tipoNorma": "2", "numero": "", "anioSancion": "", "dependencia": "",
        "diaPubDesde": "01", "mesPubDesde": f"{mes:02d}", "anioPubDesde": str(anio),
        "diaPubHasta": str(ultimo_dia), "mesPubHasta": f"{mes:02d}", "anioPubHasta": str(anio),
        "texto": "renuncia",
    }
    r = session.post(action_url, data=data, headers=HTTP_HEADERS, timeout=HTTP_TIMEOUT)
    r.raise_for_status()
    time.sleep(_GABINETE_INFOLEG_PAUSA)
    if re.search(r"No se encontraron normas", r.text, re.IGNORECASE):
        return []
    m = re.search(r"Encontradas?[:\s]+(\d+)", r.text, re.IGNORECASE)
    total = int(m.group(1)) if m else None
    filas = _gabinete_parsear_filas(r.text)
    pagina = 2
    while total and len(filas) < total and pagina <= max_paginas:
        rp = session.post(action_url, data={"desplazamiento": "AP", "irAPagina": str(pagina)},
                          headers=HTTP_HEADERS, timeout=HTTP_TIMEOUT)
        rp.raise_for_status()
        time.sleep(_GABINETE_INFOLEG_PAUSA)
        nuevas = _gabinete_parsear_filas(rp.text)
        if not nuevas:
            break
        filas.extend(nuevas)
        pagina += 1
    return filas


def _gabinete_resumen_norma(session, norma_id: str, cache: dict) -> str:
    """E2 del detector: el campo Resumen COMPLETO de verNorma.do (el listado
    lo trunca antes del cargo). Cacheado en disco de inmediato — una norma
    publicada no cambia, y el caché evita repagar los GET en cada corrida
    (mismo criterio que la caché permanente por acta de Diputados)."""
    if norma_id in cache:
        return cache[norma_id]
    r = session.get(f"https://servicios.infoleg.gob.ar/infolegInternet/verNorma.do?id={norma_id}",
                    headers=HTTP_HEADERS, timeout=HTTP_TIMEOUT)
    r.raise_for_status()
    time.sleep(_GABINETE_INFOLEG_PAUSA)
    txt = re.sub(r"\s+", " ", re.sub(r"<[^>]+>", " ", r.text))
    m = re.search(r"Resumen:\s*(.*?)(?:Texto completo de la norma|Esta norma|$)", txt)
    resumen = _norm_mayusculas(m.group(1).strip()) if m else ""
    cache[norma_id] = resumen
    _guardar_cache_gabinete_decretos(cache)
    return resumen


def _detectar_salidas_gabinete_infoleg(meses_atras: int = 1) -> list | None:
    """Detector de alerta: renuncias a cargos de rango ministerial publicadas
    en el BO en el mes corriente y los `meses_atras` previos (el decreto puede
    llegar hasta ~40 días después del hecho político — caso Ferraro). Devuelve
    [{decreto, persona, cargo, mes_bo}] dedupeado (un mismo decreto puede
    repetir el párrafo de una renuncia en su resumen), o None si InfoLeg no
    respondió — el indicador NO depende de esto para publicar (el registro
    curado manda), solo para avisar."""
    try:
        session, action_url = _gabinete_sesion_infoleg()
    except Exception:
        return None
    hoy = date.today()
    meses = []
    anio, mes = hoy.year, hoy.month
    for _ in range(meses_atras + 1):
        meses.append((anio, mes))
        anio, mes = (anio - 1, 12) if mes == 1 else (anio, mes - 1)

    cache = _cargar_cache_gabinete_decretos()
    detecciones, vistos = [], set()
    for anio, mes in meses:
        try:
            filas = _gabinete_listado_mes(session, action_url, anio, mes)
        except Exception:
            return None   # listado caído: sin señal confiable, mejor no "detectar nada"
        for fila in filas:
            desc = _norm_mayusculas(fila["descripcion"])
            if not _RE_GABINETE_TAG.search(desc) or "RENUNCIA" not in desc:
                continue
            try:
                resumen = _gabinete_resumen_norma(session, fila["id"], cache)
            except Exception:
                continue   # un detalle caído no invalida el resto del barrido
            for m_c in _RE_GABINETE_CARGO.finditer(resumen):
                clave = (fila["decreto"], m_c.group(1).strip(), m_c.group(2).strip())
                if clave in vistos:
                    continue
                vistos.add(clave)
                detecciones.append({
                    "decreto": fila["decreto"],
                    "persona": m_c.group(1).strip().title(),
                    "cargo": m_c.group(2).strip(),
                    "mes_bo": f"{anio}-{mes:02d}",
                })
    return detecciones


def _gabinete_discrepancias(registro: dict, detecciones: list) -> list:
    """Detecciones cuyo decreto NO está citado en el registro curado (ni como
    salida ni como movimiento lateral): son las que exigen curaduría. Se
    compara por número de decreto (no por persona) para que una re-salida de
    una persona ya registrada también dispare la alerta."""
    texto_registro = json.dumps(registro, ensure_ascii=False)
    decretos_registrados = {f"Decreto {n}/{a}" for n, a
                            in _RE_GABINETE_DECRETO.findall(texto_registro)}
    return [d for d in detecciones if d["decreto"] not in decretos_registrados]


def fetch_rotacion_gabinete() -> dict | None:
    """Salidas de rango ministerial (jefe de Gabinete + ministros) acumuladas
    en la ventana móvil de 12 meses que termina en el mes corriente, desde el
    registro curado (menor = mejor). Cuenta salidas políticas Y estructurales
    (ceses por banca electoral) sin distinguir — la composición se publica
    como transparencia y los casos extremos se administran con el override
    del analista (ajustes_itcp.json). No cuenta pases laterales dentro del
    gabinete ni reestructuraciones de la Ley de Ministerios.

    fecha_dato = fecha del chequeo (hoy): el registro no "vence" por falta de
    eventos — el no-evento es dato (0 salidas ese mes) — y el guard real de
    frescura es el detector InfoLeg, que avisa si aparece un decreto de
    renuncia ministerial no registrado. La última salida queda en
    `ultima_salida` para la lectura fina."""
    registro = cargar_gabinete_salidas()
    if registro is None or not isinstance(registro.get("salidas"), list):
        _warn("rotacion_gabinete", f"registro curado ausente o ilegible ({GABINETE_SALIDAS_PATH.name})")
        return None

    ym = date.today().strftime("%Y-%m")
    en_ventana = salidas_gabinete_ventana_12m(registro["salidas"], ym)
    politicas = [s for s in en_ventana if s.get("clasificacion") == "salida_politica"]
    estructurales = [s for s in en_ventana if s.get("clasificacion") == "salida_estructural_electoral"]
    ultima = max(registro["salidas"], key=lambda s: str(s.get("mes", "")), default=None)

    if en_ventana:
        lista = ", ".join(f"{s['persona']} ({s['mes']})"
                          for s in sorted(en_ventana, key=lambda s: str(s.get("mes", ""))))
        detalle = (f"{len(en_ventana)} salidas en 12 meses = {len(politicas)} políticas · "
                   f"{len(estructurales)} por bancas electorales — {lista}")
    else:
        detalle = "0 salidas de rango ministerial en los últimos 12 meses"

    resultado = {
        "valor": len(en_ventana),
        "unidad": "salidas de rango ministerial (acum. 12 meses)",
        "fuente": "Decretos de designación y renuncia — Boletín Oficial (registro curado CIGOB)",
        "fecha_dato": str(date.today()),
        "desactualizado": False,
        "salidas_politicas": len(politicas),
        "salidas_estructurales": len(estructurales),
        "ultima_salida": (f"{ultima['persona']} — {ultima['cargo']} ({ultima['mes']})"
                          if ultima else None),
        "detalle_txt": detalle,
    }

    # Detector de alerta (no bloquea: si InfoLeg falla, el registro manda)
    try:
        detecciones = _detectar_salidas_gabinete_infoleg()
    except Exception as e:
        print(f"[WARN] {CINTURON}.rotacion_gabinete: detector InfoLeg falló ({e}) — "
              "se publica igual desde el registro curado")
        detecciones = None
    if detecciones is not None:
        discrepancias = _gabinete_discrepancias(registro, detecciones)
        for d in discrepancias:
            print(f"[ALERTA] {CINTURON}.rotacion_gabinete: {d['decreto']} (BO {d['mes_bo']}) "
                  f"acepta la renuncia de {d['persona']} al cargo de {d['cargo']} y NO figura "
                  f"en el registro curado — revisar data/politica/gabinete_salidas.json")
        if discrepancias:
            resultado["detector_decretos_sin_registrar"] = [
                f"{d['decreto']} — {d['persona']} ({d['cargo']})" for d in discrepancias]

    return resultado


UMBRAL_FRESCURA_COHESION = 10  # días SIN una corrida exitosa (no sin votos nuevos)


def _cohesion_desactualizada(cache_previo: dict | None, corrida_actual: dict | None,
                              umbral_dias: int = UMBRAL_FRESCURA_COHESION) -> bool:
    """True solo si no hubo NINGUNA corrida que haya llegado al sitio en los
    últimos `umbral_dias` días — nunca por ausencia de votos nuevos (el receso
    legislativo es normal y no debe marcarse como stale)."""
    if corrida_actual is not None:
        return False
    if cache_previo is None or not cache_previo.get("corrida_exitosa_en"):
        return True
    ultima = datetime.strptime(cache_previo["corrida_exitosa_en"], "%Y-%m-%d")
    return (datetime.now() - ultima).days > umbral_dias


def _es_cohesion_legado(anterior: dict) -> bool:
    """True si `anterior` (el cache previo de cohesion_bloque) tiene la forma
    VIEJA del placeholder manual (pre-scraper) — {"valor": 78, "estado":
    "placeholder", "unidad": "% votos en línea con la posición oficial del
    bloque LLA", ...} — en vez de la forma NUEVA que devuelve
    fetch_cohesion_bloque (índice de Rice real).

    El discriminador es la AUSENCIA de "n_actas": es una clave que solo la
    corrida del scraper nuevo pone (junto con "corrida_exitosa_en"); el
    placeholder manual nunca la tuvo. Ambas formas son números 0-100 con
    significados totalmente distintos (78 = "% de diputados alineados con la
    posición oficial" vs. índice de Rice real) — sin este chequeo, un cache
    viejo se arrastraría para siempre como si fuera un dato de Rice genuino
    (hallazgo de revisión externa, ver commit)."""
    return "n_actas" not in anterior


# Pesos por cámara del compuesto bicameral de cohesión (ADR-0048): el mismo
# ratio interno 65/35 ≈ 45/25 que las dos cámaras tenían como indicadores
# separados dentro de la dimensión (ADR-0036/0047). Renormaliza si una cámara
# no tiene dato — mismo criterio que la paramétrica ante faltantes.
COHESION_PESOS_CAMARAS = {"diputados": 0.65, "senado": 0.35}


def componer_cohesion_bloque(dip: dict | None, sen: dict | None) -> dict | None:
    """Card compuesta bicameral de cohesión del bloque LLA (ADR-0048,
    revisión editorial CIGOB 2026-07-10): un solo indicador en lugar de las
    dos cards por cámara. Rice de Diputados 65% + Rice del Senado 35%,
    renormalizado sobre las cámaras con dato. Los componentes por cámara
    quedan en la card (mismo patrón que el Fondo de Cese en gestión) y son
    la fuente del carry-forward por cámara de la próxima corrida
    (_anterior_camara los lee de vuelta)."""
    camaras = {"diputados": dip, "senado": sen}
    con_dato = {c: e for c, e in camaras.items()
                if e is not None and e.get("valor") is not None}
    if not con_dato:
        return None
    peso_total = sum(COHESION_PESOS_CAMARAS[c] for c in con_dato)
    valor = round(sum(e["valor"] * COHESION_PESOS_CAMARAS[c]
                      for c, e in con_dato.items()) / peso_total, 1)
    fechas = [e["fecha_dato"] for e in con_dato.values() if e.get("fecha_dato")]
    return {
        "valor": valor,
        "unidad": ("% cohesión (índice de Rice bicameral: Diputados 65% + Senado 35%, "
                   "promedio actas divididas últimos 90 días)"),
        "fuente": "Votaciones nominales de Diputados y Senado — elaboración CIGOB",
        "fecha_dato": max(fechas) if fechas else None,
        "n_actas": sum(e.get("n_actas") or 0 for e in con_dato.values()),
        # desactualizado solo si TODAS las cámaras que aportan están
        # desactualizadas: con una cámara fresca el compuesto tiene
        # información nueva (la composición por cámara queda visible abajo).
        "desactualizado": all(e.get("desactualizado") for e in con_dato.values()),
        "componentes": {
            c: {k: e.get(k) for k in ("valor", "fecha_dato", "n_actas",
                                       "corrida_exitosa_en", "desactualizado")}
            for c, e in camaras.items() if e is not None
        },
    }


def _anterior_camara(indicadores_anteriores: dict, camara: str) -> dict | None:
    """Último dato conocido de una cámara para el carry-forward del compuesto.
    Forma nueva: dentro de "componentes" de la card compuesta. Formas de
    migración (cache anterior a ADR-0048): cohesion_bloque era la card de
    Diputados sola y cohesion_bloque_senado la del Senado."""
    comp = indicadores_anteriores.get("cohesion_bloque")
    if comp is not None and isinstance(comp.get("componentes"), dict):
        return comp["componentes"].get(camara)
    if camara == "diputados":
        return comp
    return indicadores_anteriores.get("cohesion_bloque_senado")


def _entrada_camara(resultado: dict | None, anterior: dict | None) -> tuple:
    """(entrada, corrida_ok) por cámara para el compuesto — las MISMAS tres
    ramas que tenían las dos cards separadas: corrida con votos nuevos /
    corrida exitosa sin votos en la ventana (receso: se reusa el último valor
    sin marcarlo desactualizado) / sin corrida (cache, con el chequeo de
    staleness de _cohesion_desactualizada). corrida_ok = la corrida de HOY
    llegó al sitio (insumo del conteo de frescura del colector)."""
    if resultado is not None and resultado.get("valor") is not None:
        return resultado, True
    if resultado is not None and anterior is not None:
        return ({**anterior, "desactualizado": False,
                 "corrida_exitosa_en": resultado["corrida_exitosa_en"]}, True)
    if anterior is not None:
        return ({**anterior,
                 "desactualizado": _cohesion_desactualizada(anterior, resultado)}, False)
    return None, resultado is not None


def _valor_itcp(nombre: str, entry: dict):
    """Valor a puntuar en el ITCP para un indicador ya fresco/cacheado.

    protestas_caba es la ÚNICA excepción: puntúa sobre "var_vs_2023" (% de
    variación de eventos ACLED en CABA contra la base 2023), NO sobre "valor"
    (el conteo crudo de eventos acumulado 12 meses, que gestion.fetch_protestas_caba
    devuelve y que puede estar en cientos — ver docstring de itcp.BANDAS_ITCP).
    Cualquier otro indicador puntúa directo sobre su propio "valor"."""
    if nombre == "protestas_caba":
        return entry.get("var_vs_2023")
    return entry.get("valor")


def _resultado_utilizable(nombre: str, resultado: dict | None) -> bool:
    """True si el resultado de un fetcher tiene el valor que efectivamente se
    usa en el ITCP para `nombre` (vía _valor_itcp), no solo un "valor" crudo
    presente. Cierra un borde latente (auditoría 2026-07-08): protestas_caba
    puntúa sobre var_vs_2023; si var_vs_2023 fuera None con "valor" presente
    (base_2023 == 0), el indicador se contaría como fresco sin aportar
    realmente al índice. Para el resto de los indicadores es equivalente a
    `resultado.get("valor") is not None` (_valor_itcp devuelve "valor" directo)."""
    return resultado is not None and _valor_itcp(nombre, resultado) is not None


def calcular_itcp_cinturon(indicadores: dict) -> dict | None:
    """ITCP 0-100 (ver scripts/itcp.py) sobre los indicadores del índice, a
    partir de valores YA PERSISTIDOS (fetch fresco del colector o
    cache["indicadores"] existente) — no depende de red. Extraída de main()
    para que generar_informe.py pueda recomputar el ITCP con el itcp.py
    vigente sin volver a pegarle a InfoLeg/HCDN/etc."""
    ajustes = itcp.cargar_ajustes(AJUSTES_ITCP_PATH, datetime.now().strftime("%Y-%m"))
    valores = {}
    for nombre, entry in indicadores.items():
        valor = _valor_itcp(nombre, entry)
        if valor is not None:
            valores[nombre] = valor
    return itcp.calcular_itcp(valores, ajustes)


def _anotar_indicadores_itcp(indicadores: dict, resultado: dict | None) -> None:
    """Marca cada indicador con su rol en el ITCP: los del índice llevan
    puntaje, dimensión y peso efectivo; el resto queda como contexto (mismo
    patrón que gestion.anotar_indicadores para el ITCG). Desde ADR-0048 el
    ITCP tiene contexto declarado (itcp.INDICADORES_CONTEXTO:
    rotacion_gabinete y protestas_caba — se publican, no puntúan); sus bandas
    siguen en itcp.BANDAS_ITCP como referencia histórica, por eso el
    override explícito después del fallback (mismo esquema que macro/ITCM)."""
    por_indicador = {}
    if resultado:
        for dkey, dim in resultado["dimensiones"].items():
            for ikey, info in dim["indicadores"].items():
                por_indicador[ikey] = {
                    "en_indice": True,
                    "dimension": dkey,
                    "puntaje_itcp": info["puntaje_aplicado"],
                    "puntaje_banda": info["puntaje_banda"],
                    "peso_efectivo": info["peso_efectivo"],
                }
    for nombre, ind in indicadores.items():
        if nombre in por_indicador:
            ind.update(por_indicador[nombre])
        else:
            ind["en_indice"] = nombre in itcp.BANDAS_ITCP  # del índice pero sin dato
            if nombre in itcp.INDICADORES_CONTEXTO:
                ind["en_indice"] = False


# ── Main ──────────────────────────────────────────────────────────────────────

def main() -> None:
    cache_anterior         = load_cache()
    indicadores_anteriores = cache_anterior.get("indicadores", {})

    frescos: dict = {}
    frescos_count = 0

    colectores = [
        ("votometro_ventaja_lla",         fetch_votometro),
        ("ratio_dnu",                     fetch_ratio_dnu),
        ("brecha_obra_publica",           fetch_brecha_obra_publica),
        ("apoyo_empresario",              fetch_apoyo_empresario),
        ("conflictividad_nacional",       fetch_conflictividad_nacional),
        ("jornadas_individuales_no_trabajadas_12m",
                                             fetch_jornadas_individuales_no_trabajadas),
        ("movilizacion_cepa",             fetch_cepa_movilizacion),
        ("iaf_transferencias",            fetch_iaf_transferencias),
        ("eficacia_legislativa",          fetch_eficacia_legislativa),
        ("veto_quorum",                   fetch_veto_quorum),
        ("comisiones_caidas",             fetch_comisiones_caidas),
        ("adhesion_reformas_provincial",  fetch_adhesion_reformas_provincial),
        ("derrotas_legislativas",         fetch_derrotas_legislativas),
        ("rotacion_gabinete",             fetch_rotacion_gabinete),
        ("protestas_caba",                fetch_protestas_caba),
    ]

    for nombre, fetcher in colectores:
        resultado = fetcher()
        if _resultado_utilizable(nombre, resultado):
            frescos[nombre] = _sellar(resultado)
            frescos_count += 1
        elif nombre in indicadores_anteriores:
            frescos[nombre] = {**indicadores_anteriores[nombre], "desactualizado": True}

    # Compuesto bicameral de cohesión (ADR-0048): una sola card desde las dos
    # corridas por cámara. Cada cámara conserva sus tres ramas de fallback
    # (fresco / receso / cache) vía _entrada_camara; el compuesto cuenta como
    # fresco solo si AMBAS corridas llegaron a su sitio — con una sola cámara
    # fresca el valor sale igual (renormalizado) pero el exit code del
    # colector reporta frescura mixta, como cuando eran dos cards.
    resultado_cohesion = fetch_cohesion_bloque()
    anterior_dip = _anterior_camara(indicadores_anteriores, "diputados")
    if anterior_dip is not None and _es_cohesion_legado(anterior_dip):
        # Cache heredado del viejo placeholder manual (78, "% votos alineados
        # con la posición oficial") — NO es un dato de Rice genuino, tratarlo
        # como ausente para no corromper el ITCP con un valor de significado
        # distinto (hallazgo de revisión externa).
        anterior_dip = None
    entrada_dip, dip_ok = _entrada_camara(resultado_cohesion, anterior_dip)

    resultado_cohesion_senado = fetch_cohesion_bloque_senado()
    anterior_sen = _anterior_camara(indicadores_anteriores, "senado")
    entrada_sen, sen_ok = _entrada_camara(resultado_cohesion_senado, anterior_sen)

    compuesto_cohesion = componer_cohesion_bloque(entrada_dip, entrada_sen)
    if compuesto_cohesion is not None:
        frescos["cohesion_bloque"] = _sellar(compuesto_cohesion)
        if dip_ok and sen_ok:
            frescos_count += 1

    # bloqueo_sostenido corre DESPUÉS de cohesión a propósito (ADR-0069): su
    # clasificador usa como universo de ids el caché de actas de Diputados
    # que el walk de cohesión acaba de refrescar, y el registro de eventos
    # que fetch_derrotas_legislativas (en la lista de arriba) ya actualizó.
    resultado_bloqueo = fetch_bloqueo_sostenido()
    if _resultado_utilizable("bloqueo_sostenido", resultado_bloqueo):
        frescos["bloqueo_sostenido"] = _sellar(resultado_bloqueo)
        frescos_count += 1
    elif "bloqueo_sostenido" in indicadores_anteriores:
        frescos["bloqueo_sostenido"] = {**indicadores_anteriores["bloqueo_sostenido"],
                                        "desactualizado": True}

    # desafios_legislativos sale del MISMO registro y la MISMA ventana que el
    # bloqueo (ADR-0089): corre inmediatamente después para que ambos vean el
    # registro en idéntico estado. Son numerador y denominador de la misma
    # razón —desafiadas y sostenidas—, así que leerlos en momentos distintos
    # los haría inconsistentes entre sí.
    resultado_desafios = fetch_desafios_legislativos()
    if _resultado_utilizable("desafios_legislativos", resultado_desafios):
        frescos["desafios_legislativos"] = _sellar(resultado_desafios)
        frescos_count += 1
    elif "desafios_legislativos" in indicadores_anteriores:
        frescos["desafios_legislativos"] = {**indicadores_anteriores["desafios_legislativos"],
                                            "desactualizado": True}

    # ADR-0168: los tres de comportamiento judicial + la producción legislativa.
    for _clave, _fn in (("produccion_legislativa", fetch_produccion_legislativa),
                        ("judicializacion", fetch_judicializacion),
                        ("velocidad_resolucion", fetch_velocidad_resolucion),
                        ("paralisis_denuncias", fetch_paralisis_denuncias)):
        _res = _fn()
        if _resultado_utilizable(_clave, _res):
            frescos[_clave] = _sellar(_res)
            frescos_count += 1
        elif _clave in indicadores_anteriores:
            frescos[_clave] = {**indicadores_anteriores[_clave], "desactualizado": True}

    resultado_judicial = fetch_cobertura_judicial()
    if _resultado_utilizable("cobertura_judicial", resultado_judicial):
        frescos["cobertura_judicial"] = _sellar(resultado_judicial)
        frescos_count += 1
    elif "cobertura_judicial" in indicadores_anteriores:
        frescos["cobertura_judicial"] = {**indicadores_anteriores["cobertura_judicial"],
                                         "desactualizado": True}

    # Detector de novedades judiciales (ADR-0140). No produce indicador ni
    # entra en `frescos`: sólo actualiza el registro de fallos a revisar. Va
    # envuelto porque no debe poder tumbar el colector — si la CSJN no
    # responde, lo que se pierde es un aviso, no un dato del índice.
    try:
        pendientes = detectar_novedades_judiciales().get("pendientes", {})
        if pendientes:
            print(f"  [i] CSJN: {len(pendientes)} fallo(s) pendientes de revisar "
                  f"→ data/politica/csjn_novedades.json")
    except Exception as e:
        print(f"  [WARN] detector de novedades judiciales no corrió ({e})")

    # Detector de postura empresaria (ADR-0149). Tampoco produce indicador: sólo
    # evita que el registro codificado a mano se quede viejo entre corridas.
    try:
        pend = detectar_novedades_empresarias().get("pendientes", {})
        if pend:
            print(f"  [i] cámaras: {len(pend)} comunicado(s) sin codificar "
                  f"→ data/politica/apoyo_empresario_novedades.json")
    except Exception as e:
        print(f"  [WARN] detector de postura empresaria no corrió ({e})")

    # alineamiento_senadores_prov comparte el mismo contrato de retorno que
    # cohesion_bloque_senado (misma sesión/descubrimiento de actas de Senado):
    # dict con "valor": None (no None a secas) cuando la corrida llegó al
    # sitio pero no hubo actas divididas en la ventana de recencia (receso
    # legislativo normal). Mismo bloque dedicado que cohesion_bloque_senado
    # en vez de la lista plana de colectores, para no marcar "desactualizado"
    # solo por ausencia de votos nuevos (reutiliza _cohesion_desactualizada,
    # que es un chequeo de staleness genérico, no específico de cohesión).
    resultado_alineamiento = fetch_alineamiento_senadores_prov()
    anterior_alineamiento = indicadores_anteriores.get("alineamiento_senadores_prov")
    if resultado_alineamiento is not None and resultado_alineamiento.get("valor") is not None:
        frescos["alineamiento_senadores_prov"] = resultado_alineamiento
        frescos_count += 1
    elif resultado_alineamiento is not None and anterior_alineamiento is not None:
        frescos["alineamiento_senadores_prov"] = {
            **anterior_alineamiento,
            "desactualizado": False,
            "corrida_exitosa_en": resultado_alineamiento["corrida_exitosa_en"],
        }
        frescos_count += 1
    elif anterior_alineamiento is not None:
        frescos["alineamiento_senadores_prov"] = {
            **anterior_alineamiento,
            "desactualizado": _cohesion_desactualizada(anterior_alineamiento, resultado_alineamiento),
        }

    resultado_itcp = calcular_itcp_cinturon(frescos)
    score = itcp.tension_de_itcp(resultado_itcp["valor"]) if resultado_itcp else 5.0
    _anotar_indicadores_itcp(frescos, resultado_itcp)

    payload = {
        "cinturon":     CINTURON,
        "generated_at": datetime.now().isoformat(),
        "score":        score,
        "itcp":         resultado_itcp,
        "indicadores":  frescos,
    }

    if frescos:
        save_cache(payload)
        total = len(INDICADORES_ESPERADOS)
        print(f"[OK] {CINTURON}: score={score} frescos={frescos_count}/{total}")

    if frescos_count == len(INDICADORES_ESPERADOS):
        sys.exit(0)
    elif frescos_count > 0:
        sys.exit(1)
    else:
        sys.exit(2)


if __name__ == "__main__":
    main()
