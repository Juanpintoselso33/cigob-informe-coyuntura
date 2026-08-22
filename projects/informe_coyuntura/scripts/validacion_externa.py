"""validacion_externa.py — Validación de constructo del ITVC y el ITCM (ADR-0019, D6).

Paso 9 del Handbook JRC/OCDE ("links to other variables"): si el índice mide
lo que dice medir, debería co-moverse con variables externas relacionadas que
NO lo componen. Dos estudios:
  * ITVC (condiciones materiales de la vida cotidiana) contra el CONSUMO medido:
    ventas en supermercados a precios constantes, serie desestacionalizada del
    INDEC — correlación positiva esperada. Fue el ICC de UTDT hasta jul-2026;
    se reemplazó porque el ICC ES un componente del ITVC (6,75%), lo que
    obligaba a publicar un índice artificial «sin ICC», y porque un tercio del
    peso del ITVC correlaciona NEGATIVO contra él (ADR-0155). El ICC queda como
    contraste DISCRIMINANTE: mide si la percepción sigue a las condiciones.
  * ITCM (tensión macroeconómica, reconstrucción mensual desde las series de
    componentes con puntaje interpolado) contra el ÍNDICE LÍDER de la UTDT
    (marcha de la actividad) — correlación POSITIVA esperada: menos tensión
    macro, más actividad. Fue el riesgo país hasta jul-2026; se reemplazó
    porque no validaba en primeras diferencias (ADR-0154).

Para que la comparación no sea circular (el ICC es un componente del ITVC,
7,5% del peso), la serie del ITVC se recalcula EXCLUYENDO al ICC — la
renormalización estándar del motor absorbe la ausencia.

Cómo se reconstruye la serie mensual del ITVC (dic-2023 → hoy):
- Componentes transformados (itvc_alimentos/tarifas/ipi/isac/endeudamiento):
  ya son índices base-100 en series.json.
- Componentes de rebase directo (brecha, ICC, pluriempleo, carne, motos):
  se rebasea TODA la serie contra su promedio 4T-2023.
- Anuales (informalidad, inseguridad): rebase anual + forward-fill mensual
  (regla del doc: "último dato disponible").
- Cada mes se agrega con itvc.calcular_itvc (misma renormalización que el
  índice publicado); un componente entra con su último dato ≤ mes.

Correlaciones (Pearson): niveles y primeras diferencias, contemporáneas y
con ±1 rezago. Salida: output/validacion_externa.json + resumen legible.

Uso: python scripts/validacion_externa.py
"""
import io
import json
import statistics
import sys
from datetime import date, datetime, timezone
from pathlib import Path

import requests

sys.path.insert(0, str(Path(__file__).parent))
sys.stdout.reconfigure(encoding="utf-8")

import itcg
import itcm
import itcp
import itvc
import parametrica
import publicar

# Las ventas en supermercados ya no se bajan acá: desde ADR-0225 son un
# COMPONENTE del ITCIS y las baja su colector
# (`vida_cotidiana/collectors/indec_supermercados.py`), como cualquier otro
# componente. Tener dos caminos de descarga para la misma serie era
# exactamente la clase de duplicación que deja al índice y a su contraste
# leyendo bases distintas sin que nada avise.
#
# Lo que SÍ hay que conservar de aquella decisión, porque vuelve a morder cada
# vez que alguien suaviza una serie a mano: la variante ORIGINAL del INDEC
# tiene una estacionalidad enorme y aplanarla con una media móvil de 12 meses
# la atrasa medio año — probado, y el atraso INVIERTE el signo de la
# correlación (−0,514 con media móvil contra +0,563 con la desestacionalizada
# del organismo). Se usa siempre la desestacionalizada que publica la fuente.

# Los otros dos canales de consumo, para el panel del ITVC (ADR-0159).
CONSUMO_MAYORISTAS_ID = "456.1_VENTAS_PREADA_0_M_44_40"
CONSUMO_SHOPPINGS_ID = "458.1_VENTAS_TOTADA_0_M_52_56"

# Volúmenes FÍSICOS consumidos por los hogares (ADR-0163). Entran al panel del
# ITVC como contraste externo y son los que arman su factor común: al medirse en
# unidades físicas no llevan deflactor, así que no comparten insumo con
# `ipc_alimentos`, que sí es componente del índice. Traen estacionalidad fuerte
# y se ajustan antes de entrar (ver `desestacionalizar.py`).
# Respuesta del capital privado al programa de transformación (ADR-0164), para
# la familia del ITCG, que tenía UNA sola estadística propia. Cuenta Capital y
# Financiera Cambiaria del BCRA, mensual desde 2003.
#
# Se usan SUBCUENTAS DE NO RESIDENTES y no el total: el total de la cuenta
# equivale a la variación de reservas, y `reservas_bcra` es componente del ITCM.
# Por la misma regla queda afuera «formación de activos externos», que sale del
# mismo balance cambiario que `desequilibrio_monetario` (también componente).
#
# Son FLUJOS en millones de dólares que cruzan el cero: no se rebasean a 100
# —dividir por un promedio cercano a cero no significa nada— y no hace falta,
# porque la correlación es invariante a la escala y el factor estandariza.
CAPITAL_PRIVADO_IDS = {
    "inversion_directa_externa": "182.1_C_K_FINANCTES_0_M_51",
    "inversion_portafolio_externa": "182.1_C_K_FINC_CTES_0_M_50",
    "financiamiento_externo_privado": "182.1_C_K_FINC_CRED_0_M_52",
}

CONSUMO_FISICO_IDS = {
    "electricidad_residencial": "367.3_DEMANDA_REIAL__19",   # CAMMESA
    "gas_residencial": "364.3_RESIDENCIAIAL__11",            # Secretaría de Energía
    "transporte_pasajeros": "302.3_TRANSP_PASSAJ_0_S_29",    # INDEC
    "ventas_naftas": "38.3_N_1994_M_6",                      # Secretaría de Energía
}
SERIES_API = "https://apis.datos.gob.ar/series/api/series"
MERVAL_YAHOO_URL = "https://query1.finance.yahoo.com/v8/finance/chart/%5EMERV"
CCL_URL = "https://api.argentinadatos.com/v1/cotizaciones/dolares/contadoconliqui"
EPU_LATAM_URL = ("https://www.bde.es/f/webbe/SES/AnalisisEconomico/AnalisisEconomico/"
                 "America_latina/Publicaciones/EPU_LATAM.xlsx")

ROOT = Path(__file__).resolve().parents[1]
SERIES = ROOT / "web" / "src" / "data" / "series.json"
SALIDA = ROOT / "output" / "validacion_externa.json"
BASE_MESES = ("2023-10", "2023-11", "2023-12")

# componente → (clave de serie, invertido, anual, ya_es_indice)
# ADR-0154 saca `indice_lider` y `endeudamiento_familiar`: dejaron de integrar
# el ITVC, así que la reconstrucción histórica y la matriz de redundancia no
# pueden seguir midiéndolos — medirían una composición que el índice ya no
# publica. El líder no desaparece del script: pasó a validador externo del
# ITCM, más abajo.
COMPONENTES = {
    "ipc_alimentos":          ("itvc_alimentos", False, False, True),   # ya base-100 (ADR-0033: relativo al IPC)
    "peso_tarifas":           ("itvc_tarifas", False, False, True),
    "alquiler_real":          ("itvc_alquiler", False, False, True),  # ADR-0111
    # ADR-0218: cantidad de empleadores PyME (SRT). Serie en unidades, no
    # base-100, así que la reconstrucción la rebasea como a las demás.
    "mortalidad_pymes":       ("mortalidad_pymes", False, False, False),
    "despacho_cemento":       ("itvc_isac", False, False, True),
    "pobreza_nowcast":        ("itvc_pobreza", False, False, True),   # ADR-0153
    "mora_familias":          ("mora_familias", True, False, False),   # ADR-0067 (2026-07-15):
    # separada del compuesto de endeudamiento; nivel B100 vs 4T-2023 INVERTIDO
    # (más mora = peor) — la reconstrucción quedó un día desactualizada tras el
    # split y la dimensión vulnerabilidad renormalizaba sobre endeudamiento solo
    # (hallazgo de la revisión ITVC↔ICC 2026-07-16: el pico reconstruido de
    # oct-2025 ignoraba la mora disparándose en plena crisis)
    "carga_servicio_deuda_hogares": (
        "carga_servicio_deuda_hogares", True, False, False),  # ADR-0231
    "brecha_salario_cbt":     ("brecha_salario_cbt", False, False, False),
    "icc_utdt":               ("icc_utdt", False, False, False),
    "pluriempleo":            ("pluriempleo", True, False, False),
    # ADR-0130: empleo registrado privado (SIPA). NO invertido — más empleo es
    # mejor. Entra a la reconstrucción como los demás componentes de rebase.
    "empleo_registrado":      ("empleo_registrado", False, False, False),
    # ADR-0217: puntúa el consumo TOTAL de carnes, no la vacuna sola. Su
    # serie se reconstruye desde la faena del INDEC y YA llega en base 100.
    "consumo_carnes_total":   ("consumo_carnes_total", False, False, True),
    # ADR-0224: el que puntúa es la motorización TOTAL —autos + motos per
    # cápita—, no cada vehículo por su lado. Su serie YA llega en base 100 y
    # con el móvil de 12 meses aplicado a la suma, así que entra como la de
    # carnes: sin rebasear y sin pasar por MOVIL12.
    "motorizacion_total":     ("motorizacion_total", False, False, True),
    # ADR-0225: ventas en supermercados a precios constantes, serie
    # desestacionalizada del INDEC. Rebase normal a 4T-2023 y sin móvil 12m,
    # igual que en `itvc.indices_desde_series` — la fuente ya la desestacionalizó.
    "consumo_supermercados":  ("consumo_supermercados", False, False, False),
    "informalidad":           ("informalidad", True, True, False),
    # ADR-0219: invertido, igual que informalidad y pluriempleo.
    "trabajo_independiente":  ("trabajo_independiente", True, False, False),
    "inseguridad":            ("inseguridad", True, False, False),      # IVI mensual (ADR-0032)
    "sentimiento_digital":    ("sentimiento_digital", True, False, False),  # ADR-0034
}
# Bases DECLARADAS distintas del 4T-2023 (misma regla que publicar):
BASES_PROPIAS = {"inseguridad": ("2024-01",)}   # IVI reanudado ene-2024 (ADR-0032)
# Componentes que entran por acumulado móvil de 12 meses porque su flujo
# mensual crudo tiene estacionalidad fuerte y contra una base fija mediría
# calendario. Tiene que ser la MISMA lista que aplica
# `itvc.indices_desde_series`, o la serie reconstruida y el índice vivo dejan
# de ser el mismo índice.
#
# Quedó VACÍA en ADR-0224: los dos que estaban —motos (ADR-0024) y autos
# (ADR-0223)— se fundieron en `motorizacion_total`, cuyo colector aplica la
# ventana móvil a la SUMA de los dos y entrega la serie ya rebaseada. Se deja
# el mecanismo en pie porque la próxima serie de flujo crudo lo va a necesitar.
MOVIL12: set[str] = set()
ITVC_TECHO = 140.0                              # winsorización asimétrica (ADR-0033)
# ADR-0224: la excepción al techo, acotada a un componente. Tiene que ser la
# MISMA que `itvc.WINSOR_EXENTOS` — si divergen, la serie reconstruida y el
# índice vivo dejan de ser el mismo índice, que es justo lo que esta
# reconstrucción existe para poder afirmar.
TECHO_EXENTOS = frozenset({"motorizacion_total"})


def _mensual(serie: list) -> dict:
    return {p["fecha"][:7]: p["valor"] for p in serie}


def cargar_series() -> dict:
    """Combina el snapshot publicado con los CSV locales recién descargados.

    **Único punto de lectura de series de este módulo.** La validación corre
    ANTES que publicar.py en el pipeline, así que `series.json` todavía es el
    del día anterior: los CSV que acaba de escribir descargar_series.py tienen
    que prevalecer, y el snapshot sólo aporta los puntos históricos que ya no
    estén en los CSV.

    Hasta 2026-07-25 sólo la reconstrucción del ITCM pasaba por acá y el resto
    leía `series.json` crudo. Eso hacía que la matriz de redundancia se
    calculara con las series de ayer y que el test que la compara contra una
    reconstrucción viva fallara en el cron casi todas las noches
    ("la matriz publicada mide 70 pares y la reconstrucción da 62"), con la
    falla apareciendo recién en pytest, tres pasos después del cálculo.
    """
    acumuladas = json.loads(SERIES.read_text(encoding="utf-8"))
    for clave, puntos_frescos in publicar.build_series().items():
        por_fecha = {p["fecha"]: p for p in acumuladas.get(clave) or []}
        por_fecha.update({p["fecha"]: p for p in puntos_frescos})
        acumuladas[clave] = [por_fecha[fecha] for fecha in sorted(por_fecha)]
    return acumuladas


def _cargar_series_itcm() -> dict:
    """Alias histórico de `cargar_series` (los tests lo monkeypatchean)."""
    return cargar_series()


def _movil12(vals: dict) -> dict:
    """Acumulado móvil de 12 meses consecutivos (ADR-0024: desestacionaliza
    flujos con calendario fuerte, ej. motos)."""
    yms = sorted(vals)
    out = {}
    for i in range(11, len(yms)):
        win = yms[i - 11:i + 1]
        a0, m0 = int(win[0][:4]), int(win[0][5:7])
        af, mf = int(win[-1][:4]), int(win[-1][5:7])
        if (af * 12 + mf) - (a0 * 12 + m0) == 11:
            out[win[-1]] = sum(vals[k] for k in win) / 12.0
    return out


def _rebase(vals: dict, invertido: bool, anual: bool, base_meses: tuple = None) -> dict:
    """Serie {ym: valor} → {ym: índice base-100 vs 4T-2023 (o base declarada)}."""
    if anual:
        # base = valor del año 2023; el índice anual se asigna al mes de enero
        # del dato y el forward-fill mensual lo propaga
        base = next((v for ym, v in sorted(vals.items()) if ym[:4] == "2023"), None)
        if not base:
            return {}
        return {ym: round((base / v if invertido else v / base) * 100.0, 1)
                for ym, v in vals.items() if v}
    base_vals = [vals[m] for m in (base_meses or BASE_MESES) if vals.get(m)]
    if not base_vals:
        return {}
    base = sum(base_vals) / len(base_vals)
    return {ym: round((base / v if invertido else v / base) * 100.0, 1)
            for ym, v in vals.items() if v and ym >= "2023-10"}


def _meses(desde: str, hasta: str) -> list:
    out, (y, m) = [], (int(desde[:4]), int(desde[5:7]))
    while f"{y}-{m:02d}" <= hasta:
        out.append(f"{y}-{m:02d}")
        m += 1
        if m > 12:
            m, y = 1, y + 1
    return out


def _indices_itvc_por_componente() -> dict:
    """{componente: {YYYY-MM: índice base 100}} con todo ya resuelto.

    Lo comparten la reconstrucción de la serie del ITVC y su matriz de
    redundancia, para que no puedan divergir en qué componentes miran ni en
    cómo los rebasan — el mismo motivo por el que existe
    `_valores_itcm_por_mes`.
    """
    series = cargar_series()
    indices_por_comp = {}
    for comp, (skey, invertido, anual, ya_rebaseada) in COMPONENTES.items():
        vals = _mensual(series.get(skey) or [])
        if comp in MOVIL12:
            vals = _movil12(vals)          # estacionalidad fuerte del flujo crudo
        idx = (vals if ya_rebaseada
               else _rebase(vals, invertido, anual, BASES_PROPIAS.get(comp)))
        # winsorización asimétrica del ADR-0033: mismo techo que publicar, y
        # la misma excepción acotada del ADR-0224
        indices_por_comp[comp] = (
            dict(idx) if comp in TECHO_EXENTOS
            else {ym: min(v, ITVC_TECHO) for ym, v in idx.items()})
    return indices_por_comp


def _valores_itvc_por_mes() -> dict:
    """{YYYY-MM: {componente: índice}} — la vista por mes de lo anterior.

    Arrastra el último dato disponible de cada componente (doc IV.2.1), que es
    lo que hace el índice publicado: sin eso, un mes cualquiera tendría sólo los
    componentes que publicaron justo ese mes.
    """
    por_comp = _indices_itvc_por_componente()
    ult = max(max(v) for v in por_comp.values() if v)
    out = {}
    for ym in _meses("2023-12", ult):
        punto = {}
        for comp, vals in por_comp.items():
            previos = [k for k in vals if k <= ym]
            if previos:
                punto[comp] = vals[max(previos)]
        out[ym] = punto
    return out


def construir_series_itvc() -> tuple:
    """(serie ITVC completa, serie ITVC sin ICC, serie ICC) mensuales."""
    series = cargar_series()
    indices_por_comp = _indices_itvc_por_componente()
    ult = max(max(v) for v in indices_por_comp.values() if v)
    itvc_full, itvc_sin_icc = {}, {}
    for ym in _meses("2023-12", ult):
        punto = {}
        for comp, vals in indices_por_comp.items():
            previos = [k for k in vals if k <= ym]
            if previos:
                punto[comp] = vals[max(previos)]     # último dato disponible (doc IV.2.1)
        r = itvc.calcular_itvc(punto)
        if r:
            itvc_full[ym] = r["valor"]
        r2 = itvc.calcular_itvc({k: v for k, v in punto.items() if k != "icc_utdt"})
        if r2:
            itvc_sin_icc[ym] = r2["valor"]
    icc = _mensual(series.get("icc_utdt") or [])
    return itvc_full, itvc_sin_icc, icc


def _valores_itcm_por_mes() -> dict:
    """{YYYY-MM: {indicador: valor}} con los componentes del ITCM listos para
    el motor (ventanas móviles y equivalencias ya resueltas). Lo comparten la
    reconstrucción del índice y la matriz de redundancia interna, para que no
    puedan divergir en qué componentes miran."""
    series = _cargar_series_itcm()
    m = lambda k: _mensual(series.get(k) or [])
    ipc_mm = m("ipc_total")               # ya publicada en % m/m (04-jul-2026)
    rem = m("rem_ipc_12m")                # % anual → equivalente mensual
    saldo = m("saldo_comercial")          # M USD mensual → suma móvil 12m
    directos = {k: m(k) for k in ("idm", "desequilibrio_monetario", "recaudacion",
                                  "reservas_bcra", "idc", "credito_privado",
                                  "emae_ia", "emae_difusion", "ipi_manufacturero", "tcrm",
                                  # ADR-0071 / ADR-0072: ambos tienen serie
                                  # mensual desde dic-2023 y entran a la
                                  # reconstrucción como valores directos.
                                  "costo_financiamiento_tesoro",
                                  "resultado_primario")}

    def saldo_12m(ym):
        yms = sorted(saldo)
        if ym not in yms or yms.index(ym) < 11:
            return None
        win = yms[yms.index(ym) - 11:yms.index(ym) + 1]
        a0, m0 = int(win[0][:4]), int(win[0][5:7])
        af, mf = int(win[-1][:4]), int(win[-1][5:7])
        return sum(saldo[k] for k in win) if (af * 12 + mf) - (a0 * 12 + m0) == 11 else None

    return {
        ym: {
            "ipc_total": ipc_mm.get(ym),
            "rem_ipc_12m": rem.get(ym),   # crudo: lo transforma el motor
            "saldo_comercial_12m": saldo_12m(ym),
            **{k: v.get(ym) for k, v in directos.items()},
        }
        for ym in _meses("2023-12", max(ipc_mm))
    }


LINEA_BASE_YM = "2023-12"


def linea_base_itcm(serie_itcm: dict) -> dict | None:
    """El ITCM del mes del traspaso, con su cobertura declarada (ADR-0106).

    La auditoría de macro observó que el índice puntúa el estado actual contra
    anclas fijas, de modo que "un mes de diciembre de 2023 y un mes de hoy se
    evalúan con la misma tabla" — correcto para medir tensión vigente, pero deja
    sin responder la mitad de la pregunta declarada, que incluye avanzar
    respecto de lo recibido en la transición.

    El valor sale de la MISMA reconstrucción que ya se usa para validar el
    índice contra su ancla externa, así que la línea de base y la serie publicada
    no pueden divergir: son el mismo cálculo.

    La cobertura se emite porque el mes del traspaso es justamente el que peor
    cubierto está —varias series arrancan con el mandato— y publicar el número
    sin decir sobre qué porción del índice se calculó lo haría parecer más firme
    de lo que es.
    """
    valores = _valores_itcm_por_mes().get(LINEA_BASE_YM) or {}
    total = con_dato = 0.0
    sin_dato = []
    for dim in itcm.DIMENSIONES_ITCM.values():
        for ind, peso_ind in dim["indicadores"].items():
            peso = dim["peso"] * peso_ind
            total += peso
            if valores.get(ind) is None:
                sin_dato.append(ind)
            else:
                con_dato += peso
    valor = serie_itcm.get(LINEA_BASE_YM)
    if valor is None or not total:
        return None
    return {
        "periodo": LINEA_BASE_YM,
        "valor": valor,
        "cobertura": round(con_dato / total, 3),
        "sin_dato": sorted(sin_dato),
    }


# Piso de cobertura de las series reconstruidas (ADR-0197). Un mes por debajo
# no se publica: no es "el índice con ruido", es otro índice —el de los pocos
# componentes que llegaron— presentado con el mismo nombre.
PISO_COBERTURA = 0.60


def _cobertura_de_peso(r: dict) -> float:
    """Fracción del peso NOMINAL del índice que efectivamente tiene dato.

    No es lo mismo que contar dimensiones con algún dato, que era el criterio
    anterior del ITCP. El motor renormaliza DENTRO de cada dimensión, así que
    una dimensión con un solo indicador vivo de cinco sigue aportando su peso
    entero y la cobertura dimensional da 100%. Medido de las dos maneras, el
    ITCG de jul-2026 daba 100% dimensional y 44,8% de peso real: el criterio
    viejo lo dejaba pasar.
    """
    return sum(d["peso"] * sum(i["peso"] for i in d["indicadores"].values())
               for d in r["dimensiones"].values())


def _ultimo_mes_completo() -> str:
    """El mes calendario cerrado. El mes en curso reconstruye con lo que haya
    llegado hasta hoy, que es una muestra arbitraria del mes."""
    hoy = datetime.now(timezone.utc)
    return f"{hoy.year - 1}-12" if hoy.month == 1 else f"{hoy.year}-{hoy.month - 1:02d}"


def _serie_con_piso(nombre: str, valores_por_mes: dict, calcular,
                    hasta: str | None = None) -> dict:
    """Reconstruye la serie y descarta los meses que no llegan al piso.

    Se descarta en vez de publicar-con-advertencia porque el consumidor de esto
    es una correlación: un punto con 30% de cobertura pesa lo mismo que uno con
    94% dentro del r, y ninguna nota al pie corrige eso.
    """
    out = {}
    for ym, valores in valores_por_mes.items():
        if hasta and ym > hasta:
            print(f"  [i] serie {nombre}: {ym} excluido por mes en curso (incompleto)")
            continue
        r = calcular(valores)
        if not r:
            continue
        cobertura = _cobertura_de_peso(r)
        if cobertura < PISO_COBERTURA:
            print(f"  [i] serie {nombre}: {ym} excluido por cobertura insuficiente "
                  f"({cobertura:.0%} del peso del índice con datos)")
            continue
        out[ym] = r["valor"]
    return out


def construir_serie_itcm() -> dict:
    """Serie mensual del ITCM reconstruida desde las series de componentes
    (mismo motor, puntaje interpolado, sin overrides del analista): todos los
    componentes tienen serie salvo IAI/ICIP, que faltan y el motor renormaliza.
    Reservas netas solo desde jun-2024 (límite de fuente documentado).

    Sin tope de mes en curso, a diferencia de ITCG/ITCP: las fuentes del ITCM
    (INDEC, BCRA mensual, Hacienda) publican por mes cerrado, así que el último
    mes reconstruible ya es un mes completo. El piso de cobertura sí se aplica,
    aunque hoy no recorte nada (mínimo histórico 73,4%): es la red por si una
    fuente se cae y el índice queda armado sobre la mitad de sus componentes."""
    return _serie_con_piso("ITCM", _valores_itcm_por_mes(), itcm.calcular_itcm)


# Pares cuyo acoplamiento es DE DISEÑO, con el motivo. No son hallazgos: que
# correlacionen alto es la construcción funcionando, y presentarlos junto a los
# demás induce a leer como defecto lo que es intencional.
ACOPLADOS_POR_DISENO = {
    frozenset(("ipc_total", "rem_ipc_12m")):
        "el REM es la inflación esperada: un pronóstico del IPC, en la misma "
        "dimensión y a propósito, para leer la misma magnitud en dos momentos",
    frozenset(("credito_privado", "idc")):
        "los dos se construyen sobre depósitos y préstamos del sistema "
        "bancario; la superposición está declarada desde su diseño",
    frozenset(("desafios_legislativos", "bloqueo_sostenido")):
        "son el denominador y la tasa de la misma razón: cuántas normas propias "
        "desafió el Congreso, y qué proporción de ellas el Ejecutivo logró "
        "sostener. Se los mantiene separados porque responden preguntas "
        "distintas —cuánto lo confrontan y cuánto aguanta—, pero no son dos "
        "confirmaciones independientes",
}


def matriz_redundancia(escala, dimensiones: dict, valores_por_mes: dict,
                       por_diseno: dict | None = None,
                       umbral: float = 0.7) -> dict:
    """Correlación de Pearson entre los PUNTAJES mensuales de los componentes
    de un índice (sección IV.3 de la auditoría de jul-2026).

    Se correlacionan puntajes y no valores crudos a propósito: el puntaje es lo
    que efectivamente se promedia dentro del índice, así que es ahí donde dos
    indicadores que se mueven juntos terminan contando dos veces el mismo ciclo.
    Un par con |r| alto y en DIMENSIONES DISTINTAS es el caso que preocupa: el
    índice cree estar midiendo dos cosas y mide una.

    `por_diseno` declara los pares cuyo acoplamiento es intencional, con el
    motivo: publicarlos junto a los demás induce a leer como defecto lo que es
    construcción.

    Genérica desde ADR-0085: la usan ITCM, ITCG e ITCP con su propia escala,
    sus dimensiones y su reconstrucción de valores.
    """
    por_diseno = por_diseno or {}
    dim_de = {ind: dkey for dkey, d in dimensiones.items() for ind in d["indicadores"]}
    puntajes = {}
    for ym, valores in valores_por_mes.items():
        for ind, val in valores.items():
            if val is None or not escala.puntuable(ind):
                continue
            puntajes.setdefault(ind, {})[ym] = escala.puntaje(val, ind)

    inds = sorted(puntajes)
    matriz, pares = {}, []
    for i, a in enumerate(inds):
        for b in inds[i + 1:]:
            r, n = _pearson(puntajes[a], puntajes[b])
            if r is None:
                continue
            matriz.setdefault(a, {})[b] = r
            matriz.setdefault(b, {})[a] = r
            if abs(r) >= umbral:
                pares.append({
                    "a": a, "b": b, "r": r, "n": n,
                    "dimension_a": dim_de.get(a), "dimension_b": dim_de.get(b),
                    "misma_dimension": dim_de.get(a) == dim_de.get(b),
                    "por_diseno": por_diseno.get(frozenset((a, b))),
                })
    pares.sort(key=lambda p: -abs(p["r"]))
    todos = [r for i, a in enumerate(inds) for b in inds[i + 1:]
             if (r := matriz.get(a, {}).get(b)) is not None]

    # La MISMA matriz sobre primeras diferencias. Es el test que separa
    # co-tendencia de co-movimiento: dos series que sólo suben correlacionan
    # cerca de 1 aunque no compartan información, y en un índice con
    # contadores acumulados (varios del ITCG) eso infla la redundencia
    # aparente. Sobre diferencias, la tendencia común se cancela y queda lo
    # que se mueve junto MES A MES, que es lo que de verdad se cuenta dos
    # veces al promediar.
    difs = {ind: _difs(serie) for ind, serie in puntajes.items()}
    r_difs = [r for i, a in enumerate(inds) for b in inds[i + 1:]
              if (r := _pearson(difs[a], difs[b])[0]) is not None]
    dif_resumen = {
        "n_pares": len(r_difs),
        "r_abs_medio": round(statistics.mean(abs(r) for r in r_difs), 3) if r_difs else None,
        "share_altos": round(sum(1 for r in r_difs if abs(r) >= umbral) / len(r_difs), 3) if r_difs else None,
    }
    return {
        "diferencias": dif_resumen,
        "umbral": umbral,
        "n_indicadores": len(inds),
        "n_pares": len(todos),
        "r_abs_medio": round(statistics.mean(abs(r) for r in todos), 3) if todos else None,
        "share_altos": round(sum(1 for r in todos if abs(r) >= umbral) / len(todos), 3) if todos else None,
        "share_bajos": round(sum(1 for r in todos if abs(r) < 0.3) / len(todos), 3) if todos else None,
        "matriz": matriz,
        "pares_altos": pares,
        "pares_cruzados": sum(1 for p in pares if not p["misma_dimension"]),
        # El número que realmente importa: acoplados, de dimensiones distintas
        # y sin una razón de diseño que lo explique.
        "pares_no_explicados": sum(1 for p in pares
                                   if not p["misma_dimension"] and not p["por_diseno"]),
    }


def matriz_redundancia_itcm(umbral: float = 0.7) -> dict:
    return matriz_redundancia(itcm.ESCALA_ITCM, itcm.DIMENSIONES_ITCM,
                              _valores_itcm_por_mes(), ACOPLADOS_POR_DISENO, umbral)


def matriz_redundancia_itcg(umbral: float = 0.7) -> dict:
    escala = parametrica.Escala(itcg.BANDAS_ITCG,
                                getattr(itcg, "ANCLAS_ITCG", None),
                                getattr(itcg, "TRANSFORMACIONES_ITCG", None))
    return matriz_redundancia(escala, itcg.DIMENSIONES_ITCG,
                              _valores_itcg_por_mes(), ACOPLADOS_POR_DISENO, umbral)


class _EscalaIdentidad:
    """Adaptador para que el ITVC pueda usar la matriz de los otros tres.

    Los índices por bandas convierten un valor crudo en puntaje 0-100 y es ese
    puntaje el que se promedia. El ITVC no: sus componentes YA son índices base
    100 = 4T-2023, y el número que se promedia es el índice mismo. Así que la
    conversión correcta es la identidad — no una escala ausente.

    Mantiene el contrato de `parametrica.Escala` (`puntuable` y `puntaje`) para
    no tener que ramificar `matriz_redundancia`, que es genérica desde ADR-0085.
    """

    def __init__(self, componentes):
        self._comp = set(componentes)

    def puntuable(self, indicador: str) -> bool:
        return indicador in self._comp

    def puntaje(self, valor, indicador: str) -> float:
        return float(valor)


def matriz_redundancia_itvc(umbral: float = 0.7) -> dict:
    """Redundancia interna del ITVC (ADR-0108).

    La auditoría de vida cotidiana pidió expresamente comprobar si
    `patentamiento_motos` «aporta señal independiente» del ICC, dado su peso
    marginal y su ambigüedad de constructo (confianza del consumidor vs. acceso
    al crédito prendario). Ésta es la medición que responde esa pregunta, y de
    paso las mismas dudas sobre `consumo_carne`.

    ADR-0224 cerró esa pregunta por otro lado: motos y autos dejaron de ser
    componentes y se fundieron en `motorizacion_total`. Con eso desaparece de
    la matriz el par autos↔motos, que ADR-0223 había tenido que declarar como
    el único del cinturón por encima de 0,7 al destendenciar (+0,801) — y que
    daba tan alto justamente porque el techo aplanaba a motos y comparaba a
    autos contra una recta.

    ADVERTENCIA DE LECTURA: los componentes entran winsorizados al techo de
    ADR-0033, igual que en el índice publicado. Un componente clavado en el
    techo pierde varianza, y sin varianza no hay correlación que calcular — su
    fila puede salir vacía o subestimada. Es una limitación real de la medición,
    no un resultado. `motorizacion_total` es la única excepción (TECHO_EXENTOS,
    ADR-0224): entra sin recortar, así que su fila SÍ es comparable con las
    demás — y conviene tenerlo presente al leer la matriz, porque no todas las
    filas están medidas en las mismas condiciones.
    """
    comp = {i for d in itvc.DIMENSIONES_ITVC.values() for i in d["indicadores"]}
    dims = {k: {"indicadores": d["indicadores"]} for k, d in itvc.DIMENSIONES_ITVC.items()}
    return matriz_redundancia(_EscalaIdentidad(comp), dims,
                              _valores_itvc_por_mes(), ACOPLADOS_POR_DISENO, umbral)


def matriz_redundancia_itcp(umbral: float = 0.7) -> dict:
    escala = parametrica.Escala(itcp.BANDAS_ITCP,
                                getattr(itcp, "ANCLAS_ITCP", None),
                                getattr(itcp, "TRANSFORMACIONES_ITCP", None))
    return matriz_redundancia(escala, itcp.DIMENSIONES_ITCP,
                              _valores_itcp_por_mes(), ACOPLADOS_POR_DISENO, umbral)




ITCG_SERIES = [
    "cepo_mulc", "apertura_comercial", "desregulacion_normativa",
    "reduccion_estado", "gasto_funcionamiento", "masa_salarial",
    "reestructuracion_organismos", "fal_modernizacion_laboral",
    "litigiosidad_laboral", "privatizaciones", "rigi_inversiones",
    "concesiones_infraestructura", "asistencia_directa",
    "protocolo_antipiquetes", "libertad_opcion_salud",
]


def construir_serie_itcg() -> dict:
    """Serie mensual del ITCG reconstruida desde las series de componentes
    (mismo motor, puntaje interpolado, sin overrides del analista): 14 de los
    15 componentes tienen serie con historia; el protocolo antipiquetes recién
    acumula y el motor renormaliza.

    PISO DE COBERTURA Y MES EN CURSO (2026-08-12, ADR-0197). El ITCP tenía las
    dos defensas desde 2026-07-09 y el ITCG no tenía ninguna, así que la cola de
    la serie venía publicando meses armados sobre una fracción del índice: jul
    con 44,8% del peso y ago con 29,2%, contra una mediana histórica de 94%.
    Los dos entraban a la correlación contra el Merval como puntos de pleno
    derecho.

    No era ruido: los componentes que faltan en la cola son sistemáticamente los
    que puntúan alto (apertura_comercial 14%, reduccion_estado 10,9%), así que
    la ausencia empuja el índice para abajo. Recalculando jun-2026 —cobertura
    plena, índice 81,8— con sólo los componentes que sobrevivían en agosto da
    66,9, prácticamente el 65,2 que se publicaba como agosto. La "caída" de 16
    puntos era el faltante, no la gestión."""
    return _serie_con_piso("ITCG", _valores_itcg_por_mes(), itcg.calcular_itcg,
                           hasta=_ultimo_mes_completo())


# Indicadores del ITCG cuya SERIE guarda una magnitud distinta de la que
# puntúa el índice, así que no pueden entrar a la reconstrucción histórica
# (ADR-0086). No es lo mismo que "no tienen serie": la tienen, pero mide otra
# cosa y puntuarla contra las bandas del indicador da un número sin sentido.
ITCG_SERIE_NO_COMPARABLE = {
    # La card puntúa el % de inversión aprobada sobre el pipeline (bandas
    # 0-60+); la serie guarda el MONTO del pipeline en millones de dólares.
    # Puntuar 31.192 contra bandas de porcentaje daba 100 en todos los meses
    # desde ene-2025 y 10 antes: un escalón binario que no existió.
    "rigi_inversiones": "serie en M USD vs banda en % (ADR-0086)",
}


def _valores_itcg_por_mes() -> dict:
    """{YYYY-MM: {indicador: valor crudo}} del ITCG. Lo comparten la
    reconstrucción y la matriz de redundancia, por la misma razón que en el
    ITCM: que no puedan divergir en qué componentes miran (ADR-0082)."""
    series = cargar_series()
    valores_por_comp = {k: _mensual(series.get(k) or [])
                        for k in ITCG_SERIES if k not in ITCG_SERIE_NO_COMPARABLE}
    ult = max(max(v) for v in valores_por_comp.values() if v)
    return {ym: {k: v.get(ym) for k, v in valores_por_comp.items()}
            for ym in _meses("2023-12", ult)}


# Composición post ADR-0052 (2026-07-11): conflictividad_nacional (ACLED
# país, serie real de 30 meses) reemplaza a movilizacion_cepa (2 puntos,
# acumulado YTD) en la dimensión conflicto_social — la reconstrucción gana
# una pata mensual completa desde dic-2023. Antes, post ADR-0048
# (2026-07-10): cohesion_bloque ya es la serie del COMPUESTO bicameral
# 65/35 (una sola clave); cohesion_bloque_senado, rotacion_gabinete,
# protestas_caba y ahora movilizacion_cepa están fuera del índice — no
# entran a la reconstrucción aunque sus series sigan existiendo como
# contexto.
# Los componentes se DERIVAN de las dimensiones del índice, no se listan a
# mano (ADR-0082, aplicado acá el 2026-07-19). La lista escrita a mano que
# había antes ya había divergido: seguía nombrando a derrotas_legislativas
# —fuera del índice desde ADR-0089— y no incluía a los dos indicadores nuevos,
# así que la matriz de redundancia publicaba pares de un índice que ya no
# existía. Es el mismo bug que ADR-0082 fue a arreglar en el ITCM; el ITCP se
# había quedado con su versión.
#
# Un componente sin serie en disco no rompe nada: queda con {} y la
# reconstrucción renormaliza, igual que con veto_quorum antes de su primer
# período o bloqueo_sostenido antes de mar-2024.
# Indicadores del ITCP cuya SERIE es ANUAL y por eso no puede entrar a la
# reconstrucción mensual (ADR-0169). Mismo criterio que ITCG_SERIE_NO_COMPARABLE
# (ADR-0086): la serie existe y el indicador puntúa bien desde su card, pero
# interpolada a mensual aporta once ceros y un salto por año — en primeras
# diferencias eso es ruido con forma de escalón, no señal.
#
# Medido el 31-jul-2026 con el contrafáctico de ADR-0095: sacar estos dos de la
# reconstrucción mejora el ITCP↔EPU en diferencias de −0,366 a −0,405, MÁS que
# sacar los cuatro indicadores nuevos juntos (−0,402). En niveles el efecto es
# el contrario y por eso hay que mirar las dos métricas.
ITCP_SERIE_ANUAL = {
    "judicializacion": "serie anual (SAIJ), un punto por año — ADR-0169",
    "velocidad_resolucion": "serie anual (anuario CSJN), un punto por año — ADR-0169",
}

ITCP_SERIES = [k for d in itcp.DIMENSIONES_ITCP.values() for k in d["indicadores"]
               if k not in ITCP_SERIE_ANUAL]

# MÁSCARA DE ERA para eficacia_legislativa en la reconstrucción (ADR-0070,
# 2026-07-16): la cohorte madura del indicador (expedientes PE publicados
# 12-24 meses antes del mes evaluado) recién es 100% de la gestión actual
# cuando t−730d ≥ 10-dic-2023, o sea desde DIC-2025. Antes de eso el
# indicador mide la cartera de la gestión ANTERIOR muriendo con el cambio de
# congreso (todo 2024: expedientes 2022-2023 de Fernández; el 74→26 de
# puntaje reconstruido en 2024 era ese artefacto, no capital de esta
# gestión). El criterio es A PRIORI —composición de la cohorte, la misma
# doctrina que excluye dic-2023 de toda la reconstrucción (ver docstring de
# construir_serie_itcp)— y NO una calibración contra el benchmark: los meses
# de cohorte mixta (dic-2024→nov-2025) también se excluyen porque siguen
# ponderados mayormente por expedientes pre-gestión. Solo afecta la serie
# reconstruida de validación; la card publicada no se toca (hoy su cohorte
# ya es 100% de esta gestión).
EFICACIA_COHORTE_100PCT_MILEI_DESDE = "2025-12"


def construir_serie_itcp() -> dict:
    """Serie mensual del ITCP reconstruida desde las series de componentes
    (mismo motor, puntaje interpolado, sin overrides del analista) — bastante
    más ruidosa que la de ITCM/ITCG porque la cobertura histórica real de
    política es dispareja:
    - Con historia mensual sólida desde dic-2023: votometro_ventaja_lla,
      eficacia_legislativa, (desde 2026-07-09, ADR-0046)
      derrotas_legislativas —cuya serie completa se deriva del registro
      versionado de eventos— y (desde 2026-07-15, ADR-0058) ratio_dnu, que
      pasó de un punto por año calendario a ventana móvil de 365 días
      recalculada al fin de cada mes.
    - veto_quorum llega por período legislativo (pocos puntos, no un valor
      por mes) e iaf_transferencias es un dato anual (dic-dic): solo
      "prenden" en los meses exactos en que hay dato — el resto del tiempo el
      motor renormaliza sin ellos, igual que ITCM/ITCG con sus faltantes.
    - Desde 2026-07-09 la cobertura mejoró de verdad: cohesion_bloque
      (desde ADR-0048 la serie del compuesto bicameral 65/35, construida
      sobre las dos series por cámara de ADR-0039/0041) y
      alineamiento_senadores_prov (ADR-0038) tienen ~29-31 puntos mensuales
      reales, y adhesion_reformas_provincial 24 (fechas investigadas a mano,
      ADR-0044) — las dimensiones "alianzas territoriales" y "cohesión
      interna" ya no quedan renormalizadas sobre casi nada.
    - protestas_caba y rotacion_gabinete NO entran (contexto desde
      ADR-0048); la transformación var_vs_2023 que protestas necesitaba
      se fue con ellos.

    PISO DE COBERTURA (2026-07-09; desde 2026-08-12 medido por peso de
    indicador y no por dimensión, ADR-0197): los meses que no llegan al 60%
    del peso del índice se EXCLUYEN de la serie. Hallazgo original de la
    auditoría de aquel día: el mes en curso (parcial)
    quedaba reconstruido solo con "poder legislativo" e "imagen y voto"
    (40% del peso, renormalizado al 100%) y daba un valor artefacto que
    saltaba decenas de puntos según qué serie tuviera o no un punto en ese
    mes (55,0 en la corrida anterior, 26,4 en la de hoy, para el MISMO
    mes) — sin par EPU todavía no contaminaba la correlación, pero lo iba
    a hacer apenas EPU publicara ese mes. El piso solo recorta esos meses
    de cola parcial: toda la historia 2024-01→jun-2026 tiene cobertura
    ≥75% y no se toca.

    ARRANQUE EN 2024-01 (2026-07-09, revisión conceptual de la celda
    de la matriz): dic-2023 se excluye de la reconstrucción.
    No por cobertura (pasaba el piso) sino por composición degenerada: los
    componentes de ventana anual que "prenden" ese mes describen el año
    2023 COMPLETO de la gestión anterior — iaf_transferencias dic-dic 2023
    (transferencias de todo 2023 vs 2022; cuando se decidió también pesaba
    protestas_caba, hoy fuera del índice). El punto resultante era un salto
    de composición, no de política, y metía ruido justo en el arranque de
    todas las correlaciones.

    MÁSCARA DE ERA PARA EFICACIA (2026-07-16, ADR-0070): la misma doctrina,
    aplicada por componente — eficacia_legislativa se excluye de la
    reconstrucción hasta nov-2025 inclusive porque su cohorte madura
    (12-24m) recién es 100% de esta gestión desde dic-2025 (ver la
    constante EFICACIA_COHORTE_100PCT_MILEI_DESDE arriba). Y desde ADR-0069
    entra bloqueo_sostenido (tasa de normas desafiadas en pie), que le da a
    la dimensión "poder legislativo" una pata que sí mide 2024: los vetos
    sostenidos de sep/oct-2024 y la supervivencia del DNU 70."""
    series = cargar_series()
    m = lambda k: _mensual(series.get(k) or [])
    directos = {k: m(k) for k in ITCP_SERIES}
    ult = max(max(v) for v in directos.values() if v)
    # tope: el último mes CALENDARIO COMPLETO. El mes en curso (parcial)
    # reconstruye con la cobertura justa y sin par EPU — recortarlo no
    # pierde nada (el artefacto original entró vía rotacion_gabinete, que ya
    # no puntúa, pero el criterio del mes completo sigue valiendo solo).
    hoy = datetime.now(timezone.utc)
    ult_completo = f"{hoy.year - 1}-12" if hoy.month == 1 else f"{hoy.year}-{hoy.month - 1:02d}"
    ult = min(ult, ult_completo)
    # El piso pasó de medirse por dimensión a medirse por PESO (2026-08-12,
    # ADR-0197). Es estrictamente más estricto y hoy no recorta ningún mes del
    # ITCP —su mínimo histórico es 66,8% del peso—, pero deja de haber dos
    # definiciones de "cobertura" en el mismo archivo: la dimensional daba 100%
    # en meses con menos de la mitad del índice cargado.
    return _serie_con_piso("ITCP", _valores_itcp_por_mes(directos, ult),
                           itcp.calcular_itcp)


def _valores_itcp_por_mes(directos: dict | None = None, ult: str | None = None) -> dict:
    """{YYYY-MM: {indicador: valor crudo}} del ITCP, con la máscara de era de
    la eficacia legislativa (ADR-0070) ya aplicada. Compartido por la
    reconstrucción y la matriz de redundancia (ADR-0082)."""
    if directos is None or ult is None:
        series = cargar_series()
        directos = {k: _mensual(series.get(k) or []) for k in ITCP_SERIES}
        ult = max(max(v) for v in directos.values() if v)
        hoy = datetime.now(timezone.utc)
        ult_completo = (f"{hoy.year - 1}-12" if hoy.month == 1
                        else f"{hoy.year}-{hoy.month - 1:02d}")
        ult = min(ult, ult_completo)
    out = {}
    for ym in _meses("2024-01", ult):
        valores = {k: v.get(ym) for k, v in directos.items()}
        if ym < EFICACIA_COHORTE_100PCT_MILEI_DESDE:
            valores["eficacia_legislativa"] = None   # máscara de era (ADR-0070)
        out[ym] = valores
    return out


def _sin_tendencia(serie: dict) -> dict:
    """Residuos de la serie contra una recta en el tiempo.

    Es la cuenta que separa co-movimiento de tendencia compartida, y en esta
    muestra separa mucho: el Merval publicaba +0,751 en niveles contra el ITCG
    y destendenciado queda en +0,067 (ADR-0226). Sin este número, un lector no
    tiene forma de saber cuánto de una correlación alta es que las dos series
    suben durante treinta y un meses.
    """
    meses = sorted(serie)
    n = len(meses)
    if n < 3:
        return {}
    xs = list(range(n))
    ys = [serie[m] for m in meses]
    mx, my = sum(xs) / n, sum(ys) / n
    var = sum((x - mx) ** 2 for x in xs)
    if var == 0:
        return {}
    b = sum((x - mx) * (y - my) for x, y in zip(xs, ys)) / var
    a = my - b * mx
    return {m: ys[i] - (a + b * xs[i]) for i, m in enumerate(meses)}


def _serie_datos_gob(sid: str) -> dict:
    """{YYYY-MM: valor} de una serie de datos.gob.ar."""
    r = requests.get(SERIES_API, params={"ids": sid, "limit": 5000,
                                         "format": "json"}, timeout=60)
    r.raise_for_status()
    return {f[:7]: float(v) for f, v in r.json()["data"] if v is not None}


def _rebase_4t23(serie: dict) -> dict:
    base = [serie[m] for m in BASE_MESES if m in serie]
    if len(base) < len(BASE_MESES):
        raise ValueError("falta algún mes del 4T-2023 para la base")
    prom = sum(base) / len(base)
    return {m: round(100.0 * v / prom, 2) for m, v in serie.items()}


def fetch_merval_usd_mensual() -> dict:
    """{YYYY-MM: Merval en USD} — cierre mensual del índice Merval (Yahoo
    Finance, ^MERV) sobre el CCL promedio del mes (ArgentinaDatos). Es el par
    convergente PROPIO del ITCG (ADR-0031): el mercado de acciones pricea la
    transformación estructural — reformas ejecutadas, empresas que valen más."""
    r = requests.get(MERVAL_YAHOO_URL, params={"range": "3y", "interval": "1mo"},
                     headers={"User-Agent": "Mozilla/5.0"}, timeout=60)
    r.raise_for_status()
    res = r.json()["chart"]["result"][0]
    cierres = {datetime.fromtimestamp(t, tz=timezone.utc).strftime("%Y-%m"): c
               for t, c in zip(res["timestamp"],
                               res["indicators"]["quote"][0]["close"]) if c}
    r2 = requests.get(CCL_URL, headers={"User-Agent": "Mozilla/5.0"}, timeout=60)
    r2.raise_for_status()
    por_mes = {}
    for d in r2.json():
        if d.get("venta"):
            por_mes.setdefault(d["fecha"][:7], []).append(float(d["venta"]))
    ccl = {ym: sum(v) / len(v) for ym, v in por_mes.items()}
    return {ym: round(cierres[ym] / ccl[ym], 1) for ym in sorted(cierres)
            if ym in ccl and ym >= "2023-11"}



def fetch_epu_argentina_mensual() -> dict:
    """Promedio mensual del EPU (Economic Policy Uncertainty) de Argentina,
    columna EPU_ARG_local (basado en diarios locales) de la hoja data_LATAM
    del dataset EPU_LATAM del Banco de España + SECMCA (misma familia
    metodológica de minería de texto que Baker/Bloom/Davis, no un precio de
    mercado): es el par externo del ITCP. Correlación NEGATIVA esperada — más
    capital político (menos tensión del cinturón), menos incertidumbre de
    política percibida en la prensa."""
    import io as _io
    import openpyxl
    r = requests.get(EPU_LATAM_URL, headers={"User-Agent": "Mozilla/5.0"}, timeout=60)
    r.raise_for_status()
    wb = openpyxl.load_workbook(_io.BytesIO(r.content), data_only=True)
    ws = wb["data_LATAM"]
    filas = ws.iter_rows(values_only=True)
    encabezado = next(filas)
    col = encabezado.index("EPU_ARG_local")
    out = {}
    for fila in filas:
        fecha, valor = fila[0], fila[col]
        if fecha and valor is not None:
            out[fecha.strftime("%Y-%m")] = round(float(valor), 1)
    return out


def _pearson(a: dict, b: dict) -> tuple:
    """(r, n) sobre los meses comunes. r = None cuando la correlación no es
    calculable, que ocurre en dos casos:

      * menos de 6 meses en común (muestra insuficiente);
      * alguna de las dos series es CONSTANTE — la correlación es indefinida
        (varianza cero en el denominador) y `statistics.correlation` lanza
        StatisticsError.

    El segundo caso es alcanzable y no teórico: un indicador puede quedar
    saturado en un extremo de su banda durante toda la ventana de solape, y
    varios pasan más del 60% de los meses en un extremo. Sin esta guarda, un
    solo par no calculable abortaba la corrida ENTERA de validación y dejaba
    sin actualizar el snapshot publicado. Un par que no se puede calcular debe
    reportarse como tal, no tumbar a los otros noventa.
    """
    comunes = sorted(set(a) & set(b))
    if len(comunes) < 6:
        return None, len(comunes)
    xs = [a[m] for m in comunes]
    ys = [b[m] for m in comunes]
    if len(set(xs)) < 2 or len(set(ys)) < 2:
        return None, len(comunes)
    return round(statistics.correlation(xs, ys), 3), len(comunes)


HTTP_TIMEOUT = 20
HTTP_HEADERS = {"User-Agent": "CIGOB-InformeCoyuntura/1.0"}

CONSTRUYA_URL = "https://www.ieric.org.ar/wp-content/uploads/{anio}/{mes:02d}/Indice-Construya.xlsx"


def fetch_construya_mensual() -> dict:
    """Variación interanual mensual del Índice Construya (volumen de ventas de
    insumos de la construcción de sus fabricantes líderes, base jun-2002 = 100).

    Es la pata de CONDUCTA del contraste: mide volumen físico efectivamente
    vendido, no expectativas declaradas.

    El archivo se publica en el espejo del IERIC bajo una ruta con año y mes que
    rota, así que se prueban los últimos meses hacia atrás."""
    import openpyxl

    hoy = date.today()
    intentos = []
    contenido = None
    for atras in range(0, 6):
        anio, mes = hoy.year, hoy.month - atras
        while mes <= 0:
            mes += 12
            anio -= 1
        url = CONSTRUYA_URL.format(anio=anio, mes=mes)
        try:
            r = requests.get(url, headers=HTTP_HEADERS, timeout=HTTP_TIMEOUT)
            tipo = r.headers.get("content-type", "")
            if r.status_code == 200 and "spreadsheet" in tipo.lower():
                contenido = r.content
                break
            intentos.append(f"{anio}-{mes:02d}: HTTP {r.status_code}")
        except Exception as e:
            intentos.append(f"{anio}-{mes:02d}: {e}")
    if contenido is None:
        raise ValueError("Construya no descargable — " + " · ".join(intentos))

    ws = openpyxl.load_workbook(io.BytesIO(contenido), data_only=True)["Indice Construya"]
    out = {}
    for i in range(5, ws.max_row + 1):
        fecha, var = ws.cell(i, 1).value, ws.cell(i, 3).value
        if hasattr(fecha, "year") and isinstance(var, (int, float)):
            # La planilla mezcla fracción (0,12) y porcentaje (12,0) según el
            # tramo; se normaliza a porcentaje.
            valor = float(var) * 100 if abs(float(var)) < 3 else float(var)
            out[f"{fecha.year}-{fecha.month:02d}"] = round(valor, 2)
    if len(out) < 24:
        raise ValueError(f"Construya: sólo {len(out)} meses parseados")
    return out


def _ma12(s: dict) -> dict:
    """Promedio móvil de 12 meses de un {YYYY-MM: valor}, para poder comparar
    contra series que ya vienen suavizadas."""
    meses = sorted(s)
    return {meses[i]: round(sum(s[meses[j]] for j in range(i - 11, i + 1)) / 12, 3)
            for i in range(11, len(meses))}


def _serie_indicador(nombre: str) -> dict:
    """{YYYY-MM: valor} de un indicador, leído de los CSV de `output/series/`.

    No de `series.json`: ese archivo lo escribe publicar.py, que en el pipeline
    corre DESPUÉS de este script, así que acá siempre se estaría leyendo la
    versión del día anterior — y para un indicador recién incorporado, una que
    todavía no lo contiene. Los CSV, en cambio, los deja descargar_series.py
    inmediatamente antes."""
    import csv

    out = {}
    for archivo in sorted((ROOT / "output" / "series").glob("*.csv")):
        with archivo.open(encoding="utf-8", newline="") as fh:
            for fila in csv.reader(fh):
                if len(fila) >= 3 and fila[1] == nombre:
                    try:
                        out[fila[0][:7]] = float(fila[2])
                    except ValueError:
                        continue
    return out


def _difs(s: dict) -> dict:
    yms = sorted(s)
    return {yms[i]: round(s[yms[i]] - s[yms[i - 1]], 2) for i in range(1, len(yms))}


def _lag(s: dict, k: int) -> dict:
    """Serie corrida k meses hacia adelante (k>0: s adelanta al comparador)."""
    yms = sorted(s)
    return {yms[i + k]: s[yms[i]] for i in range(len(yms) - k)} if k > 0 else s


# ── La brecha de obra pública cambia de signo según el gobierno (ADR-0095) ───
# Períodos presidenciales, para medir el indicador contra el EPU por separado en
# cada uno. El hallazgo que motiva el corte: la brecha acompaña a la
# incertidumbre de política con las dos administraciones anteriores y se INVIERTE
# con la actual, porque el recorte de obra pública dejó de ser un síntoma de
# gobierno en problemas para ser el programa de gobierno.
GOBIERNOS = [
    ("Macri (desde el inicio de la serie)", "2017-11", "2019-11"),
    ("Alberto Fernández", "2019-12", "2023-11"),
    ("Milei", "2023-12", "9999-99"),
]


def _corr_brecha_por_gobierno(epu: dict) -> dict:
    """{gobierno: {r, n}} del PUNTAJE de la brecha contra el EPU.

    Se correlaciona el puntaje y no el valor crudo porque es el puntaje lo que
    entra al índice: si el signo se da vuelta, se da vuelta ahí."""
    escala = parametrica.Escala(itcp.BANDAS_ITCP, getattr(itcp, "ANCLAS_ITCP", None))
    crudo = _serie_indicador("brecha_obra_publica")
    puntajes = {m: escala.puntaje(v, "brecha_obra_publica") for m, v in crudo.items()}
    out = {}
    for etiqueta, desde, hasta in GOBIERNOS:
        tramo = {m: v for m, v in puntajes.items() if desde <= m <= hasta}
        r, n = _pearson(tramo, epu)
        out[etiqueta] = {"r": r, "n": n}
    return out


def _serie_itcp_sin(dimension: str) -> dict:
    """Reconstrucción del ITCP dejando afuera una dimensión entera, para poder
    publicar cuánto de la validación externa aporta o resta cada una."""
    import copy
    originales = itcp.DIMENSIONES_ITCP
    recortadas = {k: v for k, v in copy.deepcopy(originales).items() if k != dimension}
    global ITCP_SERIES
    guardadas = ITCP_SERIES
    itcp.DIMENSIONES_ITCP = recortadas
    ITCP_SERIES = [k for d in recortadas.values() for k in d["indicadores"]
                   if k not in ITCP_SERIE_ANUAL]
    try:
        return construir_serie_itcp()
    finally:
        itcp.DIMENSIONES_ITCP = originales
        ITCP_SERIES = guardadas


def main():
    itvc_full, itvc_sin, icc = construir_series_itvc()
    # generated_at: sin sello no había forma de notar que este archivo dejó de
    # commitearse. El pipeline lo regeneraba cada noche, publicar.py le sacaba
    # las correlaciones para el snapshot y después se descartaba, así que la
    # copia versionada quedó cinco días atrás sin que nada avisara — y con ella
    # la matriz de redundancia, que hacía fallar un test que se leyó todo ese
    # tiempo como "staleness que resuelve la próxima corrida" (ADR-0177).
    resultados = {"_meta": {"adr": "0019 Decisión 6",
                            # Con offset, como el sello del snapshot (ADR-0203):
                            # test_salidas_versionadas_frescas RESTA los dos, y
                            # naive menos aware es TypeError.
                            "generated_at": datetime.now().astimezone().isoformat(),
                            "nota": "ITVC sin ICC para evitar circularidad (el ICC pesa 7,5% del ITVC)"}}
    print(f"serie ITVC reconstruida: {len(itvc_full)} meses "
          f"({min(itvc_full)} → {max(itvc_full)}) · último: {itvc_full[max(itvc_full)]}")
    resultados["serie_itvc"] = itvc_full
    resultados["serie_itvc_sin_icc"] = itvc_sin

    # ADR-0225: el ITCIS YA NO TIENE ANCLA ÚNICA. Las ventas en supermercados,
    # que lo eran desde ADR-0155, pasaron a componente del índice; y ninguna de
    # las candidatas que quedan sostiene un titular (el desarrollo está en el
    # ADR). El contraste del cinturón es el panel y su factor común, que se
    # calculan más abajo como para los otros dos socioeconómicos.
    #
    # El ICC queda como contraste DISCRIMINANTE, que es lo único que era: mide
    # si la percepción sigue a las condiciones materiales. Necesita la variante
    # sin ICC porque el ICC sí compone el índice.
    pares = {}
    pares.update({
        "discriminante: ITVC sin ICC vs ICC (niveles)": (itvc_sin, icc),
        "discriminante: ITVC sin ICC vs ICC (diferencias)": (_difs(itvc_sin), _difs(icc)),
    })
    resultados["correlaciones"] = {}
    print("\ncorrelaciones (Pearson):")
    for nombre, (a, b) in pares.items():
        r, n = _pearson(a, b)
        resultados["correlaciones"][nombre] = {"r": r, "n": n}
        print(f"  {nombre}: r = {r}  (n = {n})")

    # ── ITCM vs Índice Líder (correlación positiva esperada) ───────────────
    serie_itcm = construir_serie_itcm()
    print(f"\nserie ITCM reconstruida: {len(serie_itcm)} meses "
          f"({min(serie_itcm)} → {max(serie_itcm)}) · último: {serie_itcm[max(serie_itcm)]}")
    resultados["serie_itcm"] = serie_itcm
    base = linea_base_itcm(serie_itcm)
    if base:
        resultados["linea_base_itcm"] = base
        print(f"línea de base {base['periodo']}: ITCM = {base['valor']} "
              f"(cobertura {base['cobertura']:.0%}"
              + (f", sin dato: {', '.join(base['sin_dato'])}" if base["sin_dato"] else "")
              + ")")
    # Índice Líder de la UTDT: el ANCLA de validación del ITCM (ADR-0154 y sus
    # enmiendas). Reemplazó al indicador de mercado que había antes, que
    # correlacionaba fuerte en niveles y ~0 en primeras diferencias.
    #
    # Llega acá desde el ITVC, donde integraba la dimensión de empleo y no
    # correspondía: mide el ciclo de la ACTIVIDAD, no una condición de la vida
    # cotidiana. Cubre un hueco real — el ancla anterior daba r ≈ −0,08 en
    # primeras diferencias, o sea que fuera de la tendencia común no validaba.
    #
    # Se publican los adelantos en las dos direcciones a propósito, porque el
    # resultado va en contra de lo que sugiere el nombre del índice: el que
    # adelanta es el ITCM, no el líder. La lectura correcta se documenta en el
    # ADR y no se puede afirmar «el líder anticipa al ITCM».
    try:
        lider = _serie_indicador("indice_lider")
        if not lider:
            raise ValueError("la serie indice_lider no está en output/series/")
        resultados["indice_lider_mensual"] = lider
        pares_l = {
            "niveles (ITCM vs índice líder)": (serie_itcm, lider),
            "primeras diferencias (ITCM vs líder)": (_difs(serie_itcm), _difs(lider)),
            "líder adelantado 1 mes vs ITCM": (serie_itcm, _lag(lider, 1)),
            "líder adelantado 3 meses vs ITCM": (serie_itcm, _lag(lider, 3)),
            "ITCM adelantado 1 mes vs líder": (_lag(serie_itcm, 1), lider),
        }
        resultados["correlaciones_itcm"] = resultados.get("correlaciones_itcm", {})
        # Puntos de giro: el régimen que corresponde a un compuesto ECONÓMICO
        # con serie de referencia (ADR-0158). La correlación de Pearson sobre
        # niveles, con muestra corta y tendencia común, dice poco; lo que se
        # valida acá es si el ciclo del índice gira cuando gira el de la
        # referencia, y con cuánto adelanto.
        import puntos_de_giro as pdg
        resultados["giros_itcm"] = pdg.analisis(serie_itcm, lider)
        # Qué ventana de la referencia entró de verdad al cálculo (ADR-0226).
        # No es metadato decorativo: recortar la referencia a la ventana del
        # índice le fabrica giros en el borde y mueve la concordancia
        # publicada, y sin este registro el cambio no deja rastro en ningún
        # lado. Hay un test que exige historia previa suficiente.
        resultados["giros_itcm"]["referencia_ventana"] = {
            "desde": min(lider), "hasta": max(lider), "n": len(lider)}
        # ¿el compuesto yerra menos que cada una de sus partes? Es el criterio
        # con el que la OCDE justifica usar un compuesto y no los sueltos.
        _por_mes = _valores_itcm_por_mes()
        _comp = {}
        for _mes, _vals in _por_mes.items():
            for _c, _v in _vals.items():
                if _v is not None:
                    _comp.setdefault(_c, {})[_mes] = _v
        resultados["senales_itcm"] = pdg.compuesto_vs_componentes(serie_itcm, _comp, lider)
        g = resultados["giros_itcm"]
        print("puntos de giro ITCM vs índice líder:")
        print(f"  concordancia de fase: {g['concordancia']}  (n = {g['n_meses']} meses)")
        print(f"  giros del ITCM: {len(g['giros'])} · provisorios: {g['provisorios']}"
              f" · apareados y confirmados: {g['apareados']}")
        s = resultados["senales_itcm"]
        if s.get("evaluables"):
            print(f"  señales falsas/perdidas — compuesto: {s['compuesto']['total']} · "
                  f"componentes peores: {s['peores']}, iguales: {s['iguales']}, "
                  f"mejores: {s['mejores']} (de {s['evaluables']})")
        if g["desfase_medio"] is not None:
            print(f"  desfase medio (sólo confirmados): {g['desfase_medio']:+} meses")
        print("correlaciones ITCM vs índice líder (Pearson, positiva = válida):")
        for nombre, (a, b) in pares_l.items():
            r, n = _pearson(a, b)
            resultados["correlaciones_itcm"][nombre] = {"r": r, "n": n}
            print(f"  {nombre}: r = {r}  (n = {n})")
    except Exception as e:
        print(f"[WARN] índice líder no disponible: {e}")

    # ── Redundancia INTERNA del ITCM (auditoría jul-2026, IV.3) ────────────
    red = matriz_redundancia_itcm()
    resultados["redundancia_itcm"] = red
    # ADR-0085: la misma medición para gestión y política. Un fallo en una no
    # debe tumbar la corrida entera de validación.
    # ADR-0108: el ITVC se suma con escala identidad (sus componentes ya son
    # índices base 100, no puntajes de banda).
    for sigla, fn in (("itcg", matriz_redundancia_itcg), ("itcp", matriz_redundancia_itcp),
                      ("itvc", matriz_redundancia_itvc)):
        try:
            otra = fn()
            resultados[f"redundancia_{sigla}"] = otra
            print(f"redundancia interna {sigla.upper()}: {otra['n_indicadores']} indicadores, "
                  f"{otra['n_pares']} pares · |r| medio {otra['r_abs_medio']} niveles / "
                  f"{otra['diferencias']['r_abs_medio']} en cambios mes a mes")
        except Exception as e:
            print(f"[WARN] redundancia {sigla.upper()}: {e}")
    print(f"\nredundancia interna ITCM: {red['n_indicadores']} indicadores, "
          f"{red['n_pares']} pares · |r| medio {red['r_abs_medio']} · "
          f"{red['share_altos']:.0%} sobre {red['umbral']} "
          f"({red['pares_cruzados']} de ellos entre dimensiones distintas)")
    for p in red["pares_altos"][:5]:
        cruz = "" if p["misma_dimension"] else "  [dimensiones distintas]"
        print(f"  r = {p['r']:+.3f}  {p['a']} × {p['b']}{cruz}")

    # ── ITCG vs ICG UTDT (confianza en el gobierno; positiva esperada) ─────
    serie_itcg = construir_serie_itcg()
    print(f"\nserie ITCG reconstruida: {len(serie_itcg)} meses "
          f"({min(serie_itcg)} → {max(serie_itcg)}) · último: {serie_itcg[max(serie_itcg)]}")
    resultados["serie_itcg"] = serie_itcg
    series_json = cargar_series()
    icg = _mensual(series_json.get("icg_utdt") or [])
    pares_g = {}
    try:
        merval = fetch_merval_usd_mensual()
        resultados["merval_usd_mensual"] = merval
        # El Merval sigue midiéndose y sigue en el panel, pero YA NO se publica
        # como el par que valida al índice (ADR-0226): pricea lo que el mercado
        # espera de la ejecución, no la ejecución. Sus tres números son la
        # prueba de por qué dejó de encabezar.
        pares_g.update({
            "niveles (ITCG vs Merval USD)": (serie_itcg, merval),
            "primeras diferencias (ITCG vs Merval USD)": (_difs(serie_itcg), _difs(merval)),
            # El destendenciado es el que sostiene que el Merval dejó de
            # encabezar (ADR-0226): se calcula acá y no se escribe a mano en
            # la prosa, para que envejezca con los datos.
            "niveles sin tendencia (ITCG vs Merval USD)":
                (_sin_tendencia(serie_itcg), _sin_tendencia(merval)),
        })
    except Exception as e:
        print(f"[WARN] Merval USD no disponible: {e}")
    if icg:
        # Discriminante: el ITCG mide ejecución ACUMULATIVA, no popularidad —
        # la divergencia con el ciclo de confianza política es esperable.
        pares_g.update({
            "niveles (ITCG vs ICG)": (serie_itcg, icg),
            "primeras diferencias (ITCG vs ICG)": (_difs(serie_itcg), _difs(icg)),
        })
    if pares_g:
        resultados["correlaciones_itcg"] = {}
        print("correlaciones ITCG (Pearson):")
        for nombre, (a, b) in pares_g.items():
            r, n = _pearson(a, b)
            resultados["correlaciones_itcg"][nombre] = {"r": r, "n": n}
            print(f"  {nombre}: r = {r}  (n = {n})")

    # ── ITCP vs EPU Argentina (incertidumbre de política; negativa esperada) ──
    serie_itcp = construir_serie_itcp()
    print(f"\nserie ITCP reconstruida: {len(serie_itcp)} meses "
          f"({min(serie_itcp)} → {max(serie_itcp)}) · último: {serie_itcp[max(serie_itcp)]}")
    resultados["serie_itcp"] = serie_itcp
    try:
        epu = fetch_epu_argentina_mensual()
        resultados["epu_argentina_mensual"] = epu
        pares_p = {
            "niveles (ITCP vs EPU Argentina)": (serie_itcp, epu),
            "primeras diferencias (ITCP vs EPU)": (_difs(serie_itcp), _difs(epu)),
            "ITCP adelantado 1 mes vs EPU": (_lag(serie_itcp, 1), epu),
            "EPU adelantado 1 mes vs ITCP": (serie_itcp, _lag(epu, 1)),
        }
        resultados["correlaciones_itcp"] = {}
        print("correlaciones ITCP (Pearson, negativa = válida):")
        for nombre, (a, b) in pares_p.items():
            r, n = _pearson(a, b)
            resultados["correlaciones_itcp"][nombre] = {"r": r, "n": n}
            print(f"  {nombre}: r = {r}  (n = {n})")

        # Cuánto de la correlación aporta o resta la dimensión empresaria, que
        # es la más nueva y la que el EPU no cubre (ADR-0095). Se publica el
        # contrafáctico, no se esconde la caída.
        try:
            sin_priv = _serie_itcp_sin("sector_privado")
            r_sin, n_sin = _pearson(sin_priv, epu)
            resultados["correlaciones_itcp"]["niveles, sin la dimensión de sector privado"] = {
                "r": r_sin, "n": n_sin}
            print(f"  niveles SIN sector privado: r = {r_sin}  (n = {n_sin})")
        except Exception as e:
            print(f"  [WARN] contrafáctico sin sector privado: {e}")

        try:
            porgob = _corr_brecha_por_gobierno(epu)
            resultados["brecha_obra_publica_por_gobierno"] = porgob
            print("  brecha de obra pública vs EPU, por gobierno:")
            for g, x in porgob.items():
                print(f"    {g}: r = {x['r']}  (n = {x['n']})")
        except Exception as e:
            print(f"  [WARN] brecha por gobierno: {e}")
    except Exception as e:
        print(f"[WARN] EPU Argentina no disponible: {e}")

    # ── brecha de obra pública vs Índice Construya (ADR-0088) ────────────────
    # Contraste percepción/conducta: la brecha es lo que las constructoras
    # DICEN esperar; Construya es el volumen de insumos que efectivamente se
    # vende. Si las expectativas se hunden y el volumen no cae, la tensión es
    # discursiva; si caen juntas, es material.
    try:
        construya = fetch_construya_mensual()
        resultados["construya_var_ia_mensual"] = construya
        brecha = _serie_indicador("brecha_obra_publica")
        # La brecha ya viene promediada a 12 meses por construcción, así que
        # Construya se suaviza igual antes de comparar: correlacionar una serie
        # suavizada contra una cruda mide en buena parte la diferencia de
        # suavizado y atenúa el resultado (r baja de 0,79 a 0,26 sólo por eso).
        construya12 = _ma12(construya)
        pares_c = {
            "niveles (brecha obra pública vs Construya var. i.a., ambas 12m)": (brecha, construya12),
            "primeras diferencias (brecha vs Construya)": (_difs(brecha), _difs(construya12)),
        }
        resultados["correlaciones_brecha_obra_publica"] = {}
        print("correlaciones brecha obra pública ↔ Construya (positiva = válida):")
        for nombre, (a, b) in pares_c.items():
            r, n = _pearson(a, b)
            resultados["correlaciones_brecha_obra_publica"][nombre] = {"r": r, "n": n}
            print(f"  {nombre}: r = {r}  (n = {n})")
    except Exception as e:
        print(f"[WARN] Índice Construya no disponible: {e}")

    # ── Dispersión de los componentes del ITVC (ADR-0160) ──────────────────
    # El índice casi no se mueve porque sus componentes se compensan. El neto
    # solo, sin la dispersión al lado, dice "sin cambios" donde el dato dice
    # "no cambió en neto pero se recompuso fuerte por dentro".
    try:
        import statistics as _st
        _por_mes = _valores_itvc_por_mes()
        _serie_disp = {}
        for _mes in sorted(_por_mes):
            _vals = {c: v for c, v in _por_mes[_mes].items() if v is not None}
            if len(_vals) < 8:
                continue
            _r = itvc.calcular_itvc(dict(_vals))
            if not _r:
                continue
            _lo = min(_vals, key=_vals.get)
            _hi = max(_vals, key=_vals.get)
            _serie_disp[_mes] = {
                "itvc": _r["valor"],
                "rango": round(_vals[_hi] - _vals[_lo], 1),
                "desvio": round(_st.pstdev(_vals.values()), 1),
                "min": {"componente": _lo, "valor": round(_vals[_lo], 1)},
                "max": {"componente": _hi, "valor": round(_vals[_hi], 1)},
                "n": len(_vals),
            }
        if _serie_disp:
            _ms = sorted(_serie_disp)
            _pri, _ult = _serie_disp[_ms[0]], _serie_disp[_ms[-1]]
            resultados["dispersion_itvc"] = {
                "serie": _serie_disp,
                "primero": {"mes": _ms[0], **_pri},
                "ultimo": {"mes": _ms[-1], **_ult},
                "movimiento_neto": round(abs(_ult["itvc"] - _pri["itvc"]), 1),
            }
            print("")
            print("dispersión de los componentes del ITVC:")
            print(f"  rango {_pri['rango']} ({_ms[0]}) → {_ult['rango']} ({_ms[-1]})"
                  f" · desvío {_pri['desvio']} → {_ult['desvio']}"
                  f" · movimiento NETO del índice: {resultados['dispersion_itvc']['movimiento_neto']}")
    except Exception as e:
        print(f"[WARN] dispersión del ITVC no disponible: {e}")

    # ── Panel de validación socioeconómica (ADR-0159) ──────────────────────
    # El ITVC/ITCG/ITCP no tienen serie de referencia: se comparan contra VARIAS
    # estadísticas relacionadas y las diferencias se explican. El ITCM tiene su
    # propio régimen (puntos de giro, ADR-0158) y no entra acá.
    try:
        import panel_validacion as pnl
        series_json = cargar_series()
        panel = {
            "merval_usd": resultados.get("merval_usd_mensual") or {},
            "epu_argentina": resultados.get("epu_argentina_mensual") or {},
            "indice_lider": resultados.get("indice_lider_mensual") or {},
            "icg_utdt": _mensual(series_json.get("icg_utdt") or []),
            "clima_electoral": _mensual(series_json.get("clima_electoral") or []),
        }
        for clave, sid in (("consumo_mayoristas", CONSUMO_MAYORISTAS_ID),
                           ("consumo_shoppings", CONSUMO_SHOPPINGS_ID)):
            try:
                panel[clave] = _rebase_4t23(_serie_datos_gob(sid))
            except Exception as e:
                print(f"[WARN] panel: {clave} no disponible: {e}")
        # Volúmenes físicos del hogar: se desestacionalizan ANTES de entrar. Sin
        # eso, el primer componente del panel sería la estación del año y no la
        # condición material de los hogares. Se informa la amplitud estacional
        # antes y después para que se vea que el ajuste hizo algo — y cuánto
        # queda sin explicar, que en el gas no es poco.
        import desestacionalizar as _des
        for clave, sid in CONSUMO_FISICO_IDS.items():
            try:
                cruda = _serie_datos_gob(sid)
                antes = _des.amplitud_estacional(cruda)
                ajustada = _des.desestacionalizar(cruda)
                panel[clave] = _rebase_4t23(ajustada)
                resultados.setdefault("estacionalidad_panel", {})[clave] = {
                    "antes_pct": antes,
                    "despues_pct": _des.amplitud_estacional(ajustada),
                    "n_meses": len(cruda),
                }
            except Exception as e:
                print(f"[WARN] panel: {clave} no disponible: {e}")
        # Flujos de capital privado: se desestacionalizan pero NO se rebasean
        # (cruzan el cero). Deliberadamente NO se acumulan a 12 meses: un flujo
        # acumulado queda casi monótono y correlaciona ~0,96 contra cualquier
        # índice que también suba, que es la trampa que ADR-0159 ya documentó
        # con `indice_salarios_publico`. Medido acá: acumulando 12 meses el
        # financiamiento externo daba 0,962 en niveles y 0,038 mes a mes.
        for clave, sid in CAPITAL_PRIVADO_IDS.items():
            try:
                cruda = _serie_datos_gob(sid)
                antes = _des.amplitud_estacional(cruda)
                ajustada = _des.desestacionalizar(cruda)
                panel[clave] = ajustada
                resultados.setdefault("estacionalidad_panel", {})[clave] = {
                    "antes_pct": antes,
                    "despues_pct": _des.amplitud_estacional(ajustada),
                    "n_meses": len(cruda),
                }
            except Exception as e:
                print(f"[WARN] panel: {clave} no disponible: {e}")
        # Huella de frescura de cada ancla, para que el gate pueda vigilarlas
        # (ADR-0176). Las series crudas del panel no se persisten —son insumo,
        # no salida— así que sin esto no hay forma de saber desde afuera si un
        # ancla se congeló: participa igual en el factor común, con su última
        # observación de hace meses, y las correlaciones se publican como si
        # nada. Le pasó al ICG de la UTDT, que estuvo congelado hasta que lo
        # encontró un aviso lateral y no un chequeo (ADR-0175).
        # Se registran TODAS las declaradas en pnl.FAMILIA, incluidas las que
        # quedaron vacías: un ancla ausente tiene que ser visible como ausente,
        # no desaparecer del registro.
        # `avanzo` = cuándo esta ancla publicó por última vez un período NUEVO.
        # El rezago absoluto mezcla dos cosas distintas: el atraso INHERENTE de
        # la fuente (INDEC publica el consumo con 3 meses) y el congelamiento.
        # Por eso su tope tiene que ser generoso y tarda en avisar. Medir cuándo
        # avanzó las separa: una fuente sana con 96 días de rezago estructural
        # avanza igual todos los meses, y si deja de hacerlo se ve en semanas
        # en vez de meses (ADR-0178).
        previas = {}
        if SALIDA.exists():
            try:
                previas = (json.loads(SALIDA.read_text(encoding="utf-8"))
                           .get("panel_anclas") or {})
            except Exception as e:
                print(f"  [WARN] no se pudo leer el panel_anclas previo: {e}")
        hoy_iso = date.today().isoformat()
        anclas = {}
        for nombre in pnl.FAMILIA:
            serie = panel.get(nombre)
            if not serie:
                anclas[nombre] = None
                continue
            ultimo = max(serie)
            prev = previas.get(nombre) or {}
            # Si no hay registro previo, el primer avance se fecha HOY: no se
            # puede saber cuándo avanzó de verdad y suponerlo viejo daría una
            # falla inventada en la primera corrida.
            avanzo = hoy_iso if prev.get("ultimo") != ultimo else (prev.get("avanzo") or hoy_iso)
            anclas[nombre] = {"ultimo": ultimo, "n": len(serie), "avanzo": avanzo}
        resultados["panel_anclas"] = anclas
        vacias = [k for k, v in resultados["panel_anclas"].items() if v is None]
        if vacias:
            print(f"  [WARN] anclas del panel sin datos: {', '.join(sorted(vacias))}")

        indices = {"itvc": itvc_full, "itcg": serie_itcg, "itcp": serie_itcp}
        resultados["panel_validacion"] = {}
        print("")
        print("panel de validación socioeconómica:")
        for sig, serie in indices.items():
            if not serie:
                continue
            perf = pnl.perfil(sig, serie, panel)
            perf["lectura"] = pnl.lectura(perf)
            resultados["panel_validacion"][sig] = perf
            n, d = perf["niveles"], perf["diferencias"]
            print(f"  {sig.upper()}: propias {perf['n_propias']} · ajenas {perf['n_ajenas']}"
                  f" · niveles {n['convergente']}/{n['discriminante']} (brecha {n['brecha']})"
                  f" · difs {d['convergente']}/{d['discriminante']} (brecha {d['brecha']})")
            f = perf.get("factor")
            if f:
                print(f"    factor común ({f['n_series']} series, explica "
                      f"{f['varianza_explicada']}%): niveles {f['r_niveles']}"
                      f" · difs {f['r_diferencias']}"
                      f" · mejor sola {f['mejor_sola_niveles']}/{f['mejor_sola_diferencias']}"
                      f" · cargas {f['cargas']}")
    except Exception as e:
        print(f"[WARN] panel de validación no disponible: {e}")

    SALIDA.write_text(json.dumps(resultados, indent=2, ensure_ascii=False), encoding="utf-8")
    print(f"\n[OK] {SALIDA}")


if __name__ == "__main__":
    main()
