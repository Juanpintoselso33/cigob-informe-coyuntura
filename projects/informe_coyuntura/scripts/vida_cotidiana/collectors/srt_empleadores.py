"""Empleadores PyME activos — Superintendencia de Riesgos del Trabajo (SRT).

Mide el **cierre neto de empresas**: cuántas unidades productivas de hasta 50
trabajadores siguen teniendo al menos una persona declarada con cobertura de
ART. Cuando una PyME cierra, quiebra o despide a toda su nómina, el contrato
con la ART se rescinde casi en el acto, así que la baja aparece en el mes.

Por qué la SRT y no otra fuente (ADR-0218): la base de empleadores de OEDE
—que sería la equivalente por el lado de AFIP— dejó de actualizarse en octubre
de 2023, o sea justo antes del mandato que el informe evalúa. La SRT publica su
serie histórica todos los meses y llega a mayo de 2026.

## El archivo

`Serie_historica_Segun_Tamaño_de_la_nomina_del_empleador - UP.xlsx`, hoja
**"Cuadro 4.2: Parte empleadora"**: una fila por tramo de nómina y una columna
por mes, desde julio de 1996. Los tramos son 1 · 2 · 3 a 5 · 6 a 10 · 11 a 25 ·
26 a 40 · 41 a 50 · 51 a 100 · 101 a 500 · 501 a 1500 · 1501 a 2500 · 2501 a
5000 · Más de 5000.

El parser se ancla en el TEXTO de cada tramo y en las celdas de fecha de la
cabecera, nunca en posiciones: si la SRT agrega un tramo o mueve una fila, los
que busca siguen encontrándose, y si desaparece alguno el colector falla en voz
alta en vez de devolver una suma incompleta.
"""
import datetime
import io
import logging

import openpyxl
import requests

logger = logging.getLogger(__name__)

SRT_XLSX_URL = (
    "https://www.srt.gob.ar/estadisticas/series/co/up/"
    "Serie_historica_Segun_Tama%C3%B1o_de_la_nomina_del_empleador%20-%20UP.xlsx"
)
HTTP_HEADERS = {"User-Agent": "Mozilla/5.0 (compatible; CIGOB-Monitor/1.0)"}
HTTP_TIMEOUT = 180
HOJA = "Cuadro 4.2"

# El recorte PyME del documento de fuentes: hasta 50 trabajadores. Es el 95,6%
# de los empleadores del sistema, así que la serie del tramo describe
# prácticamente al universo — pero se suma explícitamente y no se toma el total,
# para que el indicador diga lo que su nombre promete y no dependa de que la
# proporción se mantenga.
TRAMOS_PYME = ("1", "2", "3 a 5", "6 a 10", "11 a 25", "26 a 40", "41 a 50")
# Contraste declarado: las grandes, para poder decir si el fenómeno es de las
# PyMEs o de toda la economía.
TRAMOS_GRANDES = ("501 a 1500", "1501 a 2500", "2501 a 5000", "Más de 5000")


def _serie_por_tramo(contenido: bytes) -> dict:
    """{tramo: {'YYYY-MM': cantidad}} desde el Cuadro 4.2."""
    wb = openpyxl.load_workbook(io.BytesIO(contenido), read_only=True, data_only=True)
    if HOJA not in wb.sheetnames:
        raise ValueError(f"el XLSX de la SRT ya no trae la hoja «{HOJA}»: {wb.sheetnames}")
    filas = list(wb[HOJA].iter_rows(values_only=True))

    columnas = None
    for fila in filas[:12]:
        fechas = [(i, c) for i, c in enumerate(fila)
                  if isinstance(c, (datetime.datetime, datetime.date))]
        if len(fechas) > 100:          # la cabecera real trae ~360 meses
            columnas = fechas
            break
    if not columnas:
        raise ValueError("no se encontró la fila de meses en el Cuadro 4.2")

    out = {}
    for fila in filas:
        etiqueta = str(fila[0]).strip() if fila[0] is not None else ""
        if etiqueta not in TRAMOS_PYME + TRAMOS_GRANDES:
            continue
        out[etiqueta] = {
            f"{f.year}-{f.month:02d}": fila[i]
            for i, f in columnas
            if i < len(fila) and isinstance(fila[i], (int, float))
        }
    faltan = set(TRAMOS_PYME + TRAMOS_GRANDES) - set(out)
    if faltan:
        raise ValueError(f"el Cuadro 4.2 ya no trae los tramos {sorted(faltan)}")
    return out


def _sumar(por_tramo: dict, tramos) -> dict:
    meses = set.intersection(*[set(por_tramo[t]) for t in tramos])
    return {m: sum(por_tramo[t][m] for t in tramos) for m in meses}


def parsear_empleadores(contenido: bytes) -> dict:
    por_tramo = _serie_por_tramo(contenido)
    pyme = _sumar(por_tramo, TRAMOS_PYME)
    grandes = _sumar(por_tramo, TRAMOS_GRANDES)
    if len(pyme) < 200:
        raise ValueError(f"la serie de la SRT trae sólo {len(pyme)} meses")
    ultimo = max(pyme)
    return {
        "mes": ultimo,
        "pyme": pyme[ultimo],
        "grandes": grandes.get(ultimo),
        "serie_pyme": pyme,
        "serie_grandes": grandes,
    }


def fetch_empleadores_pyme() -> dict:
    r = requests.get(SRT_XLSX_URL, headers=HTTP_HEADERS, timeout=HTTP_TIMEOUT)
    r.raise_for_status()
    d = parsear_empleadores(r.content)
    logger.info("empleadores PyME OK: %s → %s", d["mes"], d["pyme"])
    return d
