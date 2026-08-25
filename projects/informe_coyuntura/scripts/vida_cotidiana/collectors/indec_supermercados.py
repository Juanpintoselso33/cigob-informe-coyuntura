"""Ventas en supermercados a precios constantes — INDEC, serie desestacionalizada.

Componente del ITCIS desde ADR-0225. **Antes era el ancla de validación
externa** del cinturón (ADR-0155), y el movimiento es exactamente el que sacó
al ICC en su momento: una serie que mide condiciones materiales del hogar es
un componente del índice, no un juez del índice.

## Qué mide, y por qué es distinto de todo lo demás que hay adentro

Es el único componente que mide **volumen efectivamente comprado**. Los otros
diecisiete miden ingreso (cuánto entra), precio (cuánto cuesta), empleo (de
dónde viene el ingreso), mora (qué no se paga), percepción (qué se opina) o
victimización. Ninguno mira lo que el hogar se llevó de la góndola.

Medido antes de incorporarlo: el 43% de su nivel y el **82% de su movimiento
mes a mes** no los puede reproducir el ITCIS con sus seis dimensiones juntas.

## La serie sale del INDEC, no de su espejo (ADR-0256)

`Cuadro 1` de `serie_supermercados.xlsx`, columna **serie desestacionalizada**:
el mismo número que publica el INDEC en su informe técnico. Se toma la
desestacionalizada de la fuente y no se suaviza acá — la regla que dejó
ADR-0155 después de fabricarse una divergencia con una media móvil propia, que
atrasa la serie medio año y da vuelta el signo.

Hasta ADR-0256 la serie venía de la API de datos.gob.ar, que es un **espejo**
de esta misma planilla y le llega con unas dos semanas de atraso. Eso encadenaba
dos rezagos y dejaba la card un mes entera por detrás de lo publicado: el 25 de
agosto de 2026 el INDEC ya había sacado junio (el 21) y la web seguía mostrando
mayo. El espejo se quedó como **contraste**, no como fuente.

Lo que NO cubre, y va declarado en la ficha: comercio registrado de cadenas de
supermercados. No ve el comercio informal ni el almacén de barrio.

## El rezago, medido y no copiado

Medido sobre las **14 publicaciones** del calendario del INDEC entre julio de
2025 y agosto de 2026: el mes M sale entre 48 y 57 días después de terminado
(mediana 53), con intervalos de 23 a 34 días entre una publicación y la
siguiente. Como `fecha_dato` es el día 1 del mes de referencia, el último punto
disponible nace con 78-86 días y llega como mucho a **116** la víspera de la
publicación siguiente. Por eso el tope del gate es 130 y no el default de 110
(ver `gate_calidad.MAX_DIAS`): deja margen sobre ese techo medido y sigue
avisando si se saltea un mes entero, que llevaría el rezago a ~146.

## La trampa del sitio del INDEC

`indec.gob.ar` contesta **200 con la misma cáscara HTML de 37 KB** para
cualquier ruta que no exista. Un `raise_for_status()` no distingue el archivo
de la nada, así que el contenido se verifica por su firma antes de abrirlo: si
no arranca con la firma de un zip, no es un xlsx y el colector levanta.
"""
import datetime as dt
import io
import logging
import re
import statistics

import requests

logger = logging.getLogger(__name__)

# Constantes propias y no importadas de `config`, por el mismo motivo que en
# `dnrpa_autos` y `trabajo_independiente`: este módulo se carga POR RUTA desde
# los tests y desde `descargar_series.py`, y ahí `import config` resuelve al
# `config.py` de la raíz del proyecto, que es otro archivo.
HTTP_HEADERS = {
    "User-Agent": ("Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) "
                   "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124 Safari/537.36"),
}
HTTP_TIMEOUT = 60

# La URL es estable y sin hash, a diferencia del informe técnico en PDF
# (`/uploads/informesdeprensa/super_08_262444C24851.pdf`). Se la descubre desde
# `/Nivel4/Tema/3/1/34`, la vista parcial que arma la página de Supermercados.
FUENTE_XLSX = "https://www.indec.gob.ar/ftp/cuadros/economia/serie_supermercados.xlsx"
HOJA = "Cuadro 1"
FILA_ENCABEZADO = 3          # 'Período' | 'Serie original' | 'Serie desestacionalizada (1)'
FILA_PRIMER_DATO = 7

# El espejo, que ya no manda pero sigue mirando.
ESPEJO_API = "https://apis.datos.gob.ar/series/api/series/"
ESPEJO_ID = "455.1_VENTAS_PREADA_0_M_44_44"

# Los tres meses de la base del cinturón. Si la serie no los trae, el rebase a
# 4T-2023 no se puede hacer y el componente quedaría anclado contra otra cosa
# sin que nada avise: es la falla que este colector tiene que hacer ruidosa.
BASE_MESES = ("2023-10", "2023-11", "2023-12")

# La base del índice se LEE de la fuente, no se escribe acá. La card la
# rotulaba «2004 = 100» y la Encuesta de Supermercados vigente usa **base
# 2017=100**: la serie ni siquiera empieza antes de enero de 2017 (ADR-0243).
# El título del Cuadro 1 la declara igual que lo hacía `units` en la API.
_RE_BASE = re.compile(r"base\s*(?:a[ñn]o\s*)?(\d{4})\s*=\s*100", re.IGNORECASE)

# El mismo título declara la cobertura — «Enero 2017 – junio 2026» —, o sea el
# último mes que la planilla dice traer. Es una segunda afirmación de la fuente,
# independiente de las filas, y sirve para agarrar una lectura truncada: si el
# parser se comiera las últimas filas, el número seguiría siendo plausible.
_MESES = {m: i for i, m in enumerate(
    ("enero", "febrero", "marzo", "abril", "mayo", "junio", "julio", "agosto",
     "septiembre", "octubre", "noviembre", "diciembre"), 1)}
_RE_COBERTURA = re.compile(
    r"([A-Za-zÁÉÍÓÚáéíóú]+)\s+(\d{4})\s*[–\-—]\s*([A-Za-zÁÉÍÓÚáéíóú]+)\s+(\d{4})")

# Umbral del contraste contra el espejo, sobre variaciones mes a mes en puntos
# porcentuales. Calibrado con la superposición real del 25-ago-2026 (113 meses):
# la columna correcta da 0,041 pp de mediana; «serie original» da 4,92 y
# «tendencia-ciclo» 0,84. 0,30 deja veinte veces de margen contra el error que
# el contraste existe para agarrar —leer la columna de al lado— sin saltar por
# las revisiones del ajuste estacional, que son lo que la diferencia mide
# cuando todo está bien.
CONTRASTE_MAX_PP = 0.30


class FuenteINDECError(RuntimeError):
    """El INDEC no entregó una planilla utilizable.

    Tiene clase propia para que el clasificador de `aviso_slack.py` la lea como
    fuente caída y no como error de código: son dos avisos distintos y ADR-0175
    costó cuatro días de serie congelada por confundirlos."""


def base_declarada(titulo: str) -> int:
    """Año base del índice, leído del título del cuadro.

    Falla si el título no declara una base reconocible. Es a propósito: el
    rótulo anterior —«2004 = 100»— no vino de ningún lado verificable y
    sobrevivió porque nadie tenía contra qué compararlo."""
    m = _RE_BASE.search(titulo or "")
    if not m:
        raise ValueError(
            f"{HOJA}: el título no declara la base del índice; "
            "sin eso el rótulo de la card sería una suposición")
    return int(m.group(1))


def cobertura_declarada(titulo: str) -> str:
    """Último mes que el título dice cubrir, como 'YYYY-MM'."""
    m = _RE_COBERTURA.search(titulo or "")
    if not m:
        raise ValueError(
            f"{HOJA}: el título no declara el período que cubre; sin eso no hay "
            "contra qué chequear que la lectura llegó hasta el final")
    mes = _MESES.get(m.group(3).lower())
    if not mes:
        raise ValueError(f"{HOJA}: mes final no reconocido en «{m.group(0)}»")
    return f"{int(m.group(4)):04d}-{mes:02d}"


def _descargar_xlsx() -> bytes:
    """La planilla del INDEC, verificada por su firma y no por el código HTTP.

    `indec.gob.ar` devuelve 200 y una cáscara HTML para cualquier ruta que no
    exista, así que un 200 no prueba nada."""
    try:
        r = requests.get(FUENTE_XLSX, headers=HTTP_HEADERS, timeout=HTTP_TIMEOUT)
        r.raise_for_status()
    except requests.RequestException as e:
        raise FuenteINDECError(f"{FUENTE_XLSX}: no se pudo descargar ({e})") from e
    contenido = r.content
    if not contenido[:4] == b"PK\x03\x04":
        raise FuenteINDECError(
            f"{FUENTE_XLSX}: la respuesta no es un xlsx ({len(contenido)} bytes, "
            f"empieza con {contenido[:16]!r}). El sitio del INDEC contesta 200 con "
            "una cáscara HTML cuando la ruta no existe: probablemente movieron el "
            "archivo. Se descubre de nuevo en https://www.indec.gob.ar/Nivel4/Tema/3/1/34")
    return contenido


def _columna_desestacionalizada(ws) -> int:
    """La columna se busca por su encabezado, no por su posición.

    El Cuadro 1 trae tres series pegadas —original, desestacionalizada y
    tendencia-ciclo— con encabezados combinados. Un índice fijo convierte una
    columna agregada río arriba en un cambio de indicador silencioso: el número
    seguiría siendo un índice de ventas plausible."""
    for c in range(1, ws.max_column + 1):
        v = ws.cell(FILA_ENCABEZADO, c).value
        if isinstance(v, str) and v.strip().lower().startswith("serie desestacionalizada"):
            return c
    raise ValueError(
        f"{HOJA}: no hay ninguna columna «Serie desestacionalizada» en la fila "
        f"{FILA_ENCABEZADO}; el cuadro cambió de forma")


def _serie_del_libro(contenido: bytes) -> tuple[dict, int, str]:
    """(serie {'YYYY-MM': valor}, año base, último mes declarado en el título)."""
    import openpyxl
    try:
        wb = openpyxl.load_workbook(io.BytesIO(contenido), data_only=True)
    except Exception as e:
        raise FuenteINDECError(f"{FUENTE_XLSX}: no se pudo abrir la planilla ({e})") from e
    if HOJA not in wb.sheetnames:
        raise ValueError(f"la planilla no trae la hoja «{HOJA}» (tiene {wb.sheetnames})")
    ws = wb[HOJA]
    titulo = ws.cell(1, 1).value
    col = _columna_desestacionalizada(ws)
    serie = {}
    for fila in ws.iter_rows(min_row=FILA_PRIMER_DATO, values_only=True):
        periodo, valor = fila[0], fila[col - 1]
        if not isinstance(periodo, dt.datetime) or valor is None:
            continue
        try:
            serie[periodo.strftime("%Y-%m")] = float(valor)
        except (TypeError, ValueError):
            continue
    return serie, base_declarada(titulo), cobertura_declarada(titulo)


def _serie_del_espejo() -> dict | None:
    """La misma serie en la API de datos.gob.ar, o `None` si no contesta.

    Es un contraste, no una fuente: nunca levanta. Que el espejo esté caído no
    es motivo para no publicar lo que el INDEC ya publicó."""
    try:
        r = requests.get(ESPEJO_API,
                         params={"ids": ESPEJO_ID, "format": "json", "limit": 5000},
                         headers=HTTP_HEADERS, timeout=HTTP_TIMEOUT)
        r.raise_for_status()
        filas = r.json().get("data") or []
        serie = {f[0][:7]: float(f[1]) for f in filas if f[1] is not None}
        return serie or None
    except Exception as e:                                    # noqa: BLE001
        logger.info("espejo de datos.gob.ar no disponible para el contraste (%s)", e)
        return None


def _contrastar(serie: dict, espejo: dict) -> float | None:
    """Divergencia mediana en variación m/m, en pp. Levanta si se pasa del tope.

    Compara **variaciones**, no niveles: si el INDEC rebasea el índice, los
    niveles se corren enteros y las variaciones no se mueven. Lo que este
    contraste tiene que agarrar es haber leído la columna equivocada, no un
    cambio de base — que el rebase a 4T-2023 absorbe solo."""
    def mom(s):
        k = sorted(s)
        return {k[i]: s[k[i]] / s[k[i - 1]] - 1
                for i in range(1, len(k)) if s[k[i - 1]]}

    a, b = mom(serie), mom(espejo)
    comunes = sorted(set(a) & set(b))
    if len(comunes) < 24:
        logger.info("contraste omitido: sólo %d meses en común con el espejo",
                    len(comunes))
        return None
    dif = statistics.median(abs(a[m] - b[m]) * 100 for m in comunes)
    if dif > CONTRASTE_MAX_PP:
        raise ValueError(
            f"{HOJA}: la serie leída diverge del espejo de datos.gob.ar en "
            f"{dif:.3f} pp de variación mensual mediana sobre {len(comunes)} meses "
            f"(tope {CONTRASTE_MAX_PP}). Las dos publican la MISMA serie: esto no "
            "es una revisión, es una columna distinta")
    return dif


def fetch_consumo_supermercados() -> dict:
    """{'consumo_supermercados': {'valor', 'fecha', 'unidad'}, 'serie': {...}}.

    La card es el último punto de la MISMA serie que alimenta el índice: no hay
    dos caminos que puedan divergir. `descargar_series.py` reusa esta función.

    La unidad también sale de acá (ADR-0243), leída del título del cuadro:
    card, serie y web dejan de tener cada una su copia del rótulo.
    """
    serie, anio_base, ult_declarado = _serie_del_libro(_descargar_xlsx())
    if not serie:
        raise ValueError(f"{HOJA}: la planilla no devolvió ningún punto")

    faltan = [m for m in BASE_MESES if m not in serie]
    if faltan:
        raise ValueError(
            f"{HOJA}: faltan meses de la base 4T-2023 ({', '.join(faltan)}); "
            "sin ellos el rebase del componente mediría contra otra base")

    primero = min(serie)
    if int(primero[:4]) < anio_base:
        raise ValueError(
            f"{HOJA}: la serie arranca en {primero} pero la base declarada "
            f"es {anio_base}=100; una de las dos cosas cambió")

    ult = max(serie)
    if ult != ult_declarado:
        raise ValueError(
            f"{HOJA}: el título declara cobertura hasta {ult_declarado} y la "
            f"lectura llegó a {ult}; la planilla y su encabezado no coinciden")

    espejo = _serie_del_espejo()
    dif = _contrastar(serie, espejo) if espejo else None

    unidad = f"índice ({anio_base} = 100, desestacionalizado)"
    logger.info("consumo_supermercados OK: %s = %.1f (%d meses, %s%s)",
                ult, serie[ult], len(serie), unidad,
                f", contraste {dif:.3f} pp" if dif is not None else "")
    return {
        "consumo_supermercados": {"valor": round(serie[ult], 1), "fecha": ult,
                                  "unidad": unidad, "anio_base": anio_base},
        "serie": serie,
    }
