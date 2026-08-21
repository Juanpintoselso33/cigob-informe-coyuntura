"""
descargar_series.py — Descarga series históricas por cinturón a CSV
Salida: output/series/macro.csv | politica.csv | vida_cotidiana.csv | gestion.csv
Columnas: fecha, indicador, valor, fuente
"""
import sys
import csv
import io
import json
import re
import time
import signal
import calendar
import functools
import requests
import urllib3
from contextlib import contextmanager
from datetime import datetime, timedelta, date
from pathlib import Path

sys.path.insert(0, str(Path(__file__).parent))
import comarb    # recaudación provincial (Convenio Multilateral) + el cálculo de la base imponible
import macro     # reutiliza el parser SDDS y las constantes del balance (reservas netas)
import gestion   # reutiliza el lector del sheet oficial del RIGI + fechas del BO
import politica  # reutiliza la reconstrucción histórica del Votómetro

sys.stdout.reconfigure(encoding="utf-8")
urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)

OUTPUT_DIR = Path("output/series")
OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

HTTP_TIMEOUT = 20
HTTP_HEADERS = {"User-Agent": "CIGOB-InformeCoyuntura/1.0"}
INDEC_BASE   = "https://apis.datos.gob.ar/series/api/series/"
BCRA_BASE    = "https://api.bcra.gob.ar/estadisticas/v4.0/Monetarias"
INDICADORES_SUSTITUIDOS = {
    "desequilibrio_monetario": {"presion_dolarizacion", "dolarizacion_depositos"},
}

# ── Presupuesto de tiempo por indicador (ADR-0173) ────────────────────────────
# El pipeline no tenía ninguna defensa contra una fuente LENTA. Las que fallan
# están cubiertas hace rato (exit codes como dato, ADR-0133, carry-forward);
# una que tarda se comía el presupuesto entero del job y dejaba sin correr al
# gate, a los tests y al commit. El 5-ago-2026 dos corridas murieron así, con
# Trends haciendo backoff: 30 y 45 minutos, cero publicado.
#
# El presupuesto va acá y no en el workflow a propósito. `write_csv` escribe UN
# CSV POR CINTURÓN a medida que los termina, así que matar el script desde
# afuera deja unos cinturones frescos y otros de ayer — cards nuevas contra
# series viejas, que es el falso G3 documentado en CLAUDE.md (2026-07-09). Por
# indicador, en cambio, el corte cae dentro del try/except que ya existe y sólo
# ese indicador conserva sus filas previas.
PRESUPUESTO_INDICADOR_DEFAULT = 300      # 5 min: holgado para una fuente sana
PRESUPUESTO_INDICADOR = {
    # Caminan actas de a una desde el id más reciente: son lentas por diseño,
    # no por estar colgadas.
    "cohesion_bloque": 900,
    "alineamiento_senadores_prov": 900,
    "veto_quorum": 600,
    "comisiones_caidas": 600,
    # Arma la canasta de Trends por tandas, que es justo lo que rate-limitea.
    "sentimiento_digital": 600,
}


class TiempoAgotado(Exception):
    """Un indicador agotó su presupuesto. Es una condición de FUENTE (lenta),
    no un error del script: se trata igual que una fuente caída."""


@contextmanager
def presupuesto(segundos: int):
    """Corta el bloque a los `segundos`. Usa SIGALRM, que interrumpe el código
    bloqueado de verdad; un timeout por hilo no puede matar al worker y lo
    dejaría corriendo, que es exactamente lo que se quiere evitar.

    Sin SIGALRM (Windows, donde se desarrolla) no hay presupuesto y el bloque
    corre entero: el pipeline corre en ubuntu-latest, que es donde importa."""
    if segundos <= 0 or not hasattr(signal, "SIGALRM"):
        yield
        return

    def _vencido(_signum, _frame):
        raise TiempoAgotado(f"presupuesto de {segundos}s agotado")

    previo = signal.signal(signal.SIGALRM, _vencido)
    signal.alarm(segundos)
    try:
        yield
    finally:
        signal.alarm(0)
        signal.signal(signal.SIGALRM, previo)


def _filas_previas(cinturon: str, indicadores: set) -> list:
    """Filas que el CSV del cinturón ya tenía para `indicadores`.

    Sin esto, un fetcher caído no aporta filas y la escritura completa
    (merge=False) borra esa serie del CSV entero: el indicador desaparece del
    gráfico en silencio en vez de conservar su último valor bueno. Es la misma
    forma del bug de sentimiento_digital que documenta CLAUDE.md, del otro lado
    del pipeline."""
    path = OUTPUT_DIR / f"{cinturon}.csv"
    if not path.exists() or not indicadores:
        return []
    with open(path, newline="", encoding="utf-8-sig") as f:
        return [r for r in list(csv.reader(f))[1:] if len(r) > 1 and r[1] in indicadores]


def fetch_indec(series_id: str, limit: int = 48) -> list:
    r = requests.get(INDEC_BASE, params={"ids": series_id, "format": "json",
                     "limit": limit, "sort": "desc"},
                     headers=HTTP_HEADERS, timeout=HTTP_TIMEOUT)
    r.raise_for_status()
    return [[row[0], row[1]] for row in r.json()["data"] if row[1] is not None]


def fetch_indec_var_mensual(series_id: str, limit: int = 48) -> list:
    """Serie histórica de la variación m/m % de una serie INDEC de índices:
    (idx_t / idx_{t-1} − 1)×100. Misma fórmula que indec_series._var_mensual, para
    todo el histórico. Devuelve [[YYYY-MM-DD, var_pct]] ascendente."""
    data = sorted(fetch_indec(series_id, limit), key=lambda x: x[0])
    return [[f1, round((v1 / v0 - 1) * 100, 2)]
            for (f0, v0), (f1, v1) in zip(data, data[1:]) if v0]


def fetch_indec_x100(series_id: str, limit: int = 40) -> list:
    """Serie INDEC multiplicada por 100 (tasas EPH que la API devuelve en proporción,
    p. ej. 0,368 → 36,8 %). Misma transformación que la card. [[fecha, %]] ascendente."""
    return [[f, round(v * 100, 1)]
            for f, v in sorted(fetch_indec(series_id, limit), key=lambda x: x[0])]


def _meses_desde_asuncion() -> int:
    """Meses entre dic-2023 (asunción del gobierno) y hoy: la ventana por
    defecto de las series macro (regla de backfill ADR-0012 — toda serie debe
    cubrir el mandato completo, no una ventana fija que envejece)."""
    hoy = date.today()
    return (hoy.year - 2023) * 12 + hoy.month - 11


def fetch_bcra(var_id: int, dias: int = 960) -> list:
    desde = (datetime.today() - timedelta(days=dias)).strftime("%Y-%m-%d")
    r = requests.get(f"{BCRA_BASE}/{var_id}", params={"desde": desde},
                     headers=HTTP_HEADERS, timeout=HTTP_TIMEOUT, verify=False)
    r.raise_for_status()
    detalle = r.json()["results"][0]["detalle"]
    return sorted([[d["fecha"], d["valor"]] for d in detalle],
                  key=lambda x: x[0], reverse=True)


def write_csv(
    nombre: str,
    rows: list,
    merge: bool = False,
    eliminar_indicadores: set[str] | None = None,
):
    """Escribe un cinturón y permite purgar claves sustituidas durante el merge."""
    path = OUTPUT_DIR / f"{nombre}.csv"
    if merge and path.exists():
        tocados = {r[1] for r in rows} | (eliminar_indicadores or set())
        with open(path, newline="", encoding="utf-8-sig") as f:
            existentes = list(csv.reader(f))[1:]   # sin encabezado
        rows = [r for r in existentes if r[1] not in tocados] + rows
        rows.sort(key=lambda x: (x[1], x[0]), reverse=True)
    with open(path, "w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        writer.writerow(["fecha", "indicador", "valor", "unidad", "fuente"])
        writer.writerows(rows)
    print(f"[OK] {path}  ({len(rows)} filas)")


def fetch_saldo_ica(limit: int = 48) -> list:
    """Saldo comercial mensual derivado de las series ICA (expo − impo).
    La serie de saldo directa (164.3_SOTALTAL_0_0_8) corre con ~14 meses de
    rezago; las ICA están frescas a ~2 meses (mismo criterio que macro.py)."""
    expo = fetch_indec("74.3_IET_0_M_16", limit)
    impo_por_fecha = dict(fetch_indec("74.3_IIT_0_M_25", limit))
    return [[fecha, round(valor - impo_por_fecha[fecha], 1)]
            for fecha, valor in expo if fecha in impo_por_fecha]


def fetch_saldo_12m_serie(limit: int = 60) -> list:
    """Saldo comercial ACUMULADO MÓVIL de 12 meses — la métrica del titular
    del indicador (barrido 5/12: la serie mensual bajo un titular acumulado
    confundía la lectura). La mensual sigue publicándose como saldo_comercial,
    insumo de la validación externa del ITCM. [[YYYY-MM-01, M USD]]."""
    mensual = dict(sorted(fetch_saldo_ica(limit)))
    fechas = list(mensual)
    out = []
    for i in range(11, len(fechas)):
        win = fechas[i - 11:i + 1]
        a0, m0 = int(win[0][:4]), int(win[0][5:7])
        af, mf = int(win[-1][:4]), int(win[-1][5:7])
        if (af * 12 + mf) - (a0 * 12 + m0) == 11:      # ventana sin huecos
            out.append([fechas[i], round(sum(mensual[f] for f in win), 1)])
    return out


def _tesoro_por_mes() -> dict:
    """Mapa {YYYY-MM: dep. Tesoro en USD (M)} del Balance Consolidado del BCRA,
    para componer la serie histórica de reservas netas."""
    import xlrd
    r = requests.get(macro.BCRA_BALANCE_URL, headers=HTTP_HEADERS, timeout=60, verify=False)
    sh = xlrd.open_workbook(file_contents=r.content).sheet_by_name("B.C.R.A.")
    out = {}
    for row in range(27, sh.nrows):
        ym = sh.cell_value(row, 0)
        dep = sh.cell_value(row, macro.BAL_COL_DEP_GOB_ME)
        tc  = sh.cell_value(row, macro.BAL_COL_TC)
        if all(isinstance(x, (int, float)) for x in (ym, dep, tc)) and tc > 0:
            anio = int(ym); mes = round((ym - anio) * 100)
            if 1 <= mes <= 12:
                out[f"{anio}-{mes:02d}"] = dep / tc / 1000.0
    return out


def fetch_reservas_netas_serie(meses: int | None = None) -> list:
    """Serie mensual de reservas NETAS "a secas" = SDDS estricto + dep. Tesoro +
    Bopreal 12m (la misma fórmula que el indicador en macro.py), parseando las
    últimas `meses` planillas SDDS (por defecto, desde dic-2023). LÍMITE DE LA
    FUENTE: el BCRA borra las planillas viejas (404 antes de jun-2024) y no
    están en Wayback (verificado jul-2026) → la serie arranca en jun-2024;
    los meses sin planilla parseable se saltean."""
    meses = meses or _meses_desde_asuncion()
    tesoro = _tesoro_por_mes()
    out = []
    y, m = datetime.today().year, datetime.today().month
    for _ in range(meses):
        nombre = f"temp{m:02d}{y % 100:02d}"
        try:
            r = requests.get(macro.SDDS_URL_BASE.format(nombre), headers=HTTP_HEADERS,
                             timeout=60, verify=False)
            if r.status_code == 200 and len(r.content) > 50000:
                s = macro._parse_sdds_content(r.content)
                if s and s["fecha"]:
                    _, mm, yy = s["fecha"].split("/")
                    ym = f"20{yy}-{mm}"
                    netas = s["netas"] + tesoro.get(ym, 0.0) + abs(s["bopreal_12m"])
                    out.append([f"20{yy}-{mm}-01", round(netas, 0)])
        except Exception:
            pass
        m -= 1
        if m == 0: m = 12; y -= 1
    return out


def fetch_tcrm_serie(meses: int | None = None) -> list:
    """Serie mensual del ITCRM oficial del BCRA (base 17-dic-2015=100), de la
    planilla ITCRMSerie.xlsx. Reemplaza la serie INDEC 116.3_TCRMA, discontinuada
    en dic-2024. Devuelve los últimos `meses` como [[YYYY-MM-01, valor]]."""
    meses = meses or _meses_desde_asuncion()
    serie = macro.fetch_itcrm_serie()  # [(YYYY-MM-01, valor)] ascendente
    return [[ym, val] for ym, val in serie[-meses:]]


def fetch_idm_serie(meses: int | None = None) -> list:
    """Serie mensual del Índice de Desequilibrio Monetario (brecha i.a. real entre
    M3 y M2 privado), con la misma fórmula que el indicador en macro.py. Devuelve
    los últimos `meses` como [[YYYY-MM-01, gap_pp]]."""
    meses = meses or _meses_desde_asuncion()
    serie = macro._idm_serie_mensual(meses_hist=meses + 4)  # [(YYYY-MM, gap, m3, m2)] asc.
    return [[f"{ym}-01", gap] for ym, gap, _m3, _m2 in serie[-meses:]]


def fetch_desequilibrio_monetario_serie(
    meses: int | None = None,
) -> list:
    """Serie de tensión 0–100 construida por el mismo helper que alimenta el titular.

    Arranca en abril de 2025 y no en diciembre de 2023 como el resto: el
    componente de fuga sólo tiene lectura interpretable desde la apertura del
    cepo a personas humanas (ver MES_INICIO en desequilibrio_monetario.py)."""
    meses = meses or _meses_desde_asuncion()
    serie = macro._desequilibrio_monetario_serie_mensual(
        meses_hist=meses + 4,
    )
    return [
        [f'{fila["mes"]}-01', fila["tension"]]
        for fila in serie
    ][-meses:]


def fetch_idc_serie(meses: int | None = None) -> list:
    """Serie mensual del Índice de Capacidad Prestable, con la misma fórmula que el
    indicador en macro.py (_idc_componentes) sobre stocks de fin de mes. Devuelve
    los últimos `meses` como [[YYYY-MM-01, idc]]."""
    meses = meses or _meses_desde_asuncion()
    serie = macro._idc_serie_mensual(meses_hist=meses + 2)  # [(YYYY-MM, idc)] asc.
    return [[f"{ym}-01", idc] for ym, idc in serie[-meses:]]


def fetch_iai_serie(meses: int | None = None) -> list:
    """Serie histórica del IAI (inversión física: ISAC + BK importados, 65/35),
    con la misma fórmula que macro.py. [[YYYY-MM-01, valor]]."""
    meses = meses or _meses_desde_asuncion()
    return [[f"{ym}-01", v] for ym, v in macro._iai_serie_mensual(meses=meses)]


def fetch_icip_serie(meses: int | None = None) -> list:
    """Serie histórica del ICIP (inversión digital: servicios tech + productividad),
    con la misma fórmula que macro.py. [[YYYY-MM-01, valor]]."""
    meses = meses or _meses_desde_asuncion()
    return [[f"{ym}-01", v] for ym, v in macro._icip_serie_mensual(meses=meses)]


def descargar(cinturon: str, indec_series: list, bcra_vars: list, derivadas: list = (), solo_indicador: str | None = None):
    """Arma el CSV de un cinturón. Con `solo_indicador`, corre SOLO ese
    indicador (buscado por nombre en las tres listas) y hace MERGE con el
    CSV existente (preserva las filas de los indicadores no tocados) en vez
    de sobreescribir el cinturón entero -- así se puede refrescar un
    indicador puntual sin arrastrar al resto de su cinturón a quedar
    desincronizado card↔serie por una corrida parcial (ver CLAUDE.md,
    sección "Publishing data 'now'", y el hallazgo de gate_calidad G3 del
    2026-07-09)."""
    rows = []
    fallidos = set()

    def _correr(nombre, unidad, fuente, fetch):
        """Corre un fetcher con su presupuesto y acumula filas. Devuelve True si
        aportó datos; si no, deja el nombre en `fallidos` para que sus filas
        previas se arrastren en vez de borrarse."""
        try:
            with presupuesto(PRESUPUESTO_INDICADOR.get(nombre, PRESUPUESTO_INDICADOR_DEFAULT)):
                data = fetch()
            for fecha, valor in data:
                rows.append([fecha, nombre, valor, unidad, fuente])
            print(f"  [OK] {nombre}: {len(data)} puntos  ({data[-1][0]} → {data[0][0]})")
            return True
        except TiempoAgotado as e:
            fallidos.add(nombre)
            print(f"  [LENTO] {nombre}: {e} -- se conservan las filas anteriores")
        except Exception as e:
            fallidos.add(nombre)
            print(f"  [ERR] {nombre}: {e} -- se conservan las filas anteriores")
        return False

    for nombre, unidad, fuente, fetch_fn in derivadas:
        if solo_indicador and nombre != solo_indicador:
            continue
        _correr(nombre, unidad, fuente, fetch_fn)

    for sid, nombre, unidad, fuente in indec_series:
        if solo_indicador and nombre != solo_indicador:
            continue
        _correr(nombre, unidad, fuente, lambda sid=sid: fetch_indec(sid))

    for var_id, nombre, unidad, fuente in bcra_vars:
        if solo_indicador and nombre != solo_indicador:
            continue
        _correr(nombre, unidad, fuente, lambda var_id=var_id: fetch_bcra(var_id))

    # Un indicador que no pudo bajarse conserva lo que ya tenía. Antes no
    # aportaba filas y la escritura completa lo borraba del CSV: la serie
    # desaparecía del gráfico sin que nada avisara.
    if fallidos:
        previas = _filas_previas(cinturon, fallidos)
        rows += previas
        conservados = sorted({r[1] for r in previas})
        perdidos = sorted(fallidos - set(conservados))
        if conservados:
            print(f"  [CARRY] {len(previas)} filas conservadas de: {', '.join(conservados)}")
        if perdidos:
            # Ojo con lo que afirma este aviso: publicar.py arma series.json
            # juntando TODOS los CSV de output/series, así que un indicador sin
            # filas acá puede tener su serie igual desde otro archivo. Decir
            # "queda sin serie" a secas manda a buscar un problema que puede no
            # existir — pasó el 5-ago-2026 con icg_utdt, que se publicaba desde
            # un CSV huérfano mientras su fetcher venía roto (ADR-0175).
            print(f"  [AVISO] {', '.join(perdidos)}: falló y {cinturon}.csv no tenía filas "
                  f"suyas -- no aporta serie por esta vía. Si tampoco viene de otro CSV, "
                  f"el indicador queda sin serie")

    rows.sort(key=lambda x: (x[1], x[0]), reverse=True)
    eliminados = (
        INDICADORES_SUSTITUIDOS.get(solo_indicador, set())
        if rows
        else set()
    )
    write_csv(
        cinturon,
        rows,
        merge=solo_indicador is not None,
        eliminar_indicadores=eliminados,
    )


# ── Definición de series por cinturón ─────────────────────────────────────────

# Las tres series INDEC de macro se publican TRANSFORMADAS a la métrica del
# titular (barrido macro 04-jul-2026: la bajada cruda mostraba el NIVEL del
# IPC bajo un titular en % m/m, el EMAE en fracción y la recaudación nominal
# bajo un titular real — mismatch valor/serie en el modal).
MACRO_INDEC = []

IPC_NIVEL_ID = "148.3_INIVELNAL_DICI_M_26"


def _mes_previo(ym: str, n: int = 1) -> str:
    """'2026-03' → '2026-02'. Aritmética de CALENDARIO, no de posición."""
    anio, mes = int(ym[:4]), int(ym[5:7])
    total = anio * 12 + (mes - 1) - n
    return f"{total // 12}-{total % 12 + 1:02d}"


def _var_mensual(niveles: dict) -> dict:
    """{YYYY-MM: % m/m} exigiendo que los dos meses sean CONSECUTIVOS.

    La versión anterior dividía posiciones adyacentes de la lista ordenada. Con
    la serie completa da lo mismo, pero si la fuente saltea un mes —o si se
    acota la ventana descargada— el cociente entre dos observaciones separadas
    por dos meses se publicaba como variación mensual, sin que nada avisara.
    Detectado por auditoría de código (18-jul-2026)."""
    return {
        ym: round((v / niveles[_mes_previo(ym)] - 1) * 100, 2)
        for ym, v in niveles.items()
        if niveles.get(_mes_previo(ym))
    }


def fetch_ipc_mm_serie() -> list:
    """Inflación mensual (% m/m) derivada del nivel del IPC nacional — la
    métrica del titular. La curva de desinflación completa. [[YYYY-MM-01, %]]."""
    niveles = {f[:7]: v for f, v in fetch_indec(IPC_NIVEL_ID, limit=60) if v}
    return [[f"{ym}-01", v] for ym, v in sorted(_var_mensual(niveles).items())]


def fetch_emae_ia_serie() -> list:
    """EMAE variación i.a. en % (la serie INDEC viene en fracción). [[YYYY-MM-01, %]]."""
    return [[f, round(v * 100, 2)] for f, v in sorted(fetch_indec("143.3_ICE_SERVIA_2004_A_25", limit=60))]


def fetch_emae_difusion_serie() -> list:
    """Difusión sectorial del EMAE: % de los 15 sectores que crecen i.a.

    Misma construcción que puntúa en la ficha (macro._emae_difusion_por_mes),
    reconstruida desde dic-2023. La serie completa arranca en 2005 —los
    sectores publican desde 2004 y el interanual necesita 12 meses— así que el
    backfill del mandato está cubierto de punta a punta, sin huecos.
    [[YYYY-MM-01, %]]."""
    difusion, _ = macro._emae_difusion_por_mes(limit=5000)
    return [[f"{ym}-01", v] for ym, v in sorted(difusion.items()) if ym >= "2023-12"]


def fetch_ipi_serie() -> list:
    """IPI manufacturero: variación i.a. suavizada a 3 meses (ADR-0076), la
    misma construcción que puntúa en la ficha. [[YYYY-MM-01, %]]."""
    serie = macro._ipi_ia_por_mes()
    return [[f"{ym}-01", v] for ym, v in sorted(serie.items()) if ym >= "2023-12"]


def fetch_ipc_nucleo_serie() -> list:
    """IPC núcleo nacional, % m/m derivado del nivel — acompaña al IPC general
    en el gráfico comparado del modal (ADR-0077). No puntúa: sirve para que un
    mes de corrección tarifaria no se confunda con uno de núcleo alta.
    [[YYYY-MM-01, %]]."""
    niveles = {f[:7]: v for f, v in fetch_indec("148.3_INUCLEONAL_DICI_M_19", limit=60) if v}
    return [[f"{ym}-01", v] for ym, v in sorted(_var_mensual(niveles).items())
            if ym >= "2023-12"]


POBREZA_INDEC_ID = "64.2_POBLACION_NUA_0_0_34_74"   # INDEC EPH, % personas, TOTAL, semestral


def _una_sola_vez_por_corrida(fn):
    """Memoiza un fetcher sin argumentos: la corrida lo pide más de una vez y
    la fuente se lee una sola.

    Existe por un caso medido, no por prolijidad. `itvc_pobreza` se construye
    a partir de las OTRAS dos series de pobreza, así que `pobreza_nowcast` y
    `pobreza_indec` se pedían dos veces cada una por corrida. La del nowcast
    baja y parsea TODOS los informes PDF publicados de la UTDT —su propio
    docstring avisa que es caro y por eso no vive en el colector diario— y
    tarda más de seis minutos y medio por pasada. Se pagaba dos veces.

    Devuelve una copia: el llamador recibe una lista propia y no puede
    ensuciar la de los demás.
    """
    @functools.lru_cache(maxsize=1)
    def _cacheado():
        return fn()

    @functools.wraps(fn)
    def _wrapper():
        return [list(x) for x in _cacheado()]
    _wrapper.cache_clear = _cacheado.cache_clear
    return _wrapper


@_una_sola_vez_por_corrida
def fetch_pobreza_indec_serie() -> list:
    """Tasa OFICIAL de pobreza (INDEC, EPH continua): % de personas bajo la
    línea, total de aglomerados urbanos, semestral desde 2003 (ADR-0114).

    Acompaña al nowcast de la UTDT en su gráfico. No puntúa: es la referencia
    autorizada y la que da la historia larga —el nowcast sólo publica informes
    desde 2025—, pero llega dos veces al año y con rezago.

    La API entrega la proporción (0,282); se publica en puntos porcentuales
    para que las dos curvas del gráfico compartan unidad. [[YYYY-MM-01, %]]."""
    # la API devuelve del más nuevo al más viejo; el resto de las series del
    # proyecto van ascendentes y los gráficos lo asumen
    return sorted([[f"{f[:7]}-01", round(v * 100.0, 1)]
                   for f, v in fetch_indec(POBREZA_INDEC_ID, limit=200) if v is not None])


def fetch_itvc_pobreza() -> list:
    """I_PO: pobreza rebaseada a 4T-2023, INVERTIDA. 100 = la pobreza de la
    transición; por encima de 100 hay MENOS pobreza que entonces, que es mejora.

    Se invierte como los otros componentes «al revés» del ITVC (mora,
    informalidad, subocupación): más pobreza = peor, así que el cociente pone la
    base arriba.

    **El empalme y su costo, declarados.** El nivel sale del nowcast mensual de
    la UTDT, que sólo publica desde ene-2025 y por eso NO tiene 4T-2023. La base
    sale de la serie oficial del INDEC, que sí llega: 2º semestre de 2023, el que
    contiene al 4T. Las dos miden lo mismo pero **no coinciden en nivel, y la
    diferencia no tiene signo constante**: en los tres semestres que se solapan
    el nowcast está 2,3 puntos por debajo del INDEC (1S-2025), 0,5 por debajo
    (2S-2025) y 2,0 por ENCIMA (1S-2026). No es un sesgo corregible. Sobre una
    base de 40,1 puntos eso es hasta un 5,7% de error, que se propaga a todos los
    puntos del componente. Se acepta porque la alternativa —usar el semestral
    para todo— tira a la basura la resolución mensual, y la regla del proyecto es
    que teniendo el dato mensual gana el mensual. Pero va en la ficha.

    [[YYYY-MM-01, índice]]."""
    indec = dict(fetch_pobreza_indec_serie())
    base = indec.get("2023-07-01")            # 2º semestre 2023 (contiene el 4T)
    if not base:
        raise ValueError("itvc_pobreza: sin base del INDEC para el 2º semestre de 2023")
    return [[f, round(100.0 * base / v, 1)]
            for f, v in fetch_pobreza_nowcast_serie() if v]


POBREZA_NOWCAST_SERIE_STORE = Path(__file__).resolve().parents[1] / "data" / "vida" / "pobreza_nowcast_serie.json"


@_una_sola_vez_por_corrida
def fetch_pobreza_nowcast_serie() -> list:
    """Serie del Nowcast de Pobreza de la UTDT, un punto por informe mensual
    (ADR-0113). STORE persistente, mismo patrón que `fetch_ivi_serie` (misma
    editorial, UTDT): cada informe se lee una sola vez en su vida. Antes, esta
    serie bajaba y parseaba TODOS los PDF publicados EN CADA CORRIDA —medido:
    ~140s hoy, y en otra corrida superó los 400s sin terminar, según la red—
    porque no existía otro modo de reconstruir la serie completa. Memoizada
    además (ver `_una_sola_vez_por_corrida`): `itvc_pobreza` la vuelve a pedir
    dentro de la misma corrida.

    **El riesgo que un store URL-keyed corre, y por qué no es silencioso acá.**
    El `fname` de cada informe es un timestamp de subida, NO un hash de
    contenido (a diferencia del IVI, cuyo docstring llama "hash-based" a sus
    URLs) — así que nada impide, en principio, que la UTDT reemplace el PDF
    de una URL ya vista sin cambiar la URL. Si eso pasara con un store liso,
    el valor quedaría congelado y nadie se enteraría. El propio colector
    documenta que las revisiones del autor llegan como informe NUEVO ("si dos
    informes estiman el mismo semestre, gana el más nuevo"), nunca como un
    reemplazo del mismo archivo — así que la lectura de arriba es que el
    reemplazo silencioso no debería ocurrir. Pero "no debería" no es
    "detectado si ocurre", así que cada corrida hace un HEAD por URL YA
    CONOCIDA (barato: ~10s hoy para 23 informes vs. descargar y parsear los
    23 PDF) y compara el Content-Length contra el guardado. Si cambió, se
    re-lee ESE informe puntual y se avisa por consola — un reemplazo se
    vuelve una alerta, no un número mudo.

    Rechazado:
    - Hash de contenido propio: para calcularlo hay que bajar el PDF entero,
      que es exactamente el costo que se quiere evitar; el Content-Length del
      HEAD da una señal equivalente gratis.
    - Revalidar sólo los últimos K informes: deja al resto de la historia sin
      cobertura a cambio de un ahorro que la medición de arriba dice que no
      hace falta (~10s por los 23 informes completos).
    - Una bandera de re-lectura forzada: ya existe gratis y sin código nuevo
      — borrar este store.

    La regla "gana el más nuevo" (ADR-0153) sigue viva: la serie se
    reconstruye recorriendo TODOS los fname conocidos en orden cronológico
    (mismo criterio de `_listar_informes`) y pisando `serie[periodo]` en ese
    orden, así que un informe publicado después sigue ganando el semestre que
    comparte con uno anterior, venga o no del store.
    [[YYYY-MM-01, %]]."""
    sys.path.insert(0, str(Path(__file__).parent / "vida_cotidiana"))
    sys.path.insert(0, str(Path(__file__).parent / "vida_cotidiana" / "collectors"))
    from utdt_nowcast_pobreza import _listar_informes, _leer_informe, _huecos, NOWCAST_DESCARGA

    store = json.loads(POBREZA_NOWCAST_SERIE_STORE.read_text(encoding="utf-8-sig")) \
        if POBREZA_NOWCAST_SERIE_STORE.exists() else {"_meta": {}, "informes": {}}
    nuevos = cache = cambios = 0
    tocado = False   # backfill de content_length sin re-leer -- ver más abajo
    try:
        fnames = _listar_informes()   # ascendente, más viejo -> más nuevo
        for fname in fnames:
            previo = store["informes"].get(fname)
            try:
                head = requests.head(NOWCAST_DESCARGA + fname, headers=HTTP_HEADERS,
                                      timeout=HTTP_TIMEOUT, verify=False, allow_redirects=True)
                largo = int(head.headers["Content-Length"]) if "Content-Length" in head.headers else None
            except Exception:
                largo = None   # sin red para verificar: se trata como sin cambios (fail-open)
            if previo is not None:
                if largo is not None and previo.get("content_length") is None:
                    # el primer HEAD de este informe falló en su momento y quedó
                    # sin fingerprint -- se completa ahora sin gastar una relectura,
                    # así no queda ciego a cambios futuros para siempre.
                    previo["content_length"] = largo
                    tocado = True
                if largo is None or previo.get("content_length") is None \
                        or largo == previo["content_length"]:
                    cache += 1
                    continue
                print(f"  [ALERTA] Nowcast pobreza: {fname} cambió de tamaño "
                      f"({previo['content_length']} -> {largo} bytes) -- mismo URL, "
                      "contenido distinto; se vuelve a leer")
            try:
                d = _leer_informe(fname)
            except Exception as e:
                print(f"  [WARN] Nowcast pobreza: informe no procesado ({fname[-24:]}): {str(e)[:60]}")
                continue
            store["informes"][fname] = {
                "periodo": d["periodo"] if d else None,
                "valor": d["valor"] if d else None,
                "content_length": largo,
            }
            if previo is None:
                nuevos += 1
            else:
                cambios += 1
        store["_meta"] = {
            "fuente": "UTDT — Nowcast de Pobreza (González-Rozada), informes mensuales PDF",
            "actualizado": datetime.today().strftime("%Y-%m-%d"),
            "nota": ("Cada informe se lee una sola vez en su vida (ADR-0113); cada "
                     "corrida valida las URL ya conocidas con un HEAD (Content-Length) "
                     "y sólo re-lee la que cambió de tamaño — ver docstring de "
                     "fetch_pobreza_nowcast_serie en descargar_series.py."),
        }
        if nuevos or cambios or tocado:
            POBREZA_NOWCAST_SERIE_STORE.write_text(
                json.dumps(store, indent=2, ensure_ascii=False), encoding="utf-8")
        etiqueta = f", {cambios} cambios detectados" if cambios else ""
        print(f"  [OK] Nowcast pobreza: {nuevos} informes nuevos, {cache} del store{etiqueta} "
              f"(serie: {len(store['informes'])} informes)")
    except Exception as e:
        print(f"  [WARN] Nowcast pobreza: listado UTDT no disponible ({str(e)[:60]}); serie del store")

    serie = {}
    for fname in sorted(store["informes"]):
        info = store["informes"][fname]
        if info.get("valor") is not None:
            serie[info["periodo"]] = info["valor"]
    huecos = _huecos(serie)
    if huecos:
        print(f"  [WARN] Nowcast pobreza: huecos en la serie {huecos} -- probable "
              "informe mal fechado, no un mes sin publicar (ADR-0153)")
    return [[f"{ym}-01", v] for ym, v in sorted(serie.items())]


CUENTA_CORRIENTE_ID = "160.2_TL_CUENNTE_0_T_22"   # INDEC, balanza de pagos, trimestral


def fetch_cuenta_corriente_serie() -> list:
    """Cuenta corriente de la balanza de pagos, ACUMULADA 4 TRIMESTRES (ADR-0080).

    Acompaña al saldo comercial en su gráfico —no puntúa: es trimestral y llega
    con más rezago— y existe para mostrar lo que el saldo de bienes no ve.

    Se acumula a cuatro trimestres a propósito: el indicador que acompaña es el
    saldo comercial de DOCE MESES, y compararlo contra un trimestre suelto
    mezclaría escalas. Con las dos en base anual el contraste se lee directo.

    Fuente INDEC (devengado), no el balance cambiario del BCRA (flujos
    efectivos por el mercado de cambios): son conceptos distintos y bajo
    restricciones cambiarias divergen mucho. La cuenta corriente propiamente
    dicha es la del INDEC. [[YYYY-MM-01, M USD]]."""
    trim = sorted((f[:7], v) for f, v in fetch_indec(CUENTA_CORRIENTE_ID, limit=200)
                  if v is not None)
    return [[f"{trim[i][0]}-01", round(sum(x[1] for x in trim[i - 3:i + 1]), 1)]
            for i in range(3, len(trim))
            if trim[i][0] >= "2023-12"]


def fetch_recaudacion_real_serie() -> list:
    """Base imponible REAL desestacionalizada, 100 = 4T-2023 (nacional DGI +
    provincial COMARB).

    Reemplaza a la variación interanual con promedio móvil 3 meses que usaba el
    indicador hasta el 29-jul-2026. El cálculo vive en `comarb`, no acá, porque
    la card lo necesita idéntico: los factores estacionales dependen de la
    ventana, así que dos implementaciones —o dos ventanas— divergirían y G3
    fallaría. Misma disciplina que `apoyo_empresario_serie`.
    [[YYYY-MM-01, índice]]."""
    nominal = {f[:7]: v for f, v in fetch_indec(macro.INDEC_RECAUDACION_ID,
                                                limit=comarb.LIMITE_MESES) if v}
    ipc = {f[:7]: v for f, v in fetch_indec(IPC_NIVEL_ID, limit=comarb.LIMITE_MESES) if v}
    serie = comarb.base_imponible_real_sa(nominal, ipc)
    if not serie:
        raise ValueError("recaudación: sin ventana suficiente para desestacionalizar")
    return [[f"{ym}-01", v] for ym, v in sorted(serie.items())]
def fetch_credito_privado_serie() -> list:
    """Serie mensual de la variación i.a. REAL de los préstamos al sector
    privado (BCRA var. 26 fin de mes, deflactada por el IPC nivel) — la misma
    métrica del indicador credito_privado (ADR-0022). [[YYYY-MM-01, %]]."""
    fin_mes = {}
    for f, v in sorted(fetch_bcra(26, dias=1350)):
        fin_mes[f[:7]] = v
    ipc = {f[:7]: v for f, v in fetch_indec("148.3_INIVELNAL_DICI_M_26", limit=60)}
    out = []
    for ym in sorted(fin_mes):
        if ym < "2023-12":
            continue
        prev = f"{int(ym[:4]) - 1}{ym[4:]}"
        if prev in fin_mes and ym in ipc and prev in ipc and fin_mes[prev] and ipc[prev]:
            nominal = fin_mes[ym] / fin_mes[prev] - 1.0
            infl = ipc[ym] / ipc[prev] - 1.0
            out.append([f"{ym}-01", round(((1 + nominal) / (1 + infl) - 1) * 100.0, 1)])
    return out


def fetch_costo_financiamiento_tesoro_serie() -> list:
    """Serie mensual de la tasa REAL que paga el Tesoro por colocar deuda en
    pesos (ADR-0071): TIREA promedio ponderada de las colocaciones del mes,
    deflactada por la inflación esperada a 12 meses del REM.

    Arranca en dic-2023 (primera licitación de la gestión, la LEDE S18E4 a TEM
    8,66%). Los meses sin colocaciones a tasa fija en pesos —enero y febrero de
    2024, cuando todo lo emitido fue CER— quedan fuera de la serie: no se
    inventa un dato que no existe."""
    tirea = macro._tirea_mensual(anios=4)
    rem = macro._rem_12m_por_mes(dias=1400)
    out = []
    for ym in sorted(tirea):
        if ym < "2023-12" or ym not in rem:
            continue
        t = tirea[ym][0]
        real = ((1.0 + t) / (1.0 + rem[ym] / 100.0) - 1.0) * 100.0
        out.append([f"{ym}-01", round(real, 2)])
    return out


def fetch_resultado_primario_serie() -> list:
    """Serie mensual del superávit primario del SPN como porcentaje de la
    recaudación, ambos acumulados 12 meses (ADR-0072). Desde dic-2023."""
    serie = macro._superavit_sobre_recaudacion_12m()
    return [[f"{ym}-01", round(v, 2)] for ym, v in sorted(serie.items())
            if ym >= "2023-12"]


MACRO_DERIVADAS = [
    ("ipc_total", "% mensual", "INDEC (derivado del nivel del IPC)", fetch_ipc_mm_serie),
    ("emae_ia", "% i.a.", "INDEC/datos.gob.ar", fetch_emae_ia_serie),
    ("emae_difusion", "% de sectores en crecimiento i.a.",
     "INDEC — EMAE apertura sectorial (vía datos.gob.ar)", fetch_emae_difusion_serie),
    ("ipi_manufacturero", "% i.a. (promedio 3 meses)", "INDEC — IPI manufacturero (vía datos.gob.ar)", fetch_ipi_serie),
    # acompaña al IPC general en el modal, no puntúa (ADR-0077)
    ("ipc_nucleo", "% mensual", "INDEC — IPC núcleo nacional (vía datos.gob.ar)", fetch_ipc_nucleo_serie),
    # La unidad decía "% i.a. real" y la serie devuelve un ÍNDICE desde el
    # 29-jul-2026, cuando `fetch_recaudacion_real_serie` reemplazó la variación
    # interanual por la base imponible real desestacionalizada. El metadato
    # quedó tres semanas describiendo la métrica anterior; ningún gate lo veía
    # porque card y serie coinciden en el número (88,2 contra 88,2) y G3 sólo
    # compara números.
    ("recaudacion", "índice (100 = 4T-2023)", "Sec. Hacienda (recaudación DGI) + IPC (deflactor)", fetch_recaudacion_real_serie),
    ("credito_privado", "% i.a. real", "BCRA (préstamos privados) + IPC INDEC", fetch_credito_privado_serie),
    ("costo_financiamiento_tesoro", "% real anual", "Sec. de Finanzas (colocaciones) + BCRA (REM)", fetch_costo_financiamiento_tesoro_serie),
    ("resultado_primario", "% de la recaudación (12m)", "Sec. de Hacienda — IMIG + recaudación", fetch_resultado_primario_serie),
    ("saldo_comercial", "M USD", "INDEC/datos.gob.ar (ICA expo−impo)", fetch_saldo_ica),
    ("saldo_comercial_12m", "M USD (acum. 12 meses)", "INDEC — ICA (vía datos.gob.ar)", fetch_saldo_12m_serie),
    # acompaña al saldo comercial en el modal, no puntúa (ADR-0080)
    ("cuenta_corriente", "M USD (acum. 4 trimestres)", "INDEC — balanza de pagos (vía datos.gob.ar)", fetch_cuenta_corriente_serie),
    ("reservas_bcra", "M USD netas", "BCRA Planilla SDDS + Balance (a secas)", fetch_reservas_netas_serie),
    ("tcrm", "índice (base dic-2015)", "BCRA ITCRM", fetch_tcrm_serie),
    # bilaterales oficiales de la misma planilla (descarga memoizada): el
    # gráfico comparado del modal del TCRM, como lo presenta el propio BCRA
    ("tcrm_bilateral_brasil", "índice (base dic-2015)", "BCRA ITCRM",
     lambda: [[f, v] for f, v in macro.fetch_itcrm_bilateral("brasil")][-32:]),
    ("tcrm_bilateral_eeuu", "índice (base dic-2015)", "BCRA ITCRM",
     lambda: [[f, v] for f, v in macro.fetch_itcrm_bilateral("eeuu")][-32:]),
    ("idm", "pp (brecha i.a. real)", "BCRA (M3/M2 privado) + IPC INDEC", fetch_idm_serie),
    ("desequilibrio_monetario", "pts de tensión (0-100)",
     "BCRA (M2 transaccional privado, circulante, depósitos privados y Mercado de Cambios)",
     fetch_desequilibrio_monetario_serie),
    ("idc", "σ vs. su historia", "BCRA (BADLAR/depósitos/préstamos) + IPC INDEC", fetch_idc_serie),
    ("iai", "% i.a. ponderado", "INDEC (ISAC + bienes de capital importados)", fetch_iai_serie),
    ("icip", "% i.a. ponderado", "INDEC (servicios informática + productividad)", fetch_icip_serie),
]
MACRO_BCRA = [
    (7,  "badlar",             "% anual",  "BCRA"),
    (29, "rem_ipc_12m",        "% anual",  "BCRA"),
    (26, "prestamos_privados", "M ARS",    "BCRA"),
    (15, "base_monetaria",     "M ARS",    "BCRA"),
    (5,  "tc_mayorista",       "ARS/USD",  "BCRA"),
]

POLITICA_INDEC = []
# Cinturón político: sin series INDEC, pero el Votómetro reconstruye su histórico.


def fetch_votometro_serie() -> list:
    """Serie histórica de la brecha LLA−PJ del Votómetro (reconstruida desde
    encuestasRaw). [[YYYY-MM-01, gap]]."""
    return [[f"{ym}-01", g] for ym, g in politica.votometro_serie_mensual()]


def fetch_iaf_serie() -> list:
    """Serie ANUAL de la variación real i.a. de transferencias federales totales
    (RON Hacienda, transferencias EJECUTADAS del año calendario — el punto
    YYYY-12-01 es el acumulado de ESE año cerrado, no el presupuesto del
    siguiente), deflactada por la inflación PROMEDIO anual del índice IPC de
    INDEC (ADR-0065) — misma fórmula que el indicador. Confiable desde 2018
    (primer año con promedio completo de la base dic-2016). [[YYYY-12-01, %]]."""
    import csv
    import io
    r = requests.get(politica.RON_CSV_URL, headers=HTTP_HEADERS, timeout=HTTP_TIMEOUT)
    r.raise_for_status()
    rd = csv.reader(io.StringIO(r.text), delimiter=";")
    next(rd)
    tot = {}
    for row in rd:
        if len(row) < 5:
            continue
        if row[1].strip().lower() in politica.RON_NO_PROVINCIA:   # ADR-0066
            continue
        try:
            tot[int(row[0])] = tot.get(int(row[0]), 0.0) + float(row[4].replace(",", "."))
        except ValueError:
            continue
    ipc = politica._ipc_promedio_indec()
    out = []
    for y in sorted(tot):
        if y - 1 in tot and tot[y - 1] and y in ipc:
            var_nom = tot[y] / tot[y - 1] - 1.0
            var_real = (1.0 + var_nom) / (1.0 + ipc[y]) - 1.0
            out.append([f"{y}-12-01", round(var_real * 100.0, 1)])
    return out


def fetch_ratio_dnu_serie() -> list:
    """Serie MENSUAL del ratio DNUs/leyes (InfoLeg), ventana móvil de 365 días
    al fin de cada mes desde dic-2023 — misma fórmula que el indicador
    (ADR-0058; antes era un punto por año calendario, no comparable mes a
    mes). InfoLeg no expone un dump con fecha por registro como el CKAN de
    HCDN, así que cada mes requiere su propia consulta al buscador (una de
    leyes + una de DNUs), reutilizando la misma sesión.
    [[YYYY-MM-01, ratio]]."""
    s = requests.Session()
    rh = s.get(politica.INFOLEG_HOME, headers=HTTP_HEADERS, timeout=HTTP_TIMEOUT)
    rh.raise_for_status()
    m = re.search(r'action="(/infolegInternet/[^"]+)"', rh.text)
    if not m:
        raise ValueError("InfoLeg: no se encontró el form action")
    au = "https://servicios.infoleg.gob.ar" + m.group(1)
    out = []
    for ym, cutoff_iso, fin_iso in _hcdn_ventanas_12m():
        desde = date.fromisoformat(cutoff_iso)
        hasta = date.fromisoformat(fin_iso)
        try:
            leyes = politica._infoleg_session_count(s, au, "1", desde, hasta)
            if not leyes:
                continue
            dnus = politica._infoleg_session_count(s, au, "2", desde, hasta,
                                                     texto="necesidad y urgencia")
        except Exception as e:
            print(f"  [WARN] ratio_dnu serie {ym}: {e}")
            continue
        out.append([f"{ym}-01", round(dnus / leyes, 3)])
    return out


def _hcdn_ventanas_12m() -> list:
    """[(YYYY-MM, cutoff_iso, fin_iso)] con ventana móvil de 365 días al fin de
    cada mes, desde dic-2023 hasta hoy."""
    out = []
    y, m = 2023, 12
    hoy = date.today()
    while (y, m) <= (hoy.year, hoy.month):
        fin = min(date(y, m, calendar.monthrange(y, m)[1]), hoy)
        out.append((f"{y}-{m:02d}", (fin - timedelta(days=365)).isoformat(), fin.isoformat()))
        m += 1
        if m > 12:
            m = 1; y += 1
    return out


def fetch_eficacia_serie() -> list:
    """Serie mensual del % de proyectos de ley PE MADUROS (publicados 12-24
    meses antes de cada mes) que se convirtieron en ley, evaluado a fin de
    ese mes — misma fórmula que el indicador tras ADR-0061 (cohorte madura)
    y ADR-0062 (numerador desde el dataset leyes-sancionadas, que cubre las
    sanciones del Senado; denominador sin comunicaciones administrativas —
    ver fetch_eficacia_legislativa en politica.py). Reutiliza
    _hcdn_ventanas_12m() reinterpretando su `cutoff` como el LÍMITE SUPERIOR
    de la cohorte de publicación. Reproducible: la sanción se acota por
    SANCION_DEFINITIVA al CIERRE de ese mes histórico, no hasta hoy, para
    que un punto ya publicado no cambie retroactivamente solo porque el
    proyecto finalmente se sancionó más tarde. [[YYYY-MM-01, %]]."""
    raw_pe = (politica._hcdn_paginate(politica.HCDN_PROYECTOS_RID, q="-PE-")
              + politica._hcdn_paginate(politica.HCDN_PROYECTOS_RID, q="-JGM-"))
    pe = [(r["PROYECTO_ID"], str(r.get("PUBLICACION_FECHA", ""))[:10]) for r in raw_pe
          if r.get("PROYECTO_ID")
          and "PROYECTO DE LEY" in str(r.get("TIPO", "")).upper()
          and (politica._RE_PE_EXP.search(r.get("EXP_DIPUTADOS", "") or "")
               or politica._RE_PE_EXP.search(r.get("EXP_SENADO", "") or ""))]
    raw_leyes = politica._hcdn_paginate(politica.HCDN_LEYES_SANC_RID)
    san = [(str(r.get("PROYECTO_ID", "")).strip(), str(r.get("SANCION_DEFINITIVA", ""))[:10])
           for r in raw_leyes if r.get("PROYECTO_ID")]
    out = []
    for ym, cohorte_hasta, fin in _hcdn_ventanas_12m():
        cohorte_desde = (date.fromisoformat(cohorte_hasta) - timedelta(days=365)).isoformat()
        pe_ids = {pid for pid, f in pe if cohorte_desde <= f <= cohorte_hasta}
        # "NA" en SANCION_DEFINITIVA queda excluido solo ('N' > dígitos en str)
        san_ids = {pid for pid, f in san if f <= fin}
        if pe_ids:
            out.append([f"{ym}-01", round(len(pe_ids & san_ids) / len(pe_ids) * 100.0, 1)])
    return out


def fetch_veto_quorum_serie() -> list:
    """Serie MENSUAL del % de sesiones legislativas de Diputados que quedaron en
    minoría, ventana móvil de 12 meses calendario (ADR-0091).

    Antes era ANUAL —un punto por período legislativo, tres puntos en total— y
    puntuaba una tasa cuyo denominador se reseteaba cada marzo. Ahora usa la
    misma ventana y la misma regla que la card (`politica._veto_quorum_tasa_12m`),
    con una sola descarga de sesiones para todos los meses.
    [[YYYY-MM-DD, %]]."""
    try:
        sesiones = politica._hcdn_sesiones_legislativas()
    except Exception as e:
        print(f"  [WARN] veto_quorum serie: {e} -- serie omitida")
        return []
    out = []
    for fin_mes in _fines_de_mes(date(2023, 12, 1), date.today(), incluir_hoy=True):
        tasa = politica._veto_quorum_tasa_12m(sesiones, fin_mes)
        if tasa is not None:
            out.append([fin_mes.strftime("%Y-%m-%d"), tasa[0]])
    return out


def fetch_comisiones_serie() -> list:
    """Serie mensual del % de proyectos con dictamen (Orden del Día) SIN
    sanción, ventana móvil de 12 meses — misma fórmula que el indicador.
    [[YYYY-MM-01, %]]."""
    dictamenes = []
    for y in range(2023, date.today().year + 1):
        try:
            dictamenes += politica._hcdn_paginate(politica.HCDN_DICTAMENES_RID, q=str(y))
        except Exception as e:
            print(f"  [WARN] comisiones serie dictámenes {y}: {e}")
    od = [(str(r["EXPEDIENTE"]).strip(), str(r.get("FECHA", ""))[:10]) for r in dictamenes
          if r.get("EXPEDIENTE") and "orden" in str(r.get("TIPO", "")).lower()]
    raw_san = politica._hcdn_paginate(politica.HCDN_MOVIMIENTOS_RID, q="SANCION")
    san = [(str(r.get("PROYECTO_ID", "")).strip(), str(r.get("FECHA", ""))[:10])
           for r in raw_san if r.get("PROYECTO_ID")]
    out = []
    for ym, cutoff, fin in _hcdn_ventanas_12m():
        od_ids = {e for e, f in od if cutoff <= f <= fin}
        san_ids = {p for p, f in san if cutoff <= f <= fin}
        if od_ids:
            pct = round((len(od_ids) - len(od_ids & san_ids)) / len(od_ids) * 100.0, 1)
            out.append([f"{ym}-01", pct])
    return out


COHESION_BLOQUE_STORE = Path(__file__).resolve().parents[1] / "data" / "politica" / "cohesion_bloque_serie.json"
COHESION_BLOQUE_SENADO_STORE = Path(__file__).resolve().parents[1] / "data" / "politica" / "cohesion_bloque_senado_serie.json"
ALINEAMIENTO_SENADORES_STORE = Path(__file__).resolve().parents[1] / "data" / "politica" / "alineamiento_senadores_serie.json"


def _serie_cohesion_cacheada(store_path: Path, fetch_fn, anio_inicio: int, nombre: str) -> list:
    """Serie ANUAL de cohesión de bloque (índice de Rice) con caché
    persistente por año -- hallazgo real 2026-07-08: cada año requiere
    scrapear TODAS sus actas divididas una por una (~80/año en Senado), y
    repetir esto para 2023..hoy en CADA corrida diaria es lento y expone la
    serie a la intermitencia del sitio fuente (un timeout tira abajo un año
    entero). Mismo patrón ya usado para motos/carne/snic/ivi/sentimiento en
    este archivo: los años CERRADOS (anteriores al actual) son inmutables
    una vez bajados -- no se vuelven a pedir. El año EN CURSO SIEMPRE se
    re-pide (pueden sumarse actas nuevas); si el pedido de hoy falla, se
    degrada al último valor cacheado de ese año en vez de perder el punto
    más nuevo de la serie (que es justo el que dispara el gate G3 si falta).
    `fetch_fn` sigue la firma de politica.fetch_cohesion_bloque[_senado]
    (anio, dias_ventana). [[YYYY-01-01, % cohesión]]."""
    try:
        cache = json.loads(store_path.read_text(encoding="utf-8-sig"))
        if not isinstance(cache, dict):
            cache = {}
    except (OSError, json.JSONDecodeError):
        cache = {}

    hoy_anio = date.today().year
    for anio in range(anio_inicio, hoy_anio + 1):
        clave = str(anio)
        if anio < hoy_anio and clave in cache:
            continue   # año cerrado ya cacheado -- inmutable, no se vuelve a pedir
        try:
            resultado = fetch_fn(anio=anio, dias_ventana=366)
        except Exception as e:
            print(f"  [WARN] {nombre} {anio}: {e}")
            continue
        if resultado and resultado.get("valor") is not None:
            cache[clave] = resultado["valor"]
            store_path.write_text(json.dumps(cache, indent=2, ensure_ascii=False), encoding="utf-8")
        elif clave not in cache:
            print(f"  [WARN] {nombre} {anio}: sin dato y sin cache previo -- se omite")

    return [[f"{anio}-01-01", cache[str(anio)]]
            for anio in range(anio_inicio, hoy_anio + 1) if str(anio) in cache]


def fetch_cohesion_bloque_serie(anio_inicio: int = 2023) -> list:
    """Serie ANUAL de cohesión del bloque LLA en Diputados (índice de Rice
    promedio): un punto por año desde `anio_inicio`, con dias_ventana=366
    para cubrir TODAS las actas divididas del año sin depender de la fecha de
    corrida — mismo criterio que el indicador cohesion_bloque (Tarea 6).
    Caché persistente por año, ver _serie_cohesion_cacheada.
    [[YYYY-01-01, % cohesión]]."""
    return _serie_cohesion_cacheada(COHESION_BLOQUE_STORE, politica.fetch_cohesion_bloque,
                                     anio_inicio, "cohesion_bloque")


def fetch_cohesion_bloque_senado_serie(anio_inicio: int = 2023) -> list:
    """Serie ANUAL de cohesión del bloque LLA en el Senado (índice de Rice
    promedio): mismo patrón que fetch_cohesion_bloque_serie (Diputados) —
    un punto por año desde `anio_inicio`, con dias_ventana=366 para cubrir
    TODAS las actas divididas del año sin depender de la fecha de corrida.
    Indicador COMPLEMENTARIO (otra cámara), no reemplaza a cohesion_bloque.
    Caché persistente por año, ver _serie_cohesion_cacheada.
    [[YYYY-01-01, % cohesión]]."""
    return _serie_cohesion_cacheada(COHESION_BLOQUE_SENADO_STORE, politica.fetch_cohesion_bloque_senado,
                                     anio_inicio, "cohesion_bloque_senado")


COHESION_SENADO_ACTAS_STORE = Path(__file__).resolve().parents[1] / "data" / "politica" / "cohesion_bloque_senado_actas.json"


def _actas_cohesion_senado_cacheadas(anio_inicio: int) -> dict:
    """Cachea el detalle CRUDO por acta (politica.fetch_cohesion_bloque_senado_actas_anio)
    por año -- mismo criterio que _actas_alineamiento_cacheadas: años
    CERRADOS inmutables, el año en curso se re-pide siempre.
    {anio (str): [detalle de fetch_cohesion_bloque_senado_actas_anio]}."""
    try:
        cache = json.loads(COHESION_SENADO_ACTAS_STORE.read_text(encoding="utf-8-sig"))
        if not isinstance(cache, dict):
            cache = {}
    except (OSError, json.JSONDecodeError):
        cache = {}

    hoy_anio = date.today().year
    for anio in range(anio_inicio, hoy_anio + 1):
        clave = str(anio)
        if anio < hoy_anio and clave in cache:
            continue   # año cerrado ya cacheado -- inmutable, no se vuelve a pedir
        try:
            detalle = politica.fetch_cohesion_bloque_senado_actas_anio(anio)
        except Exception as e:
            print(f"  [WARN] cohesion_bloque_senado_actas {anio}: {e}")
            continue
        if detalle is not None:
            cache[clave] = detalle
            COHESION_SENADO_ACTAS_STORE.write_text(
                json.dumps(cache, indent=2, ensure_ascii=False), encoding="utf-8")
            print(f"  [OK] cohesion_bloque_senado_actas {anio}: {len(detalle)} actas con señal")
        elif clave not in cache:
            print(f"  [WARN] cohesion_bloque_senado_actas {anio}: sin dato y sin cache previo -- se omite")

    return cache


def fetch_cohesion_bloque_senado_mensual(anio_inicio: int = 2023, dias_ventana: int = 90) -> list:
    """Serie MENSUAL de cohesion_bloque_senado: un punto por fin de mes desde
    diciembre de `anio_inicio` hasta hoy, cada uno una ventana rolling de
    `dias_ventana` días (mismo criterio "Continua (90d)" que el valor live
    de la card) -- mismo patrón que fetch_alineamiento_senadores_prov_mensual
    (ver ese docstring y ADR-0038). Reusa el detalle crudo por acta cacheado
    por año (_actas_cohesion_senado_cacheadas) -- una sola pasada de
    scraping por año, no una por mes. [[YYYY-MM-DD, valor]]."""
    cache = _actas_cohesion_senado_cacheadas(anio_inicio)
    detalle = [fila for detalle_anio in cache.values() for fila in detalle_anio]
    if not detalle:
        return []

    puntos = []
    for fin_mes in _fines_de_mes(date(anio_inicio, 12, 1), date.today(), incluir_hoy=True):
        referencia = datetime(fin_mes.year, fin_mes.month, fin_mes.day)
        resultado = politica._agregar_cohesion_ventana(detalle, referencia, dias_ventana)
        if resultado:
            puntos.append([fin_mes.strftime("%Y-%m-%d"), resultado["valor"]])
    return puntos


COHESION_DIPUTADOS_ACTAS_STORE = Path(__file__).resolve().parents[1] / "data" / "politica" / "cohesion_bloque_diputados_actas.json"


def _actas_cohesion_diputados_cacheadas(anio_inicio: int) -> dict:
    """Cachea el detalle CRUDO por acta de Diputados
    (politica.fetch_cohesion_bloque_diputados_actas_anio) por año -- mismo
    criterio que _actas_cohesion_senado_cacheadas: años CERRADOS inmutables,
    el año en curso se re-pide siempre. A diferencia de Senado, cada llamada
    a fetch_cohesion_bloque_diputados_actas_anio ya camina sobre la caché
    PERMANENTE por acta de politica.py (_acta_diputados_cacheada) -- este
    caché por año evita además volver a RECORRER (aunque sea solo lookups de
    diccionario) los años cerrados en cada corrida.
    {anio (str): [detalle de fetch_cohesion_bloque_diputados_actas_anio]}."""
    try:
        cache = json.loads(COHESION_DIPUTADOS_ACTAS_STORE.read_text(encoding="utf-8-sig"))
        if not isinstance(cache, dict):
            cache = {}
    except (OSError, json.JSONDecodeError):
        cache = {}

    hoy_anio = date.today().year
    for anio in range(anio_inicio, hoy_anio + 1):
        clave = str(anio)
        if anio < hoy_anio and clave in cache:
            continue   # año cerrado ya cacheado -- inmutable, no se vuelve a pedir
        try:
            detalle = politica.fetch_cohesion_bloque_diputados_actas_anio(anio)
        except Exception as e:
            print(f"  [WARN] cohesion_bloque_diputados_actas {anio}: {e}")
            continue
        if detalle is not None:
            cache[clave] = detalle
            COHESION_DIPUTADOS_ACTAS_STORE.write_text(
                json.dumps(cache, indent=2, ensure_ascii=False), encoding="utf-8")
            print(f"  [OK] cohesion_bloque_diputados_actas {anio}: {len(detalle)} actas con señal")
        elif clave not in cache:
            print(f"  [WARN] cohesion_bloque_diputados_actas {anio}: sin dato y sin cache previo -- se omite")

    return cache


def fetch_cohesion_bloque_diputados_mensual(anio_inicio: int = 2023, dias_ventana: int = 90) -> list:
    """Serie MENSUAL de cohesion_bloque (Diputados): un punto por fin de mes
    desde diciembre de `anio_inicio` hasta hoy, cada uno una ventana rolling
    de `dias_ventana` días (mismo criterio "Continua (90d)" que el valor live
    de la card) -- mismo patrón que fetch_cohesion_bloque_senado_mensual
    (ADR-0039) y fetch_alineamiento_senadores_prov_mensual (ADR-0038), ahora
    viable para Diputados gracias a la caché permanente por acta (ADR-0040
    follow-up, 2026-07-09): sin ella, reconstruir 2023..hoy caminaba TODA la
    historia desde el id más reciente en cada una de las 4 llamadas anuales
    (~53min extra medidos en el pipeline completo). Reusa el detalle crudo
    por acta cacheado por año (_actas_cohesion_diputados_cacheadas).
    [[YYYY-MM-DD, valor]]."""
    cache = _actas_cohesion_diputados_cacheadas(anio_inicio)
    detalle = [fila for detalle_anio in cache.values() for fila in detalle_anio]
    if not detalle:
        return []

    puntos = []
    for fin_mes in _fines_de_mes(date(anio_inicio, 12, 1), date.today(), incluir_hoy=True):
        referencia = datetime(fin_mes.year, fin_mes.month, fin_mes.day)
        resultado = politica._agregar_cohesion_ventana(detalle, referencia, dias_ventana)
        if resultado:
            puntos.append([fin_mes.strftime("%Y-%m-%d"), resultado["valor"]])
    return puntos


def fetch_cohesion_bloque_compuesta_mensual(anio_inicio: int = 2023, dias_ventana: int = 90) -> list:
    """Serie MENSUAL del compuesto bicameral de cohesión (ADR-0048): por cada
    fin de mes, Rice de Diputados 65% + Rice del Senado 35%
    (politica.COHESION_PESOS_CAMARAS — la misma fórmula que la card), sobre
    las dos series mensuales por cámara ya cacheadas. Si un mes tiene dato de
    una sola cámara (el Senado arranca en feb-2024), renormaliza igual que la
    card. [[YYYY-MM-DD, valor]]."""
    dip = dict(fetch_cohesion_bloque_diputados_mensual(anio_inicio, dias_ventana))
    sen = dict(fetch_cohesion_bloque_senado_mensual(anio_inicio, dias_ventana))
    pesos = politica.COHESION_PESOS_CAMARAS
    out = []
    for fecha in sorted(set(dip) | set(sen)):
        con_dato = {c: v for c, v in (("diputados", dip.get(fecha)),
                                       ("senado", sen.get(fecha))) if v is not None}
        peso_total = sum(pesos[c] for c in con_dato)
        valor = sum(v * pesos[c] for c, v in con_dato.items()) / peso_total
        out.append([fecha, round(valor, 1)])
    return out


def fetch_alineamiento_senadores_prov_serie(anio_inicio: int = 2023) -> list:
    """Serie ANUAL de alineamiento_senadores_prov (reemplaza a
    gobernadores_alineamiento). Caché persistente por año, ver
    _serie_cohesion_cacheada."""
    return _serie_cohesion_cacheada(ALINEAMIENTO_SENADORES_STORE, politica.fetch_alineamiento_senadores_prov,
                                     anio_inicio, "alineamiento_senadores_prov")


ALINEAMIENTO_SENADORES_ACTAS_STORE = Path(__file__).resolve().parents[1] / "data" / "politica" / "alineamiento_senadores_actas.json"


def _actas_alineamiento_cacheadas(anio_inicio: int) -> dict:
    """Cachea el detalle CRUDO por acta (politica.fetch_alineamiento_senadores_actas_anio)
    por año -- mismo criterio que _serie_cohesion_cacheada: los años
    CERRADOS son inmutables una vez bajados, el año en curso se re-pide
    siempre. {anio (str): [detalle de fetch_alineamiento_senadores_actas_anio]}.
    Existe para poder derivar la serie MENSUAL (fetch_alineamiento_senadores_prov_mensual)
    sin volver a scrapear el Senado en cada corrida -- el año completo se
    pide una sola vez, la ventana de 90 días se recorta después en Python."""
    try:
        cache = json.loads(ALINEAMIENTO_SENADORES_ACTAS_STORE.read_text(encoding="utf-8-sig"))
        if not isinstance(cache, dict):
            cache = {}
    except (OSError, json.JSONDecodeError):
        cache = {}

    hoy_anio = date.today().year
    for anio in range(anio_inicio, hoy_anio + 1):
        clave = str(anio)
        if anio < hoy_anio and clave in cache:
            continue   # año cerrado ya cacheado -- inmutable, no se vuelve a pedir
        try:
            detalle = politica.fetch_alineamiento_senadores_actas_anio(anio)
        except Exception as e:
            print(f"  [WARN] alineamiento_senadores_actas {anio}: {e}")
            continue
        if detalle is not None:
            cache[clave] = detalle
            ALINEAMIENTO_SENADORES_ACTAS_STORE.write_text(
                json.dumps(cache, indent=2, ensure_ascii=False), encoding="utf-8")
            print(f"  [OK] alineamiento_senadores_actas {anio}: {len(detalle)} actas con señal")
        elif clave not in cache:
            print(f"  [WARN] alineamiento_senadores_actas {anio}: sin dato y sin cache previo -- se omite")

    return cache


def _fines_de_mes(desde: date, hasta: date, incluir_hoy: bool = False) -> list:
    """Lista de fechas de fin de mes (inclusive) entre `desde` y `hasta`,
    redondeando `desde` al fin de SU mes.

    Con `incluir_hoy=True` agrega `hasta` como punto final cuando no es ya un
    fin de mes. Eso es lo que alinea la serie con la card en los indicadores de
    ventana móvil: la card evalúa su ventana en `date.today()` y la serie, sin
    este punto, terminaba en el último mes CERRADO, mirando una ventana corrida
    un mes. Card y serie quedaban permanentemente desfasadas y sólo coincidían
    por casualidad los días en que las dos ventanas caían igual.

    El desfase no es cosmético: `desafios_legislativos` publicaba 3 en el
    titular (ventana sep-2025→ago-2026) contra 10 en el último punto del
    gráfico (ago-2025→jul-2026), porque los 7 desafíos de agosto de 2025 salen
    de la ventana todos juntos al rodar el mes. Puntúan 90 y 43,6: la página se
    contradecía a sí misma arriba y abajo del mismo indicador (ADR-0172)."""
    puntos = []
    anio, mes = desde.year, desde.month
    while True:
        fin_mes = date(anio, mes, calendar.monthrange(anio, mes)[1])
        if fin_mes > hasta:
            break
        puntos.append(fin_mes)
        anio, mes = (anio + 1, 1) if mes == 12 else (anio, mes + 1)
    if incluir_hoy and (not puntos or puntos[-1] != hasta):
        puntos.append(hasta)
    return puntos


def fetch_alineamiento_senadores_prov_mensual(anio_inicio: int = 2023, dias_ventana: int = 90) -> list:
    """Serie MENSUAL de alineamiento_senadores_prov: un punto por fin de mes
    desde diciembre de `anio_inicio` hasta hoy, cada uno una ventana rolling
    de `dias_ventana` días (mismo criterio "Continua (90d)" que el valor
    live de la card) -- a diferencia de la serie ANUAL de arriba
    (fetch_alineamiento_senadores_prov_serie, un punto por año con ventana
    de 366 días), esta da resolución mensual real para poder evaluar si las
    anclas del ITCP (PROVISIONAL, ver itcp.BANDAS_ITCP) siguen siendo
    razonables con más recorrido de datos. Reusa el detalle crudo por acta
    cacheado por año (_actas_alineamiento_cacheadas) -- una sola pasada de
    scraping por año, no una por mes. [[YYYY-MM-DD, valor]]."""
    cache = _actas_alineamiento_cacheadas(anio_inicio)
    detalle = [fila for detalle_anio in cache.values() for fila in detalle_anio]
    if not detalle:
        return []

    puntos = []
    for fin_mes in _fines_de_mes(date(anio_inicio, 12, 1), date.today(), incluir_hoy=True):
        referencia = datetime(fin_mes.year, fin_mes.month, fin_mes.day)
        resultado = politica._agregar_alineamiento_ventana(detalle, referencia, dias_ventana)
        if resultado:
            puntos.append([fin_mes.strftime("%Y-%m-%d"), resultado["valor"]])
    return puntos


ADHESION_REFORMAS_FECHAS_PATH = Path(__file__).resolve().parents[1] / "data" / "politica" / "adhesion_reformas_provincial_fechas.json"


def fetch_adhesion_reformas_provincial_serie() -> list:
    """Serie MENSUAL de adhesion_reformas_provincial: % de provincias (sobre
    24) que ya habían adherido al RIGI a cada fin de mes, reconstruida desde
    ADHESION_REFORMAS_FECHAS_PATH (fechas investigadas A MANO, 2026-07-09,
    ADR-0044) -- a diferencia de Diputados/Senado, no existe un endpoint
    único con fecha de adhesión por provincia: cada provincia adhirió con su
    propia ley, publicada en su propio Boletín Oficial provincial, así que
    esta serie no se puede reconstruir con un scraper genérico. Reemplaza el
    punto único que existía antes (el dato SÍ era un STOCK sin historia
    reconstruible -- dejó de serlo el día que se investigaron las fuentes
    provinciales una por una).

    Las provincias que aparecen HOY en la tabla MAGyP pero no tienen fecha
    investigada (adhesiones nuevas posteriores a esta investigación) NO
    entran al histórico -- solo al valor live de la card
    (politica.fetch_adhesion_reformas_provincial, que sigue releyendo la
    tabla MAGyP fresca en cada corrida). Por eso adhesion_reformas_provincial
    está en G3_EXCEPCIONES de gate_calidad.py: mientras todas las provincias
    adheridas tengan fecha conocida (el caso de hoy, 16/16) card y serie
    coinciden exacto; el día que aparezca una provincia nueva sin fecha
    investigada, van a dejar de coincidir hasta que se la investigue a mano
    y se agregue a ADHESION_REFORMAS_FECHAS_PATH.
    [[YYYY-MM-DD, % de provincias]]."""
    provincias_actuales = politica._provincias_adheridas_rigi()
    if not provincias_actuales:
        return []
    try:
        crudo = json.loads(ADHESION_REFORMAS_FECHAS_PATH.read_text(encoding="utf-8-sig"))
    except (OSError, json.JSONDecodeError):
        return []
    fechas = {k: v["fecha"] for k, v in crudo.items() if not k.startswith("_")}
    conocidas = {p: datetime.strptime(fechas[p], "%Y-%m-%d").date()
                 for p in provincias_actuales if p in fechas}
    desconocidas = provincias_actuales - set(fechas)
    if desconocidas:
        print(f"  [WARN] adhesion_reformas_provincial: sin fecha investigada para {sorted(desconocidas)} "
              "-- excluidas del histórico mensual, solo cuentan en el valor live de la card")
    if not conocidas:
        return []

    puntos = []
    desde = min(conocidas.values()).replace(day=1)
    for fin_mes in _fines_de_mes(desde, date.today()):
        cuenta = sum(1 for f in conocidas.values() if f <= fin_mes)
        puntos.append([fin_mes.strftime("%Y-%m-%d"), round(cuenta / 24.0 * 100.0, 1)])
    return puntos


def fetch_derrotas_legislativas_mensual() -> list:
    """Serie MENSUAL de derrotas_legislativas: para cada fin de mes desde
    dic-2023, cuántas derrotas legislativas consumadas (vetos insistidos por
    ambas cámaras + decretos rechazados en el recinto bajo la ley 26.122)
    acumulan los 12 meses calendario que terminan en ese mes — la MISMA
    ventana y conteo que la card (politica._derrotas_conteo_12m). Se deriva
    determinísticamente del registro versionado de eventos
    (data/politica/derrotas_legislativas_eventos.json, semilla verificada a
    mano + detección incremental de politica.fetch_derrotas_legislativas, que
    corre ANTES en el mismo pipeline nocturno): sin red, sin re-scrapear por
    mes. Backfill completo dic-2023→hoy disponible desde el día uno.
    [[YYYY-MM-DD, conteo]]."""
    registro = politica._cargar_derrotas_registro()
    if registro is None:
        print(f"  [WARN] derrotas_legislativas: registro de eventos ausente o ilegible "
              f"({politica.DERROTAS_EVENTOS_PATH}) -- serie omitida")
        return []
    eventos = politica._derrotas_eventos(registro)
    puntos = []
    for fin_mes in _fines_de_mes(date(2023, 12, 1), date.today(), incluir_hoy=True):
        total, _, _, _ = politica._derrotas_conteo_12m(eventos, fin_mes)
        puntos.append([fin_mes.strftime("%Y-%m-%d"), total])
    return puntos


def fetch_bloqueo_sostenido_mensual() -> list:
    """Serie MENSUAL de bloqueo_sostenido (ADR-0069): para cada fin de mes
    desde dic-2023, % de normas DESAFIADAS en el recinto (insistencias de
    veto votadas + decretos votados bajo la ley 26.122) en los 12 meses
    calendario que terminan en ese mes que seguían EN PIE al cierre — la
    MISMA ventana y regla que la card (politica._bloqueo_tasa_12m, que evalúa
    la caída al corte del mes histórico: un punto ya publicado no cambia
    retroactivamente porque la norma cayera después). Se deriva
    determinísticamente del registro versionado de eventos (el mismo de
    derrotas_legislativas, que fetch_bloqueo_sostenido actualiza ANTES en el
    pipeline nocturno): sin red, sin re-scrapear por mes. Los meses sin
    desafíos en ventana no generan punto (sin denominador no hay tasa: el
    primer desafío real es el DNU 70/2023 en mar-2024, así que la serie
    arranca ahí y el motor renormaliza antes — igual que veto_quorum entre
    períodos). Backfill completo dic-2023→hoy desde el día uno.
    [[YYYY-MM-DD, %]]."""
    registro = politica._cargar_derrotas_registro()
    if registro is None:
        print(f"  [WARN] bloqueo_sostenido: registro de eventos ausente o ilegible "
              f"({politica.DERROTAS_EVENTOS_PATH}) -- serie omitida")
        return []
    desafios = politica._bloqueo_desafios(registro)
    puntos = []
    for fin_mes in _fines_de_mes(date(2023, 12, 1), date.today(), incluir_hoy=True):
        tasa = politica._bloqueo_tasa_12m(desafios, fin_mes)
        if tasa is not None:
            puntos.append([fin_mes.strftime("%Y-%m-%d"), tasa[0]])
    return puntos


def fetch_desafios_legislativos_mensual() -> list:
    """Serie MENSUAL de desafios_legislativos (ADR-0089): conteo de normas del
    Ejecutivo DESAFIADAS en el recinto en los 12 meses calendario que terminan
    en cada mes, desde dic-2023.

    Sale del mismo `_bloqueo_tasa_12m` que la tasa de bloqueo —posición 1 de la
    tupla, donde la tasa es la 0— así que las dos series ven exactamente el
    mismo conjunto de eventos y la misma ventana. Que compartan la llamada no
    es economía de código: son el denominador y el numerador de la misma razón,
    y calcularlos por separado los dejaría inconsistentes entre sí.
    A diferencia de la tasa, acá el mes sin desafíos SÍ genera punto: cero
    desafíos es información (el Congreso no confrontó), no ausencia de dato.
    [[YYYY-MM-DD, conteo]]."""
    registro = politica._cargar_derrotas_registro()
    if registro is None:
        print(f"  [WARN] desafios_legislativos: registro de eventos ausente o ilegible "
              f"({politica.DERROTAS_EVENTOS_PATH}) -- serie omitida")
        return []
    desafios = politica._bloqueo_desafios(registro)
    puntos = []
    for fin_mes in _fines_de_mes(date(2023, 12, 1), date.today(), incluir_hoy=True):
        tasa = politica._bloqueo_tasa_12m(desafios, fin_mes)
        puntos.append([fin_mes.strftime("%Y-%m-%d"), float(tasa[1]) if tasa else 0.0])
    return puntos


def fetch_rotacion_gabinete_serie() -> list:
    """Serie MENSUAL de rotacion_gabinete: salidas de rango ministerial (JGM +
    ministros) acumuladas en la ventana móvil de 12 meses que termina en cada
    mes, desde dic-2023 (asunción) hasta el mes CORRIENTE inclusive —
    determinística desde el registro curado data/politica/gabinete_salidas.json
    (misma familia que adhesion_reformas_provincial_fechas.json: dataset
    curado y versionado, sin red). Los movimientos laterales dentro del
    gabinete y las reestructuraciones de la Ley de Ministerios no cuentan
    (viven en claves aparte del registro, ver su _meta).

    El mes corriente SÍ se incluye (fechado al día 1, como las series INDEC):
    la ventana de 12 meses calendario que termina en el mes en curso es
    exactamente la métrica del valor live de la card
    (politica.fetch_rotacion_gabinete), así que card y serie[-1] coinciden
    por construcción y el G3 del gate no necesita excepción — un no-evento en
    lo que va del mes es dato, no un parcial. [[YYYY-MM-01, salidas 12m]]."""
    registro = politica.cargar_gabinete_salidas()
    if registro is None or not isinstance(registro.get("salidas"), list):
        print("[WARN] rotacion_gabinete_serie: registro curado ausente o ilegible")
        return []
    salidas = registro["salidas"]
    puntos = []
    anio, mes = 2023, 12
    hoy = date.today()
    while (anio, mes) <= (hoy.year, hoy.month):
        ym = f"{anio}-{mes:02d}"
        puntos.append([f"{ym}-01", len(politica.salidas_gabinete_ventana_12m(salidas, ym))])
        anio, mes = (anio + 1, 1) if mes == 12 else (anio, mes + 1)
    return puntos


def fetch_cepa_movilizacion_serie(max_paginas: int = 40) -> list:
    """Backfill histórico de movilizacion_cepa: escanea hasta `max_paginas`
    páginas de centrocepa.com.ar/documentos/informes (10 informes por
    página) buscando TODOS los links con "conflictividad"/"conflictos-laborales"
    en la URL -- no solo el más reciente, a diferencia de
    politica.fetch_cepa_movilizacion(). Verificado en vivo (2026-07-08): con
    40 páginas se cubren de sobra los ~4 informes de este tipo publicados
    hasta la fecha (CEPA recién empezó a publicarlos a fines de 2025 -- no
    hay nada más atrás que buscar). Reusa politica._extraer_cifra_cepa /
    politica._fecha_informe_cepa, pero NO se limita a saltear los informes
    donde _extraer_cifra_cepa devuelve None (ausencia de cualquier patrón
    conocido) -- también filtra por el campo "rama" que esa función expone
    (hallazgo en vivo 2026-07-08): el informe 748 ("2 años de Milei") acumula
    desde ene-2024, sin la frase-gatillo "desde inicios del año en curso" que
    usa el indicador vigente, PERO su cuerpo también trae una oración con un
    promedio mensual ("promediando 24 casos por mes") que sí matchea la rama
    m_mes de _extraer_cifra_cepa (rama anterior a esta tarea, no se modifica
    acá) -- por lo tanto NO devuelve None, devuelve {"rama": "m_mes", ...}.
    Esa cifra es una TASA mensual sobre una ventana ene24-sep25, mientras que
    el indicador vigente (rama m_tot, "acumulados desde inicios del año en
    curso") es un CONTEO acumulado -- dos escalas incompatibles
    (CEPA_MAX_CASOS_MES=80 vs. CEPA_MAX_CONFLICTOS_TOT=200) que no deben
    mezclarse en la misma serie. Por eso acá, a diferencia de
    fetch_cepa_movilizacion() (que no discrimina entre ramas porque solo ve
    UN informe a la vez y ese informe siempre viene de m_tot en la
    práctica), se descartan explícitamente las lecturas rama="m_mes": solo
    se conservan lecturas rama="m_tot" ("conflictos acumulados"), homogéneas
    con el valor vigente publicado hoy (50.5 = 101/200*100, informe 809).
    [[YYYY-MM-DD, índice 0-100]]."""
    try:
        from bs4 import BeautifulSoup
    except ImportError:
        print("[WARN] movilizacion_cepa_serie: beautifulsoup4 no disponible")
        return []

    links = []
    for page in range(max_paginas):
        page_url = politica.CEPA_INFORMES_URL if page == 0 else f"{politica.CEPA_INFORMES_URL}?start={page * 10}"
        try:
            r = requests.get(page_url, headers=HTTP_HEADERS, timeout=HTTP_TIMEOUT)
            r.raise_for_status()
        except requests.RequestException:
            break
        soup = BeautifulSoup(r.text, "html.parser")
        for a in soup.find_all("a", href=True):
            href_lower = a["href"].lower()
            if any(kw in href_lower for kw in ("conflictividad", "conflictos-laborales")):
                links.append(a["href"])

    out = []
    vistos = set()
    for href in links:
        if href in vistos:
            continue
        vistos.add(href)
        informe_url = ("https://centrocepa.com.ar" + href) if href.startswith("/") else href
        try:
            r2 = requests.get(informe_url, headers=HTTP_HEADERS, timeout=HTTP_TIMEOUT)
            r2.raise_for_status()
        except requests.RequestException as e:
            print(f"[WARN] movilizacion_cepa_serie: {informe_url}: {e}")
            continue
        cifra_info = politica._extraer_cifra_cepa(r2.text)
        if cifra_info is None or cifra_info["rama"] != "m_tot":
            continue   # sin match, o tasa mensual (m_mes) -- no comparable con el conteo acumulado, ver docstring de _extraer_cifra_cepa
        fecha = politica._fecha_informe_cepa(r2.text)
        out.append([fecha, cifra_info["valor"]])

    out.sort(key=lambda x: x[0])
    return out


def fetch_conflictividad_nacional_mensual() -> list:
    """Serie mensual de conflictividad_nacional (ADR-0052): % de variación
    del acumulado 12 meses de eventos Protests+Riots de ACLED en TODO el
    país contra el total 2023, mes a mes desde dic-2023 (primera ventana
    12m íntegramente post-asunción, comparable con la base). Lee la serie
    "mensual_nacional" del store que llena gestion.actualizar_protestas_caba()
    (una sola descarga de ~8 MB por corrida, memo por proceso en gestion).
    El último mes se EXCLUYE si el archivo ACLED no llega a fin de mes —
    misma regla que politica.fetch_conflictividad_nacional(), así el último
    punto de la serie coincide con la card (G3). No se emiten puntos
    pre-dic-2023: la cobertura ACLED pre-2020 no es confiable y 2023 es la
    base, no un punto de la curva. [[YYYY-MM-01, % vs 2023]]."""
    import json as _json
    if not gestion.PROTESTAS_STORE_PATH.exists():
        return []
    store = _json.loads(gestion.PROTESTAS_STORE_PATH.read_text(encoding="utf-8"))
    mensual = store.get("mensual_nacional", {})
    if not mensual:
        return []
    hasta = store.get("_meta", {}).get("hasta_semana", "")
    yms = sorted(mensual)
    if hasta and yms and hasta[:7] == yms[-1]:
        a, m = int(hasta[:4]), int(hasta[5:7])
        if int(hasta[8:10]) < calendar.monthrange(a, m)[1]:
            yms = yms[:-1]
    base_2023 = sum(v for ym, v in mensual.items() if ym.startswith("2023"))
    if not base_2023:
        return []
    out = []
    for i, ym in enumerate(yms):
        if ym < "2023-12" or i < 11:
            continue
        acum = sum(mensual[y] for y in yms[i - 11:i + 1])
        out.append([f"{ym}-01", round((acum / base_2023 - 1.0) * 100.0, 1)])
    return out


POLITICA_DERIVADAS = [
    ("votometro_ventaja_lla", "pp (brecha LLA−PJ)", "Votómetro CIGOB", fetch_votometro_serie),
    ("iaf_transferencias", "% i.a. real", "RON Hacienda + IPC INDEC (dic-dic)", fetch_iaf_serie),
    ("ratio_dnu", "DNUs por ley (12m móviles)", "InfoLeg", fetch_ratio_dnu_serie),
    ("desafios_legislativos", "normas desafiadas en el recinto (12m)",
     "Actas de Diputados y Senado + InfoLeg — elaboración CIGOB",
     fetch_desafios_legislativos_mensual),
    # La serie sale del MISMO cálculo que la card (politica.cobertura_judicial_serie):
    # el padrón ancla el nivel y los registros de designaciones y renuncias lo
    # mueven mes a mes, así que card y serie no pueden divergir (ADR-0126).
    # ADR-0168. Las cuatro salen del MISMO origen que su card, así que card y
    # serie no pueden divergir (el patrón contrario causó ADR-0086 y ADR-0087).
    # produccion_legislativa la recalcula el colector contra HCDN; las otras
    # tres leen el relevamiento versionado, que es también lo que lee la card.
    ("produccion_legislativa", "leyes sancionadas (12m móviles)",
     "Cámara de Diputados — dataset de leyes sancionadas",
     lambda: [[f"{ym}-01", v]
              for ym, v in sorted(politica.produccion_legislativa_serie().items())]),
    ("paralisis_denuncias", "sesiones de las comisiones de control (12m móviles)",
     "Consejo de la Magistratura — archivo de notas de las comisiones",
     lambda: [[f"{ym}-01", v]
              for ym, v in sorted(politica.paralisis_denuncias_serie().items())]),
    # Las dos anuales se publican con el punto en enero del año que describen:
    # el rezago declarado en REZAGO_MESES_ITCP (9 y 12 meses) es el que avisa
    # que estas dos no son pulso de hoy (ADR-0092).
    ("judicializacion", "% de sumarios con medida cautelar (Federal + Nacional)",
     "SAIJ — buscador de jurisprudencia",
     lambda: [[f"{a}-01-01", v]
              for a, v in sorted(politica.judicializacion_serie().items())]),
    ("velocidad_resolucion", "% de expedientes resueltos sobre ingresados",
     "CSJN — anuario estadístico",
     lambda: [[f"{a}-12-31", v] for a, v in sorted(
         politica._leer_store(politica.CSJN_FUENTES_PATH)["velocidad_de_resolucion"]
         ["serie_historica_completa"]["tasa_resolucion_pct"].items())]),
    ("cobertura_judicial", "% de cargos de juez con juez designado",
     "Ministerio de Justicia — padrón, designaciones y renuncias (datos.jus.gob.ar)",
     lambda: [[f"{ym}-01", v]
              for ym, v in sorted(politica.cobertura_judicial_serie()[0].items())]),
    # La serie la calcula el propio colector: la card de politica.py devuelve el
    # último punto de esta misma lista, así que no pueden divergir (ADR-0088;
    # el patrón contrario causó ADR-0086 y ADR-0087 el día anterior).
    ("brecha_obra_publica", "pp (obra pública − privada, 12m móviles)",
     "INDEC · Encuesta Cualitativa de la Construcción (ISAC, Cuadro 7.1)",
     politica.brecha_obra_publica_serie),
    # Mismo patrón: la card es el último punto de esta lista (ADR-0150).
    ("apoyo_empresario", "saldo de postura (−1 a +1, 12m móviles)",
     "Comunicados de AEA y UIA — codificación CIGOB",
     politica.apoyo_empresario_serie),
    ("eficacia_legislativa", "% proyectos PE aprobados (12m móviles)", "Cámara de Diputados (datos abiertos)", fetch_eficacia_serie),
    ("veto_quorum", "% sesiones en minoría (12m móviles)", "Cámara de Diputados (datos abiertos)", fetch_veto_quorum_serie),
    ("comisiones_caidas", "% con dictamen sin sanción (12m móviles)", "Cámara de Diputados (datos abiertos)", fetch_comisiones_serie),
    ("derrotas_legislativas", "derrotas del Ejecutivo en el recinto (12m móviles)",
     "InfoLeg + actas del Senado — elaboración CIGOB",
     fetch_derrotas_legislativas_mensual),
    ("bloqueo_sostenido", "% normas desafiadas en el recinto que siguen en pie (12m móviles)",
     "Actas de Diputados y Senado + InfoLeg — elaboración CIGOB",
     fetch_bloqueo_sostenido_mensual),
    # cohesion_bloque es el COMPUESTO bicameral desde 2026-07-10 (ADR-0048):
    # las dos series mensuales por cámara (fetch_cohesion_bloque_diputados_mensual
    # y fetch_cohesion_bloque_senado_mensual, cada una con su caché permanente
    # por acta — ver ADR-0040/0041 para por qué eso importa, ~53min de
    # diferencia) siguen siendo los insumos, pero al CSV va una sola serie
    # 65/35 — la card del Senado dejó de existir como indicador propio, así
    # que su serie separada se purga del CSV en la primera corrida sin merge.
    ("cohesion_bloque", "% cohesión (índice de Rice bicameral 65/35)",
     "Votaciones nominales de Diputados y Senado — elaboración CIGOB",
     fetch_cohesion_bloque_compuesta_mensual),
    ("adhesion_reformas_provincial", "% de jurisdicciones (sobre 24) adheridas al RIGI",
     "Tabla de provincias adheridas — Ministerio de Agricultura, Ganadería y Pesca",
     fetch_adhesion_reformas_provincial_serie),
    ("rotacion_gabinete", "salidas de rango ministerial (acum. 12 meses)",
     "Decretos de designación y renuncia — Boletín Oficial (registro curado CIGOB)",
     fetch_rotacion_gabinete_serie),
    # conflictividad_nacional (ADR-0052): la pata puntuante de
    # conflicto_social — ACLED país entero, misma descarga que protestas.
    ("conflictividad_nacional", "% var. eventos de protesta y disturbios en el país vs 2023 (12m móviles)",
     "ACLED — agregado semanal por provincia (elaboración CIGOB)",
     fetch_conflictividad_nacional_mensual),
    # movilizacion_cepa: seguimiento interno desde 2026-07-11 (ADR-0052) —
    # la serie se sigue guardando como contraste, igual que rotacion_gabinete.
    ("movilizacion_cepa", "Índice de conflictividad social (0-100)",
     "Centro CEPA — informes de conflictividad (elaboración CIGOB)",
     fetch_cepa_movilizacion_serie),
    ("alineamiento_senadores_prov", "% votos de senadores no-LLA alineados con LLA",
     "Votaciones nominales del Senado — elaboración CIGOB",
     fetch_alineamiento_senadores_prov_mensual),
    # protestas_caba NO se registra acá: ya está en GESTION_DERIVADAS
    # (fetch_protestas_serie) y build_series() en publicar.py fusiona TODOS
    # los CSV de output/series/ en un único dict keyed por indicador — la
    # clave "protestas_caba" ya queda disponible para el ITCP de política sin
    # duplicar la descarga (~8 MB de ACLED) ni la lógica de scraping.
]

VIDA_INDEC = [
    # NIVEL del IPC bajo clave propia: es un INSUMO (deflactor del crédito de
    # consumo en publicar), no la serie de la card ipc_total (que publica el
    # % m/m desde macro) — con la misma clave, la fusión de CSVs las mezclaba.
    ("148.3_INIVELNAL_DICI_M_26", "ipc_nivel",    "indice base dic-2016=100", "INDEC/datos.gob.ar"),
    ("42.3_EPH_PUNTUATAL_0_M_30", "desocupacion", "%",                        "INDEC/datos.gob.ar"),
]
# Derivadas: la variación m/m % de cada serie INDEC de índices (misma métrica que
# muestra la card del indicador), reconstruida para todo el histórico disponible.
VIDA_DERIVADAS = [
    ("ipc_alimentos",    "% m/m",           "INDEC serie 146.3", lambda: fetch_indec_var_mensual("146.3_IALIMENNAL_DICI_M_45")),
    ("peso_tarifas",     "% m/m regulados", "INDEC serie 148.3", lambda: fetch_indec_var_mensual("148.3_IREGULANAL_DICI_M_22")),
]
def fetch_icc_serie(meses: int = 60) -> list:
    """Serie histórica del ICC UTDT: parsea TODAS las filas del XLS oficial (col 0 fecha,
    col 1 índice), no solo la última como el indicador. Reusa el scraper del colector de
    vida. Devuelve los últimos `meses` como [[YYYY-MM-01, icc]] ascendente."""
    import xlrd
    sys.path.insert(0, str(Path(__file__).parent / "vida_cotidiana"))
    sys.path.insert(0, str(Path(__file__).parent / "vida_cotidiana" / "collectors"))
    from utdt_icc import _get_latest_xls_fname
    from config import UTDT_ICC_DOWNLOAD_BASE
    r = requests.get(UTDT_ICC_DOWNLOAD_BASE + _get_latest_xls_fname(),
                     headers=HTTP_HEADERS, timeout=HTTP_TIMEOUT, verify=False)
    r.raise_for_status()
    wb = xlrd.open_workbook(file_contents=r.content)
    ws = wb.sheets()[0]
    out = []
    for i in range(ws.nrows):
        fc, vc = ws.cell(i, 0), ws.cell(i, 1)
        if fc.ctype == xlrd.XL_CELL_DATE and vc.ctype == xlrd.XL_CELL_NUMBER:
            t = xlrd.xldate_as_tuple(fc.value, wb.datemode)
            out.append([f"{t[0]}-{t[1]:02d}-01", round(vc.value, 1)])
    return out[-meses:]


VIDA_DERIVADAS.append(
    ("icc_utdt", "índice", "UTDT (ICC, serie XLS)", fetch_icc_serie)
)


def _utdt_xls(listado_url):
    """Contenido del XLS más reciente de una página de serie histórica de la UTDT.

    Todas siguen el mismo patrón —el listado enlaza el archivo por `fname`— así
    que el scraper del ICC sirve para el resto con sólo cambiar la URL."""
    r = requests.get(listado_url, headers=HTTP_HEADERS, timeout=HTTP_TIMEOUT, verify=False)
    r.raise_for_status()
    fnames = sorted(set(re.findall(r'download\.php\?fname=([^"\'&]+)', r.text)))
    if not fnames:
        raise ValueError(f"UTDT: sin link de descarga en {listado_url}")
    x = requests.get("https://www.utdt.edu/download.php?fname=" + fnames[0],
                     headers=HTTP_HEADERS, timeout=HTTP_TIMEOUT * 2, verify=False)
    x.raise_for_status()
    return x.content


def _utdt_niveles(listado_url) -> dict:
    """{YYYY-MM: nivel} del XLS de una serie histórica de la UTDT.

    Todas las planillas comparten el layout (col 0 fecha, col 1 valor), así que
    el parseo va una sola vez: había tres copias del mismo loop en este archivo
    y de ahí salían las divergencias entre la card y su serie."""
    import xlrd
    wb = xlrd.open_workbook(file_contents=_utdt_xls(listado_url))
    ws = wb.sheets()[0]
    out = {}
    for i in range(ws.nrows):
        fc, vc = ws.cell(i, 0), ws.cell(i, 1)
        if fc.ctype == xlrd.XL_CELL_DATE and vc.ctype == xlrd.XL_CELL_NUMBER:
            t = xlrd.xldate_as_tuple(fc.value, wb.datemode)
            out[f"{t[0]}-{t[1]:02d}"] = vc.value
    return out


def fetch_indice_lider_serie(meses: int = 60) -> list:
    """NIVEL del Índice Líder de la UTDT — el mismo número que publica la card.

    Segundo caso del mismo defecto que `alquiler_real` (auditoría de UI
    29-jul-2026): la card puntúa —13% de la dimensión empleo del ITVC— y su
    serie tenía un solo punto porque nunca se registró acá. El componente
    rebaseado (`itvc_lider`) sí tenía historia, pero es otra escala: 100 =
    4T-2023, no el nivel que muestra la card.

    Redondeo a 1 decimal, igual que la card en publicar.py, para que G3 cierre.

    Acotada a `meses` como la del ICC: el XLS de la UTDT arranca en 1993 y son
    402 puntos. No es que sobre historia, es que la sparkline de la card
    comprime 33 años y el movimiento reciente —lo que la card informa— queda
    invisible. 60 meses quedan muy por encima del piso de dic-2023.
    [[YYYY-MM-01, índice]]."""
    sys.path.insert(0, str(Path(__file__).parent / "vida_cotidiana"))
    from config import UTDT_IL_LISTADO
    niveles = sorted(_utdt_niveles(UTDT_IL_LISTADO).items())
    return [[f"{ym}-01", round(v, 1)] for ym, v in niveles][-meses:]


def fetch_itvc_lider() -> list:
    """I_IL (ADR-0112): Índice Líder de la UTDT rebaseado a 4T-2023.

    Es el único componente PROSPECTIVO del cinturón: está construido para
    anticipar puntos de giro del ciclo, mientras el resto describe lo que ya
    pasó. Se rebasea igual que los demás (100 = 4T-2023) y NO se invierte:
    un líder más alto anticipa mejor actividad, que es mejora.
    [[YYYY-MM-01, índice]]."""
    sys.path.insert(0, str(Path(__file__).parent / "vida_cotidiana"))
    from config import UTDT_IL_LISTADO
    return _itvc_rebase(_utdt_niveles(UTDT_IL_LISTADO))


VIDA_DERIVADAS.append(
    ("itvc_lider", "índice (100 = 4T-2023)", "UTDT — Índice Líder (serie XLS)", fetch_itvc_lider)
)
# Unidad y fuente COPIADAS de la card (publicar.py)
VIDA_DERIVADAS.append(
    ("indice_lider", "índice", "UTDT — Índice Líder (CIF)", fetch_indice_lider_serie)
)
VIDA_DERIVADAS.append(
    ("pobreza_nowcast", "% de personas", "UTDT — Nowcast de Pobreza (informes PDF)",
     fetch_pobreza_nowcast_serie)
)
VIDA_DERIVADAS.append(
    ("pobreza_indec", "% de personas", "INDEC — EPH continua (vía datos.gob.ar)",
     fetch_pobreza_indec_serie)
)
# Componente del ITVC: pobreza rebaseada a 4T-2023 e invertida. Va DESPUES de los
# dos fetchers de pobreza porque los usa a los dos (nivel del nowcast, base del INDEC).
VIDA_DERIVADAS.append(
    ("itvc_pobreza", "índice (100 = 4T-2023)",
     "UTDT — Nowcast de Pobreza (nivel) + INDEC EPH (base 2º sem. 2023)", fetch_itvc_pobreza)
)


SENTIMIENTO_SERIE_STORE = Path(__file__).resolve().parents[1] / "data" / "vida" / "sentimiento_serie.json"


def fetch_sentimiento_serie() -> list:
    """Serie MENSUAL del sentimiento digital (ADR-0034 + ADR-0222): canasta de
    los 6 términos en VENTANA FIJA 2021→hoy, resolución mensual nativa.

    El cálculo vive en el colector (`vida_cotidiana/collectors/trends.py`) y no
    acá: la card y la serie salen del MISMO store, así que no pueden discrepar
    ni gastar dos rondas de pedidos contra una fuente con rate limit. Cada
    término se consulta solo y se rebasa contra su propio 4T-2023 dentro de esa
    consulta —el escalar de Trends se cancela en el cociente—, y la canasta es
    el promedio simple de los seis índices. El mes en curso se descarta.
    [[YYYY-MM-01, índice base 100 = 4T-2023]]."""
    sys.path.insert(0, str(Path(__file__).parent / "vida_cotidiana"))
    sys.path.insert(0, str(Path(__file__).parent / "vida_cotidiana" / "collectors"))
    import trends as _t
    store = _t.fetch_sentimiento_store(SENTIMIENTO_SERIE_STORE)
    mensual = store.get("mensual") or {}
    if not mensual:
        # Se levanta a propósito en vez de devolver []: `_correr` trata el error
        # como fuente caída y `_filas_previas` conserva las filas que el CSV ya
        # tenía. Devolviendo una lista vacía, la escritura completa borraría la
        # serie del gráfico sin que nada avisara — que es el modo de falla que
        # documenta `_filas_previas`, en este mismo archivo.
        raise ValueError("Trends sin datos y sin store utilizable "
                         f"(formato {_t.SENTIMIENTO_ESQUEMA})")
    viejos = [kw for kw, t in store.get("terminos", {}).items()
              if t.get("actualizado") != datetime.today().strftime("%Y-%m-%d")]
    if viejos:
        print(f"  [WARN] sentimiento: {len(viejos)} término(s) del store previo "
              f"({', '.join(viejos)}); la canasta llega hasta {max(mensual)}")
    return [[f"{ym}-01", v] for ym, v in sorted(mensual.items())]


VIDA_DERIVADAS.append(
    ("sentimiento_digital", "índice (100 = 4T-2023)",
     "Google Trends (6 términos, ventana fija 2021→, ADR-0222)", fetch_sentimiento_serie)
)


# La serie de indice_intencion_migratoria se descargaba acá hasta ADR-0205.
# El indicador se retiró con el cinturón de espíritu de época, así que el
# nocturno dejó de bajarla: lo ya recolectado queda congelado en
# data/vida/intencion_migratoria_serie.json.


def fetch_endeudamiento_serie() -> list:
    """Serie del stock nominal de crédito de consumo (personales 114 + tarjeta 115, BCRA)
    en billones de pesos — mismo valor que el headline del indicador (la variación real
    i.a. que puntúa se muestra aparte en el box de score). [[YYYY-MM-01, billones]]."""
    def mensual(vid):
        out = {}
        for f, v in sorted(fetch_bcra(vid, dias=1300)):
            out[f[:7]] = v
        return out
    per, tar = mensual(114), mensual(115)
    return [[f"{ym}-01", round((per[ym] + tar[ym]) / 1e6, 2)]
            for ym in sorted(per) if ym in tar]


VIDA_DERIVADAS.append(
    ("endeudamiento_familiar", "billones de pesos (consumo)", "BCRA API v4.0 (personales + tarjeta)", fetch_endeudamiento_serie)
)

VIDA_DERIVADAS += [
    ("informalidad", "%", "INDEC EPH (52.2, trimestral)", lambda: fetch_indec_x100("52.2_ASDJ_0_0_37")),
    ("pluriempleo", "%", "INDEC EPH (47.2, trimestral)", lambda: fetch_indec_x100("47.2_ECTSDT_0_T_47")),
    # Empleo registrado privado (ADR-0130): asalariados del sector privado
    # declarados al SIPA, miles de personas, mensual. El rebase B100 contra el
    # 4T-2023 lo hace publicar.py, igual que el resto de los componentes.
    ("empleo_registrado", "miles de puestos",
     "Min. de Capital Humano — SIPA (vía datos.gob.ar)",
     lambda: [[f, round(v, 1)] for f, v in
              sorted(fetch_indec("151.1_AARIADODAD_2012_M_31", limit=200))
              if f >= "2023-10"]),
]


# ── Series ITVC-B100 (vida cotidiana, doc 260702 — ADR-0018) ──────────────────
# Componentes del ITVC que requieren TRANSFORMACIÓN (nivel relativo al salario,
# nivel desestacionalizado rebaseado, deuda real): se emiten como series YA
# rebaseadas a 100 = promedio 4T-2023 (oct-nov-dic), la línea base del doc.
# La base se calcula DINÁMICAMENTE de la misma serie (nunca hardcodeada).

ITVC_RIPTE_ID      = "158.1_REPTE_0_0_5"                # RIPTE mensual ($)
ITVC_IPI_DESEST_ID = "453.1_SERIE_DESEADA_0_0_24_58"    # IPI nivel desestacionalizado
ITVC_ISAC_DESEST_ID = "33.2_ISAC_SIN_EDAD_0_M_23_56"    # ISAC nivel desestacionalizado
ITVC_BASE_MESES = ("2023-10", "2023-11", "2023-12")


def _nivel_mensual(sid: str, limit: int = 60) -> dict:
    """{YYYY-MM: valor} de una serie mensual de datos.gob.ar (vía fetch_indec)."""
    return {f[:7]: v for f, v in fetch_indec(sid, limit=limit)}


def _base_t423(serie: dict) -> float:
    """Promedio del 4T-2023 de {YYYY-MM: valor}; exige los tres meses."""
    faltan = [m for m in ITVC_BASE_MESES if m not in serie]
    if faltan:
        raise ValueError(f"base 4T-2023 incompleta (faltan {faltan})")
    return sum(serie[m] for m in ITVC_BASE_MESES) / 3.0


def _itvc_rebase(serie: dict, invertido: bool = False) -> list:
    """Serie {ym: valor} → [[YYYY-MM-01, índice]] con 100 = promedio 4T-2023.
    invertido=True para métricas 'más alto = peor' (I = base/valor)."""
    base = _base_t423(serie)
    out = []
    for ym in sorted(serie):
        if ym < "2023-10" or not serie[ym]:
            continue
        indice = (base / serie[ym] if invertido else serie[ym] / base) * 100.0
        out.append([f"{ym}-01", round(indice, 1)])
    return out


def _itvc_relativo_salario(sid_precio: str) -> list:
    """Componente de precios del ITVC: nivel del índice de precios RELATIVO al
    salario (RIPTE), rebaseado. I(t) = (P_base/P_t) × (RIPTE_t/RIPTE_base) × 100:
    >100 = los precios subieron MENOS que los salarios desde el 4T-2023."""
    precio = _nivel_mensual(sid_precio)
    ripte  = _nivel_mensual(ITVC_RIPTE_ID)
    p_base, r_base = _base_t423(precio), _base_t423(ripte)
    out = []
    for ym in sorted(set(precio) & set(ripte)):
        if ym < "2023-10" or not precio[ym] or not ripte[ym]:
            continue
        indice = (p_base / precio[ym]) * (ripte[ym] / r_base) * 100.0
        out.append([f"{ym}-01", round(indice, 1)])
    return out


def fetch_itvc_alimentos() -> list:
    """I_IA REDISEÑADO (ADR-0033): encarecimiento RELATIVO de la comida —
    IPC alimentos contra el IPC general, rebaseado a 4T-2023. >100 = la comida
    sube MENOS que el resto de los precios (alivia la canasta de los hogares
    de menores ingresos); <100 = la comida encarece por encima del promedio.
    La versión anterior (IPC alimentos vs RIPTE) correlacionaba r=0,985 con la
    brecha salario/CBT: un tercio del ITVC contaba dos veces el mismo ratio
    salario/comida. Esta métrica es la pregunta de PRECIOS pura, independiente
    del salario. [[YYYY-MM-01, índice]]."""
    sys.path.insert(0, str(Path(__file__).parent / "vida_cotidiana"))
    from config import INDEC_SERIES
    alim = _nivel_mensual(INDEC_SERIES["ipc_alimentos"])
    gen = _nivel_mensual("148.3_INIVELNAL_DICI_M_26")
    a_base, g_base = _base_t423(alim), _base_t423(gen)
    out = []
    for ym in sorted(set(alim) & set(gen)):
        if ym < "2023-10" or not alim[ym] or not gen[ym]:
            continue
        indice = (a_base / alim[ym]) * (gen[ym] / g_base) * 100.0
        out.append([f"{ym}-01", round(indice, 1)])
    return out


def fetch_alquiler_real_serie() -> list:
    """% m/m del IPC-GBA «alquiler de la vivienda» — la misma cuenta que la card.

    La card era la ÚNICA de las 25 de la home sin sparkline, y no era un problema
    de UI: `alquiler_real` nunca se registró acá, así que su serie tenía un solo
    punto —el valor del día— cuando la regla del proyecto pide backfill desde
    dic-2023 (auditoría de UI 29-jul-2026). El componente del ITVC
    (`itvc_alquiler`) sí tenía historia, pero **mide otra cosa**: encarecimiento
    relativo al nivel general de GBA, no la variación mensual del alquiler. No
    son intercambiables.

    Se deriva del NIVEL con `_var_mensual`, que exige meses CONSECUTIVOS, en vez
    de dividir posiciones adyacentes de la lista descargada: si INDEC saltea un
    mes, el cociente entre dos observaciones separadas por dos meses se
    publicaría como variación mensual sin que nada avise (el mismo error que
    corrigió la auditoría del 18-jul-2026).

    Misma fórmula y mismo redondeo que la card, para que G3 cierre:
    `(idx_t / idx_{t-1} − 1) × 100` a dos decimales.
    [[YYYY-MM-01, % m/m]]."""
    sys.path.insert(0, str(Path(__file__).parent / "vida_cotidiana"))
    from config import INDEC_SERIES
    niveles = _nivel_mensual(INDEC_SERIES["ipc_alquiler_gba"])
    return [[f"{ym}-01", v] for ym, v in sorted(_var_mensual(niveles).items())]


def fetch_itvc_alquiler() -> list:
    """I_AL (ADR-0111): encarecimiento RELATIVO del alquiler — IPC-GBA alquiler
    de la vivienda contra el nivel general de GBA, rebaseado a 4T-2023.
    >100 = el alquiler sube MENOS que el resto de los precios; <100 = encarece
    por encima del promedio.

    Misma construcción que `fetch_itvc_alimentos`: es la pregunta de precios
    pura, independiente del salario, para no repetir el ratio que ya mide la
    brecha salario/CBT.

    Deflactado con el nivel general de GBA y no con el nacional: la única
    apertura de alquiler que publica INDEC es la de GBA, y dividir un precio de
    GBA por un índice nacional mezclaría dos plazas en el mismo cociente.
    [[YYYY-MM-01, índice]]."""
    sys.path.insert(0, str(Path(__file__).parent / "vida_cotidiana"))
    from config import INDEC_SERIES
    alq = _nivel_mensual(INDEC_SERIES["ipc_alquiler_gba"])
    gen = _nivel_mensual(INDEC_SERIES["ipc_gba_general"])
    a_base, g_base = _base_t423(alq), _base_t423(gen)
    out = []
    for ym in sorted(set(alq) & set(gen)):
        if ym < "2023-10" or not alq[ym] or not gen[ym]:
            continue
        indice = (a_base / alq[ym]) * (gen[ym] / g_base) * 100.0
        out.append([f"{ym}-01", round(indice, 1)])
    return out


def fetch_itvc_tarifas() -> list:
    """I_PT: peso de los servicios regulados en el salario (IPC Regulados nivel
    vs RIPTE), 100 = 4T-2023."""
    sys.path.insert(0, str(Path(__file__).parent / "vida_cotidiana"))
    from config import INDEC_SERIES
    return _itvc_relativo_salario(INDEC_SERIES["ipc_regulados"])


def fetch_trabajo_independiente_serie() -> list:
    """Participación del trabajo independiente en el empleo registrado, en %.
    Misma serie para la card y para el índice (ADR-0219)."""
    sys.path.insert(0, str(Path(__file__).parent / "vida_cotidiana"))
    from collectors.trabajo_independiente import fetch_trabajo_independiente
    d = fetch_trabajo_independiente()
    return [[f"{ym}-01", v] for ym, v in sorted(d["serie"].items())
            if ym >= "2019-01"]


def fetch_empleadores_pyme_serie() -> list:
    """Cantidad mensual de empleadores de hasta 50 trabajadores con cobertura
    de ART (SRT). Es la MISMA serie que alimenta la card y el índice: el
    rebase a 100 = 4T-2023 lo hace `itvc.rebase_de_serie`, así que no hay dos
    números distintos para lo mismo (ADR-0218)."""
    sys.path.insert(0, str(Path(__file__).parent / "vida_cotidiana"))
    from collectors.srt_empleadores import fetch_empleadores_pyme
    d = fetch_empleadores_pyme()
    return [[f"{ym}-01", v] for ym, v in sorted(d["serie_pyme"].items())
            if ym >= "2019-01"]


def fetch_itvc_ipi() -> list:
    """I_IPI: nivel del IPI manufacturero DESESTACIONALIZADO, 100 = 4T-2023."""
    return _itvc_rebase(_nivel_mensual(ITVC_IPI_DESEST_ID))


def fetch_itvc_isac() -> list:
    """I_ISC: nivel del ISAC DESESTACIONALIZADO, 100 = 4T-2023 (la serie
    original tiene un desplome estacional en dic-23 que contaminaría la base)."""
    return _itvc_rebase(_nivel_mensual(ITVC_ISAC_DESEST_ID))


BCRA_INF_BANCOS_ANEXO = ("https://www.bcra.gob.ar/archivos/Pdfs/"
                         "PublicacionesEstadisticas/informes/InfBanc_Anexo.xlsx")


def _anexo_bancos_familias() -> tuple:
    """(deuda_consumo {ym: saldo M$}, mora {ym: % ponderado}) del corte FAMILIAS
    del anexo del Informe sobre Bancos (hoja 'Calidad de Cartera (por líneas)',
    personales + tarjetas: saldos e irregularidad, mensual 2010→). El layout se
    detecta por etiquetas (no por número de fila): cada bloque tiene una fila de
    fechas seguida de filas etiquetadas; el bloque de ratios tiene valores <100
    y el de saldos, millones."""
    import openpyxl
    r = requests.get(BCRA_INF_BANCOS_ANEXO, headers=HTTP_HEADERS,
                     timeout=HTTP_TIMEOUT * 3, verify=False)
    r.raise_for_status()
    wb = openpyxl.load_workbook(io.BytesIO(r.content), read_only=True, data_only=True)
    hoja = next(s for s in wb.sheetnames if "por l" in s.lower())   # "(por líneas)"
    filas = list(wb[hoja].iter_rows(values_only=True))

    def _es_fecha(c):
        return hasattr(c, "year") and hasattr(c, "month")

    # Anclado en el título de sección: "2. Familias - Total" → los DOS bloques
    # de fechas siguientes son el ratio de irregularidad y el saldo (en ese
    # orden, cada uno con filas etiquetadas Personales / Tarjetas de crédito).
    inicio = next(i for i, f in enumerate(filas)
                  if str((f or [""])[0] or "").strip().lower().startswith("2. familias"))
    bloques = []
    i = inicio + 1
    while i < len(filas) and len(bloques) < 2:
        f = filas[i]
        if sum(1 for c in f if _es_fecha(c)) > 50:
            fechas = {j: f"{c.year}-{c.month:02d}" for j, c in enumerate(f) if _es_fecha(c)}
            etiquetas = {}
            for k in range(i + 1, min(i + 10, len(filas))):
                lbl = str(filas[k][0] or "").strip().lower()
                if not lbl or lbl.startswith("fuente"):
                    if lbl.startswith("fuente"):
                        break
                    continue
                etiquetas[lbl] = filas[k]
            bloques.append((fechas, etiquetas))
        i += 1
    if len(bloques) < 2:
        raise ValueError("anexo Informe sobre Bancos: bloques de familias no encontrados")
    ratios, saldos = bloques

    def _serie(bloque, contiene):
        fechas, etiquetas = bloque
        lbl = next(l for l in etiquetas if contiene in l)
        out = {}
        for j, ym in fechas.items():
            try:
                v = float(etiquetas[lbl][j])
            except (TypeError, ValueError, IndexError):
                continue
            out[ym] = v
        return out

    s_per, s_tar = _serie(saldos, "personales"), _serie(saldos, "tarjeta")
    r_per, r_tar = _serie(ratios, "personales"), _serie(ratios, "tarjeta")
    deuda, mora = {}, {}
    for ym in sorted(set(s_per) & set(s_tar) & set(r_per) & set(r_tar)):
        total = s_per[ym] + s_tar[ym]
        if total > 0:
            deuda[ym] = total
            mora[ym] = (r_per[ym] * s_per[ym] + r_tar[ym] * s_tar[ym]) / total
    return deuda, mora


def fetch_itvc_endeudamiento() -> list:
    """Crédito de consumo de FAMILIAS (personales + tarjetas, anexo del
    Informe sobre Bancos) en términos REALES — stock puro, base 4T-2023:

        I_EC(t) = 100 × (Deuda_real_t / Deuda_real_4T23)

    ADR-0067: el factor de mora que multiplicaba este índice (I_EC del doc
    260702) salió a un indicador propio (`mora_familias`) — mantenerlo acá
    la contaría dos veces (el mismo doble conteo que ADR-0033 eliminó entre
    brecha y alimentos). Deuda real creciendo = acceso al crédito; el estrés
    de pago lo mide la mora por separado, en la misma dimensión."""
    deuda, _mora = _anexo_bancos_familias()
    ipc = _nivel_mensual("148.3_INIVELNAL_DICI_M_26", limit=220)
    real = {ym: deuda[ym] / ipc[ym] for ym in sorted(set(deuda) & set(ipc))}
    base_real = _base_t423(real)
    out = []
    for ym in sorted(real):
        if ym < "2023-10":
            continue
        indice = 100.0 * (real[ym] / base_real)
        out.append([f"{ym}-01", round(indice, 1)])
    return out


def fetch_mora_serie() -> list:
    """Mora del crédito familiar (ADR-0067): % de la cartera de consumo de
    FAMILIAS (personales + tarjetas) en situación irregular, ponderado por el
    saldo de cada línea — anexo del Informe sobre Bancos, mismo corte que el
    endeudamiento. Serie desde 2021-07 (contexto pre-mandato, como la brecha);
    el rebase B100 invertido (más mora = peor) lo hace publicar._itvc_indices.
    [[YYYY-MM-01, %]]."""
    _deuda, mora = _anexo_bancos_familias()
    return [[f"{ym}-01", round(v, 2)] for ym, v in sorted(mora.items())
            if ym >= "2021-07" and v]


IVI_SERIE_STORE = Path(__file__).resolve().parents[1] / "data" / "vida" / "ivi_serie.json"
IVI_ARCHIVO_URL = "https://www.utdt.edu/listado_contenidos.php?id_item_menu=23763"
IVI_ULTIMO_URL = "https://www.utdt.edu/listado_contenidos.php?id_item_menu=2156"
IVI_MESES = {"ENERO": 1, "FEBRERO": 2, "MARZO": 3, "ABRIL": 4, "MAYO": 5, "JUNIO": 6,
             "JULIO": 7, "AGOSTO": 8, "SEPTIEMBRE": 9, "SETIEMBRE": 9, "OCTUBRE": 10,
             "NOVIEMBRE": 11, "DICIEMBRE": 12}


def _ivi_parse_pdf(content: bytes) -> tuple | None:
    """(YYYY-MM, ivi_pct) desde la primera página del informe mensual del
    LICIP: encabezado 'MES AÑO' + 'XX.X % de los hogares'. Patrón verificado
    en informes de 2020, 2025 y 2026."""
    import io as _io
    from pypdf import PdfReader
    t = " ".join(PdfReader(_io.BytesIO(content)).pages[0].extract_text().split())
    m = re.search(r"([A-ZÑ]+)\s+(20\d\d)", t)
    v = re.search(r"(\d{1,2}[.,]\d)\s*%\s*de los hogares", t)
    if not (m and v and m.group(1).upper() in IVI_MESES):
        return None
    ym = f"{m.group(2)}-{IVI_MESES[m.group(1).upper()]:02d}"
    return ym, float(v.group(1).replace(",", "."))


def fetch_ivi_serie() -> list:
    """IVI — Índice de Victimización del LICIP (UTDT): % de hogares de 40
    centros urbanos que sufrieron al menos un delito en los últimos 12 meses
    (denunciado o no — captura la cifra negra), encuesta MENSUAL (ADR-0032:
    reemplaza al SNIC anual como métrica del ITVC; el SNIC queda de contraste).
    La ventana de 12 meses de la pregunta desestacionaliza por construcción.
    STORE persistente: cada corrida descubre los PDFs del archivo + el último
    informe, parsea solo los no procesados y acumula. [[YYYY-MM-01, %]]."""
    store = json.loads(IVI_SERIE_STORE.read_text(encoding="utf-8-sig")) \
        if IVI_SERIE_STORE.exists() else {"_meta": {}, "procesados": [], "mensual": {}}
    try:
        urls = set()
        for pagina in (IVI_ARCHIVO_URL, IVI_ULTIMO_URL):
            r = requests.get(pagina, headers=HTTP_HEADERS, timeout=60)
            r.raise_for_status()
            urls |= set(re.findall(r'https://www\.utdt\.edu/download\.php\?fname=[^"&]+\.pdf', r.text))
        nuevos = sorted(urls - set(store["procesados"]))
        for url in nuevos:
            try:
                rp = requests.get(url, headers=HTTP_HEADERS, timeout=90)
                rp.raise_for_status()
                parsed = _ivi_parse_pdf(rp.content)
                if parsed:
                    ym, v = parsed
                    store["mensual"][ym] = v
                store["procesados"].append(url)
            except Exception as e:
                print(f"  [WARN] IVI: PDF no procesado ({url[-30:]}): {str(e)[:50]}")
        if nuevos:
            store["_meta"] = {"fuente": "UTDT — LICIP, Índice de Victimización (informes mensuales PDF)",
                              "actualizado": datetime.today().strftime("%Y-%m-%d"),
                              "nota": ("Encuesta mensual, 40 centros urbanos, ventana 12 meses. "
                                       "PDFs con URL por hash: se descubren desde el listado y "
                                       "se procesan una sola vez (ADR-0032).")}
            IVI_SERIE_STORE.write_text(
                json.dumps(store, indent=2, ensure_ascii=False), encoding="utf-8")
            print(f"  [OK] IVI: {len(nuevos)} PDFs nuevos procesados "
                  f"(serie: {len(store['mensual'])} meses)")
    except Exception as e:
        print(f"  [WARN] IVI: listado UTDT no disponible ({str(e)[:60]}); serie del store")
    return [[f"{ym}-01", v] for ym, v in sorted(store["mensual"].items())]


SNIC_SERIE_STORE = Path(__file__).resolve().parents[1] / "data" / "vida" / "snic_serie.json"


def fetch_inseguridad_serie() -> list:
    """Serie ANUAL de hechos delictivos del SNIC (total país), con el MISMO
    criterio de suma que el colector — VALIDADO contra los totales oficiales
    (2016: 1,59M · 2023: 2,4M · dinámica 2025 = informe 01-jun-2026). Con
    STORE PERSISTENTE (barrido vida 10/13): cloud-snic sufre caídas de días
    enteros; si el host responde se refresca el store, si no, la serie sale
    del store — la web nunca se queda sin serie (antes, el apagón hacía que
    el fallback del histórico inyectara totales anuales con fechas de
    corrida). El CSV oficial se revisa retroactivamente: cada refresco pisa
    la serie completa. [[YYYY-12-01, hechos]]."""
    import csv as _csv
    sys.path.insert(0, str(Path(__file__).parent / "vida_cotidiana"))
    from config import SNIC_CSV
    store = json.loads(SNIC_SERIE_STORE.read_text(encoding="utf-8-sig"))
    try:
        r = requests.get(SNIC_CSV, headers=HTTP_HEADERS, timeout=HTTP_TIMEOUT * 3)
        r.raise_for_status()
        texto = r.content.decode("utf-8", errors="replace")
        sep = ";" if ";" in texto.split("\n", 1)[0] else ","   # el CSV cambió a ';' en 2026
        por_anio = {}
        for row in _csv.DictReader(io.StringIO(texto), delimiter=sep):
            anio = (row.get("anio") or row.get("year") or row.get("Anio") or "").strip()
            if not anio.isdigit():
                continue
            try:
                hechos = int(float(row.get("cantidad_hechos") or row.get("hechos")
                                    or row.get("cantidad") or 0))
            except ValueError:
                continue
            por_anio[anio] = por_anio.get(anio, 0) + hechos
        por_anio = {a: v for a, v in por_anio.items() if int(a) >= 2014 and v > 0}
        if len(por_anio) >= 8:                       # descarga sana → refrescar store
            store["anual"] = {a: por_anio[a] for a in sorted(por_anio)}
            store["_meta"]["actualizado"] = datetime.today().strftime("%Y-%m-%d")
            SNIC_SERIE_STORE.write_text(
                json.dumps(store, indent=2, ensure_ascii=False), encoding="utf-8")
    except Exception as e:
        print(f"  [WARN] inseguridad: cloud-snic caído ({str(e)[:60]}); serie del store "
              f"(al {store['_meta'].get('actualizado')})")
    return [[f"{a}-12-01", v] for a, v in sorted(store["anual"].items())]


CARNE_SERIE_STORE = Path(__file__).resolve().parents[1] / "data" / "vida" / "carne_serie.json"
MOTOS_SERIE_STORE = Path(__file__).resolve().parents[1] / "data" / "vida" / "motos_serie.json"


# Faena mensual en toneladas de las tres carnes (INDEC, vía la API de series de
# datos.gob.ar). Verificadas el 2026-08-20: las tres llegan a 2026-06 y arrancan
# mucho antes del 4T-2023, que es la base del ITCIS.
FAENA_TONELADAS = {
    "vacuna":  "40.3_VT_0_M_17",
    "porcina": "40.3_PT_0_M_18",
    "aviar":   "40.3_AT_0_M_14",
}
# Población urbana total (INDEC, EPH). Trimestral y proyectada en línea recta
# —crece exactamente 100 mil por trimestre—, así que interpolar y extender con
# su propia pendiente no inventa nada que la fuente no diga.
POBLACION_ID = "461.3_POBLACION_ANO_AEA_T_28_3"


def _poblacion_mensual(pob: list):
    """Interpola la serie trimestral a meses y la extiende con su pendiente."""
    puntos = sorted((f[:7], v) for f, v in pob)
    ym = lambda x: int(x[:4]) * 12 + int(x[5:7])

    def en(mes: str) -> float:
        t = ym(mes)
        if t <= ym(puntos[0][0]):
            return puntos[0][1]
        for (f0, v0), (f1, v1) in zip(puntos, puntos[1:]):
            if ym(f0) <= t <= ym(f1):
                return v0 + (v1 - v0) * (t - ym(f0)) / (ym(f1) - ym(f0))
        (f0, v0), (f1, v1) = puntos[-2], puntos[-1]
        return v1 + (v1 - v0) / (ym(f1) - ym(f0)) * (t - ym(f1))

    return en


def fetch_carnes_total_serie() -> list:
    """Consumo TOTAL de carnes per cápita (vacuna + aviar + porcina), índice
    base 100 = promedio del 4T-2023.

    Es el componente que PUNTÚA en el ITCIS (ADR-0217): mide el acceso a
    proteína cárnica sin confundir sustitución con empobrecimiento, que es lo
    que la vacuna sola no puede distinguir.

    Se reconstruye desde la FAENA en toneladas del INDEC y no desde el tablero
    de SAGYP: el tablero publica el nivel per cápita ya calculado, pero es una
    foto del mes que se pisa en cada edición, así que no tiene historia contra
    la cual rebasear. La faena sí, y desde 2009.

    Dos supuestos, declarados:

    - La faena es PRODUCCIÓN. Sin netear exportaciones el nivel no es consumo
      aparente. No importa acá: el índice se lee contra su propia base, así que
      lo que pesa es la evolución, no el nivel — y el nivel per cápita oficial
      se sigue publicando en la card, que sale de SAGYP.
    - El pasaje a per cápita usa la población urbana total del INDEC,
      interpolada a meses. Es una proyección en línea recta de la propia
      fuente, no una estimación nuestra.
    """
    crudas = {}
    for carne, sid in FAENA_TONELADAS.items():
        crudas[carne] = {f[:7]: v for f, v in fetch_indec(sid, limit=240)}
    meses = sorted(set.intersection(*[set(d) for d in crudas.values()]))
    if len(meses) < 24:
        raise RuntimeError(f"faena: sólo {len(meses)} meses en común entre las tres carnes")
    total = {m: sum(crudas[c][m] for c in crudas) for m in meses}

    # Promedio móvil de 12 meses: la misma ventana con la que SAGYP publica su
    # per cápita, y la que saca la estacionalidad fuerte de la faena.
    movil = {meses[i]: sum(total[m] for m in meses[i - 11:i + 1])
             for i in range(11, len(meses))}

    pob = _poblacion_mensual(fetch_indec(POBLACION_ID, limit=80))
    per_capita = {m: movil[m] / pob(m) for m in movil}

    base_meses = [m for m in ("2023-10", "2023-11", "2023-12") if m in per_capita]
    if len(base_meses) < 3:
        raise RuntimeError("faena: la serie no llega al 4T-2023, que es la base del índice")
    base = sum(per_capita[m] for m in base_meses) / len(base_meses)

    return [[f"{m}-01", round(per_capita[m] / base * 100, 1)]
            for m in sorted(per_capita) if m >= "2023-01"]


def fetch_carne_serie() -> list:
    """Serie MENSUAL del consumo de carne per cápita (promedio móvil 12m,
    CICCRA) desde oct-2023 (línea base del ITVC). Los PDFs mensuales se bajan
    una sola vez y quedan cacheados por mes en data/vida/carne_serie.json
    (~33 PDFs solo la primera corrida). Los informes de 2023 usan sufijo 'b'
    (separata económica); 2024→ van sin sufijo. [[YYYY-MM-01, kg/hab/año]]."""
    sys.path.insert(0, str(Path(__file__).parent / "vida_cotidiana"))
    sys.path.insert(0, str(Path(__file__).parent / "vida_cotidiana" / "collectors"))
    from ciccra import _url_pdf, _extraer_per_capita
    try:
        cache = json.loads(CARNE_SERIE_STORE.read_text(encoding="utf-8-sig"))
    except (OSError, json.JSONDecodeError):
        cache = {}
    hoy = date.today()
    fin = hoy.replace(day=1) - timedelta(days=1)          # dato hasta ~2 meses atrás
    fin = fin.replace(day=1) - timedelta(days=1)
    y, m = 2023, 10
    while (y, m) <= (fin.year, fin.month):
        ym = f"{y}-{m:02d}"
        if ym not in cache:
            url = _url_pdf(y, m)
            valor = None
            for u in (url, url.replace(f"-{y}-", f"b-{y}-", 1)):
                try:
                    r = requests.get(u, headers=HTTP_HEADERS, timeout=HTTP_TIMEOUT * 3)
                    if r.status_code == 200 and r.content[:4] == b"%PDF":
                        valor = _extraer_per_capita(r.content)
                        if valor is not None:
                            break
                except requests.RequestException:
                    continue
            # None también se cachea (mes sin informe parseable: no reintentar
            # cada corrida; borrar la clave del store para forzar reintento).
            cache[ym] = valor
            CARNE_SERIE_STORE.write_text(
                json.dumps(cache, indent=1, ensure_ascii=False), encoding="utf-8")
        m += 1
        if m > 12:
            m = 1; y += 1
    # El informe del MES PASADO (M−1) sale en los primeros días del mes
    # corriente y la card (fetch_ciccra) lo levanta apenas existe: la serie
    # también debe intentarlo, pero SIN persistir None (el None permanente es
    # la razón por la que el bucle principal corta en M−2) — si el PDF todavía
    # no salió, se reintenta en la corrida siguiente. Sin esto, el titular
    # avanza un mes antes que la serie y el gate G3 corta el pipeline.
    ult = hoy.replace(day=1) - timedelta(days=1)
    ym_ult = f"{ult.year}-{ult.month:02d}"
    if cache.get(ym_ult) is None:
        url = _url_pdf(ult.year, ult.month)
        for u in (url, url.replace(f"-{ult.year}-", f"b-{ult.year}-", 1)):
            try:
                r = requests.get(u, headers=HTTP_HEADERS, timeout=HTTP_TIMEOUT * 3)
                if r.status_code == 200 and r.content[:4] == b"%PDF":
                    valor = _extraer_per_capita(r.content)
                    if valor is not None:
                        cache[ym_ult] = valor
                        CARNE_SERIE_STORE.write_text(
                            json.dumps(cache, indent=1, ensure_ascii=False), encoding="utf-8")
                        break
            except requests.RequestException:
                continue
    return [[f"{ym}-01", v] for ym, v in sorted(cache.items()) if v]


def fetch_motos_serie() -> list:
    """Serie mensual de patentamientos de motovehículos vía la API de CAFAM
    (la misma fuente del colector; expone meses históricos — verificado
    jul-2026). Desde nov-2022: 11 meses antes de la línea base del ITVC para
    que el rebase por acumulado móvil de 12 meses (ADR-0024, motos tienen
    estacionalidad fuerte: enero ≈ 2× junio) tenga ventanas completas en el
    4T-2023. [[YYYY-MM-01, unidades]]."""
    sys.path.insert(0, str(Path(__file__).parent / "vida_cotidiana"))
    from config import CAFAM_API
    hoy = date.today()
    fin = (hoy.replace(day=1) - timedelta(days=1))     # último mes completo
    out = []
    y, m = 2022, 11
    while (y, m) <= (fin.year, fin.month):
        try:
            r = requests.get(CAFAM_API, params={"month_start": m, "month_end": m,
                                                "year": y, "type": "TODOS"},
                             headers=HTTP_HEADERS, timeout=HTTP_TIMEOUT)
            r.raise_for_status()
            total = sum(p["count"] for p in r.json().get("provinces", []))
            if total > 0:
                out.append([f"{y}-{m:02d}-01", total])
        except Exception as e:
            print(f"  [WARN] motos serie {y}-{m:02d}: {e}")
        m += 1
        if m > 12:
            m = 1; y += 1
    return out


def fetch_motos_serie_cached() -> list:
    """Serie mensual de motos con cache historico persistente."""
    import importlib.util
    config_path = Path(__file__).parent / "vida_cotidiana" / "config.py"
    spec = importlib.util.spec_from_file_location("_vida_cotidiana_config", config_path)
    vida_config = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(vida_config)
    CAFAM_API = vida_config.CAFAM_API

    def _load_cache() -> dict:
        try:
            data = json.loads(MOTOS_SERIE_STORE.read_text(encoding="utf-8-sig"))
            if isinstance(data, dict):
                return {k: int(v) for k, v in data.items() if v is not None}
        except (OSError, json.JSONDecodeError, TypeError, ValueError):
            pass

        cache = {}
        csv_path = OUTPUT_DIR / "vida_cotidiana.csv"
        try:
            with open(csv_path, newline="", encoding="utf-8-sig") as f:
                for row in csv.DictReader(f):
                    if row.get("indicador") == "patentamiento_motos":
                        ym = str(row.get("fecha", ""))[:7]
                        try:
                            cache[ym] = int(float(row.get("valor") or 0))
                        except ValueError:
                            continue
        except OSError:
            pass
        return cache

    cache = _load_cache()
    hoy = date.today()
    fin = hoy.replace(day=1) - timedelta(days=1)
    y, m = 2022, 11
    while (y, m) <= (fin.year, fin.month):
        ym = f"{y}-{m:02d}"
        if ym not in cache:
            try:
                r = requests.get(CAFAM_API, params={"month_start": m, "month_end": m,
                                                    "year": y, "type": "TODOS"},
                                 headers=HTTP_HEADERS, timeout=min(HTTP_TIMEOUT, 10))
                r.raise_for_status()
                total = sum(p["count"] for p in r.json().get("provinces", []))
                if total > 0:
                    cache[ym] = total
                    MOTOS_SERIE_STORE.write_text(
                        json.dumps(cache, indent=2, ensure_ascii=False), encoding="utf-8")
            except Exception as e:
                print(f"  [WARN] motos serie {ym}: {e}")
        m += 1
        if m > 12:
            m = 1; y += 1
    if cache:
        MOTOS_SERIE_STORE.write_text(
            json.dumps({k: cache[k] for k in sorted(cache)}, indent=2, ensure_ascii=False),
            encoding="utf-8")
    return [[f"{ym}-01", cache[ym]] for ym in sorted(cache)]


# Componentes transformados del ITVC-B100 (100 = promedio 4T-2023, ADR-0018)
VIDA_DERIVADAS += [
    ("itvc_alimentos", "índice (100 = 4T-2023)", "INDEC IPC Alimentos + IPC general (elab. CIGOB)", fetch_itvc_alimentos),
    ("itvc_tarifas", "índice (100 = 4T-2023)", "INDEC IPC Regulados + RIPTE (elab. CIGOB)", fetch_itvc_tarifas),
    ("itvc_alquiler", "índice (100 = 4T-2023)", "INDEC IPC-GBA alquiler + nivel general GBA (elab. CIGOB)", fetch_itvc_alquiler),
    # La card de alquiler, con historia. Va acá y no en el literal de
    # VIDA_DERIVADAS de más arriba porque ese se evalúa antes de que existan
    # estas funciones. Unidad y fuente COPIADAS de la card (publicar.py): si
    # divergen, la ficha pública dice una cosa y el CSV de la serie otra.
    ("alquiler_real", "% m/m alquileres",
     "INDEC — IPC-GBA alquiler de la vivienda (vía datos.gob.ar)", fetch_alquiler_real_serie),
    # ADR-0218: `mortalidad_pymes` deja de ser el IPI industrial y pasa a medir
    # lo que su nombre promete — el cierre neto de PyMEs. Una sola serie para la
    # card y para el índice; `itvc_ipi` se retira.
    ("trabajo_independiente", "% del empleo registrado",
     "SIPA — autónomos y monotributo sobre el total de trabajo registrado",
     fetch_trabajo_independiente_serie),
    ("mortalidad_pymes", "empleadores (hasta 50 trabajadores)",
     "SRT — serie histórica de partes empleadoras por tamaño de nómina",
     fetch_empleadores_pyme_serie),
    ("itvc_isac", "índice (100 = 4T-2023)", "INDEC ISAC desestacionalizado", fetch_itvc_isac),
    ("itvc_endeudamiento", "índice real (100 = 4T-2023)", "BCRA Informe sobre Bancos (familias) + IPC INDEC", fetch_itvc_endeudamiento),
    ("mora_familias", "% de cartera irregular (familias)", "BCRA — Informe sobre Bancos (personales + tarjetas)", fetch_mora_serie),
    ("patentamiento_motos", "unidades/mes", "CAFAM API (histórico mensual)", fetch_motos_serie_cached),
    # inseguridad = IVI mensual (ADR-0032); el SNIC anual sigue como serie de
    # contraste bajo clave propia (sin card: alimenta la ficha y validaciones)
    ("inseguridad", "% de hogares víctimas (12 meses)", "UTDT — IVI (LICIP)", fetch_ivi_serie),
    ("inseguridad_snic", "hechos/año (total país)", "SNIC (CSV oficial, suma anual)", fetch_inseguridad_serie),
    ("consumo_carne", "kg/hab/año (PM 12m)", "CICCRA (informes mensuales, caché local)", fetch_carne_serie),
    ("consumo_carnes_total", "índice base 100 = 4T-2023",
     "INDEC — faena de vacunos, porcinos y aves (toneladas), per cápita",
     fetch_carnes_total_serie),
    # La MISMA métrica que muestra la card del indicador (ISAC general nivel
    # s.e., serie 33.2 desestacionalizada): card y modal comparten fuente.
    ("despacho_cemento", "índice ISAC (desest.)", "INDEC ISAC (33.2, s.e.)",
     lambda: [[f, round(v, 2)] for f, v in sorted(fetch_indec(ITVC_ISAC_DESEST_ID, limit=60))]),
]


def fetch_brecha_serie() -> list:
    """Serie de la brecha salario/CBT = RIPTE / Canasta Básica Total, ALINEADA por mes
    (mismo mes en ambas series; el indicador live toma el último de cada una, que a veces
    son meses distintos). Cuántas canastas cubre el salario imponible promedio.
    [[YYYY-MM-01, canastas]]."""
    sys.path.insert(0, str(Path(__file__).parent / "vida_cotidiana"))
    from config import RIPTE_CSV, INDEC_SERIES
    r = requests.get(RIPTE_CSV, headers=HTTP_HEADERS, timeout=HTTP_TIMEOUT)
    r.raise_for_status()
    ripte = {}
    for line in r.text.strip().split("\n"):
        p = line.split(",")
        if len(p) >= 2 and p[0][:4].isdigit():
            ripte[p[0][:7]] = float(p[1])
    cbt = {f[:7]: v for f, v in fetch_indec(INDEC_SERIES["cbt"], limit=60)}
    comun = sorted(set(ripte) & set(cbt))
    return [[f"{ym}-01", round(ripte[ym] / cbt[ym], 2)] for ym in comun if cbt[ym]]


VIDA_DERIVADAS.append(
    ("brecha_salario_cbt", "canastas (RIPTE/CBT)", "INDEC (RIPTE + Canasta Básica Total)", fetch_brecha_serie)
)

GESTION_INDEC = [
    ("149.1_SOR_PUBICO_OCTU_0_14",   "indice_salarios_publico", "indice base oct-2016=100", "INDEC/datos.gob.ar"),
    # (33.4 ISAC insumo cemento se retiró jul-2026: no lo consumía ningún
    #  indicador ni el front — la card de construcción usa el ISAC general
    #  33.2 s.e., ver despacho_cemento arriba.)
]


def fetch_rigi_serie() -> list:
    """Serie histórica del RIGI: inversión aprobada ACUMULADA (US$ M) por mes,
    reconstruida con la fecha de sanción (BO) de cada proyecto aprobado. La
    plataforma oficial solo publica la foto actual; esto arma la evolución."""
    proy = gestion.rigi_proyectos_aprobados()   # [(fecha_iso, nombre, inv)] ordenado
    cum, por_mes = 0.0, {}
    for fecha, _nombre, inv in proy:
        cum += inv
        por_mes[fecha[:7]] = round(cum)         # acumulado al último proyecto del mes
    # Antes del primer proyecto aprobado la inversión RIGI era 0 (el régimen
    # se reglamentó en ago-2024): la serie arranca en dic-2023 con ceros.
    y, m = 2023, 12
    primero = min(por_mes) if por_mes else None
    while primero and f"{y}-{m:02d}" < primero:
        por_mes[f"{y}-{m:02d}"] = 0
        m += 1
        if m > 12:
            m = 1; y += 1
    return [[f"{ym}-01", v] for ym, v in sorted(por_mes.items())]


def fetch_infoleg_serie(texto: str) -> list:
    """Serie mensual del conteo ACUMULADO de normas InfoLeg que contienen `texto`,
    desde dic-2023 hasta el fin de cada mes (mismo conteo que el indicador, evaluado
    a cada corte). [[YYYY-MM-01, count]]."""
    out = []
    y, mo = 2023, 12
    today = date.today()
    while (y, mo) <= (today.year, today.month):
        last = min(date(y, mo, calendar.monthrange(y, mo)[1]), today)
        try:
            c = gestion._infoleg_post(
                texto=texto, tipo_norma="",
                fecha_desde=("01", "12", "2023"),
                fecha_hasta=(last.strftime("%d"), last.strftime("%m"), last.strftime("%Y")),
            )
            out.append([f"{y}-{mo:02d}-01", c])
        except Exception:
            pass
        mo += 1
        if mo > 12:
            mo = 1; y += 1
    return out


def fetch_reduccion_serie() -> list:
    """Serie de la variación % de la dotación APN vs dic-2023 (misma fórmula y
    fuente que el indicador: XLSX mensual del INDEC). [[YYYY-MM-01, var%]]."""
    # dotacion_apn_series() devuelve también METADATOS bajo claves con guion
    # bajo (`_empresas`, `_total`, ADR-0097) cuyo valor es un dict, no un
    # número. El colector los saca con .pop(); acá hay que filtrarlos igual o
    # la resta revienta con "unsupported operand type(s) for -: 'dict'".
    serie = {ym: v for ym, v in gestion.dotacion_apn_series().items()
             if not ym.startswith("_")}
    base = serie["2023-12"]
    return [[f"{ym}-01", round((v - base) / base * 100, 2)]
            for ym, v in sorted(serie.items()) if ym >= "2023-12"]


def fetch_tdps_serie() -> list:
    """Serie mensual del TDPS (% del devengado pagado directo a personas,
    partida 5.1.4, sobre el total del inciso 5) del programa de ingreso social
    vigente en cada momento: Potenciar Trabajo (2023, jur. 85 prog. 38) y sus
    sucesores Volver al Trabajo + Acompañamiento Social (2024–). API
    Presupuesto Abierto; sin token devuelve [] (la serie queda en el último
    snapshot). Meses de transición con devengado ínfimo se omiten."""
    token = gestion._pa_token()
    if not token:
        return []
    cols = ["impacto_presupuestario_mes", "inciso_id", "principal_id",
            "parcial_id", "credito_devengado"]
    por_mes: dict = {}
    for anio in range(2023, date.today().year + 1):
        if anio == 2023:
            bloques = [[{"column": "jurisdiccion_id", "operator": "equal", "value": "85"},
                        {"column": "programa_id", "operator": "equal", "value": "38"}]]
        else:
            bloques = [[{"column": "actividad_desc", "operator": "like", "value": f"%{a}%"}]
                       for a in gestion.TDPS_ACTIVIDADES]
        for filtros in bloques:
            try:
                rows = gestion._pa_credito(token, anio, filtros, columns=cols)
            except Exception as e:
                print(f"  [WARN] tdps serie {anio}: {e}")
                continue
            for r in rows:
                try:
                    if int(r.get("inciso_id") or 0) != 5:
                        continue
                    ym = f"{anio}-{int(r.get('impacto_presupuestario_mes') or 0):02d}"
                    dev = float(r.get("credito_devengado") or 0)
                except (TypeError, ValueError):
                    continue
                d = por_mes.setdefault(ym, {"directo": 0.0, "total": 0.0})
                d["total"] += dev
                if int(r.get("principal_id") or 0) == 1 and int(r.get("parcial_id") or 0) == 4:
                    d["directo"] += dev
    return [[f"{ym}-01", round(100.0 * d["directo"] / d["total"], 1)]
            for ym, d in sorted(por_mes.items()) if d["total"] > 1000.0]


def _serie_var_real_vs_2023(nominal_ids: list) -> list:
    """Serie mensual de la variación % REAL vs el MISMO MES de 2023 (misma
    fórmula que el indicador: deflactada por IPC, mismo mes evita el sesgo
    del aguinaldo) para la suma de una o más series nominales de datos.gob.ar.
    [[YYYY-MM-01, var%]] desde 2024-01."""
    nominal: dict = {}
    for sid in nominal_ids:
        for ym, v in gestion._indec_nivel_mensual(sid, limit=48).items():
            nominal[ym] = nominal.get(ym, 0.0) + v
    ipc = gestion._indec_nivel_mensual(gestion.IPC_ID, limit=48)
    out = []
    for ym in sorted(nominal):
        if ym < "2024-01":
            continue
        base = f"2023-{ym[5:7]}"
        if base not in nominal or base not in ipc or ym not in ipc:
            continue
        real_t = nominal[ym] / ipc[ym]
        real_b = nominal[base] / ipc[base]
        out.append([f"{ym}-01", round((real_t / real_b - 1.0) * 100.0, 2)])
    return out


def fetch_opcion_salud_serie() -> list:
    """Serie mensual del % de usuarios de prepagas con aportes derivados
    directo (misma fórmula que el indicador): RNAS 90xxxx / total RNEMP, por
    año desde 2024 (los XLSX de la SSS existen por año, URL estable; usa
    gestion._sss_tabla, que no depende de la fila de totales — falta en los
    archivos de algunos años). El denominador RNEMP se arrastra al último mes
    disponible (cambia lento y su archivo tiene más rezago que el RNAS)."""
    import openpyxl
    derivados, usuarios = {}, {}
    for anio in range(2024, date.today().year + 1):
        for url_tpl, destino, solo_prepagas in (
                (gestion.SSS_RNAS_URL, derivados, True),
                (gestion.SSS_RNEMP_URL, usuarios, False)):
            try:
                wb = openpyxl.load_workbook(io.BytesIO(
                    gestion._http_get_resiliente(url_tpl.format(anio=anio))), data_only=True)
                ws = wb[wb.sheetnames[0]]
                cols, entidades = gestion._sss_tabla(ws)
                for col, mm in sorted(cols.items()):
                    total_mes, _ = gestion._sss_suma_mes(entidades, col)
                    if total_mes > 0:
                        suma, _ = gestion._sss_suma_mes(entidades, col, solo_prepagas=solo_prepagas)
                        destino[f"{anio}-{mm}"] = suma
            except Exception as e:
                registro = "RNAS" if url_tpl is gestion.SSS_RNAS_URL else "RNEMP"
                print(f"  [WARN] opcion salud serie {registro} {anio}: {e}")
    # Denominador: se arrastra el último RNEMP conocido; para los meses previos
    # al primer archivo mensual (el RNEMP 2024 agrupa por cuatrimestres y se
    # saltea) se usa el primero disponible — el canal era ínfimo en 2024.
    out, ult_u = [], None
    primera_u = usuarios[min(usuarios)] if usuarios else None
    for ym in sorted(derivados):
        if usuarios.get(ym):
            ult_u = usuarios[ym]
        denom = ult_u or primera_u
        if denom:
            out.append([f"{ym}-01", round(100.0 * derivados[ym] / denom, 1)])
    # dic-2023: el canal de derivación directa nace con el DNU 70/23 (dic-2023);
    # antes del primer XLSX (ene-2024) los derivados eran 0.
    if out and out[0][0] == "2024-01-01":
        out.insert(0, ["2023-12-01", 0.0])
    return out


def fetch_protestas_serie() -> list:
    """Serie mensual de eventos de protesta en CABA desde el store que llena
    gestion.actualizar_protestas_caba() (evita re-bajar los ~8 MB de ACLED:
    el colector corre antes en el pipeline). El último mes se EXCLUYE si el
    archivo ACLED no llega a fin de mes (misma regla que el indicador: un mes
    parcial dibujaría un derrumbe falso al final de la curva).
    [[YYYY-MM-01, eventos]]."""
    import json as _json
    if not gestion.PROTESTAS_STORE_PATH.exists():
        return []
    store = _json.loads(gestion.PROTESTAS_STORE_PATH.read_text(encoding="utf-8"))
    mensual = store.get("mensual", {})
    hasta = store.get("_meta", {}).get("hasta_semana", "")
    yms = sorted(mensual)
    if hasta and yms and hasta[:7] == yms[-1]:
        a, m = int(hasta[:4]), int(hasta[5:7])
        if int(hasta[8:10]) < calendar.monthrange(a, m)[1]:
            yms = yms[:-1]
    return [[f"{ym}-01", mensual[ym]] for ym in yms]


ARGENTINADATOS_CCL = "https://api.argentinadatos.com/v1/cotizaciones/dolares/contadoconliqui"


def fetch_brecha_serie() -> list:
    """Serie mensual de la brecha CCL/mayorista (misma métrica que cepo_mulc):
    promedio mensual del CCL histórico de ArgentinaDatos (el proyecto hermano
    de dolarapi, la MISMA familia de fuente que usa el colector vivo) sobre el
    promedio mensual del A3500 oficial (BCRA var. 5). Desde dic-2023. El
    último punto es el promedio del mes corriente parcial (el live es el spot
    del día — difieren de forma inmaterial)."""
    hoy = date.today()
    r = requests.get(ARGENTINADATOS_CCL,
                     headers={"User-Agent": "Mozilla/5.0 (compatible; CIGOB-Monitor/1.0)"},
                     timeout=120)
    r.raise_for_status()
    ccl: dict = {}
    for fila in r.json():
        fecha = str(fila.get("fecha", ""))
        try:
            v = float(fila.get("venta") or 0)
        except (TypeError, ValueError):
            continue
        if fecha >= "2023-12" and v > 0:
            ccl.setdefault(fecha[:7], []).append(v)
    dias = (hoy - date(2023, 11, 25)).days
    a3500 = gestion._tc_mayorista_promedio_por_mes(dias=dias)   # {ym: promedio}
    out = []
    for ym in sorted(set(ccl) & set(a3500)):
        prom_ccl = sum(ccl[ym]) / len(ccl[ym])
        if a3500[ym] > 0:
            out.append([f"{ym}-01", round((prom_ccl / a3500[ym] - 1.0) * 100.0, 2)])
    return out


def fetch_litigiosidad_serie() -> list:
    """Serie mensual de la variación % de los juicios SRT (acumulado 12 meses
    vs los 12 previos, misma fórmula que el indicador). [[YYYY-MM-01, var%]]."""
    import io as _io, openpyxl
    content = gestion._http_get_resiliente(gestion.SRT_JUICIOS_URL)
    wb = openpyxl.load_workbook(_io.BytesIO(content), data_only=True)
    hoja = next(s for s in wb.sheetnames if "TOTAL" in s.upper())
    filas = list(wb[hoja].iter_rows(values_only=True))
    meses_es = {"ene": 1, "feb": 2, "mar": 3, "abr": 4, "may": 5, "jun": 6,
                "jul": 7, "ago": 8, "sep": 9, "oct": 10, "nov": 11, "dic": 12}
    patron = re.compile(r"^([A-Za-z]{3})-(\d{4})$")
    header = next(f for f in filas
                  if sum(1 for c in f if isinstance(c, str) and patron.match(c.strip())) > 20)
    fila_total = next(f for f in filas
                      if isinstance(f[0], str) and f[0].lower().startswith("total de juicios"))
    serie = {}
    for h, v in zip(header, fila_total):
        m = patron.match(h.strip()) if isinstance(h, str) else None
        if m and m.group(1).lower() in meses_es:
            try:
                serie[f"{m.group(2)}-{meses_es[m.group(1).lower()]:02d}"] = float(v)
            except (TypeError, ValueError):
                continue
    yms = sorted(serie)
    out = []
    for i in range(23, len(yms)):
        ult12 = sum(serie[y] for y in yms[i - 11:i + 1])
        prev12 = sum(serie[y] for y in yms[i - 23:i - 11])
        if prev12:
            out.append([f"{yms[i]}-01", round((ult12 / prev12 - 1.0) * 100.0, 1)])
    return out[-60:]   # últimos 5 años (la serie arranca en 2010)


def fetch_alicuota_serie() -> list:
    """Serie mensual de la ALÍCUOTA EFECTIVA del comercio exterior
    (apertura_comercial desde el ADR-0021: la brecha salió del compuesto):
    recaudación DEX+DIM en USD por el A3500 promedio del mes, sobre el
    intercambio expo+impo del ICA. Desde dic-2023. [[YYYY-MM-01, %]]."""
    dex  = gestion._indec_nivel_mensual(gestion.DEX_ID, limit=48)
    dim  = gestion._indec_nivel_mensual(gestion.DIM_ID, limit=48)
    expo = gestion._indec_nivel_mensual(gestion.EXPO_ICA_ID, limit=48)
    impo = gestion._indec_nivel_mensual(gestion.IMPO_ICA_ID, limit=48)
    dias = (date.today() - date(2023, 11, 25)).days
    tc   = gestion._tc_mayorista_promedio_por_mes(dias=dias)
    out = []
    for ym in sorted(set(dex) & set(dim) & set(expo) & set(impo) & set(tc)):
        if ym < "2023-12" or expo[ym] + impo[ym] <= 0 or tc[ym] <= 0:
            continue
        alicuota = 100.0 * ((dex[ym] + dim[ym]) / tc[ym]) / (expo[ym] + impo[ym])
        out.append([f"{ym}-01", round(alicuota, 2)])
    return out


CONCESIONES_FECHAS_STORE = Path(__file__).resolve().parents[1] / "data" / "gestion" / "concesiones_fechas.json"


def fetch_concesiones_serie() -> list:
    """Serie mensual del % de km adjudicados de la RFC, reconstruida como
    función ESCALONADA por hitos fechados (las adjudicaciones son actos
    administrativos puntuales): cada etapa suma sus km desde su mes de
    adjudicación. El store concesiones_fechas.json trae las fechas oficiales
    verificadas (Etapa I: RESOL-2025-80-ST ene-2026 · II-A: Res. 706/2026
    may-2026) y se auto-actualiza cuando CONTRAT.AR muestra una etapa nueva
    en ADJUDICADO (mes corriente) — mismo patrón que rigi_fechas.
    Antes de la primera adjudicación el avance es 0. [[YYYY-MM-01, %]]."""
    store = json.loads(CONCESIONES_FECHAS_STORE.read_text(encoding="utf-8-sig"))
    etapas = store["etapas"]
    try:
        km = gestion._rfc_km_por_etapa()
        store["km_totales"] = round(sum(km.values()))
        hoy_ym = date.today().strftime("%Y-%m")
        for proceso, nombre, estado in gestion._contratar_procesos_rfc():
            etapa = gestion._etapa_de_proceso(nombre)
            if etapa and gestion._esta_adjudicado(estado) and etapa in km:
                if etapa not in etapas:
                    etapas[etapa] = {"fecha": hoy_ym, "km": km[etapa],
                                     "fuente": f"CONTRAT.AR {proceso} (detectado {hoy_ym})"}
                else:
                    etapas[etapa]["km"] = km[etapa]
        CONCESIONES_FECHAS_STORE.write_text(
            json.dumps(store, indent=2, ensure_ascii=False), encoding="utf-8")
    except Exception as e:
        print(f"  [WARN] concesiones serie: CONTRAT.AR/RFC no disponible ({e}); se usa el store")
    km_total = store.get("km_totales") or 0
    if not km_total:
        return []
    out = []
    y, m = 2023, 12
    hoy = date.today()
    while (y, m) <= (hoy.year, hoy.month):
        ym = f"{y}-{m:02d}"
        adj = sum(e["km"] for e in etapas.values() if e["fecha"] <= ym)
        out.append([f"{ym}-01", round(100.0 * adj / km_total, 1)])
        m += 1
        if m > 12:
            m = 1; y += 1
    return out


PRIVATIZACIONES_FECHAS_STORE = (Path(__file__).resolve().parents[1] / "data" / "gestion"
                                / "privatizaciones_fechas.json")


def fetch_privatizaciones_serie() -> list:
    """Serie mensual del avance de privatizaciones (promedio de etapas/4 × 100,
    la MISMA fórmula del indicador) reconstruida con las transiciones de etapa
    FECHADAS por norma del BO (privatizaciones_fechas.json, verificadas
    jul-2026). El mes corriente usa las etapas del store vivo
    (privatizaciones.json) y avisa si no reconcilia con las transiciones —
    señal de que falta fechar un hito nuevo. [[YYYY-MM-01, %]]."""
    fechas = json.loads(PRIVATIZACIONES_FECHAS_STORE.read_text(encoding="utf-8-sig"))["empresas"]
    live_path = Path(__file__).resolve().parents[1] / "data" / "gestion" / "privatizaciones.json"
    live = json.loads(live_path.read_text(encoding="utf-8-sig"))["empresas"]

    def etapa_en(emp: str, ym: str) -> float:
        vigente = 0.0
        for t in fechas.get(emp, []):
            if t["fecha"] <= ym:
                vigente = float(t["etapa"])
        return vigente

    empresas = sorted(live)
    hoy_ym = date.today().strftime("%Y-%m")
    for emp in empresas:
        if abs(etapa_en(emp, hoy_ym) - float(live[emp]["etapa"])) > 0.01:
            print(f"  [WARN] privatizaciones serie: {emp} etapa live "
                  f"{live[emp]['etapa']} != fechada {etapa_en(emp, hoy_ym)} — "
                  f"agregar la transición a privatizaciones_fechas.json")
    out = []
    y, m = 2023, 12
    while (y, m) <= (int(hoy_ym[:4]), int(hoy_ym[5:])):
        ym = f"{y}-{m:02d}"
        if ym == hoy_ym:
            etapas = [float(live[e]["etapa"]) for e in empresas]
        else:
            etapas = [etapa_en(e, ym) for e in empresas]
        out.append([f"{ym}-01", round(sum(etapas) / len(etapas) / 4.0 * 100.0, 1)])
        m += 1
        if m > 12:
            m = 1; y += 1
    return out


def fetch_fal_serie() -> list:
    """Serie mensual del avance del FAL por sus DOS ACTOS FUNDAMENTALES.

    Desde ADR-0142 la serie se reconstruye en la MISMA escala que la card: 50
    puntos por cada acto cumplido a fin de cada mes, leídos de `fal_hitos.json`
    (norma publicada y fechada, verificable por número en InfoLeg).

        0    hasta feb-2026    ni ley ni reglamentación
        50   desde mar-2026    Ley 27.802 sancionada (06-mar-2026)
        100  desde jun-2026    Decreto 408/2026 reglamenta (01-jun-2026)

    Es una escalera de tres peldaños y no se mueve más: los dos actos ocurrieron
    y no se deshacen. La limitación está documentada en ADR-0142 — se publica
    así por decisión editorial, sabiendo que el indicador dejó de discriminar.

    No requiere red: sale del registro de hitos.
    [[YYYY-MM-01, índice]]."""
    hitos_json = json.loads(gestion.FAL_HITOS_PATH.read_text(encoding="utf-8-sig"))
    por_norma = {h.get("norma"): h for h in hitos_json.get("construccion", [])}
    fechas_actos = [(por_norma.get(norma) or {}).get("fecha")
                    for norma, _ in gestion.FAL_ACTOS_FUNDAMENTALES]
    total = len(gestion.FAL_ACTOS_FUNDAMENTALES)

    out = []
    hoy = date.today()
    fin = hoy.replace(day=1) - timedelta(days=1)          # último mes completo
    y, m = 2023, 12
    while (y, m) <= (fin.year, fin.month):
        ult_dia = calendar.monthrange(y, m)[1]
        cierre = f"{y}-{m:02d}-{ult_dia:02d}"
        cumplidos = sum(1 for f in fechas_actos if f and f <= cierre)
        out.append([f"{y}-{m:02d}-01", round(100.0 * cumplidos / total, 1)])
        m += 1
        if m > 12:
            m, y = 1, y + 1
    return out


# El ICG no cuelga de `listado_contenidos.php` como el ICC (16458) o el Índice
# Líder (16461) —por eso no aparece sondeando ids vecinos—: su ficha vive en
# `ver_contenido.php` y la descarga en una página aparte, "Descarga de datos".
# Estas dos constantes se usaban en fetch_icg_serie sin estar definidas en
# ningún lado: la función levantaba NameError en cada corrida desde que se
# escribió, así que la serie nunca se refrescó (ADR-0175).
UTDT_ICG_LISTADO = "https://www.utdt.edu/listado_contenidos.php?id_item_menu=28756"
UTDT_ICG_REFERER = "https://www.utdt.edu/ver_contenido.php?id_contenido=1439&id_item_menu=2964"


def fetch_icg_serie() -> list:
    """Serie mensual del ICG UTDT (Índice de Confianza en el Gobierno, 0-5) —
    insumo de la validación externa del ITCG (no es indicador del cinturón).
    El XLS oficial viene TRANSPUESTO (fechas en columnas): en cada hoja se
    busca la fila de fechas y la fila rotulada ICG. El listado de descargas
    exige Referer (sin él devuelve vacío). [[YYYY-MM-01, índice]]."""
    import xlrd
    h = dict(HTTP_HEADERS, Referer=UTDT_ICG_REFERER)
    r = requests.get(UTDT_ICG_LISTADO, headers=h, timeout=HTTP_TIMEOUT, verify=False)
    r.raise_for_status()
    m = re.search(r"download\.php\?fname=([^\"'\s>]+\.xls)\b", r.text, re.IGNORECASE)
    if not m:
        raise ValueError("listado ICG sin XLS (¿cambió la página o falta Referer?)")
    r = requests.get(f"https://www.utdt.edu/download.php?fname={m.group(1)}",
                     headers=h, timeout=60, verify=False)
    r.raise_for_status()
    wb = xlrd.open_workbook(file_contents=r.content)
    out = {}
    for ws in wb.sheets():
        fila_fechas = next((i for i in range(ws.nrows)
                            if sum(1 for j in range(ws.ncols)
                                   if ws.cell(i, j).ctype == xlrd.XL_CELL_DATE) >= 5), None)
        if fila_fechas is None:
            continue
        fila_icg = next((i for i in range(fila_fechas + 1, min(fila_fechas + 4, ws.nrows))
                         if "icg" in str(ws.cell(i, 1).value).lower()
                         or "icg" in str(ws.cell(i, 0).value).lower()), fila_fechas + 1)
        for j in range(ws.ncols):
            fc, vc = ws.cell(fila_fechas, j), ws.cell(fila_icg, j)
            if fc.ctype == xlrd.XL_CELL_DATE and vc.ctype == xlrd.XL_CELL_NUMBER:
                t = xlrd.xldate_as_tuple(fc.value, wb.datemode)
                out[f"{t[0]}-{t[1]:02d}-01"] = round(float(vc.value), 3)
    return sorted([f, v] for f, v in out.items())


def fetch_protocolo_serie() -> list:
    """Serie del IRPC (reducción de cortes CABA vs 2023) por escalones ANUALES
    de los anclajes públicos de Diagnóstico Político (ADR-0025): cada año
    cerrado aporta su IRPC y el valor se arrastra mes a mes hasta el siguiente
    (regla del último dato disponible). [[YYYY-MM-01, %]]."""
    dp = json.loads((Path(__file__).resolve().parents[1] / "data" / "gestion" /
                     "dp_piquetes.json").read_text(encoding="utf-8-sig"))
    caba = {int(k): int(v) for k, v in dp["caba"].items()}
    base = caba[2023]
    hoy = date.today()
    fin = hoy.replace(day=1) - timedelta(days=1)
    out = []
    y, m = 2024, 1
    while (y, m) <= (fin.year, fin.month):
        anio_dato = y if y in caba else max(a for a in caba if 2023 < a <= y)
        irpc = round((1.0 - caba[anio_dato] / base) * 100.0, 1)
        out.append([f"{y}-{m:02d}-01", irpc])
        m += 1
        if m > 12:
            m, y = 1, y + 1
    return out


def fetch_desregulacion_serie() -> list:
    """Serie MENSUAL de ARTÍCULOS modificados o eliminados, acumulados desde
    dic-2023, según el Ministerio de Desregulación (ADR-0125 / ADR-0143).

    Misma construcción que puntúa en la ficha
    (`gestion.desregulacion_oficial_serie`): el backfill sale del gráfico del
    informe de abril-2026 y los meses posteriores del titular de cada informe.
    Todo cacheado en `data/gestion/desregulacion_oficial.json`.

    Desde ADR-0143 la unidad son ARTÍCULOS y no normas: el backfill sale de la
    Figura 5 del informe de abril-2026 ("evolución acumulada" de artículos), no
    de la Figura 1. Las dos series existen y no son empalmables — mezclarlas
    dejaría dos escalas en el mismo gráfico.
    [[YYYY-MM-01, artículos acumulados]]."""
    try:
        serie = gestion.desregulacion_oficial_serie("articulos")
    except Exception as e:
        print(f"  [WARN] desregulacion: fuente oficial ilegible ({e}) -- serie omitida")
        return []
    out = [[f"{ym}-01", float(v)] for ym, v in sorted(serie.items())]
    return out


GESTION_DERIVADAS = [
    ("protocolo_antipiquetes", "% reducción de cortes CABA vs 2023 (IRPC, anual)",
     "Diagnóstico Político (monitoreos públicos)", fetch_protocolo_serie),
    ("icg_utdt", "índice 0-5 (confianza en el gobierno)", "UTDT (ICG, serie XLS)", fetch_icg_serie),
    ("apertura_comercial", "% alícuota efectiva del comercio exterior", "ARCA (DEX+DIM) + INDEC ICA + BCRA A3500", fetch_alicuota_serie),
    ("concesiones_infraestructura", "% km adjudicados RFC", "CONTRAT.AR + RFC (hitos fechados)", fetch_concesiones_serie),
    ("privatizaciones", "% avance (etapas 0-4, cartera Ley Bases)", "BO — hitos fechados (elab. CIGOB)", fetch_privatizaciones_serie),
    ("fal_modernizacion_laboral", "Índice 0–100 (FAL)", "Boletín Oficial (menciones del FAL, Ley 27.802) + CNV (registro FCI)", fetch_fal_serie),
    ("rigi_inversiones", "US$ M aprobados (acum.)", "Min. Economía RIGI + BO (fechas de sanción)", fetch_rigi_serie),
    ("desregulacion_normativa", "artículos modificados o eliminados, acumulados desde dic-2023",
     "Min. de Desregulación y Transformación del Estado — informe mensual",
     fetch_desregulacion_serie),
    # A % calibrado (45 actos = plan completo, misma escala que el titular):
    # a diferencia de desregulación (100 normas = 100%), acá conteo ≠ %.
    # El conteo sale de gestion.serie_reestructuracion_vigentes() y no de un
    # recuento propio: cuenta los cierres VIGENTES contra el registro curado,
    # el mismo filtro que la card (ADR-0188). El total del plan también se
    # toma de la constante, no se repite el 45 acá.
    ("reestructuracion_organismos", "% de avance (proxy InfoLeg, 45 actos = 100%)",
     "InfoLeg ('disolucion' desde dic-2023, cierres vigentes)",
     lambda: [[f, round(min(100.0, v * 100.0 / gestion.ORGANISMOS_PLAN_TOTAL), 1)]
              for f, v in gestion.serie_reestructuracion_vigentes()]),
    ("reduccion_estado", "% vs dic-2023", "INDEC (dotación APN mensual)", fetch_reduccion_serie),
    ("asistencia_directa", "% TDPS (directo a personas / transferencias)", "API Presupuesto Abierto (SIDIF)", fetch_tdps_serie),
    ("gasto_funcionamiento", "% real vs mismo mes 2023", "Sec. Hacienda IMIG + IPC INDEC",
     lambda: _serie_var_real_vs_2023([gestion.FUNC_SALARIOS_ID, gestion.FUNC_OTROS_ID])),
    ("masa_salarial", "% real vs mismo mes 2023", "Sec. Hacienda AIF + IPC INDEC",
     lambda: _serie_var_real_vs_2023([gestion.REMUNERACIONES_ID])),
    ("libertad_opcion_salud", "% usuarios de prepagas con derivación directa", "SSS — RNAS/RNEMP", fetch_opcion_salud_serie),
    ("litigiosidad_laboral", "% i.a. juicios SRT (12m vs 12m)", "SRT — serie de litigiosidad", fetch_litigiosidad_serie),
    ("cepo_mulc", "% brecha CCL/mayorista (prom. mensual)", "ArgentinaDatos (CCL) + BCRA (A3500)", fetch_brecha_serie),
    ("protestas_caba", "eventos de protesta/mes (CABA)", "ACLED — agregado semanal (acleddata.com)", fetch_protestas_serie),
]


# ── Componente B — datos duros de migración real (ADR-0035) ──────────────────
# Cruce de validación del Componente A (Google Trends, intención expresada):
# flujos reales de migración argentina hacia los principales destinos. NUNCA
# se puntúa (mezclaría proxies de naturaleza distinta sin una paramétrica que
# los pondere) -- se expone como contexto duro en la card de
# indice_intencion_migratoria (espiritu_epoca.py), nunca como indicador solo.

COMPONENTE_B_STORE = Path(__file__).resolve().parents[1] / "data" / "vida" / "componente_b_migracion.json"

MESES_EN = ["JANUARY", "FEBRUARY", "MARCH", "APRIL", "MAY", "JUNE", "JULY",
            "AUGUST", "SEPTEMBER", "OCTOBER", "NOVEMBER", "DECEMBER"]


def _fiscal_year_eeuu(anio: int, mes: int) -> int:
    """Año fiscal de EEUU (oct->sep): oct/nov/dic pertenecen al FY siguiente."""
    return anio + 1 if mes >= 10 else anio


def _meses_desde(anio_inicio: int, mes_inicio: int) -> list:
    """[(anio, mes), ...] desde (anio_inicio, mes_inicio) hasta el mes anterior
    al actual (el mes en curso todavía no está publicado)."""
    hoy = datetime.today()
    out = []
    y, m = anio_inicio, mes_inicio
    while (y, m) < (hoy.year, hoy.month):
        out.append((y, m))
        m = 1 if m == 12 else m + 1
        y = y + 1 if m == 1 else y
    return out


def _fetch_visa_mes_argentina(tipo: str, anio: int, mes: int) -> int | None:
    """Descarga el Excel mensual del State Dept (NIV o IV) de un mes puntual
    y devuelve el total de emisiones a Argentina (todas las clases de visa
    sumadas). None si el archivo todavía no está publicado o falla.
    Prueba dos variantes de nombre: el State Dept publicó marzo-septiembre
    2024 con DOBLE espacio antes del guión ("MARCH 2024  - NIV...", typo real
    del sitio, verificado 2026-07-10) en vez del espacio simple habitual."""
    fy = _fiscal_year_eeuu(anio, mes)
    nombre_mes = MESES_EN[mes - 1]
    if tipo == "niv":
        carpeta = "Non-Immigrant-Statistics/MonthlyNIVIssuances"
        sufijo = "NIV Issuances by Nationality and Visa Class.xlsx"
    else:
        carpeta = "Immigrant-Statistics/MonthlyIVIssuances"
        sufijo = "IV Issuances by FSC or Place of Birth and Visa Class.xlsx"
    base = f"https://travel.state.gov/content/dam/visas/Statistics/{carpeta}/Excel/FY{fy}/"
    for espacios in (" ", "  "):
        archivo = f"{nombre_mes} {anio}{espacios}- {sufijo}"
        url = base + archivo.replace(" ", "%20")
        try:
            r = requests.get(url, headers=HTTP_HEADERS, timeout=HTTP_TIMEOUT)
            if r.status_code == 404:
                continue
            r.raise_for_status()
            import openpyxl
            wb = openpyxl.load_workbook(io.BytesIO(r.content), read_only=True, data_only=True)
            ws = wb[wb.sheetnames[0]]
            total, encontrado = 0, False
            for row in ws.iter_rows(min_row=3, values_only=True):
                if row[0] == "Argentina":
                    total += int(row[2] or 0)
                    encontrado = True
            if encontrado:
                return total
        except Exception as e:
            print(f"  [WARN] componente_b {tipo} {anio}-{mes:02d}: {str(e)[:80]}")
    return None


def fetch_componente_b_eeuu_niv_mensual(store: dict) -> dict:
    """Serie MENSUAL de visas NIV (no-inmigrante, incluye turismo) emitidas a
    argentinos -- backfill incremental desde dic-2023, solo pide los meses
    que todavía no están en el store (no re-descarga los ya guardados)."""
    mensual = dict(store.get("eeuu_niv", {}).get("mensual", {}))
    for a, m in _meses_desde(2023, 12):
        clave = f"{a}-{m:02d}"
        if clave in mensual:
            continue
        total = _fetch_visa_mes_argentina("niv", a, m)
        if total is not None:
            mensual[clave] = total
        time.sleep(0.5)   # evitar el CAPTCHA blando de travel.state.gov ante ráfagas
    return {"mensual": mensual}


def fetch_componente_b_eeuu_iv_mensual(store: dict) -> dict:
    """Serie MENSUAL de visas IV (inmigrante permanente, green card) emitidas
    a argentinos -- backfill incremental desde dic-2023. Señal más directa de
    inmigración permanente real que el NIV (que incluye turismo/negocios)."""
    mensual = dict(store.get("eeuu_iv", {}).get("mensual", {}))
    for a, m in _meses_desde(2023, 12):
        clave = f"{a}-{m:02d}"
        if clave in mensual:
            continue
        total = _fetch_visa_mes_argentina("iv", a, m)
        if total is not None:
            mensual[clave] = total
        time.sleep(0.5)   # evitar el CAPTCHA blando de travel.state.gov ante ráfagas
    return {"mensual": mensual}


def fetch_componente_b_canada_mensual() -> dict:
    """Serie MENSUAL de residentes permanentes de Canadá por ciudadanía
    argentina (IRCC, datos abiertos, CSV completo 2015->hoy). Valores
    redondeados a 5 por privacidad -- no sumar como valor exacto, sirve para
    tendencia, no para totales precisos."""
    url = "https://www.ircc.canada.ca/opendata-donneesouvertes/data/ODP-PR-Citz.csv"
    r = requests.get(url, headers=HTTP_HEADERS, timeout=HTTP_TIMEOUT * 2)
    r.raise_for_status()
    meses_map = {m: i + 1 for i, m in enumerate(
        ["Jan", "Feb", "Mar", "Apr", "May", "Jun", "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"])}
    mensual = {}
    reader = csv.DictReader(io.StringIO(r.content.decode("utf-8-sig")), delimiter="\t")
    for row in reader:
        if row.get("EN_COUNTRY_OF_CITIZENSHIP") == "Argentina":
            mes = meses_map.get(row["EN_MONTH"])
            if mes:
                mensual[f"{row['EN_YEAR']}-{mes:02d}"] = int(row["TOTAL"])
    return {"mensual": mensual}


def fetch_componente_b_espana_anual() -> dict:
    """Serie ANUAL de argentinos que adquirieron la nacionalidad española
    (INE, Estadística de Adquisiciones de Nacionalidad Española de
    Residentes, tabla 15800 -- API pública, sin scraping)."""
    r = requests.get("https://servicios.ine.es/wstempus/js/ES/DATOS_TABLA/15800",
                      params={"nult": 15}, headers=HTTP_HEADERS, timeout=HTTP_TIMEOUT * 2)
    r.raise_for_status()
    anual = {}
    for serie in r.json():
        if serie["Nombre"].strip() == "Total Nacional. Dato base. Argentina. Total.":
            for punto in serie["Data"]:
                anual[str(punto["Anyo"])] = int(punto["Valor"])
            break
    return {"anual": anual}


def fetch_componente_b_italia_aire_anual() -> dict:
    """Serie ANUAL de argentinos que adquirieron la ciudadanía italiana
    (ISTAT, balance demográfico AIRE, columna 'Acquisizioni di cittadinanza',
    Paese=Argentina, Sesso=Totale -- ZIP/CSV descargable, sin scraping)."""
    import zipfile
    r = requests.get("https://demo.istat.it/data/aire/AIRE_it.zip",
                      headers=HTTP_HEADERS, timeout=HTTP_TIMEOUT * 2)
    r.raise_for_status()
    anual = {}
    with zipfile.ZipFile(io.BytesIO(r.content)) as z:
        with z.open("AIRE_it.csv") as f:
            reader = csv.DictReader(io.TextIOWrapper(f, encoding="utf-8-sig"), delimiter=";")
            for row in reader:
                if row["Paese"] == "Argentina" and row["Sesso"] == "Totale":
                    anual[row["Anno"]] = int(row["Acquisizioni di cittadinanza"])
    return {"anual": anual}


def fetch_componente_b_chile_anual() -> dict:
    """Serie ANUAL de residencias definitivas OTORGADAS a argentinos en Chile
    (SERMIG, ex-DEM, datos abiertos -- microdato RD-Resueltas, filtrado
    PAÍS=Argentina y TIPO_RESUELTO=Otorga, serie completa desde 2000)."""
    pagina = requests.get("https://serviciomigraciones.cl/estudios-migratorios/datos-abiertos",
                          headers=HTTP_HEADERS, timeout=HTTP_TIMEOUT)
    pagina.raise_for_status()
    m = re.search(r'href="(https://serviciomigraciones\.cl/[^"]*RD-Resueltas[^"]*\.xlsx)"',
                  pagina.text)
    if not m:
        raise ValueError("no se encontró el link de RD-Resueltas en datos-abiertos")
    r = requests.get(m.group(1), headers=HTTP_HEADERS, timeout=HTTP_TIMEOUT * 3)
    r.raise_for_status()
    import openpyxl
    wb = openpyxl.load_workbook(io.BytesIO(r.content), read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    anual, header = {}, None
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0:
            header = row
            continue
        d = dict(zip(header, row))
        if d.get("PAÍS") == "Argentina" and d.get("TIPO_RESUELTO") == "Otorga":
            anio = str(d["AÑO"])
            anual[anio] = anual.get(anio, 0) + int(d["Total"])
    return {"anual": anual}


def fetch_componente_b_store() -> dict:
    """Arma/actualiza data/vida/componente_b_migracion.json: datos duros de
    migración real (ADR-0035) para cruzar contra el Componente A (Trends).
    Ninguna de las 6 fuentes cambia más de una vez por mes -- gateado por
    frescura igual que el store de Trends: si ya se actualizó este mes
    calendario, no vuelve a pegarle a las 6 APIs/descargas. Cada fuente falla
    de forma independiente (try/except propio) -- si una cae, las demás
    igual se actualizan y esa mantiene su último valor sano."""
    store = json.loads(COMPONENTE_B_STORE.read_text(encoding="utf-8-sig")) \
        if COMPONENTE_B_STORE.exists() else {}

    mes_actual = datetime.today().strftime("%Y-%m")
    if store.get("_meta", {}).get("actualizado", "")[:7] == mes_actual:
        return store

    fetchers = {
        "eeuu_niv": lambda: fetch_componente_b_eeuu_niv_mensual(store),
        "eeuu_iv": lambda: fetch_componente_b_eeuu_iv_mensual(store),
        "canada_pr": fetch_componente_b_canada_mensual,
        "espana_nacionalidad": fetch_componente_b_espana_anual,
        "italia_aire": fetch_componente_b_italia_aire_anual,
        "chile_residencia": fetch_componente_b_chile_anual,
    }
    fuentes_meta = {
        "eeuu_niv": "US State Dept — NIV Issuances by Nationality (mensual)",
        "eeuu_iv": "US State Dept — IV Issuances by FSC or Place of Birth (mensual)",
        "canada_pr": "IRCC — Permanent Residents by Country of Citizenship (mensual, redondeado a 5)",
        "espana_nacionalidad": "INE — Adquisiciones de Nacionalidad Española de Residentes (anual)",
        "italia_aire": "ISTAT/AIRE — Acquisizioni di cittadinanza (anual)",
        "chile_residencia": "SERMIG (ex-DEM) — Residencias Definitivas Resueltas (anual)",
    }
    fallidas = 0
    for clave, fetcher in fetchers.items():
        try:
            store[clave] = fetcher()
        except Exception as e:
            fallidas += 1
            print(f"  [WARN] componente_b.{clave}: {str(e)[:100]} -- se mantiene el valor anterior")
    # El gate de frescura solo avanza si TODAS las fuentes respondieron: si
    # se sellara "actualizado" con fuentes caídas, el gate saltearía los
    # reintentos durante el resto del mes y los datos viejos quedarían
    # presentados bajo una fecha de actualización nueva.
    meta = {"fuentes": fuentes_meta}
    if fallidas == 0:
        meta["actualizado"] = datetime.today().strftime("%Y-%m-%d")
    else:
        anterior = store.get("_meta", {}).get("actualizado")
        if anterior:
            meta["actualizado"] = anterior
        print(f"  [WARN] componente_b: {fallidas} fuente(s) fallida(s) -- "
              f"el gate mensual no avanza, la próxima corrida reintenta")
    store["_meta"] = meta
    COMPONENTE_B_STORE.parent.mkdir(parents=True, exist_ok=True)
    COMPONENTE_B_STORE.write_text(json.dumps(store, indent=2, ensure_ascii=False), encoding="utf-8")
    return store


CINTURONES_SERIES = {
    "macro": (MACRO_INDEC, MACRO_BCRA, MACRO_DERIVADAS),
    "politica": (POLITICA_INDEC, [], POLITICA_DERIVADAS),
    "vida_cotidiana": (VIDA_INDEC, [], VIDA_DERIVADAS),
    "gestion": (GESTION_INDEC, [], GESTION_DERIVADAS),
}


def _nombres_de(indec: list, bcra: list, derivadas: list) -> set:
    return ({n for _, n, *_ in indec} | {n for _, n, *_ in bcra}
            | {n for n, *_ in derivadas})


def _cinturon_de_indicador(nombre: str) -> str | None:
    for cinturon, (indec, bcra, derivadas) in CINTURONES_SERIES.items():
        if nombre in _nombres_de(indec, bcra, derivadas):
            return cinturon
    return None


if __name__ == "__main__":
    import argparse

    parser = argparse.ArgumentParser(
        description="Backfill de series históricas. Sin argumentos: los 4 "
                     "cinturones completos (comportamiento de siempre). Con "
                     "--cinturon o --indicador: corrida acotada -- NO toca "
                     "las series de los cinturones/indicadores no incluidos, "
                     "así no queda card↔serie desincronizado en el resto "
                     "(ver CLAUDE.md).")
    parser.add_argument("--cinturon", choices=sorted(CINTURONES_SERIES),
                         help="Solo este cinturón (los otros 3 quedan intactos).")
    parser.add_argument("--indicador",
                         help="Solo este indicador (el cinturón se detecta solo; "
                              "el resto del cinturón se preserva vía merge).")
    args = parser.parse_args()

    if args.indicador and not args.cinturon:
        args.cinturon = _cinturon_de_indicador(args.indicador)
        if args.cinturon is None:
            raise SystemExit(f"'{args.indicador}' no está en ningún cinturón "
                              f"(revisar INDEC/BCRA/DERIVADAS de cada uno)")

    cinturones = [args.cinturon] if args.cinturon else list(CINTURONES_SERIES)
    for c in cinturones:
        indec, bcra, derivadas = CINTURONES_SERIES[c]
        print(f"\n=== {c.upper()} ===")
        descargar(c, indec, bcra, derivadas, solo_indicador=args.indicador)

    print(f"\nCSVs en {OUTPUT_DIR.resolve()}")
