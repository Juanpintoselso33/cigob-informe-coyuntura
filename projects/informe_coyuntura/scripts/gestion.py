"""
Colector Cinturón Gestión — CIGOB
Mide la brecha entre lo PROMETIDO y lo EJECUTADO de la agenda de reformas
(dic-2023–). Puntúa con el ITCG (Índice de Tensión del Cinturón de Gestión,
paramétrica CIGOB doc 260702, jul-2026): escala 0-100, 0 = promete pero no
ejecuta (cinturón apretado), 100 = agenda ejecutándose. La tensión 0-10 del
informe es (100 − ITCG) / 10. Bandas, dimensiones y pesos en scripts/itcg.py;
overrides del analista en data/gestion/ajustes_itcg.json.

CINCO DIMENSIONES (35/25/15/15/10):
  1. Reformas económicas — cepo_mulc (brecha CCL/mayorista, dolarapi),
     apertura_comercial (ILCE: brecha inversa + alícuota efectiva DEX+DIM/ICA),
     desregulacion_normativa (proxy InfoLeg "deroga").
  2. Reforma del Estado — reduccion_estado (dotación APN, XLSX INDEC mensual),
     gasto_funcionamiento (IMIG salarios+otros, real vs mismo mes 2023),
     masa_salarial (AIF SPN remuneraciones, real vs mismo mes 2023),
     reestructuracion_organismos (InfoLeg "disolucion").
  3. Reforma laboral — fal_modernizacion_laboral (Fondo de Cese, manual:
     adopción por CCT sin fuente estructurada; CNV RG 1071/2025 como contexto).
  4. Privatizaciones e inversión — privatizaciones (etapas 0-4 por empresa,
     store data/gestion/privatizaciones.json actualizado con el BO),
     rigi_inversiones (plataforma oficial, ADR-0011),
     concesiones_infraestructura (km adjudicados de la Red Federal de
     Concesiones: CONTRAT.AR + página oficial RFC, ADR-0016).
  5. Reforma social y orden — asistencia_directa (TDPS contra la ejecución
     presupuestaria real, API Presupuesto Abierto, ADR-0015),
     protocolo_antipiquetes (reducción de cortes CABA vs 2023, manual: el
     registro histórico del GCBA está muerto; el poller GTFS-RT acumula la
     serie que lo reemplazará, ADR-0014),
     libertad_opcion_salud (derivación directa a prepagas inscriptas:
     padrones RNAS/RNEMP de la SSS, ADR-0016).

Indicadores AUTO (16): cepo_mulc, apertura_comercial, desregulacion_normativa,
  reduccion_estado, gasto_funcionamiento, masa_salarial,
  reestructuracion_organismos, fal_modernizacion_laboral (CNV),
  litigiosidad_laboral (contexto), privatizaciones (store curado),
  rigi_inversiones, concesiones_infraestructura, asistencia_directa (TDPS),
  libertad_opcion_salud, alertas_manifestacion (contexto),
  protestas_caba (contexto, ACLED).
Indicadores manuales: 0 — protocolo_antipiquetes se automatizó con los
anclajes públicos de Diagnóstico Político (ADR-0025); manuales.json queda
solo como fallback.

Nota INDEC www: el WAF de indec.gob.ar resetea el handshake TLS de Python
(fingerprinting del ClientHello); el fetch del XLSX cae a curl como fallback.
"""
import io
import os
import re
import sys
import calendar
import json
import subprocess
import requests
import logging
from datetime import datetime, date
from pathlib import Path

sys.stdout.reconfigure(encoding="utf-8")
sys.path.insert(0, str(Path(__file__).parent))

import itcg

try:  # BCRA con verify=False (cadena de certificados incompleta del BCRA)
    requests.packages.urllib3.disable_warnings()
except Exception:
    pass

# ── Paths ─────────────────────────────────────────────────────────────────────
SCRIPT_DIR           = Path(__file__).parent
PROJECT_DIR          = SCRIPT_DIR.parent
CACHE_PATH           = PROJECT_DIR / "output" / "cache" / "gestion.json"
MANUALES_PATH        = PROJECT_DIR / "data" / "gestion" / "manuales.json"
PRIVATIZACIONES_PATH = PROJECT_DIR / "data" / "gestion" / "privatizaciones.json"
AJUSTES_PATH         = PROJECT_DIR / "data" / "gestion" / "ajustes_itcg.json"
RIGI_FECHAS_PATH     = PROJECT_DIR / "data" / "gestion" / "rigi_fechas.json"

# ── URLs ──────────────────────────────────────────────────────────────────────
DOLARAPI_URL        = "https://dolarapi.com/v1/dolares"
INDEC_SERIES_BASE   = "https://apis.datos.gob.ar/series/api/series/"
INFOLEG_HOME        = "https://servicios.infoleg.gob.ar/infolegInternet/"
INDEC_DOTACION_URL  = "https://www.indec.gob.ar/ftp/cuadros/economia/serie_dotacion_apn.xlsx"
BCRA_MONETARIAS_BASE = "https://api.bcra.gob.ar/estadisticas/v4.0/monetarias"
RIGI_PORTAL_URL     = "https://www.argentina.gob.ar/economia/rigi"
# Plataforma oficial del RIGI (Min. Economía, jun-2026): el mapa interactivo se
# alimenta de un Google Sheet público. Lo leemos vía gviz CSV (sin auth).
# Pestañas: "dataset" (proyectos aprobados, fila por proyecto) · "evaluacion" (agregados).
RIGI_SHEET_ID       = "1eytHJrzUjIFOXI-P1Hx_wbmZiSqPxVle059Djdos6u8"
RIGI_SHEET_GVIZ     = "https://docs.google.com/spreadsheets/d/{sid}/gviz/tq?tqx=out:csv&sheet={tab}"

# datos.gob.ar — series verificadas 2026-07-02 (último dato may-2026 salvo IPC)
IPC_ID              = "148.3_INIVELNAL_DICI_M_26"   # IPC nivel nacional (deflactor)
EXPO_ICA_ID         = "74.3_IET_0_M_16"             # ICA exportaciones totales (M USD)
IMPO_ICA_ID         = "74.3_IIT_0_M_25"             # ICA importaciones totales (M USD)
DEX_ID              = "142.3_DEREC_2001_M_20"       # Recaudación derechos de exportación (M ARS)
DIM_ID              = "142.3_DEREC_2001_M_26"       # Recaudación derechos de importación (M ARS)
REMUNERACIONES_ID   = "379.9_GTOS_CORR_017__49_26"  # AIF SPN: remuneraciones (M ARS, mensual)
FUNC_SALARIOS_ID    = "452.2_SALARIOSIOS_0_T_8_22"  # IMIG gastos de funcionamiento: salarios
FUNC_OTROS_ID       = "452.2_OTROS_GASTNTO_0_T_27_55"  # IMIG funcionamiento: otros gastos
BCRA_TC_MAYOR_ID    = 5                             # TC mayorista de referencia A3500 (ARS/USD)

HTTP_TIMEOUT = 60
HTTP_HEADERS = {"User-Agent": "Mozilla/5.0 (compatible; CIGOB-Monitor/1.0)"}

logging.basicConfig(level=logging.WARNING, format="%(message)s")

CINTURON = "gestion"
INDICADORES_ESPERADOS = [
    # D1 — Reformas económicas fundamentales
    "cepo_mulc", "apertura_comercial", "desregulacion_normativa",
    # D2 — Reforma del Estado
    "reduccion_estado", "gasto_funcionamiento", "masa_salarial",
    "reestructuracion_organismos",
    # D3 — Reforma laboral (+ litigiosidad SRT como contexto, fuera del índice)
    "fal_modernizacion_laboral", "litigiosidad_laboral",
    # D4 — Privatizaciones e inversión
    "privatizaciones", "rigi_inversiones", "concesiones_infraestructura",
    # D5 — Reforma social y orden (+ contexto: alertas GTFS-RT y protestas ACLED)
    "asistencia_directa", "protocolo_antipiquetes", "libertad_opcion_salud",
    "alertas_manifestacion", "protestas_caba",
]

# Calibración del proxy InfoLeg de reestructuración: 18 actos = avance 40%
# (validado con estimación manual, ver docstring original); 45 = plan completo.
ORGANISMOS_PLAN_TOTAL = 45


# ── Helpers ───────────────────────────────────────────────────────────────────

def load_cache() -> dict:
    if CACHE_PATH.exists():
        with open(CACHE_PATH, encoding="utf-8") as f:
            return json.load(f)
    return {}


def save_cache(data: dict) -> None:
    CACHE_PATH.parent.mkdir(parents=True, exist_ok=True)
    with open(CACHE_PATH, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2, default=str)


def load_manuales() -> dict:
    if not MANUALES_PATH.exists():
        return {}
    with open(MANUALES_PATH, encoding="utf-8") as f:
        raw = json.load(f)
    return {k: v for k, v in raw.items() if not k.startswith("_")}


def _warn(ind: str, err: Exception) -> None:
    print(f"[WARN] {CINTURON}.{ind}: {err}. Usando fallback.")


def _manual_entry(nombre: str, manuales: dict, cache_anterior: dict) -> dict | None:
    """Entrada desde manuales.json (valor = insumo que bandea el ITCG) o cache
    anterior como último recurso. Manual ⇒ desactualizado=True (badge honesto)."""
    m = manuales.get(nombre, {})
    valor = m.get("valor", m.get("avance_pct"))   # compat con el formato viejo
    if valor is not None:
        return {
            "valor":          valor,
            "unidad":         m.get("unidad", ""),
            "fuente":         m.get("fuente", "manual"),
            "fecha_dato":     m.get("fecha_dato", ""),
            "estado":         m.get("estado", "manual"),
            "notas":          m.get("notas", ""),
            "desactualizado": True,
        }
    prev = cache_anterior.get("indicadores", {}).get(nombre, {})
    if prev and prev.get("valor") is not None:
        return {**prev, "desactualizado": True}
    return None


def _indec_serie(series_id: str, limit: int = 16) -> list:
    params = {"ids": series_id, "format": "json", "limit": limit, "sort": "desc"}
    r = requests.get(INDEC_SERIES_BASE, params=params, headers=HTTP_HEADERS, timeout=HTTP_TIMEOUT)
    r.raise_for_status()
    return r.json()["data"]


def _indec_nivel_mensual(series_id: str, limit: int = 48) -> dict:
    """{YYYY-MM: valor} de una serie mensual de datos.gob.ar."""
    return {row[0][:7]: float(row[1]) for row in _indec_serie(series_id, limit=limit)
            if row[1] is not None}


def _infoleg_post(texto: str, tipo_norma: str, fecha_desde: tuple, fecha_hasta: tuple) -> int:
    """
    Sesión POST en InfoLeg con reintento automático (2 intentos).
    Retorna el conteo de normas encontradas.
    fecha_desde / fecha_hasta: (dia, mes, anio) como strings.
    """
    last_err = None
    for attempt in range(2):
        try:
            session = requests.Session()
            r_home = session.get(INFOLEG_HOME, headers=HTTP_HEADERS, timeout=HTTP_TIMEOUT)
            r_home.raise_for_status()

            action_m = re.search(r'action="(/infolegInternet/[^"]+)"', r_home.text)
            if not action_m:
                raise ValueError("No se encontró form action URL en InfoLeg home")
            action_url = "https://servicios.infoleg.gob.ar" + action_m.group(1)

            post_data = {
                "tipoNorma":    tipo_norma,
                "numero":       "",
                "anioSancion":  "",
                "dependencia":  "",
                "diaPubDesde":  fecha_desde[0],
                "mesPubDesde":  fecha_desde[1],
                "anioPubDesde": fecha_desde[2],
                "diaPubHasta":  fecha_hasta[0],
                "mesPubHasta":  fecha_hasta[1],
                "anioPubHasta": fecha_hasta[2],
                "texto":        texto,
            }
            r = session.post(action_url, data=post_data, headers=HTTP_HEADERS, timeout=HTTP_TIMEOUT)
            r.raise_for_status()

            if re.search(r"No se encontraron normas", r.text, re.IGNORECASE):
                return 0    # cero resultados es un dato, no un error (series al inicio del mandato)
            m = re.search(r"Encontradas?[:\s]+(\d+)", r.text, re.IGNORECASE)
            if not m:
                raise ValueError("Conteo no encontrado en respuesta InfoLeg")
            return int(m.group(1))
        except Exception as e:
            last_err = e
            if attempt == 0:
                print(f"[RETRY] infoleg texto='{texto}': {e}. Reintentando...")
    raise last_err


def _http_get_resiliente(url: str) -> bytes:
    """GET binario con fallback a curl: el WAF de indec.gob.ar (www) resetea el
    handshake TLS de Python (fingerprinting del ClientHello) pero acepta curl."""
    try:
        r = requests.get(url, headers=HTTP_HEADERS, timeout=HTTP_TIMEOUT)
        r.raise_for_status()
        return r.content
    except requests.exceptions.ConnectionError:
        out = subprocess.run(["curl", "-sS", "--fail", "-L", url],
                             capture_output=True, timeout=HTTP_TIMEOUT * 2)
        if out.returncode != 0:
            raise RuntimeError(f"curl fallback: {out.stderr.decode(errors='replace')[:200]}")
        return out.stdout


# ── D1: Reformas económicas fundamentales ─────────────────────────────────────

def fetch_cepo_mulc() -> dict | None:
    """
    Brecha CCL/mayorista como proxy del cepo corporativo (giro de dividendos, capitales).
    El mayorista (referencia A3500 del BCRA) es la tasa correcta para medir la brecha
    cambiaria: el "oficial" minorista ya incluye el spread del banco y subestima la brecha.
    Blue≈0% porque el cepo minorista se levantó (abr-2025); el CCL mide la restricción
    que persiste para empresas: repatriación de utilidades, acceso al MULC para capital.
    Doc 260702: brecha sostenida <10-15% = condiciones óptimas de unificación.
    """
    try:
        r = requests.get(DOLARAPI_URL, headers=HTTP_HEADERS, timeout=HTTP_TIMEOUT)
        r.raise_for_status()
        dolares = {d["casa"]: d for d in r.json()}
        ccl       = dolares.get("contadoconliqui", {}).get("venta")
        mayorista = dolares.get("mayorista",       {}).get("venta")
        if not ccl or not mayorista:
            raise ValueError("CCL o mayorista no encontrado en respuesta")
        brecha = round((float(ccl) - float(mayorista)) / float(mayorista) * 100.0, 2)
        return {
            "valor":          brecha,
            "unidad":         "% de brecha CCL/mayorista",
            "fuente":         DOLARAPI_URL,
            "fecha_dato":     date.today().isoformat(),
            "desactualizado": False,
        }
    except Exception as e:
        _warn("cepo_mulc", e)
        return None


def _tc_mayorista_promedio_por_mes(dias: int = 100) -> dict:
    """{YYYY-MM: TC A3500 promedio del mes} desde la API del BCRA (v4.0)."""
    desde = date.fromordinal(date.today().toordinal() - dias).isoformat()
    r = requests.get(f"{BCRA_MONETARIAS_BASE}/{BCRA_TC_MAYOR_ID}",
                     params={"desde": desde, "limit": 3000},
                     headers=HTTP_HEADERS, timeout=HTTP_TIMEOUT, verify=False)
    r.raise_for_status()
    detalle = r.json()["results"][0]["detalle"]
    por_mes: dict = {}
    for d in detalle:
        por_mes.setdefault(d["fecha"][:7], []).append(float(d["valor"]))
    return {ym: sum(vs) / len(vs) for ym, vs in por_mes.items()}


# Ancla de la alícuota efectiva (ILCE, componente arancelario): 15% del
# intercambio total ≈ cierre comercial de hecho (0 puntos); 0% = libre comercio
# (100 puntos). Lineal entre ambos. Referencia: la alícuota argentina se movió
# en 4-8% del comercio total en 2016-2026 (ADR-0013).
ALICUOTA_CIERRE_PCT = 15.0


def fetch_apertura_comercial(brecha_pct: float | None = None) -> dict | None:
    """
    Alícuota efectiva del comercio exterior (ADR-0021): recaudación de
    derechos de importación + exportación (ARCA, en ARS → USD por el A3500
    promedio del mes) sobre el intercambio total del ICA (expo+impo, USD).
    0% = comercio libre de fricción arancelaria · ≥15% = cierre de hecho
    (la lineal del doc 260702, reproducida por las anclas de banda del ITCG).

    Historia: hasta jul-2026 este indicador era el ILCE (compuesto con la
    brecha cambiaria al 40%) — la brecha ya puntúa como indicador propio
    (cepo_mulc) y el compuesto la hacía pesar doble en la dimensión (doble
    conteo, ADR-0021). El canal verde aduanero (20% en el doc) sigue sin
    fuente pública estructurada. `brecha_pct` se acepta y se ignora
    (compatibilidad con el orquestador).
    """
    try:
        dex  = _indec_nivel_mensual(DEX_ID, limit=8)
        dim  = _indec_nivel_mensual(DIM_ID, limit=8)
        expo = _indec_nivel_mensual(EXPO_ICA_ID, limit=8)
        impo = _indec_nivel_mensual(IMPO_ICA_ID, limit=8)
        tc   = _tc_mayorista_promedio_por_mes()
        comunes = sorted(set(dex) & set(dim) & set(expo) & set(impo) & set(tc))
        if not comunes:
            raise ValueError("sin mes común entre recaudación DEX/DIM, ICA y TC A3500")
        ym = comunes[-1]
        recaudacion_musd = (dex[ym] + dim[ym]) / tc[ym]      # M ARS / (ARS/USD) = M USD
        comercio_musd    = expo[ym] + impo[ym]
        if comercio_musd <= 0:
            raise ValueError("intercambio comercial nulo")
        alicuota = 100.0 * recaudacion_musd / comercio_musd
        miles = lambda x: f"{x:,.0f}".replace(",", ".")
        return {
            "valor":          round(alicuota, 2),
            "unidad":         "% del intercambio (alícuota efectiva)",
            "fuente":         "ARCA (DEX+DIM) + INDEC ICA + BCRA A3500",
            "fecha_dato":     f"{ym}-01",
            "desactualizado": False,
            "detalle_txt": (f"US$ {miles(recaudacion_musd)} M recaudados por derechos de "
                            f"impo+expo sobre US$ {miles(comercio_musd)} M de intercambio "
                            f"({ym}) · canal aduanero sin fuente pública"),
        }
    except Exception as e:
        _warn("apertura_comercial", e)
        return None


# ── D1: desregulación — normas efectivamente derogadas (ADR-0096) ────────────
# La versión anterior contaba las normas cuyo TEXTO COMPLETO contenía la palabra
# "deroga", y trataba ese conteo como porcentaje de avance contra una meta
# interna de 100 normas. Dos problemas medidos el 2026-07-20:
#
#   1. Cerca de la mitad de lo que contaba no derogaba nada. La búsqueda de
#      InfoLeg matchea la palabra en cualquier parte del documento, incluidos
#      los considerandos que mencionan la derogación hecha por OTRA norma. Entre
#      lo contado había una resolución de Cancillería sobre un nombramiento
#      episcopal, que menciona un decreto derogado de la Sagrada Congregación
#      Consistorial.
#   2. El DNU 70/23 —que deroga 38 normas completas y modifica otras 32— pesaba
#      exactamente lo mismo que un decreto que elimina un trámite: uno.
#
# Ahora se cuentan NORMAS DEROGADAS, no actos derogatorios, y sólo si la
# derogación está en la parte dispositiva. El texto de una norma publicada es
# inmutable, así que el análisis de cada una se cachea para siempre y cada
# corrida sólo baja las nuevas.

INFOLEG_ANEXO = "https://servicios.infoleg.gob.ar/infolegInternet/anexos/{lo}-{hi}/{id}/norma.htm"
DESREG_STORE = PROJECT_DIR / "data" / "gestion" / "desregulacion_normas.json"

# ── Fuente oficial del indicador desde ADR-0125 ──────────────────────────────
# Ministerio de Desregulación y Transformación del Estado, informe mensual
# "Análisis de la desregulación implementada" (Unidad de Evaluación de Impacto).
# Cuenta normas de desregulación acumuladas desde el 10-dic-2023.
#
# Los nombres de archivo son IRREGULARES (_junio, _junio_1, _mayo_2026, _agos,
# _diciembre_enero, y un typo oficial en "analsisi"), así que los enlaces se
# resuelven leyendo la página — mismo patrón que las colocaciones de deuda en
# macro.py. Nunca construir la URL a mano.
DESREG_OFICIAL_PAGINA = "https://www.argentina.gob.ar/desregulacion/desregulacion-en-numeros"
DESREG_OFICIAL_STORE  = PROJECT_DIR / "data" / "gestion" / "desregulacion_oficial.json"

# El informe de abril-2026 es el último del formato con "Figura 1. Cantidad de
# normas de desregulación acumuladas": un gráfico de barras MENSUAL desde
# dic-2023. De ahí sale el backfill (ADR-0125). Es un documento publicado e
# inmutable, así que el resultado se cachea de forma permanente.
DESREG_INFORME_BACKFILL = "informe_analisis_de_la_desregulacion_implementada_abril.pdf"
DESREG_BACKFILL_DESDE   = "2023-12"

# Frontera entre considerandos y parte dispositiva. Todo lo que está antes es
# relato de antecedentes: ahí "deroga" habla de lo que hizo otro.
_DISPOSITIVA = re.compile(r"\b(RESUELVE|RESUELVEN|DECRETA|DECRETAN|DISPONE|DISPONEN)\s*:", re.I)
# Verbo derogatorio en forma imperativa/resolutiva (no "fue derogada por").
_VERBO_DEROGA = re.compile(r"[Dd]er[oó]g(?:ar|[uú]ese|an?se)|[Dd][eé]jan?se sin efecto", re.I)
# Lo que se deroga va desde una ley entera hasta "el punto 9) del apartado E)
# del artículo 20". Son objetos incommensurables, así que se separan: sólo se
# cuentan las normas COMPLETAS y las derogaciones parciales se relevan aparte,
# como contexto (ADR-0096). Si la cláusula empieza nombrando una parte, es
# parcial por más que después mencione la norma a la que pertenece.
_PARTE_DE_NORMA = re.compile(
    r"^\s*(?:el|la|los|las)?\s*"
    r"(art[ií]culos?|secci[oó]n|t[ií]tulos?|cap[ií]tulos?|puntos?|apartados?|incisos?|anexos?)",
    re.I)
# Identificador de norma: 20.680 · 20680 · 71/2020
_ID_NORMA = re.compile(
    r"(?:N[°ºo]?\.?\s*)((?:\d{1,2}\.\d{3}|\d{4,5})(?:/\d{2,4})?|\d{1,4}/\d{2,4})")
_CLAUSULA_DEROGA = re.compile(r"[Dd]er[oó]g(?:ar|[uú]ese|an?se)\s*,?\s*(.{0,220})")


def _desreg_cargar_store() -> dict:
    if DESREG_STORE.exists():
        try:
            return json.loads(DESREG_STORE.read_text(encoding="utf-8-sig"))
        except Exception:
            pass
    return {"_meta": {"descripcion":
             "Análisis por norma del indicador desregulacion_normativa (ADR-0096). "
             "El texto de una norma publicada no cambia, así que cada entrada se "
             "calcula una sola vez y se conserva. Sin esta caché cada corrida "
             "bajaría 60+ documentos de InfoLeg."},
            "normas": {}}


def _infoleg_buscar_mes(texto: str, anio: int, mes: int, session=None) -> list:
    """[(id, titulo)] de normas publicadas en ese mes cuyo texto contiene `texto`.

    Se enumera mes a mes y no de una sola vez porque el buscador no pagina de
    forma estable; con ventanas mensuales ningún mes supera una página
    (verificado sobre dic-2023 → jul-2026: máximo 5 resultados)."""
    from bs4 import BeautifulSoup
    ultimo = calendar.monthrange(anio, mes)[1]
    ses = session or requests.Session()
    r_home = ses.get(INFOLEG_HOME, headers=HTTP_HEADERS, timeout=HTTP_TIMEOUT)
    r_home.raise_for_status()
    m = re.search(r'action="(/infolegInternet/[^"]+)"', r_home.text)
    if not m:
        raise ValueError("InfoLeg: no se encontró la URL del formulario")
    url = "https://servicios.infoleg.gob.ar" + m.group(1)
    datos = {"tipoNorma": "", "numero": "", "anioSancion": "", "dependencia": "",
             "diaPubDesde": "01", "mesPubDesde": f"{mes:02d}", "anioPubDesde": str(anio),
             "diaPubHasta": f"{ultimo:02d}", "mesPubHasta": f"{mes:02d}",
             "anioPubHasta": str(anio), "texto": texto}
    r = ses.post(url, data=datos, headers=HTTP_HEADERS, timeout=HTTP_TIMEOUT)
    r.raise_for_status()
    vistos, out = [], set()
    for a in BeautifulSoup(r.text, "html.parser").find_all("a", href=True):
        g = re.search(r"verNorma\.do\?(?:resaltar=true&)?id=(\d+)", a["href"])
        if g and g.group(1) not in out:
            out.add(g.group(1))
            # el enlace trae el tipo y el número partidos por tabulaciones y
            # saltos; sin colapsar, el título sale ilegible
            titulo = re.sub(r"\s+", " ", a.get_text(" ", strip=True)).strip()
            vistos.append((g.group(1), titulo[:180]))
    return vistos


def _infoleg_listar_deroga(anio: int, mes: int) -> list:
    """[(id, 'YYYY-MM-DD')] de normas del mes cuyo texto menciona 'deroga'."""
    return [(nid, f"{anio}-{mes:02d}-01")
            for nid, _titulo in _infoleg_buscar_mes("deroga", anio, mes)]


def _infoleg_analizar_norma(norma_id: str) -> dict:
    """{dispositiva, derogadas, listado} de una norma, desde su texto completo."""
    from bs4 import BeautifulSoup
    lo = (int(norma_id) // 5000) * 5000
    url = INFOLEG_ANEXO.format(lo=lo, hi=lo + 4999, id=norma_id)
    r = requests.get(url, headers=HTTP_HEADERS, timeout=HTTP_TIMEOUT * 3)
    if r.status_code != 200:
        raise ValueError(f"texto no disponible (HTTP {r.status_code})")
    texto = re.sub(r"\s+", " ", BeautifulSoup(r.content, "html.parser").get_text(" ", strip=True))

    corte = _DISPOSITIVA.search(texto)
    cuerpo = texto[corte.end():] if corte else texto
    dispositiva = bool(_VERBO_DEROGA.search(cuerpo))

    completas, parciales = set(), 0
    if dispositiva:
        for m in _CLAUSULA_DEROGA.finditer(cuerpo):
            fragmento = m.group(1)
            if _PARTE_DE_NORMA.match(fragmento):
                parciales += 1
                continue
            ids = set(_ID_NORMA.findall(fragmento))
            if ids:
                completas |= ids
            else:
                # deroga algo sin identificador numérico (una sección de un
                # digesto, un régimen nombrado en prosa): cuenta como parcial
                parciales += 1
    return {"dispositiva": dispositiva,
            "derogadas": len(completas),
            "parciales": parciales,
            "listado": sorted(completas)}


_DESREG_MESES = {"enero": 1, "febrero": 2, "marzo": 3, "abril": 4, "mayo": 5,
                 "junio": 6, "julio": 7, "agosto": 8, "septiembre": 9,
                 "setiembre": 9, "octubre": 10, "noviembre": 11, "diciembre": 12}


def _mes_previo_ym(ym: str) -> str:
    """'2026-01' → '2025-12'. Aritmética de calendario, no de posición."""
    total = int(ym[:4]) * 12 + (int(ym[5:7]) - 1) - 1
    return f"{total // 12}-{total % 12 + 1:02d}"


def _desreg_oficial_store() -> dict:
    if DESREG_OFICIAL_STORE.exists():
        try:
            return json.loads(DESREG_OFICIAL_STORE.read_text(encoding="utf-8-sig"))
        except Exception:
            pass
    return {}


def _desreg_guardar_store(store: dict) -> None:
    DESREG_OFICIAL_STORE.parent.mkdir(parents=True, exist_ok=True)
    DESREG_OFICIAL_STORE.write_text(
        json.dumps(store, indent=1, ensure_ascii=False, sort_keys=True), encoding="utf-8")


def _desreg_informes_publicados() -> dict:
    """{nombre_archivo: url} de los informes enlazados hoy en la página oficial."""
    r = requests.get(DESREG_OFICIAL_PAGINA, headers=HTTP_HEADERS, timeout=HTTP_TIMEOUT)
    r.raise_for_status()
    out = {}
    for href in set(re.findall(r'href="([^"]*\.pdf[^"]*)"', r.text, re.I)):
        url = href if href.startswith("http") else "https://www.argentina.gob.ar" + href
        out[url.rsplit("/", 1)[-1]] = url
    return out


def _desreg_leer_informe(url: str) -> dict | None:
    """Período y cifras de portada de un informe. Cubre los DOS formatos.

    Hasta abr-2026 las cifras venían en prosa ("se registran N normas de
    desregulación que eliminan o modifican M normativas anteriores... un total
    de A artículos"); desde may-2026 vienen como fichas sueltas de portada.
    """
    from pypdf import PdfReader

    blob = requests.get(url, headers=HTTP_HEADERS, timeout=HTTP_TIMEOUT).content
    texto = "\n".join((p.extract_text() or "") for p in PdfReader(io.BytesIO(blob)).pages)
    plano = re.sub(r"[ \t\n]+", " ", texto)

    # El período sale de la FECHA DE CORTE ("al 31 de octubre de 2025"), no del
    # mes de portada: hay informes cuya portada dice "JUNIO 2025" pero cuyo
    # corte es el 4 de julio. Cuando el corte cae en los primeros días del mes,
    # el informe cubre hasta el cierre del mes ANTERIOR y se le asigna ése.
    meses = "|".join(_DESREG_MESES)
    corte = re.search(r"al\s+(\d{1,2})\s+de\s+(" + meses + r")\s+de\s+(20\d\d)",
                      plano, re.I)
    if corte:
        dia, anio = int(corte.group(1)), int(corte.group(3))
        mes = _DESREG_MESES[corte.group(2).lower()]
        if dia <= 7:
            mes -= 1
            if mes == 0:
                mes, anio = 12, anio - 1
        periodo = f"{anio}-{mes:02d}"
    else:
        m = re.search(r"\b(" + meses + r")\s+(?:de\s+)?(20\d\d)\b", plano, re.I)
        if not m:
            return None
        periodo = f"{m.group(2)}-{_DESREG_MESES[m.group(1).lower()]:02d}"

    def n(s):
        return int(s.replace(".", ""))

    prosa = re.search(
        r"se registran\s+([\d\.]+)\s+normas de desregulaci[oó]n\s+que\s+eliminan o "
        r"modifican\s+([\d\.]+)\s+normativas anteriores.{0,80}?total de\s+([\d\.]+)\s+art",
        plano, re.I | re.S)
    if prosa:
        normas, modif, arts = (n(prosa.group(i)) for i in (1, 2, 3))
    else:
        def cap(pat):
            mm = re.search(pat + r"\s*([\d\.]+)", plano, re.I)
            return n(mm.group(1)) if mm else None
        normas = cap(r"Normas de desregulaci[oó]n")
        modif  = cap(r"Normas modi[fﬁ]cadas o eliminadas")
        arts   = cap(r"Art[ií]culos modi[fﬁ]cados o eliminados")
        if not (normas and modif and arts):
            return None
    return {"periodo": periodo, "normas": normas,
            "normas_afectadas": modif, "articulos": arts}


def _desreg_backfill_grafico(url: str, ancla: int) -> dict:
    """Serie mensual desde dic-2023, medida sobre las barras de la Figura 1.

    El PDF trae las etiquetas del gráfico convertidas a curvas, así que el
    texto no se puede extraer — pero las barras son rectángulos vectoriales y
    su altura es proporcional al valor. Se calibra con el titular del propio
    informe, que es la última barra.

    Es una reconstrucción por geometría y hay que tratarla como tal: el test
    `test_gestion_desregulacion` la contrasta contra los titulares de los otros
    informes, que son independientes del gráfico (ADR-0125).
    """
    from collections import Counter

    import fitz

    blob = requests.get(url, headers=HTTP_HEADERS, timeout=HTTP_TIMEOUT).content
    doc = fitz.open(stream=blob, filetype="pdf")
    for pagina in doc:
        txt = pagina.get_text() or ""
        if not (re.search(r"Figura 1", txt, re.I) and re.search(r"acumulad", txt, re.I)):
            continue
        rects = [it[1] for d in pagina.get_drawings() for it in d["items"]
                 if it[0] == "re" and it[1].height > 1 and 3 < it[1].width < 60]
        if not rects:
            continue
        base = Counter(round(r.y1, 1) for r in rects).most_common(1)[0][0]
        barras = sorted((r for r in rects if abs(r.y1 - base) < 1.5 and r.height > 1.0),
                        key=lambda r: r.x0)
        if len(barras) < 12:
            continue
        escala = ancla / (base - barras[-1].y0)
        anio, mes = int(DESREG_BACKFILL_DESDE[:4]), int(DESREG_BACKFILL_DESDE[5:7])
        serie = {}
        for r in barras:
            serie[f"{anio}-{mes:02d}"] = round((base - r.y0) * escala)
            mes += 1
            if mes > 12:
                anio, mes = anio + 1, 1
        return serie
    raise ValueError("no se encontró la Figura 1 en el informe de backfill")


def desregulacion_oficial_serie() -> dict:
    """{YYYY-MM: normas de desregulación acumuladas} desde dic-2023.

    Backfill desde el gráfico del informe de abril-2026 + el titular de cada
    informe posterior. Todo se cachea: los informes publicados son inmutables.
    """
    store = _desreg_oficial_store()
    informes = store.setdefault("informes", {})
    disponibles = _desreg_informes_publicados()

    for nombre, url in sorted(disponibles.items()):
        if nombre in informes:
            continue
        datos = _desreg_leer_informe(url)
        if datos:
            informes[nombre] = datos

    por_periodo = {d["periodo"]: d for d in informes.values()}

    backfill = store.get("backfill_grafico")
    if not backfill:
        url = disponibles.get(DESREG_INFORME_BACKFILL)
        ancla = por_periodo.get("2026-04", {}).get("normas")
        if url and ancla:
            backfill = _desreg_backfill_grafico(url, ancla)
            store["backfill_grafico"] = backfill
    backfill = backfill or {}

    _desreg_guardar_store(store)

    serie = dict(backfill)
    for periodo, d in por_periodo.items():
        # el titular manda sobre el gráfico donde ambos existen: es el número
        # que el ministerio publicó en texto para ese mes
        if periodo > max(backfill, default=""):
            serie[periodo] = d["normas"]
    return dict(sorted(serie.items()))


def fetch_desregulacion_normativa() -> dict | None:
    """Normas de desregulación acumuladas desde el 10-dic-2023, según el
    Ministerio de Desregulación y Transformación del Estado (ADR-0125).

    Reemplaza al conteo propio sobre InfoLeg (ADR-0096). Lo que se gana es que
    el número deja de ser una construcción nuestra y pasa a ser el que publica
    el organismo responsable, verificable en su informe mensual. Lo que se
    pierde está declarado en la ficha: es el Gobierno midiendo su propio
    programa, y el criterio de qué norma cuenta como "de desregulación" es suyo.
    """
    try:
        serie = desregulacion_oficial_serie()
        if not serie:
            raise ValueError("sin informes oficiales legibles")
        periodo = max(serie)
        store = _desreg_oficial_store()
        ultimo = next((d for d in store.get("informes", {}).values()
                       if d["periodo"] == periodo), {})
        previo = serie.get(_mes_previo_ym(periodo))
        return {
            "valor":          float(serie[periodo]),
            "unidad":         "normas de desregulación acumuladas desde dic-2023",
            "fuente":         "Ministerio de Desregulación y Transformación del Estado — "
                              "Análisis de la desregulación implementada (informe mensual)",
            "fecha_dato":     f"{periodo}-01",
            "desactualizado": False,
            "normas_afectadas": ultimo.get("normas_afectadas"),
            "articulos_afectados": ultimo.get("articulos"),
            "alta_del_mes":   None if previo is None else serie[periodo] - previo,
            "detalle_txt": (
                f"{serie[periodo]} normas de desregulación acumuladas desde el 10 de "
                f"diciembre de 2023"
                + ("" if previo is None else f" ({serie[periodo] - previo:+d} en el mes)")
                + ("" if not ultimo.get("normas_afectadas") else
                   f" · modificaron o eliminaron {ultimo['normas_afectadas']} normas "
                   f"anteriores y {ultimo.get('articulos')} artículos")),
        }
    except Exception as e:
        _warn("desregulacion_normativa", e)
        return None


# ── D2: Reforma del Estado ────────────────────────────────────────────────────

_MESES_ES = {"ene": 1, "feb": 2, "mar": 3, "abr": 4, "may": 5, "jun": 6,
             "jul": 7, "ago": 8, "sep": 9, "oct": 10, "nov": 11, "dic": 12}


def _parse_ym_celda(c) -> str | None:
    """'may-26 (i)' o datetime → 'YYYY-MM' (encabezados del XLSX del INDEC)."""
    if hasattr(c, "year") and hasattr(c, "month"):
        return f"{c.year}-{c.month:02d}"
    if isinstance(c, str):
        m = re.match(r"^([a-z]{3})-(\d{2})", c.strip().lower())
        if m and m.group(1) in _MESES_ES:
            return f"20{m.group(2)}-{_MESES_ES[m.group(1)]:02d}"
    return None


def dotacion_apn_series() -> dict:
    """{YYYY-MM: dotación APN} desde el XLSX oficial del INDEC ("Dotación de
    personal de la APN, empresas y sociedades", mensual, jul-2022–). Meses
    marcados "(i)" son imputados y se revisan hacia atrás en cada publicación.
    También la usa descargar_series.py para la serie histórica del indicador."""
    import openpyxl
    content = _http_get_resiliente(INDEC_DOTACION_URL)
    wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
    hoja = next(s for s in wb.sheetnames if "cuadro 1" in s.lower())
    ws = wb[hoja]
    filas = list(ws.iter_rows(values_only=True))

    header = None
    for fila in filas[:8]:
        parsed = [_parse_ym_celda(c) for c in fila]
        if sum(1 for p in parsed if p) >= 6:
            header = parsed
            break
    if header is None:
        raise ValueError("XLSX INDEC: fila de encabezados de meses no encontrada")

    # El cuadro 1 abre la dotación en tres filas que hay que distinguir
    # (ADR-0097): "Total" = "Administración pública nacional" + "Empresas y
    # sociedades". El indicador puntúa la SEGUNDA —la planta administrativa del
    # Ejecutivo, sin las empresas del Estado— y las otras dos se leen igual para
    # publicarlas como contexto: la revisión externa preguntó cuál de los tres
    # universos se estaba usando, y la respuesta merece estar a la vista.
    def _fila(*prefijos):
        for fila in filas:
            primera = str(fila[0] or "").strip().lower()
            if any(primera.startswith(p) for p in prefijos):
                return fila
        return None

    fila_apn = _fila("administración pública nacional", "administracion publica nacional")
    if fila_apn is None:
        raise ValueError("XLSX INDEC: fila 'Administración pública nacional' no encontrada")

    def _serie_de(fila):
        if fila is None:
            return {}
        return {ym: float(v) for ym, v in zip(header, fila)
                if ym and isinstance(v, (int, float))}

    serie = _serie_de(fila_apn)
    if "2023-12" not in serie:
        raise ValueError("XLSX INDEC: baseline dic-2023 ausente de la serie")
    serie["_empresas"] = _serie_de(_fila("empresas y sociedades"))
    serie["_total"] = _serie_de(_fila("total"))
    serie["_fuerzas"] = _dotacion_fuerzas(wb)
    return serie


# Las siete entidades de fuerzas armadas y de seguridad que informan al padrón
# (ADR-0128). Están DENTRO de la dotación APN que puntúa el indicador, cosa que
# la revisión externa preguntó y que no estaba escrita en ningún lado.
_ENTES_FUERZAS = ("estado mayor", "gendarmería nacional", "gendarmeria nacional",
                  "prefectura naval", "policía de seguridad aerop",
                  "policia de seguridad aerop")


def _dotacion_fuerzas(wb) -> dict:
    """{YYYY-MM: dotación de fuerzas armadas y de seguridad} desde la hoja de
    detalle por entidad.

    Se publica como CONTEXTO, no se descuenta: las fuerzas son parte del Estado
    y sacarlas del denominador sería una decisión editorial que hay que
    justificar aparte. Lo que hace falta es que se vea.

    La hoja de detalle va uno o dos meses ATRÁS del cuadro principal, así que
    esta serie termina antes que la del indicador — por eso se publica con su
    propia fecha y no se fuerza a coincidir.
    """
    hoja = next((s for s in wb.sheetnames if s.strip().lower() == "detalle2"), None)
    if hoja is None:
        return {}
    filas = list(wb[hoja].iter_rows(values_only=True))
    if len(filas) < 4:
        return {}
    meses = {j: ym for j, c in enumerate(filas[2])
             if (ym := _parse_ym_celda(c))}
    out = {}
    for f in filas[3:]:
        entidad = str(f[1] or "").strip().lower()
        if not any(k in entidad for k in _ENTES_FUERZAS):
            continue
        for j, ym in meses.items():
            if j < len(f) and isinstance(f[j], (int, float)):
                out[ym] = out.get(ym, 0.0) + float(f[j])
    return out


def fetch_reduccion_estado() -> dict | None:
    """
    Variación % de la dotación de la Administración Pública Nacional vs el
    baseline dic-2023 (inicio del mandato). Fuente: XLSX mensual oficial del
    INDEC + Secretaría de Transformación del Estado — la métrica insignia de
    la reforma (a diferencia de la serie SIPA, excluye provincias/municipios).
    Reducción = reforma ejecutándose (bandas en itcg.BANDAS_ITCG).
    """
    try:
        serie = dotacion_apn_series()
        empresas = serie.pop("_empresas", {})
        total = serie.pop("_total", {})
        fuerzas = serie.pop("_fuerzas", {})
        ult = max(serie)
        base = serie["2023-12"]
        var_pct = round((serie[ult] - base) / base * 100.0, 2)
        miles = lambda x: f"{x:,.0f}".replace(",", ".")

        def _var(otra):
            if otra.get(ult) and otra.get("2023-12"):
                return round((otra[ult] - otra["2023-12"]) / otra["2023-12"] * 100.0, 2)
            return None

        var_empresas, var_total = _var(empresas), _var(total)

        # Fuerzas armadas y de seguridad (ADR-0128): están DENTRO del número
        # que puntúa. Su hoja de detalle va uno o dos meses atrás del cuadro
        # principal, así que se compara en el último mes que ambas tienen —
        # comparar meses distintos daría una planta civil inventada.
        var_fuerzas = var_civil = fecha_fuerzas = None
        comunes = sorted(set(serie) & set(fuerzas))
        if comunes and "2023-12" in fuerzas:
            fu = comunes[-1]
            fecha_fuerzas = fu
            var_fuerzas = round(
                (fuerzas[fu] - fuerzas["2023-12"]) / fuerzas["2023-12"] * 100.0, 2)
            civ_b = serie["2023-12"] - fuerzas["2023-12"]
            civ_u = serie[fu] - fuerzas[fu]
            if civ_b:
                var_civil = round((civ_u - civ_b) / civ_b * 100.0, 2)

        return {
            "valor":          var_pct,
            "unidad":         "% de variación vs dic-2023 (dotación APN)",
            "fuente":         "INDEC — Dotación de personal de la APN (mensual)",
            "fecha_dato":     f"{ult}-01",
            "desactualizado": False,
            "dotacion_actual": round(serie[ult]),
            "dotacion_dic23":  round(base),
            # Contexto declarado (ADR-0097): qué pasa si se mide el universo
            # completo. Las tres cifras son casi idénticas, así que la elección
            # de universo no cambia la lectura — y eso es justamente lo que
            # conviene mostrar.
            "var_empresas_publicas": var_empresas,
            "var_universo_completo": var_total,
            # ADR-0128: la revisión externa preguntó si el denominador incluye
            # fuerzas armadas y de seguridad. Sí las incluye —siete entes, ~10%
            # de la dotación— y se publica cuánto cambia la lectura sin ellas.
            "var_fuerzas_seguridad": var_fuerzas,
            "var_planta_civil": var_civil,
            "fecha_desglose_fuerzas": fecha_fuerzas,
            "detalle_txt": (
                f"{miles(serie[ult])} agentes ({ult}) vs {miles(base)} en dic-2023"
                + ("" if var_empresas is None or var_total is None else
                   f" · sin contar empresas del Estado, que por su lado varían "
                   f"{str(var_empresas).replace('.', ',')}%; el universo completo "
                   f"varía {str(var_total).replace('.', ',')}%")
                + ("" if var_civil is None else
                   f" · incluye fuerzas armadas y de seguridad, que son ~10% de la "
                   f"dotación y se redujeron menos "
                   f"({str(var_fuerzas).replace('.', ',')}%): sin ellas, la planta "
                   f"civil varía {str(var_civil).replace('.', ',')}% "
                   f"(a {fecha_fuerzas})")),
        }
    except Exception as e:
        _warn("reduccion_estado", e)
        return None


def _var_real_vs_mismo_mes_2023(nominal: dict, ipc: dict) -> tuple:
    """Variación % REAL del último mes disponible vs el MISMO MES de 2023
    (aísla inflación y estacionalidad —aguinaldos— a la vez). Devuelve
    (var_pct, ym_ultimo, ym_base)."""
    comunes = sorted(set(nominal) & set(ipc))
    if not comunes:
        raise ValueError("sin meses comunes entre la serie nominal y el IPC")
    ym = comunes[-1]
    base_ym = f"2023-{ym[5:7]}"
    if base_ym not in nominal or base_ym not in ipc:
        raise ValueError(f"mes base {base_ym} ausente (serie o IPC)")
    real_t = nominal[ym] / ipc[ym]
    real_b = nominal[base_ym] / ipc[base_ym]
    return round((real_t / real_b - 1.0) * 100.0, 2), ym, base_ym


def fetch_masa_salarial() -> dict | None:
    """
    Masa salarial del Sector Público Nacional (AIF base caja: remuneraciones),
    variación REAL vs el mismo mes de 2023 (deflactada por IPC; comparar el
    mismo mes evita el sesgo del aguinaldo). Filtra el efecto de la inflación
    sobre plantas nominales: complementa a la dotación con el costo real.
    """
    try:
        nominal = _indec_nivel_mensual(REMUNERACIONES_ID, limit=48)
        ipc     = _indec_nivel_mensual(IPC_ID, limit=48)
        var_pct, ym, base_ym = _var_real_vs_mismo_mes_2023(nominal, ipc)
        infl = ipc[ym] / ipc[base_ym]
        return {
            "valor":          var_pct,
            "unidad":         f"% de variación real vs {base_ym} (SPN remuneraciones)",
            "fuente":         "Sec. Hacienda AIF (datos.gob.ar) + IPC INDEC",
            "fecha_dato":     f"{ym}-01",
            "desactualizado": False,
            "detalle_txt":    (f"$ {nominal[ym]/1e6:,.2f} bn devengados ({ym}) vs "
                               f"$ {nominal[base_ym]/1e6:,.2f} bn ({base_ym}) · inflación del "
                               f"período ×{infl:,.1f}").replace(",", "@").replace(".", ",").replace("@", "."),
        }
    except Exception as e:
        _warn("masa_salarial", e)
        return None


def fetch_gasto_funcionamiento() -> dict | None:
    """
    Gasto de funcionamiento del Estado nacional (IMIG: salarios + otros gastos
    de funcionamiento), variación REAL vs el mismo mes de 2023. Mide la
    magnitud fiscal del aparato administrativo aislada de la inflación —
    distingue achicamiento real de licuación nominal.
    """
    try:
        salarios = _indec_nivel_mensual(FUNC_SALARIOS_ID, limit=48)
        otros    = _indec_nivel_mensual(FUNC_OTROS_ID, limit=48)
        ipc      = _indec_nivel_mensual(IPC_ID, limit=48)
        comunes  = set(salarios) & set(otros)
        total    = {ym: salarios[ym] + otros[ym] for ym in comunes}
        var_pct, ym, base_ym = _var_real_vs_mismo_mes_2023(total, ipc)
        infl = ipc[ym] / ipc[base_ym]
        return {
            "valor":          var_pct,
            "unidad":         f"% de variación real vs {base_ym} (IMIG funcionamiento)",
            "fuente":         "Sec. Hacienda IMIG (datos.gob.ar) + IPC INDEC",
            "fecha_dato":     f"{ym}-01",
            "desactualizado": False,
            "detalle_txt":    (f"$ {total[ym]/1e6:,.2f} bn devengados ({ym}) vs "
                               f"$ {total[base_ym]/1e6:,.2f} bn ({base_ym}) · inflación del "
                               f"período ×{infl:,.1f}").replace(",", "@").replace(".", ",").replace("@", "."),
        }
    except Exception as e:
        _warn("gasto_funcionamiento", e)
        return None


def fetch_reestructuracion_organismos() -> dict | None:
    """
    Avance de la reestructuración del aparato estatal, proxy InfoLeg: normas
    con "disolucion" desde dic-2023. DNU 70/23 no aparece en búsqueda de texto
    libre (no indexado); los documentos capturados son actos POSTERIORES al
    megadecreto. Calibración: 18 actos = 40% (validado manualmente); 45 = 100%.
    """
    try:
        today = date.today()
        count = _infoleg_post(
            texto="disolucion", tipo_norma="",
            fecha_desde=("01", "12", "2023"),
            fecha_hasta=(today.strftime("%d"), today.strftime("%m"), today.strftime("%Y")),
        )
        avance = round(min(100.0, count * 100.0 / ORGANISMOS_PLAN_TOTAL), 1)
        return {
            "valor":          avance,
            "unidad":         "% de avance (proxy InfoLeg)",
            "fuente":         INFOLEG_HOME,
            "fecha_dato":     today.isoformat(),
            "desactualizado": False,
            "conteo_normas":  count,
            "detalle_txt":    f"{count} actos de disolución/fusión desde dic-2023 ({ORGANISMOS_PLAN_TOTAL} = plan completo)",
        }
    except Exception as e:
        _warn("reestructuracion_organismos", e)
        return None


# ── D3: Reforma laboral ───────────────────────────────────────────────────────

CNV_FCI_URL = "https://www.cnv.gov.ar/SitioWeb/FondosComunesInversion/GetFCIPorTipo"
SRT_JUICIOS_URL = "https://www.srt.gob.ar/estadisticas/litigiosidad/series/Juicios%20-%20Total%20Sistema.xlsx"

# Escala del componente de adopción financiera del Fondo de Cese/FAL: la RG
# CNV 1071/2025 (art. 2) obliga a incluir "Cese Laboral" en la denominación
# de los PICs del régimen Ley Bases; la Ley 27.802 renombró el instrumento a
# "Fondo de Asistencia Laboral" — el detector busca ambas denominaciones.
# 10 fondos/fideicomisos registrados ≈ adopción financiera plena (provisional
# hasta que exista patrimonio consultable — ver ADR-0013).
FCI_CESE_PLENO = 10

# Cobertura del FAL vía BOLETÍN OFICIAL (ADR-0068, reemplaza la consulta por
# "fondo de cese laboral" desde dic-2023): esa frase era indistinguible del
# régimen de la CONSTRUCCIÓN (Ley 22.250, llamado así desde 1980) — 7-8
# menciones/año ANTES de la reforma vs 8,1/año después, señal neta nula; la
# serie que "crecía" era ruido de fondo acumulado. La Ley 27.802
# (Modernización Laboral, publicada mar-2026; Dto. 408/2026 reglamentó)
# renombró el instrumento a "Fondo de Asistencia Laboral" (FAL, Título II;
# entra en vigencia el 1-nov-2026 según el art. 27 del Dto. 408/2026): se
# cuenta esa frase DESDE la publicación de la ley — el corte de fecha elimina
# el ruido pre-creación (verificado: 1 mención espuria en ene-2026 queda
# afuera). Hasta nov-2026 la cobertura mide andamiaje normativo; la adopción
# financiera (CNV) está diferida por norma, por eso es 0.
# Pleno = 420 menciones acumuladas ≈ un año de negociación colectiva donde
# ~1 de cada 5 homologaciones incorpora el FAL (el MTEySS homologa ~2.000
# convenios y acuerdos/año, serie 2008-2022; las cláusulas de crisis
# llegaron al 42% anual en 2020-21 — techo empírico de difusión de una
# cláusula nueva). Provisional hasta que exista serie MTEySS de
# homologaciones con FAL.
BO_BUSQUEDA_URL = "https://www.boletinoficial.gob.ar/busquedaAvanzada/realizarBusqueda"
FAL_BO_TEXTO = "fondo de asistencia laboral"
FAL_BO_DESDE = "01/03/2026"           # publicación de la Ley 27.802
FAL_BO_MENCIONES_PLENO = 420


def _bo_conteo(texto: str, desde: str = "10/12/2023", hasta: str | None = None) -> int:
    """Conteo de normas de la Primera Sección del Boletín Oficial que
    contienen `texto` (todas las palabras, sin stemming) en el rango de
    fechas. Endpoint público sin sesión: POST con `params` JSON mínimo; el
    total viene en content.cantidad_result_seccion (por sección, rango
    completo — no por página)."""
    params = {
        "texto": texto,
        "fechaDesde": desde,
        "fechaHasta": hasta or date.today().strftime("%d/%m/%Y"),
        "todasLasPalabras": True,
        "seccion": [1],
        "tipoBusqueda": "Avanzada",
        "numeroPagina": 1,
    }
    r = requests.post(BO_BUSQUEDA_URL,
                      data={"params": json.dumps(params), "array_volver": "[]"},
                      headers=HTTP_HEADERS, timeout=60)
    r.raise_for_status()
    d = r.json()
    if d.get("error"):
        raise ValueError(f"BO búsqueda error={d.get('error')}: {d.get('mensajes')}")
    return int((d.get("content", {}).get("cantidad_result_seccion") or {}).get("1", 0))


def _cnv_fondos_cese() -> int:
    """Cantidad de FCI registrados en CNV cuya denominación refiere al
    instrumento de cese: 'CESE' (RG 1071/2025, régimen Ley Bases) o
    'ASISTENCIA LABORAL' (FAL, Ley 27.802). Registro completo vía POST JSON,
    sin auth — verificado 2026-07: 1.656 fondos, 0 bajo cualquiera de las dos
    denominaciones; el cero es un dato duro, no un faltante."""
    r = requests.post(CNV_FCI_URL, json={}, headers=HTTP_HEADERS, timeout=HTTP_TIMEOUT)
    r.raise_for_status()
    fondos = r.json()
    if not isinstance(fondos, list) or not fondos:
        raise ValueError("registro CNV de FCI vacío o con formato inesperado")
    return sum(1 for f in fondos
               if any(t in str(f.get("Text", "")).upper()
                      for t in ("CESE", "ASISTENCIA LABORAL")))


FAL_HITOS_PATH = PROJECT_DIR / "data" / "gestion" / "fal_hitos.json"

# ADR-0142: el indicador pasa a medir los DOS ACTOS FUNDAMENTALES que ponen en
# pie al Fondo, a propuesta de la revisión externa del cinturón y por decisión
# editorial explícita. Cada acto vale la mitad del índice — o sea el 25% de la
# dimensión reforma_laboral, que es lo que pidió el documento.
#
# Se reemplaza el compuesto de ADR-0098 (0,40 construcción + 0,20 vigencia +
# 0,40 adopción). El desacuerdo es de fondo y conviene dejarlo escrito: ADR-0098
# sostenía que "mientras nada rija el efecto es cero" y por eso topeaba en 30
# puntos hasta noviembre; la revisión externa sostiene que sancionar la ley y
# reglamentarla "cumple hasta aquí el cumplimiento de la promesa del Gobierno".
# Ganó la segunda lectura.
FAL_ACTOS_FUNDAMENTALES = (
    ("Ley 27.802", "Congreso de la Nación — Ley de Modernización Laboral que instaura el FAL"),
    ("Decreto 408/2026", "Poder Ejecutivo Nacional — reglamentación del Título II (FAL)"),
)


def fetch_fal_modernizacion_laboral() -> dict | None:
    """
    Avance del Fondo de Asistencia Laboral (Ley 27.802) medido por sus DOS
    ACTOS FUNDAMENTALES, 50 puntos cada uno (ADR-0142):

        1. Ley 27.802 — el Congreso instaura el FAL          (06-mar-2026)
        2. Decreto 408/2026 — el PEN lo reglamenta           (01-jun-2026)

    Los dos actos ponen las bases para el funcionamiento del Fondo y, según el
    criterio editorial adoptado, agotan lo que el Gobierno podía cumplir de la
    promesa hasta la entrada en vigencia del régimen el 1-nov-2026.

    QUÉ CAMBIÓ Y QUÉ SE PIERDE. Reemplaza al compuesto de ADR-0098 (0,40
    construcción + 0,20 vigencia + 0,40 adopción), que valía 40,2 y puntuaba
    30,8. Con los dos actos cumplidos el índice vale 100 y puntúa 100: la
    dimensión `reforma_laboral` pasa de 45,1 a 79,7 y el ITCG sube 5,2 puntos.
    No se puede invocar neutralidad — el cambio mejora el puntaje y la
    justificación es editorial, no empírica.

    Y HAY UN COSTO QUE HAY QUE MIRAR DE FRENTE: los dos actos ya ocurrieron y
    no se pueden deshacer, así que **el indicador queda fijo en 100 y ningún
    hecho futuro lo mueve** — ni la entrada en vigencia, ni que el Fondo
    funcione o no. Deja de discriminar (contra ADR-0042). La decisión de
    publicarlo así fue del editor, con el efecto a la vista.

    Todo lo que el compuesto puntuaba sigue relevándose y viaja en la card como
    CONTEXTO, sin incidir en el puntaje: fondos registrados en CNV, menciones
    del FAL en el Boletín Oficial y fecha de vigencia del régimen. Si en algún
    momento se quiere volver a un indicador vivo, los insumos están.
    """
    try:
        hitos = json.loads(FAL_HITOS_PATH.read_text(encoding="utf-8-sig"))
        hoy_iso = date.today().isoformat()
        por_norma = {h.get("norma"): h for h in hitos.get("construccion", [])}

        actos = []
        for norma, descripcion in FAL_ACTOS_FUNDAMENTALES:
            h = por_norma.get(norma)
            fecha_acto = (h or {}).get("fecha")
            actos.append({
                "norma": norma,
                "descripcion": descripcion,
                "fecha": fecha_acto,
                "cumplido": bool(fecha_acto and fecha_acto <= hoy_iso),
            })
        cumplidos = [a for a in actos if a["cumplido"]]
        indice = round(100.0 * len(cumplidos) / len(actos), 1) if actos else 0.0

        # ── Contexto: se releva pero NO puntúa (ADR-0142) ───────────────────
        n_cese = _cnv_fondos_cese()
        menciones = None
        cobertura_txt = ""
        try:
            menciones = _bo_conteo(FAL_BO_TEXTO, desde=FAL_BO_DESDE)
            cobertura_txt = f"{menciones} menciones del FAL en el BO desde mar-2026"
        except Exception as e:
            _warn("fal cobertura BO (contexto, no puntúa)", e)

        fecha_vigencia = (hitos.get("vigencia") or {}).get("fecha", "9999-12-31")
        vigente = hoy_iso >= fecha_vigencia

        return {
            "valor":          indice,
            "unidad":         "Índice 0–100 (actos fundamentales del FAL: ley + reglamentación)",
            "fuente":         "InfoLeg — Ley 27.802 y Decreto 408/2026 (normas publicadas)",
            "fecha_dato":     date.today().isoformat(),
            "desactualizado": False,
            "actos":          actos,
            "actos_cumplidos": len(cumplidos),
            "actos_totales":   len(actos),
            # contexto, no puntúa
            "fci_cese_registrados": n_cese,
            "menciones_bo":   menciones,
            "vigencia_desde": fecha_vigencia,
            "regimen_vigente": vigente,
            "detalle_txt": (
                f"{len(cumplidos)} de {len(actos)} actos fundamentales cumplidos"
                + (" (" + " · ".join(f"{a['norma']} {a['fecha']}" for a in cumplidos) + ")"
                   if cumplidos else "")
                + (" · régimen VIGENTE" if vigente else
                   f" · el régimen entra en vigencia el {fecha_vigencia}")
                + f" · contexto: {n_cese} fondos registrados en CNV"
                + (f" · {cobertura_txt}" if cobertura_txt else "")),
        }
    except Exception as e:
        _warn("fal_modernizacion_laboral", e)
        return None


def fetch_litigiosidad_laboral() -> dict | None:
    """
    Variación % de los juicios laborales del sistema de riesgos del trabajo
    (SRT), acumulado 12 meses vs los 12 previos. Desde el ADR-0023 INTEGRA el
    ITCG (reforma laboral, 30%): es el RESULTADO que la reforma persigue —
    enfriar la industria del juicio — y complementa la adopción del Fondo de
    Cese (instrumento). Proxy: no mide el canal indemnizatorio que el Fondo
    reemplaza, pero es la única serie nacional mensual pública.
    """
    try:
        import openpyxl
        content = _http_get_resiliente(SRT_JUICIOS_URL)
        wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
        hoja = next(s for s in wb.sheetnames if "TOTAL" in s.upper())
        filas = list(wb[hoja].iter_rows(values_only=True))

        meses_es = {"ene": 1, "feb": 2, "mar": 3, "abr": 4, "may": 5, "jun": 6,
                    "jul": 7, "ago": 8, "sep": 9, "oct": 10, "nov": 11, "dic": 12}
        header = next(f for f in filas
                      if sum(1 for c in f if isinstance(c, str) and re.match(r"^[A-Za-z]{3}-\d{4}$", c.strip())) > 20)
        fila_total = next(f for f in filas
                          if isinstance(f[0], str) and f[0].lower().startswith("total de juicios"))

        serie = {}
        for h, v in zip(header, fila_total):
            if isinstance(h, str) and re.match(r"^[A-Za-z]{3}-\d{4}$", h.strip()):
                mes, anio = h.strip().split("-")
                try:
                    serie[f"{anio}-{meses_es[mes.lower()]:02d}"] = float(v)
                except (ValueError, TypeError, KeyError):
                    continue
        yms = sorted(serie)
        if len(yms) < 24:
            raise ValueError("serie SRT demasiado corta para 12m vs 12m")
        ult12  = sum(serie[ym] for ym in yms[-12:])
        prev12 = sum(serie[ym] for ym in yms[-24:-12])
        if not prev12:
            raise ValueError("período base sin juicios")
        var = round((ult12 / prev12 - 1.0) * 100.0, 1)
        return {
            "valor":          var,
            "unidad":         "% variación juicios SRT (12m vs 12m previos)",
            "fuente":         "SRT — Serie histórica de litigiosidad",
            "fecha_dato":     f"{yms[-1]}-01",
            "desactualizado": False,
            "juicios_12m":    round(ult12),
            "detalle_txt":    f"{round(ult12):,} juicios en 12m (hasta {yms[-1]}) vs {round(prev12):,} previos".replace(",", "."),
        }
    except Exception as e:
        _warn("litigiosidad_laboral", e)
        return None


# ── D4: Privatizaciones e inversión ───────────────────────────────────────────

PRIVATIZACIONES_FECHAS_PATH = PROJECT_DIR / "data" / "gestion" / "privatizaciones_fechas.json"


def _privatizaciones_detalle(empresas: dict) -> list:
    """[{empresa, etapa, mecanismo, hito, norma, fecha}] ordenado por etapa.

    La `norma` es la fuente de la última transición registrada que no supera la
    etapa vigente: es el acto del Boletín Oficial que respalda dónde está hoy
    esa empresa. Si el registro de transiciones no está disponible, el detalle
    sale igual pero sin norma — mejor incompleto que ausente.
    """
    try:
        fechas = json.loads(PRIVATIZACIONES_FECHAS_PATH.read_text(
            encoding="utf-8-sig")).get("empresas", {})
    except (OSError, json.JSONDecodeError):
        fechas = {}

    out = []
    for nombre, datos in sorted(empresas.items(), key=lambda kv: -float(kv[1]["etapa"])):
        etapa = float(datos["etapa"])
        norma = fecha = None
        transiciones = [t for t in fechas.get(nombre, [])
                        if float(t.get("etapa", 0)) <= etapa]
        if transiciones:
            ultima = max(transiciones, key=lambda t: (float(t["etapa"]), t.get("fecha", "")))
            norma, fecha = ultima.get("fuente"), ultima.get("fecha")
        out.append({
            "empresa":   nombre,
            "etapa":     etapa,
            "mecanismo": datos.get("mecanismo"),
            "hito":      datos.get("hito"),
            "norma":     norma,
            "fecha":     fecha,
        })
    return out


PRIVATIZACIONES_NOVEDADES_PATH = (PROJECT_DIR / "data" / "gestion"
                                  / "privatizaciones_novedades.json")

# Términos de búsqueda por empresa. La clave es el nombre del registro; el
# valor, cómo aparece la empresa nombrada en el Boletín Oficial. Se busca por
# el término MÁS DISTINTIVO: "AySA" trae poco ruido, "SOFSE" nada, pero
# "Corredores Viales" hay que buscarlo entero o matchea cualquier obra vial.
PRIVATIZACIONES_TERMINOS = {
    "AySA": "Agua y Saneamientos Argentinos",
    "Transener": "Transener",
    "Enarsa": "Energía Argentina",
    "Intercargo": "Intercargo",
    "Corredores Viales": "Corredores Viales",
    "Belgrano Cargas": "Belgrano Cargas",
    "Nucleoeléctrica": "Nucleoeléctrica",
    "YCRT": "Yacimientos Carboníferos Río Turbio",
    "SOFSE": "Operadora Ferroviaria",
}


# Una norma que sólo NOMBRA a la empresa no es una novedad del proceso: el BO
# la menciona todo el tiempo en trámites de rutina (designaciones, licencias,
# tarifas). Se exige además un verbo del proceso privatizador. Sin este filtro
# la primera corrida devolvió 180 "novedades" en tres meses, que es exactamente
# el ruido que haría que el analista dejara de mirar la lista.
_PRIVAT_PROCESO = re.compile(
    r"privatiza|desestatiza|transferencia\s+accionar|venta\s+de\s+(?:las\s+)?acciones|"
    r"paquete\s+accionar|pliego|licitaci[oó]n\s+p[uú]blica|adjudica|"
    r"concurso\s+p[uú]blico|sujeta\s+a\s+privatizaci[oó]n|programa\s+de\s+propiedad",
    re.I)


def _infoleg_texto(norma_id: str) -> str:
    """Texto plano de una norma. Vacío si no está disponible."""
    from bs4 import BeautifulSoup
    lo = (int(norma_id) // 5000) * 5000
    url = INFOLEG_ANEXO.format(lo=lo, hi=lo + 4999, id=norma_id)
    try:
        r = requests.get(url, headers=HTTP_HEADERS, timeout=HTTP_TIMEOUT * 3)
        if r.status_code != 200:
            return ""
        return re.sub(r"\s+", " ",
                      BeautifulSoup(r.content, "html.parser").get_text(" ", strip=True))
    except Exception:
        return ""


def detectar_novedades_privatizaciones(meses_atras: int = 3) -> dict:
    """Normas nuevas del Boletín Oficial que nombran a cada empresa privatizable.

    NO clasifica ni mueve etapas: sólo avisa. La asignación de etapa es criterio
    del analista y ADR-0101 documenta un caso —Nucleoeléctrica— donde el equipo
    deliberadamente se mantuvo por debajo de lo que la norma habilitaba.
    Automatizar esa decisión borraría un juicio que hoy está publicado y se
    puede discutir; lo que sí hay que automatizar es que no se escape una norma.

    Se revisan los últimos `meses_atras` meses y se marcan como PENDIENTES las
    normas cuyo id todavía no figura en el registro de novedades revisadas. Cada
    id revisado queda anotado, así que una norma se avisa una sola vez.
    """
    try:
        store = json.loads(PRIVATIZACIONES_NOVEDADES_PATH.read_text(encoding="utf-8-sig"))
    except (OSError, json.JSONDecodeError):
        store = {}
    revisadas = store.setdefault("revisadas", {})
    pendientes = store.setdefault("pendientes", {})

    hoy = date.today()
    ventana = []
    anio, mes = hoy.year, hoy.month
    for _ in range(max(1, meses_atras)):
        ventana.append((anio, mes))
        mes -= 1
        if mes == 0:
            anio, mes = anio - 1, 12

    session = requests.Session()
    nuevas = 0
    for empresa, termino in PRIVATIZACIONES_TERMINOS.items():
        for anio_v, mes_v in ventana:
            clave_mes = f"{empresa}|{anio_v}-{mes_v:02d}"
            try:
                hallazgos = _infoleg_buscar_mes(termino, anio_v, mes_v, session=session)
            except Exception as e:
                print(f"  [WARN] privatizaciones/novedades {clave_mes}: {e}")
                continue
            for norma_id, titulo in hallazgos:
                if norma_id in revisadas:
                    continue
                # el texto de una norma publicada es inmutable: se evalúa una
                # sola vez y el veredicto queda cacheado, pase o no el filtro
                texto = _infoleg_texto(norma_id)
                m = _PRIVAT_PROCESO.search(texto)
                revisadas[norma_id] = {"empresa": empresa,
                                       "periodo": f"{anio_v}-{mes_v:02d}",
                                       "del_proceso": bool(m)}
                if not m:
                    continue
                pendientes[norma_id] = {
                    "empresa": empresa,
                    "periodo": f"{anio_v}-{mes_v:02d}",
                    "titulo": titulo,
                    "coincidencia": m.group(0).lower(),
                    "url": f"https://servicios.infoleg.gob.ar/infolegInternet/"
                           f"verNorma.do?id={norma_id}",
                }
                nuevas += 1

    store["_meta"] = {
        "descripcion": ("Normas del BO que nombran empresas privatizables. Detección "
                        "automática; la etapa la sigue asignando el analista (ADR-0129). "
                        "Sacar de 'pendientes' lo ya revisado."),
        "ultima_corrida": hoy.isoformat(),
        "nuevas_en_la_corrida": nuevas,
    }
    PRIVATIZACIONES_NOVEDADES_PATH.parent.mkdir(parents=True, exist_ok=True)
    PRIVATIZACIONES_NOVEDADES_PATH.write_text(
        json.dumps(store, indent=1, ensure_ascii=False, sort_keys=True), encoding="utf-8")
    return store


def fetch_privatizaciones() -> dict | None:
    """
    Avance de privatizaciones por ETAPAS verificables (doc 260702): cada
    empresa de la cartera Ley Bases se puntúa 0-4 (0 sin definir · 1
    preparatoria · 2 pliegos · 3 licitación/adjudicación · 4 cerrada) en
    data/gestion/privatizaciones.json, actualizado con el Boletín Oficial.
    avance = promedio(etapa)/4 × 100 — separa el anuncio del hecho consumado.
    """
    try:
        if not PRIVATIZACIONES_PATH.exists():
            raise FileNotFoundError("data/gestion/privatizaciones.json ausente")
        with open(PRIVATIZACIONES_PATH, encoding="utf-8") as f:
            store = json.load(f)
        empresas = store.get("empresas", {})
        if not empresas:
            raise ValueError("store de privatizaciones sin empresas")
        etapas = [float(e["etapa"]) for e in empresas.values()]
        etapa_prom = sum(etapas) / len(etapas)
        avance = round(etapa_prom / 4.0 * 100.0, 1)
        cerradas = [n for n, e in empresas.items() if float(e["etapa"]) >= 4]

        # Detector de novedades (ADR-0129). No debe poder tumbar el indicador:
        # si InfoLeg no responde, el avance se publica igual y la lista queda
        # vacía — lo que se pierde es un aviso, no el dato.
        try:
            novedades = detectar_novedades_privatizaciones().get("pendientes", {})
        except Exception as e:
            print(f"  [WARN] privatizaciones: detector de novedades no corrió ({e})")
            novedades = {}

        return {
            "valor":          avance,
            "unidad":         "% de avance (etapas 0-4, cartera Ley Bases)",
            "fuente":         store.get("_meta", {}).get("fuente", "Boletín Oficial — seguimiento CIGOB"),
            "fecha_dato":     store.get("_meta", {}).get("actualizado", ""),
            "desactualizado": False,
            "etapa_promedio": round(etapa_prom, 2),
            "empresas":       len(etapas),
            # etapa por empresa → gráfico de barras del modal (0 = sin definir · 4 = cerrada)
            "componentes":    {n: float(e["etapa"]) for n, e in
                               sorted(empresas.items(), key=lambda kv: -float(kv[1]["etapa"]))},
            # Detalle empresa por empresa CON la norma que respalda su etapa
            # actual (ADR-0101). Es el único indicador del cinturón sin fuente
            # en vivo —la etapa la asigna el analista— así que publicar el
            # respaldo es lo que lo blinda frente a la objeción obvia de "¿quién
            # decide en qué etapa está cada empresa?". El dato ya existía en el
            # registro versionado; lo que faltaba era exponerlo.
            "empresas_detalle": _privatizaciones_detalle(empresas),
            # ADR-0129: normas del BO que nombran a una privatizable y hablan
            # del proceso, todavía sin revisar por el analista. NO mueven la
            # etapa; sólo avisan para que no se escape ninguna.
            "novedades_pendientes": novedades,
            "detalle_txt": (f"{len(etapas)} empresas · etapa promedio "
                            f"{str(round(etapa_prom, 2)).replace('.', ',')}/4"
                            + (f" · cerradas: {', '.join(cerradas)}" if cerradas else "")
                            + (f" · {len(novedades)} norma(s) nueva(s) del Boletín "
                               f"Oficial sin revisar" if novedades else "")),
        }
    except Exception as e:
        _warn("privatizaciones", e)
        return None


def _rigi_sheet_csv(tab: str) -> list:
    """Lee una pestaña del Google Sheet oficial del RIGI vía gviz CSV."""
    url = RIGI_SHEET_GVIZ.format(sid=RIGI_SHEET_ID, tab=tab)
    r = requests.get(url, headers=HTTP_HEADERS, timeout=HTTP_TIMEOUT)
    r.raise_for_status()
    if "html" in r.headers.get("content-type", "").lower():
        raise ValueError(f"pestaña '{tab}' no accesible (devolvió HTML)")
    import csv as _csv, io as _io
    return list(_csv.reader(_io.StringIO(r.text)))


def _rigi_monto(raw: str) -> float:
    """Parsea un monto en millones de USD (es-AR: punto miles, coma decimal)."""
    s = raw.replace(".", "").replace(",", ".").strip()
    try:
        return float(s) if s else 0.0
    except ValueError:
        return 0.0


def _rigi_aprobados() -> tuple:
    """(cantidad de proyectos aprobados, inversión aprobada en M USD) desde la
    pestaña 'dataset'. Dedupe por título: un proyecto multi-provincia ocupa varias
    filas pero cuenta una vez (igual que el `Set(titulo)` del sitio oficial)."""
    rows = _rigi_sheet_csv("dataset")
    ci = {n: i for i, n in enumerate(rows[0])}
    n_idx, inv_idx = ci["nombre"], ci["inv-comprometida"]
    proyectos: dict = {}
    for r in rows[1:]:
        if not r or not r[0] or r[0] == "Provincia":   # saltea la fila de encabezado humano
            continue
        nombre = r[n_idx].strip()
        if nombre:
            proyectos.setdefault(nombre, _rigi_monto(r[inv_idx]))
    return len(proyectos), sum(proyectos.values())


def _rigi_evaluacion() -> tuple:
    """(cantidad, inversión M USD) de los proyectos en evaluación (pestaña
    agregada). El sheet oficial cambió de formato (jul-2026): los encabezados
    ahora traen 'clave Etiqueta' en la misma celda y los valores en la fila
    siguiente — se toma el primer token como clave y la primera fila numérica."""
    rows = _rigi_sheet_csv("evaluacion")
    claves = [(c or "").split()[0].strip().lower() if c else "" for c in rows[0]]
    ci = {k: i for i, k in enumerate(claves) if k}
    for fila in rows[1:]:
        try:
            return int(float(fila[ci["cantidad"]])), _rigi_monto(fila[ci["inversion"]])
        except (ValueError, IndexError, KeyError):
            continue
    raise ValueError("pestaña 'evaluacion' sin fila de valores numéricos")


def _rigi_fecha_norma(url: str) -> str | None:
    """Fecha de sanción (YYYY-MM-DD) de una norma del BO/normativa (argentina.gob.ar)."""
    try:
        rr = requests.get(url, headers=HTTP_HEADERS, timeout=HTTP_TIMEOUT)
        m = re.search(r"Sanci[oó]n[:\s]*(\d{2})-(\d{2})-(\d{4})", rr.text)
        if m:
            return f"{m.group(3)}-{m.group(2)}-{m.group(1)}"
        m = re.search(r"\d{4}-\d{2}-\d{2}", re.sub(r"<[^>]+>", " ", rr.text))
        return m.group(0) if m else None
    except Exception:
        return None


def rigi_proyectos_aprobados() -> list:
    """[(fecha_iso, nombre, inversión M USD)] de los proyectos aprobados, ordenado
    por fecha. La inversión sale del sheet (fresca); la fecha de sanción se cachea
    en RIGI_FECHAS_PATH y solo se fetchea del BO para proyectos nuevos. Reconstruye
    la serie histórica que la plataforma oficial no publica."""
    rows = _rigi_sheet_csv("dataset")
    ci = {n: i for i, n in enumerate(rows[0])}
    proy: dict = {}
    for r in rows[1:]:
        if not r or not r[0] or r[0] == "Provincia":
            continue
        nombre = r[ci["nombre"]].strip()
        if nombre and nombre not in proy:
            proy[nombre] = {"inv": _rigi_monto(r[ci["inv-comprometida"]]),
                            "enlace": r[ci["enlace"]] if ci.get("enlace") is not None else ""}
    store = {}
    if RIGI_FECHAS_PATH.exists():
        store = json.loads(RIGI_FECHAS_PATH.read_text(encoding="utf-8"))
    dirty = False
    for nombre, d in proy.items():
        if nombre not in store and d["enlace"]:
            f = _rigi_fecha_norma(d["enlace"])
            if f:
                store[nombre] = f
                dirty = True
    if dirty:
        RIGI_FECHAS_PATH.parent.mkdir(parents=True, exist_ok=True)
        RIGI_FECHAS_PATH.write_text(
            json.dumps(store, ensure_ascii=False, indent=2, sort_keys=True), encoding="utf-8")
    return sorted((store[n], n, d["inv"]) for n, d in proy.items() if n in store)


def _rigi_nota_denominador(avance: float, inv_aprobada: float,
                          pct_ant: float | None = None,
                          usd_ant: float | None = None) -> str | None:
    """Aviso cuando el porcentaje baja por el denominador y no por retroceso.

    El indicador es inversión aprobada sobre el pipeline total (aprobada + en
    evaluación). Cada proyecto grande que entra "en evaluación" agranda el
    denominador, así que el porcentaje puede caer aunque el capital aprobado
    haya crecido. Es un artefacto aritmético, no un retroceso de gestión, y la
    revisión externa pidió explicarlo cada vez que ocurre en vez de dejarlo a
    cargo del lector (ADR-0102).

    Devuelve None si el caso no se da: la nota aparece sólo cuando hace falta.
    Los valores previos se pueden pasar explícitamente (para poder probarla sin
    depender del estado del caché); si no, salen del snapshot anterior.
    """
    if pct_ant is None or usd_ant is None:
        anterior = (load_cache().get("indicadores") or {}).get("rigi_inversiones") or {}
        pct_ant = anterior.get("valor")
        usd_ant = anterior.get("inversion_aprobada_musd")
    if not isinstance(pct_ant, (int, float)) or not isinstance(usd_ant, (int, float)):
        return None
    if avance >= pct_ant or inv_aprobada <= usd_ant:
        return None
    miles = lambda x: f"{x:,.0f}".replace(",", ".")
    coma = lambda x: str(round(x, 1)).replace(".", ",")
    return (f"El porcentaje bajó de {coma(pct_ant)}% a {coma(avance)}% y aun así la "
            f"inversión aprobada CRECIÓ, de US$ {miles(usd_ant)}M a US$ {miles(inv_aprobada)}M. "
            f"No es un retroceso: el indicador mide la porción aprobada del total, y ese total "
            f"se agranda cada vez que se presenta un proyecto nuevo. Entraron US$ "
            f"{miles(inv_aprobada - usd_ant)}M de inversión aprobada y, al mismo tiempo, más "
            f"proyectos a la cola de evaluación.")


def fetch_rigi_inversiones() -> dict | None:
    """Avance del RIGI desde la PLATAFORMA OFICIAL del Ministerio de Economía
    (jun-2026, ADR-0011): puntúa la **inversión aprobada sobre el total del
    pipeline** (aprobada + en evaluación); la ficha expone el conteo de
    proyectos y los montos. No existe fuente estructurada de inversión
    EJECUTADA/desembolsada (verificado 2026-07): el pipeline aprobado es lo
    más cercano al hecho. Fallback: proxy histórico InfoLeg ('VPU')."""
    try:
        n_ap, inv_ap = _rigi_aprobados()
        n_ev, inv_ev = _rigi_evaluacion()
        inv_total = inv_ap + inv_ev
        if inv_total <= 0:
            raise ValueError("inversión total del pipeline = 0")
        avance = round(100.0 * inv_ap / inv_total, 1)
        miles = lambda x: f"{x:,.0f}".replace(",", ".")
        return {
            "valor":          avance,
            "unidad":         "% de inversión aprobada sobre el pipeline",
            "fuente":         "Ministerio de Economía — Plataforma oficial RIGI",
            "fecha_dato":     date.today().isoformat(),
            "desactualizado": False,
            "proyectos_aprobados":        n_ap,
            "inversion_aprobada_musd":    round(inv_ap),
            "proyectos_evaluacion":       n_ev,
            "inversion_evaluacion_musd":  round(inv_ev),
            "nota_denominador": _rigi_nota_denominador(avance, inv_ap),
            "detalle_txt": (f"{n_ap} proyectos aprobados (US$ {miles(inv_ap)}M) / "
                            f"{n_ev} en evaluación (US$ {miles(inv_ev)}M) → "
                            f"{str(avance).replace('.', ',')}% de la inversión total ya aprobada"),
        }
    except Exception as e:
        _warn("rigi_inversiones (plataforma oficial, cae a InfoLeg VPU)", e)
    try:  # FALLBACK histórico: conteo de Resoluciones RIGI vía InfoLeg ("VPU")
        today = date.today()
        count = _infoleg_post(
            texto="VPU", tipo_norma="3",
            fecha_desde=("01", "07", "2024"),
            fecha_hasta=(today.strftime("%d"), today.strftime("%m"), today.strftime("%Y")),
        )
        return {
            "valor":          round(min(100.0, float(count)), 1),
            "unidad":         "% (proxy: conteo de Resoluciones VPU)",
            "fuente":         INFOLEG_HOME,
            "fecha_dato":     today.isoformat(),
            "desactualizado": True,
        }
    except Exception as e:
        _warn("rigi_inversiones", e)
        return None


# ── D5 (contexto): alertas de manifestación — API Transporte GCBA ─────────────
# El endpoint histórico de cortes del GCBA (/transito/v1/eventos) está MUERTO
# (verificado 2026-07-02 con credenciales válidas: HTTP 500, comentado en el
# RAML oficial). Lo único vivo son los feeds GTFS-RT de alertas de servicio
# (colectivos y subtes): tiempo real puro, sin histórico → misma estrategia que
# los patentamientos de macro: ACUMULAR. Cada corrida upserta las alertas de
# manifestación activas en data/gestion/piquetes_alertas.json (dedupe por id y
# día); .github/workflows/piquetes-poll.yml muestrea además 2×/día en la franja
# típica de piquetes (12:00/18:00 ART). Ver ADR-0014.

BA_ALERTS_URL   = "https://apitransporte.buenosaires.gob.ar/{feed}/serviceAlerts"
BA_ALERTS_FEEDS = ("colectivos", "subtes")
PIQUETES_STORE_PATH = PROJECT_DIR / "data" / "gestion" / "piquetes_alertas.json"
ENV_PATH = PROJECT_DIR / ".env"

# GTFS-RT: cause=5 es DEMONSTRATION. El texto complementa (algunas alertas
# vienen con cause genérico 1/2 pero describen la protesta en el header).
GTFS_CAUSE_DEMONSTRATION = 5
PIQUETE_KEYWORDS = ("manifestaci", "piquete", "marcha", "protesta", "movilizaci")


def _ba_credenciales() -> tuple | None:
    """(client_id, client_secret) de la API Transporte GCBA: variables de
    entorno BA_TRANSPORTE_CLIENT_ID/SECRET (CI) o .env local gitignored."""
    cid = os.environ.get("BA_TRANSPORTE_CLIENT_ID")
    sec = os.environ.get("BA_TRANSPORTE_CLIENT_SECRET")
    if not (cid and sec) and ENV_PATH.exists():
        pares = {}
        for linea in ENV_PATH.read_text(encoding="utf-8").splitlines():
            if "=" in linea and not linea.strip().startswith("#"):
                k, _, v = linea.partition("=")
                pares[k.strip()] = v.strip().strip('"').strip("'")
        cid = cid or pares.get("BA_TRANSPORTE_CLIENT_ID")
        sec = sec or pares.get("BA_TRANSPORTE_CLIENT_SECRET")
    return (cid, sec) if cid and sec else None


def _alerta_es_manifestacion(alert: dict) -> bool:
    if alert.get("cause") == GTFS_CAUSE_DEMONSTRATION:
        return True
    textos = []
    for campo in ("header_text", "description_text"):
        for t in (alert.get(campo) or {}).get("translation", []):
            textos.append(str(t.get("text", "")))
    blob = " ".join(textos).lower()
    return any(kw in blob for kw in PIQUETE_KEYWORDS)


def actualizar_alertas_manifestacion() -> dict | None:
    """Consulta los feeds GTFS-RT de alertas (colectivos + subtes) y upserta las
    alertas de manifestación ACTIVAS en el store, keyed por día (dedupe por id:
    la misma protesta vista en dos muestreos del día cuenta una vez). Un día
    presente con dict vacío = "se muestreó y no había" (cobertura verificable).
    Devuelve {fecha, nuevas, total_dia} o None sin credenciales/red."""
    creds = _ba_credenciales()
    if not creds:
        print("[WARN] gestion.alertas_manifestacion: sin credenciales BA_TRANSPORTE_*. Salteado.")
        return None
    cid, sec = creds
    encontradas = {}
    ok_algun_feed = False
    for feed in BA_ALERTS_FEEDS:
        try:
            r = requests.get(BA_ALERTS_URL.format(feed=feed),
                             params={"json": 1, "client_id": cid, "client_secret": sec},
                             headers=HTTP_HEADERS, timeout=HTTP_TIMEOUT)
            r.raise_for_status()
            ok_algun_feed = True
            for ent in r.json().get("entity", []):
                alert = ent.get("alert", {})
                if _alerta_es_manifestacion(alert):
                    tr = (alert.get("header_text") or {}).get("translation", [])
                    texto = str(tr[0].get("text", ""))[:120] if tr else ""
                    encontradas[f"{feed}:{ent.get('id', '')}"] = texto
        except Exception as e:
            _warn(f"alertas_manifestacion ({feed})", e)
    if not ok_algun_feed:
        return None

    store = {"_meta": {"descripcion": ("Alertas GTFS-RT de manifestación (API Transporte GCBA, "
                                       "colectivos+subtes) por día. Acumulación desde jul-2026; "
                                       "dedupe por id de alerta. Día con {} = muestreado sin alertas."),
                       "acumula_desde": date.today().isoformat()},
             "dias": {}}
    if PIQUETES_STORE_PATH.exists():
        store = json.loads(PIQUETES_STORE_PATH.read_text(encoding="utf-8"))
    hoy = date.today().isoformat()
    dia = store.setdefault("dias", {}).setdefault(hoy, {})
    nuevas = 0
    for k, texto in encontradas.items():
        if k not in dia:
            dia[k] = texto
            nuevas += 1
    PIQUETES_STORE_PATH.parent.mkdir(parents=True, exist_ok=True)
    PIQUETES_STORE_PATH.write_text(
        json.dumps(store, ensure_ascii=False, indent=2, sort_keys=True), encoding="utf-8")
    return {"fecha": hoy, "nuevas": nuevas, "total_dia": len(dia)}


def fetch_alertas_manifestacion() -> dict | None:
    """
    CONTEXTO (no integra el ITCG): alertas de manifestación únicas captadas en
    el mes corriente en los feeds GTFS-RT del transporte porteño. Serie en
    construcción (acumula desde jul-2026, sin baseline 2023): cuando tenga
    historia suficiente es el candidato a automatizar protocolo_antipiquetes.
    """
    try:
        actualizar_alertas_manifestacion()
        if not PIQUETES_STORE_PATH.exists():
            return None
        store = json.loads(PIQUETES_STORE_PATH.read_text(encoding="utf-8"))
        dias = store.get("dias", {})
        if not dias:
            return None
        ym = date.today().strftime("%Y-%m")
        dias_mes = {d: a for d, a in dias.items() if d.startswith(ym)}
        total_mes = sum(len(a) for a in dias_mes.values())
        return {
            "valor":          total_mes,
            "unidad":         "alertas de manifestación (mes corriente, GTFS-RT)",
            "fuente":         "API Transporte GCBA — serviceAlerts (colectivos + subtes)",
            "fecha_dato":     max(dias),
            "desactualizado": False,
            "dias_muestreados_mes": len(dias_mes),
            "detalle_txt": (f"{total_mes} alertas en {len(dias_mes)} días muestreados de {ym} · "
                            f"acumula desde {min(dias)} (sin baseline 2023)"),
        }
    except Exception as e:
        _warn("alertas_manifestacion", e)
        return None


# ── D5: TDPS — asistencia directa (API Presupuesto Abierto, ADR-0015) ─────────
# TDPS = 100 × (pago directo a personas / total de transferencias del programa).
# "Directo" = partida 5.1.4 "Ayudas sociales a personas y asignaciones
# familiares" (clasificador por objeto del gasto); todo el resto del inciso 5
# (instituciones sin fines de lucro, cooperativas, municipios) es plata que
# llega al beneficiario a través de un tercero — las "Unidades de Gestión" que
# el Dto. 198/2024 eliminó. Programas sucesores: actividades "Volver al
# Trabajo" y "Acompañamiento Social"; baseline: Potenciar Trabajo 2023
# (jurisdicción 85, programa 38), cacheado porque el ejercicio está cerrado.

PA_CREDITO_URL = "https://www.presupuestoabierto.gob.ar/api/v1/credito"
TDPS_ACTIVIDADES = ("Volver al Trabajo", "Acompañamiento Social")
TDPS_BASELINE_PATH = PROJECT_DIR / "data" / "gestion" / "tdps_baseline_2023.json"


def _pa_token() -> str | None:
    """Token de la API Presupuesto Abierto: env PRESUPUESTO_ABIERTO_TOKEN (CI)
    o .env local gitignored."""
    token = os.environ.get("PRESUPUESTO_ABIERTO_TOKEN")
    if not token and ENV_PATH.exists():
        for linea in ENV_PATH.read_text(encoding="utf-8").splitlines():
            if linea.strip().startswith("PRESUPUESTO_ABIERTO_TOKEN="):
                token = linea.partition("=")[2].strip().strip('"').strip("'")
    return token or None


def _pa_credito(token: str, ejercicio: int, filtros: list, columns: list | None = None) -> list:
    """Consulta de crédito agrupada por partida (inciso.principal.parcial);
    `columns` permite pedir otra agrupación (ej. + mes, para la serie)."""
    body = {
        "ejercicios": [ejercicio],
        "columns": columns or ["inciso_id", "principal_id", "parcial_id",
                               "parcial_desc", "credito_devengado"],
        "filters": filtros,
    }
    r = requests.post(PA_CREDITO_URL, params={"format": "json"}, json=body,
                      headers={**HTTP_HEADERS, "Authorization": token},
                      timeout=HTTP_TIMEOUT)
    r.raise_for_status()
    return r.json()


def _tdps_de_partidas(rows: list) -> dict:
    """TDPS desde filas {inciso_id, principal_id, parcial_id, credito_devengado}:
    directo = 5.1.4 · total = inciso 5 completo. Montos en millones de ARS."""
    directo = total = 0.0
    for r in rows:
        try:
            if int(r.get("inciso_id") or 0) != 5:
                continue
            dev = float(r.get("credito_devengado") or 0)
        except (TypeError, ValueError):
            continue
        total += dev
        if int(r.get("principal_id") or 0) == 1 and int(r.get("parcial_id") or 0) == 4:
            directo += dev
    if total <= 0:
        raise ValueError("sin transferencias devengadas (inciso 5) en el período")
    return {"tdps": round(100.0 * directo / total, 1),
            "directo_musd": round(directo), "total_musd": round(total),
            "intermediado_musd": round(total - directo)}


def _tdps_baseline_2023(token: str) -> dict | None:
    """TDPS del Potenciar Trabajo 2023 (contexto de la ficha). El ejercicio está
    cerrado → se computa una vez y se cachea en data/gestion/."""
    if TDPS_BASELINE_PATH.exists():
        return json.loads(TDPS_BASELINE_PATH.read_text(encoding="utf-8"))
    try:
        rows = _pa_credito(token, 2023, [
            {"column": "jurisdiccion_id", "operator": "equal", "value": "85"},
            {"column": "programa_id", "operator": "equal", "value": "38"},
        ])
        base = _tdps_de_partidas(rows)
        base["_meta"] = ("Potenciar Trabajo 2023 (jur. 85, prog. 38), devengado del "
                         "ejercicio cerrado — API Presupuesto Abierto")
        TDPS_BASELINE_PATH.write_text(
            json.dumps(base, ensure_ascii=False, indent=2), encoding="utf-8")
        return base
    except Exception:
        return None


def fetch_asistencia_directa() -> dict | None:
    """
    TDPS — Tasa de Desintermediación de Planes Sociales (doc 260702), contra la
    EJECUCIÓN PRESUPUESTARIA REAL (API Presupuesto Abierto, base SIDIF):
    qué % del devengado de los programas sucesores del Potenciar Trabajo
    ("Volver al Trabajo" + "Acompañamiento Social") se paga directo a personas
    (partida 5.1.4) sobre el total de transferencias. Si el ejercicio corriente
    aún no tiene devengado (enero), cae al ejercicio anterior.
    """
    token = _pa_token()
    if not token:
        print("[WARN] gestion.asistencia_directa: sin PRESUPUESTO_ABIERTO_TOKEN. Usando fallback.")
        return None
    try:
        anio = date.today().year
        rows, ejercicio = [], anio
        for ej in (anio, anio - 1):
            acumulado = []
            for actividad in TDPS_ACTIVIDADES:
                acumulado += _pa_credito(token, ej, [
                    {"column": "actividad_desc", "operator": "like", "value": f"%{actividad}%"},
                ])
            if any(float(r.get("credito_devengado") or 0) > 0 for r in acumulado):
                rows, ejercicio = acumulado, ej
                break
        tdps = _tdps_de_partidas(rows)
        base = _tdps_baseline_2023(token)
        miles = lambda x: f"{x:,.0f}".replace(",", ".")
        detalle = (f"Devengado {ejercicio}: $ {miles(tdps['directo_musd'])}M directo a personas "
                   f"(5.1.4) / $ {miles(tdps['total_musd'])}M transferido"
                   + (f" · baseline Potenciar 2023: {str(base['tdps']).replace('.', ',')}%"
                      f" (con $ {miles(base['intermediado_musd'])}M vía organizaciones)" if base else ""))
        return {
            "valor":          tdps["tdps"],
            "unidad":         "TDPS: % del gasto social pagado directo (sin intermediación)",
            "fuente":         "API Presupuesto Abierto (SIDIF) — devengado por partida",
            "fecha_dato":     date.today().isoformat(),
            "desactualizado": False,
            "ejercicio":      ejercicio,
            "componentes":    {"directo_musd": tdps["directo_musd"],
                               "intermediado_musd": tdps["intermediado_musd"],
                               "tdps_2023": base["tdps"] if base else None},
            "detalle_txt":    detalle,
        }
    except Exception as e:
        _warn("asistencia_directa", e)
        return None


# ── D4: concesiones viales — CONTRAT.AR + página RFC (ADR-0016) ───────────────
# La Red Federal de Concesiones se licita por CONTRAT.AR (UOC 504, Min.
# Economía): la búsqueda avanzada del portal funciona sin login (ASP.NET
# WebForms: sesión + __VIEWSTATE y POST del formulario). El kilometraje por
# tramo/etapa sale de las tablas de la página oficial de la RFC. Los datos
# abiertos de CONTRAT.AR (CKAN) están congelados en mar-2023 — verificado.

CONTRATAR_HOME  = "https://contratar.gob.ar/"
CONTRATAR_BUSQ  = "https://contratar.gob.ar/BuscarAvanzado.aspx"
RFC_PAGE_URL    = "https://www.argentina.gob.ar/transporte/vialidad-nacional/red-federal-de-concesiones"


def _contratar_procesos_rfc() -> list:
    """[(proceso, nombre, estado)] de la búsqueda 'RED FEDERAL' en CONTRAT.AR."""
    from bs4 import BeautifulSoup
    s = requests.Session()
    s.get(CONTRATAR_HOME, headers=HTTP_HEADERS, timeout=HTTP_TIMEOUT)
    r = s.get(CONTRATAR_BUSQ, headers=HTTP_HEADERS, timeout=HTTP_TIMEOUT)
    r.raise_for_status()
    soup = BeautifulSoup(r.text, "html.parser")
    form = {}
    for campo in ("__VIEWSTATE", "__VIEWSTATEGENERATOR", "__EVENTVALIDATION"):
        tag = soup.find("input", {"name": campo})
        if tag and tag.get("value") is not None:
            form[campo] = tag["value"]
    if "__VIEWSTATE" not in form:
        raise ValueError("CONTRAT.AR: __VIEWSTATE no encontrado (¿rediseño del portal?)")
    form["ctl00$CPH1$txtNombrePliego"] = "RED FEDERAL"
    form["ctl00$CPH1$btnListarPliegoAvanzado"] = "Buscar"
    r = s.post(CONTRATAR_BUSQ, data=form, headers=HTTP_HEADERS, timeout=HTTP_TIMEOUT)
    r.raise_for_status()
    procesos = []
    for fila in BeautifulSoup(r.text, "html.parser").find_all("tr"):
        celdas = [c.get_text(" ", strip=True) for c in fila.find_all("td")]
        texto = " | ".join(celdas)
        m = re.search(r"\b(\d{3}-\d{4}-[A-Z]{3}\d{2})\b", texto)
        if m and "RED FEDERAL" in texto.upper() and len(celdas) >= 3:
            # Columnas: proceso · nombre · tipo · apertura · ESTADO · UOC
            estado = celdas[-2]
            nombre = next((c for c in celdas if "RED FEDERAL" in c.upper()), texto)
            procesos.append((m.group(1), nombre, estado))
    if not procesos:
        raise ValueError("CONTRAT.AR: la búsqueda 'RED FEDERAL' no devolvió procesos")
    return procesos


def _rfc_km_por_etapa() -> dict:
    """{etapa: km} desde las tablas de la página oficial de la RFC (una tabla
    por etapa, en orden I · II · II-B · III). Los km vienen 'es-AR' (682,28)."""
    from bs4 import BeautifulSoup
    html = _http_get_resiliente(RFC_PAGE_URL).decode("utf-8", errors="replace")
    tablas = BeautifulSoup(html, "html.parser").find_all("table")
    etiquetas = ("I", "II", "II-B", "III")

    def _km(s: str) -> float | None:
        """'682,28' (es-AR) · '681.92' (punto decimal, tabla Etapa III) · '720'."""
        s = s.strip()
        if re.match(r"^[\d\.]+,\d+$", s):
            return float(s.replace(".", "").replace(",", "."))
        if re.match(r"^\d+(\.\d+)?$", s):
            return float(s)
        return None

    km_por_etapa = {}
    for etiqueta, tabla in zip(etiquetas, tablas):
        km_total = 0.0
        for fila in tabla.find_all("tr"):
            celdas = [c.get_text(" ", strip=True) for c in fila.find_all("td")]
            if len(celdas) >= 2:
                v = _km(celdas[1])
                if v is not None:
                    km_total += v
        if km_total > 0:
            km_por_etapa[etiqueta] = round(km_total, 2)
    if len(km_por_etapa) < 3:
        raise ValueError(f"página RFC: se esperaban ≥3 etapas con km, hay {len(km_por_etapa)}")
    return km_por_etapa


def _etapa_de_proceso(nombre: str) -> str | None:
    m = re.search(r"ETAPA\s+([IVX]+(?:\s*-\s*B)?)", nombre.upper())
    return m.group(1).replace(" ", "") if m else None


def _esta_adjudicado(estado: str) -> bool:
    """CONTRAT.AR usa 'Preadjudicado' para la etapa previa a la adjudicación, y
    **'ADJUDICADO' in 'PREADJUDICADO' es True** (ADR-0087). La frontera de
    palabra distingue las dos: `\\bADJUDICADO\\b` no matchea dentro de
    PREADJUDICADO, y sigue aceptando variantes como 'Adjudicado Parcial'."""
    return bool(re.search(r"\bADJUDICADO\b", (estado or "").upper()))


def fetch_concesiones_infraestructura() -> dict | None:
    """
    Tasa de adjudicación de la Red Federal de Concesiones, en KM (doc 260702:
    'km bajo concesión adjudicada / km totales del proceso'): el estado de cada
    proceso licitatorio sale de CONTRAT.AR (sin login) y el kilometraje por
    etapa de la página oficial de la RFC. Una etapa cuenta como adjudicada
    cuando su proceso está en estado 'Adjudicado' (las etapas II-B y III
    adjudican por renglones — refinamiento pendiente, cuentan al cierre).
    """
    try:
        procesos = _contratar_procesos_rfc()
        km = _rfc_km_por_etapa()
        km_total = sum(km.values())
        adjudicadas, detalle_p = [], []
        for proceso, nombre, estado in procesos:
            etapa = _etapa_de_proceso(nombre)
            adjudicado = _esta_adjudicado(estado)
            if etapa and adjudicado and etapa in km:
                adjudicadas.append(etapa)
            detalle_p.append(f"{etapa or proceso}: {estado}")
        km_adj = sum(km[e] for e in set(adjudicadas))
        avance = round(100.0 * km_adj / km_total, 1)
        miles = lambda x: f"{x:,.0f}".replace(",", ".")
        etapa_key = {"I": "etapa_i", "II": "etapa_ii", "II-B": "etapa_ii_b", "III": "etapa_iii"}
        return {
            "valor":          avance,
            "unidad":         "% de km adjudicados / km del plan (Red Federal de Concesiones)",
            "fuente":         "CONTRAT.AR (UOC 504) + página oficial RFC",
            "fecha_dato":     date.today().isoformat(),
            "desactualizado": False,
            "km_adjudicados": round(km_adj),
            "km_totales":     round(km_total),
            "procesos":       len(procesos),
            # % adjudicado por etapa → gráfico de barras del modal
            "componentes":    {etapa_key.get(e, e): (100.0 if e in adjudicadas else 0.0)
                               for e in km},
            "detalle_txt": (f"{miles(km_adj)} de {miles(km_total)} km adjudicados · "
                            + " · ".join(sorted(set(detalle_p)))),
        }
    except Exception as e:
        _warn("concesiones_infraestructura", e)
        return None


# ── D5: libertad de opción en salud — XLSX oficiales SSS (ADR-0016) ───────────
# La SSS publica en argentina.gob.ar/sssalud/estadisticas los XLSX de
# beneficiarios por entidad (RNAS, mensual) y usuarios de prepagas (RNEMP).
# Las prepagas inscriptas como Agentes del Seguro (post DNU 70/23) llevan
# código RNAS >= 900000: sus beneficiarios son aportes DERIVADOS DIRECTO, sin
# triangulación. El viejo contador de "opciones de cambio" de sssalud.gob.ar
# sigue muerto (fingerprinting back-end) — esto lo reemplaza con más riqueza.

SSS_RNAS_URL  = ("https://www.argentina.gob.ar/sites/default/files/2020/07/"
                 "evolucion_anual_de_beneficiarios_de_agentes_del_seguro_de_salud_por_entidad_-_ano_{anio}.xlsx")
SSS_RNEMP_URL = ("https://www.argentina.gob.ar/sites/default/files/2020/07/"
                 "evolucion_anual_de_usuarios_por_entidades_de_medicina_prepaga_-_ano_{anio}.xlsx")
RNAS_PREPAGA_MIN = 900000   # códigos RNAS de prepagas inscriptas (DNU 70/23)


def _sss_xlsx(url_tpl: str) -> tuple:
    """(workbook, año) del XLSX del año corriente (fallback: año previo)."""
    import openpyxl
    anio = date.today().year
    for a in (anio, anio - 1):
        try:
            content = _http_get_resiliente(url_tpl.format(anio=a))
            return openpyxl.load_workbook(io.BytesIO(content), data_only=True), a
        except Exception:
            continue
    raise ValueError("XLSX SSS no disponible (año corriente ni previo)")


_MES_NUM = {"enero": 1, "febrero": 2, "marzo": 3, "abril": 4, "mayo": 5, "junio": 6,
            "julio": 7, "agosto": 8, "septiembre": 9, "octubre": 10, "noviembre": 11, "diciembre": 12}


def _sss_mes_columna(valor) -> str | None:
    """Mes de una columna SSS.

    Algunos RNEMP historicos no traen los 12 meses: 2024 publica
    "1er. Cuatrimestre" y "agosto y septiembre". En esos casos se usa el
    ultimo mes del rango; el consumidor arrastra el ultimo denominador conocido.
    """
    texto = str(valor or "").strip().lower()
    if not texto:
        return None
    meses = [num for mes, num in _MES_NUM.items() if mes in texto]
    if meses:
        return f"{max(meses):02d}"
    if "cuatrimestre" in texto:
        if "1" in texto or "primer" in texto:
            return "04"
        if "2" in texto or "segundo" in texto:
            return "08"
        if "3" in texto or "tercer" in texto:
            return "12"
    return None


def _sss_tabla(ws) -> tuple:
    """({col: 'MM'}, [(código, fila)]) de una hoja RNAS/RNEMP: columnas de
    meses y filas de entidades (código numérico en la col. 0). Los totales se
    computan sumando entidades — la fila 'Total de ...' FALTA en los archivos
    de algunos años (ej. 2025), así que no se depende de ella."""
    filas = list(ws.iter_rows(values_only=True))
    fila_meses = next((f for f in filas
                       if sum(1 for c in f if _sss_mes_columna(c)) >= 3), None)
    if fila_meses is None:
        raise ValueError("hoja SSS sin fila de meses")
    cols = {i: mm
            for i, m in enumerate(fila_meses)
            for mm in [_sss_mes_columna(m)]
            if mm}
    entidades = []
    for f in filas:
        try:
            entidades.append((int(str(f[0]).strip()), f))
        except (TypeError, ValueError):
            continue
    if not cols or not entidades:
        raise ValueError("hoja SSS sin meses o sin entidades")
    return cols, entidades


def _sss_suma_mes(entidades: list, col: int, solo_prepagas: bool = False) -> tuple:
    """(suma de beneficiarios/usuarios del mes, entidades con dato > 0)."""
    total, n = 0.0, 0
    for codigo, f in entidades:
        if solo_prepagas and codigo < RNAS_PREPAGA_MIN:
            continue
        try:
            v = float(f[col] or 0)
        except (TypeError, ValueError):
            continue
        if v > 0:
            total += v
            n += 1
    return total, n


def fetch_libertad_opcion_salud() -> dict | None:
    """
    Libertad de opción en salud, contra los XLSX oficiales de la SSS:
    % de los usuarios de medicina prepaga (RNEMP) cuyos aportes van DERIVADOS
    DIRECTO a la prepaga inscripta como Agente del Seguro (RNAS, códigos
    900000+, habilitados por el DNU 70/23). Antes de la reforma este canal no
    existía (todo se triangulaba por una obra social). Expone además el conteo
    de prepagas inscriptas.
    """
    try:
        wb_rnas, anio_rnas = _sss_xlsx(SSS_RNAS_URL)
        ws = wb_rnas[next(h for h in wb_rnas.sheetnames if "RNAS" in h.upper())]
        cols, entidades = _sss_tabla(ws)
        col_rnas = mm_rnas = derivados = prepagas = None
        for col, mm in sorted(cols.items()):
            total_mes, _ = _sss_suma_mes(entidades, col)
            if total_mes > 0:
                col_rnas, mm_rnas = col, mm
        if col_rnas is None:
            raise ValueError("RNAS sin ningún mes con datos")
        derivados, prepagas = _sss_suma_mes(entidades, col_rnas, solo_prepagas=True)

        wb_rnemp, anio_rnemp = _sss_xlsx(SSS_RNEMP_URL)
        ws2 = wb_rnemp[next(h for h in wb_rnemp.sheetnames if "RNEMP" in h.upper())]
        cols2, entidades2 = _sss_tabla(ws2)
        usuarios_prepagas = mm_rnemp = None
        for col, mm in sorted(cols2.items()):
            total_mes, _ = _sss_suma_mes(entidades2, col)
            if total_mes > 0:
                usuarios_prepagas, mm_rnemp = total_mes, mm
        if not usuarios_prepagas or not derivados:
            raise ValueError("RNAS o RNEMP sin datos del período")

        avance = round(100.0 * derivados / usuarios_prepagas, 1)
        miles = lambda x: f"{x:,.0f}".replace(",", ".")
        fecha = f"{anio_rnas}-{mm_rnas}-01"
        mes_rnas, mes_rnemp = f"{mm_rnas}/{anio_rnas}", f"{mm_rnemp}/{anio_rnemp}"
        return {
            "valor":          avance,
            "unidad":         "% de usuarios de prepagas con aportes derivados directo (sin triangulación)",
            "fuente":         "SSS — RNAS y RNEMP por entidad (argentina.gob.ar/sssalud/estadisticas)",
            "fecha_dato":     fecha,
            "desactualizado": False,
            "derivados":            round(derivados),
            "prepagas_inscriptas":  prepagas,
            "usuarios_prepagas":    round(usuarios_prepagas),
            "detalle_txt": (f"{miles(derivados)} beneficiarios derivados directo a {prepagas} prepagas "
                            f"inscriptas ({mes_rnas}) / {miles(usuarios_prepagas)} usuarios de "
                            f"prepagas (RNEMP {mes_rnemp}); canal inexistente pre-DNU 70/23"),
        }
    except Exception as e:
        _warn("libertad_opcion_salud", e)
        return None


# ── D5: protocolo antipiquetes — Diagnóstico Político (ADR-0025) ──────────────

DP_PIQUETES_PATH = PROJECT_DIR / "data" / "gestion" / "dp_piquetes.json"
DP_MONITOREOS_URL = "https://diagnosticopolitico.com.ar/monitoreos-politicos"


def fetch_protocolo_antipiquetes() -> dict | None:
    """
    IRPC automático (ADR-0025): reducción % de cortes en CABA vs 2023, con los
    anclajes ANUALES PÚBLICOS de Diagnóstico Político (la misma consultora que
    citaba la estimación manual; su definición de piquete coincide con la Res.
    943/23). Anclajes curados con fuente en data/gestion/dp_piquetes.json —
    mismo patrón de hitos fechados que privatizaciones/concesiones. Además
    scrapea la página pública de monitoreos como DETECTOR: si aparece un año
    nuevo, avisa para actualizar el store. Corrigió la subcalibración del
    valor manual (55% era la foto 2024; con 2025 cerrado el IRPC CABA es 74%).
    """
    try:
        dp = json.loads(DP_PIQUETES_PATH.read_text(encoding="utf-8-sig"))
        caba = {int(k): int(v) for k, v in dp["caba"].items()}
        base = caba[2023]
        ult = max(caba)
        irpc = round((1.0 - caba[ult] / base) * 100.0, 1)

        # Detector de informes nuevos (no fatal): años nacionales publicados
        # posteriores al último ancla del store.
        try:
            r = requests.get(DP_MONITOREOS_URL, headers=HTTP_HEADERS, timeout=30)
            r.raise_for_status()
            anios = {int(a) for a in re.findall(r"piquetes en (20\d{2})", r.text)}
            nuevos = sorted(a for a in anios if a > ult)
            if nuevos:
                print(f"  [AVISO] Diagnóstico Político publicó datos de {nuevos}: "
                      f"actualizar data/gestion/dp_piquetes.json")
        except Exception:
            pass

        return {
            "valor":          irpc,
            "unidad":         "% de reducción de cortes por manifestación en CABA vs 2023 (IRPC)",
            "fuente":         "Diagnóstico Político — monitoreos públicos de piquetes",
            "fecha_dato":     f"{ult}-12-31",
            "desactualizado": False,
            "componentes":    {"caba_2023": base, f"caba_{ult}": caba[ult]},
            "detalle_txt":    (f"{caba[ult]} cortes en CABA en {ult} vs {base} en 2023 "
                               f"(Diagnóstico Político; definición de piquete = Res. 943/23)"),
        }
    except Exception as e:
        _warn("protocolo_antipiquetes", e)
        return None


# ── D5 (contexto): protestas en CABA — ACLED (ADR-0017) ───────────────────────
# ACLED (acleddata.com) publica para miembros un agregado SEMANAL × provincia ×
# tipo de evento (archivo XLSX regional, actualizado semanalmente). El nivel
# "Open" de la cuenta no habilita la API de eventos (403) pero SÍ este archivo,
# vía sesión web (login Drupal). CABA cubierta desde 2018 → línea base 2023
# real. OJO: ACLED cuenta EVENTOS DE PROTESTA (marchas, concentraciones), no
# cortes de calle — complementa a protocolo_antipiquetes (que mide cortes, DP)
# y expone el contraste: la protesta no bajó, se reconvirtió a marcha sin corte.
# Atribución obligatoria: datos de ACLED (acleddata.com).

ACLED_LOGIN_URL = "https://acleddata.com/user/login"
ACLED_AGG_PAGE  = "https://acleddata.com/aggregated/aggregated-data-latin-america-caribbean"
PROTESTAS_STORE_PATH = PROJECT_DIR / "data" / "gestion" / "protestas_caba.json"


def _acled_credenciales() -> tuple | None:
    """(usuario, contraseña) de ACLED: env ACLED_USERNAME/ACLED_PASSWORD (CI)
    o .env local gitignored."""
    user = os.environ.get("ACLED_USERNAME")
    pwd  = os.environ.get("ACLED_PASSWORD")
    if not (user and pwd) and ENV_PATH.exists():
        pares = {}
        for linea in ENV_PATH.read_text(encoding="utf-8").splitlines():
            if "=" in linea and not linea.strip().startswith("#"):
                k, _, v = linea.partition("=")
                pares[k.strip()] = v.strip().strip('"').strip("'")
        user = user or pares.get("ACLED_USERNAME")
        pwd  = pwd or pares.get("ACLED_PASSWORD")
    return (user, pwd) if user and pwd else None


# ── Cortes en CABA vía API de EVENTOS de ACLED (cuenta académica UBA) ─────────
# La cuenta 69pi...@campus.economicas.uba.ar autentica OK pero nació con nivel
# Open (la API de eventos da 403). El perfil quedó Academic/UBA, que es lo que
# ACLED reevalúa de forma asíncrona. Este colector es AUTO-DESTRABANTE: corre
# cada noche en el pipeline, y el día que el nivel pase a Research construye
# solo la serie de cortes sin tocar código. Nivel Research = eventos con ~12
# meses de rezago: suficiente para el baseline 2023 y la validación de los
# anclajes DP (ADR-0025); el numerador vivo del IRPC sigue siendo DP.
ACLED_OAUTH_URL = "https://acleddata.com/oauth/token"
ACLED_READ_URL  = "https://acleddata.com/api/acled/read"
ACLED_CORTES_STORE = PROJECT_DIR / "data" / "gestion" / "acled_cortes.json"
# memo por proceso del store de protestas (ver actualizar_protestas_caba)
_PROTESTAS_STORE_MEMO: dict | None = None
# vocabulario de bloqueo en las notas (ACLED codifica en inglés)
ACLED_CORTES_KEYWORDS = ("block", "picket", "barricad", "roadblock", "corte", "piquete")


def _acled_uba_credenciales() -> tuple | None:
    """(usuario, contraseña) de la cuenta ACLED académica: env o .env
    ACLED_UBA_USERNAME/ACLED_UBA_PASSWORD."""
    user = os.environ.get("ACLED_UBA_USERNAME")
    pwd  = os.environ.get("ACLED_UBA_PASSWORD")
    if not (user and pwd) and ENV_PATH.exists():
        pares = {}
        for linea in ENV_PATH.read_text(encoding="utf-8").splitlines():
            if "=" in linea and not linea.strip().startswith("#"):
                k, _, v = linea.partition("=")
                pares[k.strip()] = v.strip().strip('"').strip("'")
        user = user or pares.get("ACLED_UBA_USERNAME")
        pwd  = pwd or pares.get("ACLED_UBA_PASSWORD")
    return (user, pwd) if user and pwd else None


def actualizar_acled_cortes() -> None:
    """Intenta la API de eventos con la cuenta UBA; con nivel Open (403) se
    saltea informando, con Research pagina los eventos de protesta de CABA
    (2018→hoy), cuenta por mes los que mencionan bloqueo en las notas y guarda
    data/gestion/acled_cortes.json con la comparación anual contra los
    anclajes DP. No puntúa: es el store de validación del IRPC."""
    creds = _acled_uba_credenciales()
    if not creds:
        print("[INFO] gestion.acled_cortes: sin credenciales ACLED_UBA_*. Salteado.")
        return
    user, pwd = creds
    try:
        r = requests.post(ACLED_OAUTH_URL,
                          data={"username": user, "password": pwd,
                                "grant_type": "password", "client_id": "acled",
                                "scope": "authenticated"},
                          headers={"User-Agent": "Mozilla/5.0"}, timeout=60)
        if r.status_code != 200:
            print(f"[WARN] gestion.acled_cortes: OAuth {r.status_code}. Salteado.")
            return
        H = {"Authorization": f"Bearer {r.json()['access_token']}",
             "User-Agent": "Mozilla/5.0"}
        eventos, pagina = [], 1
        while pagina <= 40:
            r2 = requests.get(ACLED_READ_URL,
                              params={"country": "Argentina",
                                      "admin1": "Ciudad Autonoma de Buenos Aires",
                                      "event_date": "2018-01-01|2030-01-01",
                                      "event_date_where": "BETWEEN",
                                      "limit": 1000, "page": pagina},
                              headers=H, timeout=120)
            if r2.status_code == 403:
                print("[INFO] gestion.acled_cortes: la cuenta UBA sigue en nivel Open "
                      "(403 en eventos); se reintenta en la próxima corrida.")
                return
            r2.raise_for_status()
            lote = r2.json().get("data", [])
            if not lote:
                break
            eventos.extend(lote)
            pagina += 1
        if not eventos:
            print("[WARN] gestion.acled_cortes: API de eventos accesible pero vacía.")
            return
        mensual: dict = {}
        for e in eventos:
            ym = str(e.get("event_date", ""))[:7]
            if not ym:
                continue
            fila = mensual.setdefault(ym, {"eventos": 0, "cortes": 0})
            fila["eventos"] += 1
            notas = str(e.get("notes", "")).lower()
            if any(k in notas for k in ACLED_CORTES_KEYWORDS):
                fila["cortes"] += 1
        anual = {}
        for ym, fila in mensual.items():
            a = anual.setdefault(ym[:4], {"eventos": 0, "cortes": 0})
            a["eventos"] += fila["eventos"]
            a["cortes"] += fila["cortes"]
        dp = {}
        if DP_PIQUETES_PATH.exists():
            dp = json.loads(DP_PIQUETES_PATH.read_text(encoding="utf-8-sig")).get("caba", {})
        store = {
            "_meta": {
                "fuente": "ACLED (acleddata.com) — API de eventos, cuenta académica UBA",
                "descargado": datetime.now().strftime("%Y-%m-%d"),
                "criterio": f"cortes = notas con {ACLED_CORTES_KEYWORDS}",
                "nota": ("Serie de VALIDACIÓN de los anclajes DP (no puntúa). "
                         "Nivel Research = eventos con ~12 meses de rezago."),
                "dp_caba": dp,
            },
            "anual": anual,
            "mensual": dict(sorted(mensual.items())),
        }
        ACLED_CORTES_STORE.write_text(
            json.dumps(store, indent=2, ensure_ascii=False), encoding="utf-8")
        comparacion = " · ".join(
            f"{a}: ACLED {anual[a]['cortes']} cortes vs DP {dp.get(a, '—')}"
            for a in sorted(anual) if a in {str(k) for k in dp} or a >= "2023")
        print(f"[OK] gestion.acled_cortes: ¡NIVEL RESEARCH ACTIVO! {len(eventos)} eventos "
              f"→ {ACLED_CORTES_STORE.name} | {comparacion}")
    except Exception as e:
        print(f"[WARN] gestion.acled_cortes: {e}. Salteado.")


def actualizar_protestas_caba() -> dict | None:
    """Baja el agregado semanal LatAm de ACLED (sesión web), filtra Argentina ×
    Protests/Riots, agrega por mes — CABA ("mensual") y país entero
    ("mensual_nacional", ADR-0052: insumo de conflictividad_nacional del ITCP)
    — y guarda ambas series en el store versionado (así descargar_series no
    re-baja los ~8 MB). Devuelve el store completo.

    Memo por proceso: politica.py consume este store dos veces por corrida
    (conflictividad_nacional y protestas_caba de contexto) y gestion.py una —
    sin el memo, cada llamada re-bajaría el XLSX de ~8 MB."""
    global _PROTESTAS_STORE_MEMO
    if _PROTESTAS_STORE_MEMO is not None:
        return _PROTESTAS_STORE_MEMO
    creds = _acled_credenciales()
    if not creds:
        print("[WARN] gestion.protestas_caba: sin credenciales ACLED_*. Salteado.")
        return None
    import openpyxl
    user, pwd = creds
    try:
        s = requests.Session()
        s.headers.update({"User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64)"})
        r = s.get(ACLED_LOGIN_URL, timeout=HTTP_TIMEOUT)
        m = re.search(r'name="form_build_id" value="([^"]+)"', r.text)
        if not m:
            raise ValueError("login ACLED: form_build_id no encontrado")
        s.post(ACLED_LOGIN_URL, data={"name": user, "pass": pwd, "form_build_id": m.group(1),
                                      "form_id": "user_login_form", "op": "Log in"},
               timeout=HTTP_TIMEOUT)
        page = s.get(ACLED_AGG_PAGE, timeout=HTTP_TIMEOUT).text
        m = re.search(r'href="(https://acleddata\.com/system/files/[^"]+aggregated_data[^"]+\.xlsx)"', page)
        if not m:
            raise ValueError("página de agregados ACLED sin link .xlsx (¿sesión no logueada?)")
        url_xlsx = m.group(1)
        content = s.get(url_xlsx, timeout=300).content

        wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True)
        ws = wb[wb.sheetnames[0]]
        ws.reset_dimensions()   # el export trae dimensiones rotas (dice 1×1)
        filas = ws.iter_rows(values_only=True)
        next(filas)             # WEEK|REGION|COUNTRY|ADMIN1|EVENT_TYPE|SUB_EVENT_TYPE|EVENTS|...
        mensual: dict = {}
        nacional: dict = {}
        hasta = None
        for f in filas:
            if not f or len(f) < 7 or str(f[2]) != "Argentina":
                continue
            if str(f[4]) not in ("Protests", "Riots"):
                continue
            ym = f[0].strftime("%Y-%m")
            n = int(f[6] or 0)
            nacional[ym] = nacional.get(ym, 0) + n
            if "Ciudad" in str(f[3]):
                mensual[ym] = mensual.get(ym, 0) + n
            hasta = max(hasta, f[0].date()) if hasta else f[0].date()
        if not mensual:
            raise ValueError("agregado ACLED sin filas de CABA")
        store = {"_meta": {"descripcion": ("Eventos de protesta (Protests+Riots) por mes, del "
                                           "agregado semanal de ACLED (acleddata.com): 'mensual' "
                                           "= CABA (card de contexto de gestión), "
                                           "'mensual_nacional' = Argentina entera (insumo de "
                                           "conflictividad_nacional del ITCP, ADR-0052). "
                                           "Atribución: datos de ACLED."),
                           "actualizado": date.today().isoformat(),
                           "hasta_semana": hasta.isoformat(),
                           "fuente_xlsx": url_xlsx},
                 "mensual": {ym: mensual[ym] for ym in sorted(mensual)},
                 "mensual_nacional": {ym: nacional[ym] for ym in sorted(nacional)}}
        PROTESTAS_STORE_PATH.write_text(
            json.dumps(store, ensure_ascii=False, indent=2), encoding="utf-8")
        _PROTESTAS_STORE_MEMO = store
        return store
    except Exception as e:
        _warn("protestas_caba (descarga ACLED)", e)
        return None


def fetch_protestas_caba() -> dict | None:
    """
    CONTEXTO (no integra el ITCG): eventos de protesta en CABA según ACLED,
    acumulado 12 meses (sin el mes parcial del archivo), contra el total 2023.
    Mide protesta callejera (marchas, concentraciones), NO cortes: leído junto
    a protocolo_antipiquetes expone si la protesta cayó o solo se reconvirtió.
    """
    try:
        store = actualizar_protestas_caba()
        if store is None and PROTESTAS_STORE_PATH.exists():
            store = json.loads(PROTESTAS_STORE_PATH.read_text(encoding="utf-8"))
        if not store:
            return None
        mensual = store.get("mensual", {})
        hasta = store.get("_meta", {}).get("hasta_semana", "")
        yms = sorted(mensual)
        # último mes completo: si la última semana del archivo no llega a fin de mes,
        # el mes final está parcial y se excluye de la ventana de 12.
        if hasta and yms and hasta[:7] == yms[-1]:
            import calendar as _cal
            a, m = int(hasta[:4]), int(hasta[5:7])
            if int(hasta[8:10]) < _cal.monthrange(a, m)[1]:
                yms = yms[:-1]
        if len(yms) < 12:
            raise ValueError("serie ACLED demasiado corta")
        ult12 = sum(mensual[ym] for ym in yms[-12:])
        base_2023 = sum(v for ym, v in mensual.items() if ym.startswith("2023"))
        var = round((ult12 / base_2023 - 1.0) * 100.0, 1) if base_2023 else None
        return {
            "valor":          ult12,
            "unidad":         "eventos de protesta en 12 meses (CABA, ACLED)",
            "fuente":         "ACLED — agregado semanal por provincia (acleddata.com)",
            "fecha_dato":     f"{yms[-1]}-01",
            "desactualizado": False,
            "eventos_2023":   base_2023,
            "var_vs_2023":    var,
            "detalle_txt": (f"{ult12} eventos en 12m (hasta {yms[-1]}) vs {base_2023} en 2023"
                            + (f" ({var:+.1f}%)".replace(".", ",") if var is not None else "")
                            + " — cuenta marchas y concentraciones, no cortes: contrastar con el protocolo"),
        }
    except Exception as e:
        _warn("protestas_caba", e)
        return None


# ── Scoring (ITCG — Paramétrica CIGOB jul-2026) ───────────────────────────────

def calcular_itcg_cinturon(indicadores: dict) -> dict | None:
    """ITCG 0-100 (ver scripts/itcg.py) sobre los indicadores del índice.
    Overrides del analista vigentes para el mes corriente en
    data/gestion/ajustes_itcg.json (con justificación y vencimiento)."""
    periodo = datetime.now().strftime("%Y-%m")
    ajustes = itcg.cargar_ajustes(AJUSTES_PATH, periodo)
    valores = {}
    for nombre in itcg.BANDAS_ITCG:
        v = indicadores.get(nombre, {}).get("valor")
        valores[nombre] = float(v) if isinstance(v, (int, float)) and not isinstance(v, bool) else None
    return itcg.calcular_itcg(valores, ajustes)


def anotar_indicadores(indicadores: dict, resultado: dict | None) -> None:
    """Marca cada indicador con su rol en el ITCG: los del índice llevan
    puntaje, dimensión y peso efectivo; el resto queda como contexto."""
    por_indicador = {}
    if resultado:
        for dkey, dim in resultado["dimensiones"].items():
            for ikey, info in dim["indicadores"].items():
                por_indicador[ikey] = {
                    "en_indice": True,
                    "dimension": dkey,
                    "puntaje_itcg": info["puntaje_aplicado"],
                    "puntaje_banda": info["puntaje_banda"],
                    "peso_efectivo": info["peso_efectivo"],
                }
    for nombre, ind in indicadores.items():
        if nombre in por_indicador:
            ind.update(por_indicador[nombre])
        else:
            ind["en_indice"] = nombre in itcg.BANDAS_ITCG  # del índice pero sin dato
            if nombre in itcg.INDICADORES_CONTEXTO:
                ind["en_indice"] = False
        # Las promesas cumplidas se muestran pero no puntúan (ADR-0100). Se
        # marcan SIEMPRE, hayan entrado o no por por_indicador.
        if nombre in itcg.INDICADORES_CUMPLIDOS:
            meta_cumplido = dict(itcg.INDICADORES_CUMPLIDOS[nombre])
            ind["en_indice"] = False
            # conserva su dimensión para poder mostrarse junto a sus pares
            ind["dimension"] = meta_cumplido.get("dimension")
            ind["cumplido"] = meta_cumplido


def calcular_score(indicadores: dict) -> float:
    """Tensión 0-10 del cinturón, derivada del ITCG: (100 − ITCG) / 10.
    Sin ningún indicador del índice disponible, devuelve 5.0 (neutro)."""
    resultado = calcular_itcg_cinturon(indicadores)
    return itcg.tension_de_itcg(resultado["valor"]) if resultado else 5.0


# ── Main ──────────────────────────────────────────────────────────────────────

def main() -> None:
    cache_anterior = load_cache()
    manuales       = load_manuales()

    # sonda auto-destrabante: cortes CABA vía eventos ACLED (cuenta UBA).
    # Mientras el nivel sea Open se saltea con un INFO; nunca corta la corrida.
    actualizar_acled_cortes()

    indicadores: dict = {}
    frescos_auto = 0

    # cepo primero: su brecha alimenta el componente cambiario del ILCE.
    cepo = fetch_cepo_mulc()
    brecha = float(cepo["valor"]) if cepo and isinstance(cepo.get("valor"), (int, float)) else None

    auto_fetchers = {
        "cepo_mulc":                    lambda: cepo,
        "apertura_comercial":           lambda: fetch_apertura_comercial(brecha),
        "desregulacion_normativa":      fetch_desregulacion_normativa,
        "reduccion_estado":             fetch_reduccion_estado,
        "gasto_funcionamiento":         fetch_gasto_funcionamiento,
        "masa_salarial":                fetch_masa_salarial,
        "reestructuracion_organismos":  fetch_reestructuracion_organismos,
        "fal_modernizacion_laboral":    fetch_fal_modernizacion_laboral,
        "litigiosidad_laboral":         fetch_litigiosidad_laboral,
        "privatizaciones":              fetch_privatizaciones,
        "rigi_inversiones":             fetch_rigi_inversiones,
        "concesiones_infraestructura":  fetch_concesiones_infraestructura,
        "asistencia_directa":           fetch_asistencia_directa,
        "protocolo_antipiquetes":       fetch_protocolo_antipiquetes,
        "libertad_opcion_salud":        fetch_libertad_opcion_salud,
        "alertas_manifestacion":        fetch_alertas_manifestacion,
        "protestas_caba":               fetch_protestas_caba,
    }

    for nombre in INDICADORES_ESPERADOS:
        resultado = None
        if nombre in auto_fetchers:
            resultado = auto_fetchers[nombre]()

        if resultado is not None and resultado.get("valor") is not None:
            indicadores[nombre] = resultado
            if not resultado.get("desactualizado"):
                frescos_auto += 1
        else:
            fallback = _manual_entry(nombre, manuales, cache_anterior)
            indicadores[nombre] = fallback or {
                "valor":          None,
                "unidad":         "",
                "fuente":         "pendiente",
                "fecha_dato":     "",
                "desactualizado": True,
            }

    resultado_itcg = calcular_itcg_cinturon(indicadores)
    anotar_indicadores(indicadores, resultado_itcg)
    score = itcg.tension_de_itcg(resultado_itcg["valor"]) if resultado_itcg else 5.0

    payload = {
        "cinturon":     CINTURON,
        "generated_at": datetime.now().isoformat(),
        "score":        score,
        "itcg":         resultado_itcg,
        "indicadores":  indicadores,
    }
    save_cache(payload)

    total      = len(INDICADORES_ESPERADOS)
    total_auto = len(auto_fetchers)
    con_datos  = sum(1 for v in indicadores.values() if v and v.get("valor") is not None)
    itcg_txt   = resultado_itcg["valor"] if resultado_itcg else "—"
    print(f"[OK] {CINTURON}: ITCG={itcg_txt} score={score} "
          f"auto_frescos={frescos_auto}/{total_auto} con_datos={con_datos}/{total}")

    if frescos_auto == total_auto:
        sys.exit(0)
    elif frescos_auto > 0:
        sys.exit(1)
    else:
        sys.exit(2)


if __name__ == "__main__":
    main()
